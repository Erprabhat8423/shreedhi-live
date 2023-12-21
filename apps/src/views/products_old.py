import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
import time


from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage

# Create your views here.

# products View
@login_required
def index(request):
    context = {}
    page = request.GET.get('page')
    products = SpProducts.objects.all().order_by('-id')
    
    last_product = SpProducts.objects.order_by('-id').first()
    if last_product:
        product_variants = SpProductVariants.objects.filter(product_id=last_product.id).order_by('-id')
        page = request.GET.get('page')
        variant_paginator = Paginator(product_variants, getConfigurationResult('page_limit'))
        try:
            product_variants = variant_paginator.page(page)
        except PageNotAnInteger:
            product_variants = variant_paginator.page(1)
        except EmptyPage:
            product_variants = variant_paginator.page(variant_paginator.num_pages)  
        if page is not None:
            page = page
            total_variant_pages = 0
        else:
            page = 1
            total_variant_pages = int(variant_paginator.count/getConfigurationResult('page_limit')) 
        
        if(variant_paginator.count == 0):
            variant_paginator.count = 1

        temp = total_variant_pages%variant_paginator.count
        if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
            total_variant_pages = total_variant_pages+1
        else:
            total_variant_pages = total_variant_pages

        context['total_variant_pages']            = total_variant_pages
    else:
        product_variants = {}
        context['total_variant_pages']            = 0

    
    context['products']          = products
    context['page_limit']             = getConfigurationResult('page_limit')
    context['product_variants']   = product_variants
    context['page_title'] = "Product & Variant Management"
    context['product_classes']          = SpProductClass.objects.filter(status=1)
    context['product_containers']          = SpContainers.objects.all()

    template = 'products/products.html'
    return render(request, template, context)


@login_required
def getProductVariants(request,product_id):
    if SpProducts.objects.filter(id=product_id).exists() :
        product_variants = SpProductVariants.objects.filter(product_id=product_id).order_by('-id')
        page = request.GET.get('page')
        variant_paginator = Paginator(product_variants, getConfigurationResult('page_limit'))
        try:
            product_variants = variant_paginator.page(page)
        except PageNotAnInteger:
            product_variants = variant_paginator.page(1)
        except EmptyPage:
            product_variants = variant_paginator.page(variant_paginator.num_pages)  
        if page is not None:
            page = page
        else:
            page = 1
        total_variant_pages = int(variant_paginator.count/getConfigurationResult('page_limit')) 
        
        if(variant_paginator.count == 0):
            variant_paginator.count = 1

        temp = total_variant_pages%variant_paginator.count
        if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
            total_variant_pages = total_variant_pages+1
        else:
            total_variant_pages = total_variant_pages
        
        context = {}
        context['page_limit']             = getConfigurationResult('page_limit')
        context['total_variant_pages']     = total_variant_pages
        context['product_variants']   = product_variants
        template = 'products/product-variants.html'
        return render(request, template, context)
        
    else:
        return HttpResponse('Not found')


def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#Automaticly downloads to PDF file
@login_required
def exportToPDF(request, columns):
    column_list = columns.split (",")
    context = {}
    organizations = SpOrganizations.objects.all().values().order_by('-id')
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('role-permission/organization_pdf_template.html', {'organizations': organizations, 'url': baseurl, 'columns' : column_list})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'organizations.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

@login_required
def exportToXlsx(request, columns):
    column_list = columns.split (",")
    organizations = SpOrganizations.objects.all().order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=organizations.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Organizations'
    
    # Define the titles for columns
    columns = []

    if 'org_name' in column_list:
        columns += [ 'Organization Name' ]

    if 'landline_no' in column_list:
        columns += [ 'Landline No.' ]
    
    if 'mobile_no' in column_list:
        columns += [ 'Mobile No.' ] 

    if 'email_id' in column_list:
        columns += [ 'Email Id' ]

        columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for organization in organizations:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'org_name' in column_list:
            row += [ organization.organization_name ]

        if 'landline_no' in column_list:
            row += [ organization.landline_country_code + ' ' + organization.landline_state_code + ' ' + organization.landline_number ]
        
        if 'mobile_no' in column_list:
            row += [ organization.mobile_country_code + ' ' + organization.mobile_number ] 

        if 'email_id' in column_list:
            row += [ organization.email ]       
       
        row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response


@login_required
def ajaxProductList(request):
    page = request.GET.get('page')
    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))   
    
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpOrganizations.objects.order_by('-id').first()
    
    context = {}
    context['organizations']          = organizations
    context['total_pages']            = total_pages
    context['organization_details']   = organization_details

    template = 'role-permission/ajax-products.html'
    return render(request, template, context)


@login_required
def ajaxProductVariantLists(request,product_id):
    if SpProducts.objects.filter(id=product_id).exists() :

        product_variants = SpProductVariants.objects.filter(product_id=product_id).order_by('-id')
        page = request.GET.get('page')
        variant_paginator = Paginator(product_variants, getConfigurationResult('page_limit'))
        try:
            product_variants = variant_paginator.page(page)
        except PageNotAnInteger:
            product_variants = variant_paginator.page(1)
        except EmptyPage:
            product_variants = variant_paginator.page(variant_paginator.num_pages)  
        if page is not None:
            page = page
        else:
            page = 1
        total_variant_pages = int(variant_paginator.count/getConfigurationResult('page_limit')) 
        
        if(variant_paginator.count == 0):
            variant_paginator.count = 1

        temp = total_variant_pages%variant_paginator.count
        if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
            total_variant_pages = total_variant_pages+1
        else:
            total_variant_pages = total_variant_pages
        
        context = {}
        context['page_limit']             = getConfigurationResult('page_limit')
        context['total_variant_pages']     = total_variant_pages
        context['product_variants']   = product_variants
        template = 'products/ajax-product-variant-lists.html'
        return render(request, template, context)

    else:
        return HttpResponse('Not found')

    





@login_required
def addProduct(request):
    template = 'products/add-product.html'
    context = {}
    context['product_classes']          = SpProductClass.objects.filter(status=1)
    context['product_containers']          = SpContainers.objects.all()
    context['color_codes']          = SpColorCodes.objects.filter(status=1)
    return render(request, template,context)


@login_required
def saveProduct(request):
    if request.method == "POST":
        response = {}
        try:
            product_class_id = request.POST.get('product_class_id')
            product_name = request.POST.get('product_name')
            product_container_id = request.POST.get('product_container_id')

            
            if product_class_id == '':
                response['flag'] = False
                response['message'] = "Please select product class"
            if product_name == '':
                response['flag'] = False
                response['message'] = "Please enter product name"
            if SpProducts.objects.filter(product_name=product_name).exists():
                response['flag'] = False
                response['message'] = "Product name already exists"
            
            if product_container_id == '':
                response['flag'] = False
                response['message'] = "Please select container"   
            
            else:
                product = SpProducts()
                product.product_class_id = request.POST['product_class_id']
                product.product_class_name = getModelColumnById(SpProductClass,request.POST['product_class_id'],'product_class')
                product.product_hsn = getModelColumnById(SpProductClass,request.POST['product_class_id'],'product_hsn')
                product.product_name = request.POST['product_name']
                product.container_id  = request.POST['product_container_id']
                product.container_name  = getModelColumnById(SpContainers,request.POST['product_container_id'],'container')
                product.product_color_code = request.POST['product_color_code']
                product.description = request.POST['description']
                product.status  = 1
                product.save()

                if product.id :

                    #Save Activity
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = 'New product created'
                    activity    = 'New product created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    saveActivity('Product & Variant Management', 'Create Product', heading, activity, request.user.id, user_name, 'addProduct.png', '1', 'web.png')
                    
                    response['flag'] = True
                    response['message'] = "Record has been saved successfully."
                else:
                    response['flag'] = False
                    response['message'] = "Failed to save"

            return JsonResponse(response)
        
        except Exception as e:
            response['flag'] = False
            response['message'] = str(e)
            return JsonResponse(response)
    else:
        response = {}
        response['error'] = False
        response['message'] = "Method not allowed"
        return JsonResponse(response)


@login_required
def editProduct(request,product_id):
    if request.method == "POST":
        response = {}
        try:
            if SpProducts.objects.filter(product_name=request.POST['product_name']).exclude(id=request.POST['product_id']).exists():
                response['flag'] = False
                response['message'] = "Product name already exist"
            else:
                product = SpProducts.objects.get(id=request.POST['product_id'])
                product.product_class_id = request.POST['product_class_id']
                product.product_class_name = getModelColumnById(SpProductClass,request.POST['product_class_id'],'product_class')
                product.product_hsn = getModelColumnById(SpProductClass,request.POST['product_class_id'],'product_hsn')
                product.product_name = request.POST['product_name']
                product.container_id  = request.POST['product_container_id']
                product.container_name  = getModelColumnById(SpContainers,request.POST['product_container_id'],'container')
                product.product_color_code = request.POST['product_color_code']
                product.description = request.POST['description']
                product.status  = 1
                product.save()

                if product.id :

                    #Save Activity
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = 'Product edited'
                    activity    = 'Product edited by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    saveActivity('Product & Variant Management', 'Edit Product', heading, activity, request.user.id, user_name, 'edit_product.png', '1', 'web.png')

                    response['flag'] = True
                    response['message'] = "Record has been saved successfully."
                else:
                    response['flag'] = False
                    response['message'] = "Failed to save"
        except Exception as e:
            response['error'] = False
            response['message'] = str(e)
        return JsonResponse(response)
    else:
        context = {}
        context['product']     = SpProducts.objects.get(id=product_id)
        context['product_classes']        = SpProductClass.objects.filter(status=1)
        context['product_containers']     = SpContainers.objects.all()
        context['product_units']          = SpProductUnits.objects.filter(status=1)
        context['color_codes']          = SpColorCodes.objects.filter(status=1)
        template = 'products/edit-product.html'
        return render(request, template, context)


@login_required
def addProductVariant(request,product_id):
    if request.method == "POST":
        response = {}
        try:
            product_variant = SpProductVariants()
            product_variant.product_id = request.POST['product_id']
            product_variant.product_name  = getModelColumnById(SpProducts,request.POST['product_id'],'product_name')
            product_variant.product_class_id  = getModelColumnById(SpProducts,request.POST['product_id'],'product_class_id')
            product_variant.item_sku_code = request.POST['item_code']
            product_variant.variant_quantity = request.POST['variant_qty']
            product_variant.variant_unit_id  = request.POST['variant_unit']
            product_variant.variant_name  = request.POST['variant_name']
            product_variant.variant_unit_name  = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'unit')
            product_variant.largest_unit_name = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'largest_unit')
            product_variant.variant_size  = request.POST['variant_size']
            product_variant.no_of_pouch  = request.POST['no_of_pouch']
            product_variant.container_size  = request.POST['container_size']
            product_variant.is_bulk_pack  = request.POST['variant_type']
            product_variant.included_in_scheme  = request.POST['included_in_scheme']
            product_variant.mrp  = request.POST['mrp']
            product_variant.sp_distributor  = request.POST['sp_distributor']
            product_variant.sp_superstockist  = request.POST['sp_superstockist']
            product_variant.sp_employee  = request.POST['sp_employee']
            product_variant.container_mrp  = float(request.POST['mrp']) * float(request.POST['no_of_pouch'])
            product_variant.container_sp_distributor  = float(request.POST['sp_distributor']) * float(request.POST['no_of_pouch'])
            product_variant.container_sp_superstockist  = float(request.POST['sp_superstockist']) * float(request.POST['no_of_pouch'])
            product_variant.container_sp_employee  = float(request.POST['sp_employee']) * float(request.POST['no_of_pouch'])
            product_variant.valid_from = datetime.strptime(request.POST['valid_from'], '%d/%m/%Y').strftime('%Y-%m-%d')
            product_variant.valid_to = datetime.strptime(request.POST['valid_to'], '%d/%m/%Y').strftime('%Y-%m-%d')
            product_variant.status  = 1
            product_variant.save()
            if product_variant.id:


                operational_users = SpUsers.objects.filter(Q(is_distributor=1) | Q(is_super_stockist=1)).values('id','user_type','is_distributor','is_super_stockist')
                if len(operational_users) :
                    for operational_user in operational_users :
                        user_product_variant = SpUserProductVariants()
                        user_product_variant.user_id = operational_user['id']
                        user_product_variant.product_id = request.POST['product_id']
                        user_product_variant.product_name  = getModelColumnById(SpProducts,request.POST['product_id'],'product_name')
                        user_product_variant.product_variant_id = product_variant.id
                        user_product_variant.product_class_id  = getModelColumnById(SpProducts,request.POST['product_id'],'product_class_id')
                        user_product_variant.item_sku_code = request.POST['item_code']
                        user_product_variant.variant_quantity = request.POST['variant_qty']
                        user_product_variant.variant_unit_id  = request.POST['variant_unit']
                        user_product_variant.variant_name  = request.POST['variant_name']
                        user_product_variant.variant_unit_name  = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'unit')
                        user_product_variant.largest_unit_name = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'largest_unit')
                        user_product_variant.variant_size  = request.POST['variant_size']
                        user_product_variant.no_of_pouch  = request.POST['no_of_pouch']
                        user_product_variant.container_size  = request.POST['container_size']
                        user_product_variant.is_bulk_pack  = request.POST['variant_type']
                        user_product_variant.included_in_scheme  = request.POST['included_in_scheme']
                        user_product_variant.mrp  = request.POST['mrp']
                        user_product_variant.container_mrp  = float(request.POST['mrp']) * float(request.POST['no_of_pouch'])

                        if operational_user['user_type'] == 1 :
                            user_product_variant.sp_user  = request.POST['sp_employee']
                            user_product_variant.container_sp_user  = float(request.POST['sp_employee']) * float(request.POST['no_of_pouch'])
                        else:
                            if operational_user['is_distributor'] == 1:
                                user_product_variant.sp_user = request.POST['sp_distributor']
                                user_product_variant.container_sp_user  = float(request.POST['sp_distributor']) * float(request.POST['no_of_pouch'])
                            else:
                                user_product_variant.sp_user = request.POST['sp_superstockist']
                                user_product_variant.container_sp_user  = float(request.POST['sp_superstockist']) * float(request.POST['no_of_pouch'])
                        

                        user_product_variant.valid_from = datetime.strptime(request.POST['valid_from'], '%d/%m/%Y').strftime('%Y-%m-%d')
                        user_product_variant.valid_to = datetime.strptime(request.POST['valid_to'], '%d/%m/%Y').strftime('%Y-%m-%d')
                        user_product_variant.status  = 1
                        user_product_variant.save()

                variant_images       = request.FILES.getlist('variantImage[]')
                i = 0 
                for variant_image in variant_images:
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    variant_image_name = variant_image.name
                    temp = variant_image_name.split('.')
                    variant_image_name = str(product_variant.id) + "_"+str(timestamp)+"_"+str(i)+"."+temp[(len(temp) - 1)]
                    
                    uploaded_variant_image = storage.save(variant_image_name, variant_image)
                    uploaded_variant_image = storage.url(uploaded_variant_image)
                    product_variant_image = SpProductVariantImages()
                    product_variant_image.product_variant_id = product_variant.id
                    product_variant_image.image_url = uploaded_variant_image
                    product_variant_image.save()
                    i = i+1
                    
                #Save Activity
                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'Product variant created'
                activity    = 'Product variant created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Product & Variant Management', 'Add Product Variant', heading, activity, request.user.id, user_name, 'addProduct.png', '1', 'web.png')

                response['flag'] = True
                response['message'] = "Record has been saved successfully."
            else:
                response['flag'] = False
                response['message'] = "Failed to save"

        except Exception as e:
            response['flag'] = False
            response['message'] = str(e)
        return JsonResponse(response)

    else:
        template = 'products/add-product-variant.html'
        context = {}
        context['products']     = SpProducts.objects.filter(status=1)
        context['product']     = SpProducts.objects.get(id=product_id)
        context['product_units']          = SpProductUnits.objects.filter(status=1)
        

    return render(request, template,context)

@login_required
def editProductVariant(request,product_variant_id):
    if request.method == "POST":
        response = {}
        try:
            product_variant = SpProductVariants.objects.get(id=request.POST['product_variant_id'])
            if (int(request.POST['update_confirm']) == 0) and ((float(request.POST['mrp']) !=  float(product_variant.mrp)) or (float(request.POST['sp_distributor']) !=  float(product_variant.sp_distributor)) or (float(request.POST['sp_superstockist']) !=  float(product_variant.sp_superstockist))) :
                response['flag'] = False
                response['confirm'] = True
                response['message'] = "MRP/SP.Distributor/SP.Super Stockist has been changed. Do you want save changes?"
            else:
                valid_from = datetime.strptime(request.POST['valid_from'], '%d/%m/%Y').strftime('%Y-%m-%d')
                valid_to = datetime.strptime(request.POST['valid_to'], '%d/%m/%Y').strftime('%Y-%m-%d')

                if (str(valid_from) != str(product_variant.valid_from)) or (str(valid_to) != str(product_variant.valid_to)) or  ((float(request.POST['mrp']) !=  float(product_variant.mrp)) or (float(request.POST['sp_distributor']) !=  float(product_variant.sp_distributor)) or (float(request.POST['sp_superstockist']) !=  float(product_variant.sp_superstockist))) :
                        description = ''
                        if (str(valid_from) != str(product_variant.valid_from)) :
                            description += 'From date changed from '+ str(product_variant.valid_from) +' to '+valid_from +'. '
                        if (str(valid_to) != str(product_variant.valid_to)) :
                            description += 'To date changed from '+ str(product_variant.valid_to) +' to '+valid_to +'. '
                        if float(request.POST['mrp']) !=  float(product_variant.mrp) :
                            description += 'MRP changed from '+ str(product_variant.mrp) +' to '+request.POST['mrp'] +'. '
                        if float(request.POST['sp_distributor']) !=  float(product_variant.sp_distributor) :
                            description += 'SP Distributor changed from '+ str(product_variant.sp_distributor) +' to '+request.POST['sp_distributor'] +'. '
                        if float(request.POST['sp_superstockist']) !=  float(product_variant.sp_superstockist) :
                            description += 'SP Super Stockist changed from '+ str(product_variant.sp_superstockist) +' to '+request.POST['sp_superstockist'] +'. '
                        if float(request.POST['sp_employee']) !=  float(product_variant.sp_employee) :
                            description += 'SP Employee changed from '+ str(product_variant.sp_employee) +' to '+request.POST['sp_employee'] +'. '

                        history = SpProductVariantsHistory()
                        current_user = request.user
                        history.user_id  = current_user.id
                        history.user_name  = str(current_user.first_name) + " " + current_user.middle_name+ " " + current_user.last_name
                        history.product_variant_id  = product_variant.id
                        history.mrp  = product_variant.mrp
                        history.sp_distributor  = product_variant.sp_distributor
                        history.sp_superstockist  = product_variant.sp_superstockist
                        history.sp_employee  = product_variant.sp_employee

                        history.container_mrp  = product_variant.container_mrp
                        history.container_sp_distributor  = product_variant.container_sp_distributor
                        history.container_sp_superstockist  = product_variant.container_sp_superstockist
                        history.container_sp_employee  = product_variant.container_sp_employee

                        history.valid_from = product_variant.valid_from
                        history.valid_to = product_variant.valid_to
                        history.description = description
                        history.save()
                       

                product_variant.product_id = request.POST['product_id']
                product_variant.product_name  = getModelColumnById(SpProducts,request.POST['product_id'],'product_name')
                product_variant.product_class_id  = getModelColumnById(SpProducts,request.POST['product_id'],'product_class_id')
                product_variant.item_sku_code = request.POST['item_code']
                product_variant.variant_quantity = request.POST['variant_qty']
                product_variant.variant_unit_id  = request.POST['variant_unit']
                product_variant.variant_name  = request.POST['variant_name']
                product_variant.variant_unit_name  = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'unit')
                product_variant.largest_unit_name = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'largest_unit')
                product_variant.variant_size  = request.POST['variant_size']
                product_variant.no_of_pouch  = request.POST['no_of_pouch']
                product_variant.container_size  = request.POST['container_size']
                product_variant.is_bulk_pack  = request.POST['variant_type']
                product_variant.included_in_scheme  = request.POST['included_in_scheme']
                product_variant.mrp  = request.POST['mrp']
                product_variant.sp_distributor  = request.POST['sp_distributor']
                product_variant.sp_superstockist  = request.POST['sp_superstockist']
                product_variant.sp_employee  = request.POST['sp_employee']

                product_variant.container_mrp  = float(request.POST['mrp']) * float(request.POST['no_of_pouch'])
                product_variant.container_sp_distributor  = float(request.POST['sp_distributor']) * float(request.POST['no_of_pouch'])
                product_variant.container_sp_superstockist  = float(request.POST['sp_superstockist']) * float(request.POST['no_of_pouch'])
                product_variant.container_sp_employee  = float(request.POST['sp_employee']) * float(request.POST['no_of_pouch'])

                product_variant.valid_from = valid_from
                product_variant.valid_to = valid_to
                product_variant.status  = 1
                product_variant.save()



                if product_variant.id:
                    
                    users = SpUsers.objects.raw(''' SELECT id,user_type,is_distributor,is_super_stockist FROM sp_users where 
                                                id in (SELECT user_id FROM sp_user_product_variants WHERE product_variant_id=%s)
                                                 ''', [product_variant.id])
                    print(len(users))
                    if len(users) :
                        for user in users :
                            user_product_variant = SpUserProductVariants.objects.get(user_id=user.id,product_variant_id=product_variant.id)
                            user_product_variant.user_id = user.id
                            user_product_variant.product_id = request.POST['product_id']
                            user_product_variant.product_name  = getModelColumnById(SpProducts,request.POST['product_id'],'product_name')
                            user_product_variant.product_variant_id = product_variant.id
                            user_product_variant.product_class_id  = getModelColumnById(SpProducts,request.POST['product_id'],'product_class_id')
                            user_product_variant.item_sku_code = request.POST['item_code']
                            user_product_variant.variant_quantity = request.POST['variant_qty']
                            user_product_variant.variant_unit_id  = request.POST['variant_unit']
                            user_product_variant.variant_name  = request.POST['variant_name']
                            user_product_variant.variant_unit_name  = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'unit')
                            user_product_variant.largest_unit_name = getModelColumnById(SpProductUnits,request.POST['variant_unit'],'largest_unit')
                            user_product_variant.variant_size  = request.POST['variant_size']
                            user_product_variant.no_of_pouch  = request.POST['no_of_pouch']
                            user_product_variant.container_size  = request.POST['container_size']
                            user_product_variant.is_bulk_pack  = request.POST['variant_type']
                            user_product_variant.included_in_scheme  = request.POST['included_in_scheme']
                            user_product_variant.mrp  = request.POST['mrp']
                            user_product_variant.container_mrp  = float(request.POST['mrp']) * float(request.POST['no_of_pouch'])

                            if user.user_type == 1 :
                                user_product_variant.sp_user  = request.POST['sp_employee']
                                user_product_variant.container_sp_user  = float(request.POST['sp_employee']) * float(request.POST['no_of_pouch'])
                            else:
                                if user.is_distributor == 1:
                                    user_product_variant.sp_user = request.POST['sp_distributor']
                                    user_product_variant.container_sp_user  = float(request.POST['sp_distributor']) * float(request.POST['no_of_pouch'])
                                else:
                                    user_product_variant.sp_user = request.POST['sp_superstockist']
                                    user_product_variant.container_sp_user  = float(request.POST['sp_superstockist']) * float(request.POST['no_of_pouch'])
                            

                            user_product_variant.valid_from = datetime.strptime(request.POST['valid_from'], '%d/%m/%Y').strftime('%Y-%m-%d')
                            user_product_variant.valid_to = datetime.strptime(request.POST['valid_to'], '%d/%m/%Y').strftime('%Y-%m-%d')
                            user_product_variant.status  = 1
                            user_product_variant.save()

                    if(len(request.FILES.getlist('variantImage[]'))) :
                        variant_images       = request.FILES.getlist('variantImage[]')
                        i = 0 
                        SpProductVariantImages.objects.filter(product_variant_id=request.POST['product_variant_id']).delete()

                        for variant_image in variant_images:
                            storage = FileSystemStorage()
                            timestamp = int(time.time())
                            variant_image_name = variant_image.name
                            temp = variant_image_name.split('.')
                            variant_image_name = str(product_variant.id) + "_"+str(timestamp)+"_"+str(i)+"."+temp[(len(temp) - 1)]
                            
                            uploaded_variant_image = storage.save(variant_image_name, variant_image)
                            uploaded_variant_image = storage.url(uploaded_variant_image)
                            product_variant_image = SpProductVariantImages()
                            product_variant_image.product_variant_id = product_variant.id
                            product_variant_image.image_url = uploaded_variant_image
                            product_variant_image.save()
                            i = i+1

                    #Save Activity
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = 'Product variant edited'
                    activity    = 'Product variant edited by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    saveActivity('Product & Variant Management', 'Edit Product Variant', heading, activity, request.user.id, user_name, 'edit_product.png', '1', 'web.png')

                    response['flag'] = True
                    response['message'] = "Record has been saved successfully."
                else:
                    response['flag'] = False
                    response['message'] = "Failed to save"

        except Exception as e:
            response['flag'] = False
            response['message'] = str(e)
        return JsonResponse(response)

    else:
        template = 'products/edit-product-variant.html'
        context = {}
        context['products']     = SpProducts.objects.filter(status=1)
        context['product_variant'] = product_variant = SpProductVariants.objects.get(id=product_variant_id)
        context['product_variant_unit']          = SpProductUnits.objects.get(id=product_variant.variant_unit_id)
        context['product'] = SpProducts.objects.get(id=product_variant.product_id)
        context['product_units']          = SpProductUnits.objects.filter(status=1)
        context['color_codes']          = SpColorCodes.objects.filter(status=1)
        context['product_variant_images'] = SpProductVariantImages.objects.filter(product_variant_id=product_variant_id)

    return render(request, template,context)


@login_required
def productUnitDetails(request,unit_id):
    response = {}
    if SpProductUnits.objects.filter(id=unit_id).exists():
        response['flag'] = True
        response['product_unit'] = model_to_dict(SpProductUnits.objects.get(id=unit_id))
    else:
        response['flag'] = False
        response['flag'] = "Unit not found"

    return JsonResponse(response)

@login_required
def productDetails(request,product_id):
    response = {}
    if SpProducts.objects.filter(id=product_id).exists():
        response['flag'] = True
        response['product'] = model_to_dict(SpProducts.objects.get(id=product_id))
    else:
        response['flag'] = False
        response['flag'] = "Product not found"

    return JsonResponse(response)




@login_required
def updateProductStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpProducts.objects.get(id=id)
            data.status = is_active
            data.save()
            
            SpProductVariants.objects.filter(product_id=id).update(status=is_active)
            SpUserProductVariants.objects.filter(product_id=id).update(status=is_active)

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Product status changed'
            activity    = 'Product status changed by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Edit Product', heading, activity, request.user.id, user_name, 'edit_product.png', '1', 'web.png')

            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/products')



@login_required
def updateProductVariantStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpProductVariants.objects.get(id=id)
            data.status = is_active
            data.save()
            
            SpUserProductVariants.objects.filter(product_variant_id=id).update(status=is_active)

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Product variant status changed'
            activity    = 'Product variant status changed by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Product Variant', heading, activity, request.user.id, user_name, 'edit_product.png', '1', 'web.png')

            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/products')

