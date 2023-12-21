import sys
import os
import openpyxl
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings

# List View
@login_required
def index(request):
    today   = date.today()  
    year  = today.year
    month = today.month

    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)
    
    total_opening_crates        = []
    total_dispatch_crate        = []
    total_plant_crate           = []
    total_short_for_month       = []
    total_short_since_starting  = []
    for user in user_list:
        previous_month_last_date = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        if SpPlantCrateLedger.objects.filter(user_id=user.id,updated_datetime__lt=previous_month_last_date).exists():
            opening_crates              = getLastCrateBalance(user.id, previous_month_last_date, 'normal')
            user.opening_crates         = opening_crates
        else:
            user.opening_crates         = 0 
        total_opening_crates.append(user.opening_crates)

        if getTotalCrates(user.id, year, month, 'normal', 'dispatch'):
            user.total_dispatch_crates  = getTotalCrates(user.id, year, month, 'normal', 'dispatch')
        else:
            user.total_dispatch_crates  = 0    
        total_dispatch_crate.append(user.total_dispatch_crates)

        if getTotalCrates(user.id, year, month, 'normal', 'plant'):        
            user.total_plant_crates     = getTotalCrates(user.id, year, month, 'normal', 'plant')
        else:
            user.total_plant_crates     = 0    
        total_plant_crate.append(user.total_plant_crates)

        user.short_for_month        =  int(user.total_plant_crates)-int(user.total_dispatch_crates)
        user.short_since_starting   =  int(user.opening_crates)+int(user.short_for_month)

        total_short_for_month.append(user.short_for_month)
        total_short_since_starting.append(user.short_since_starting)

        crate_lists = []
        for months in month_list:
            month_lists               = {} 
            month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            dispatch                  = getDispatchedCratesSum(user.id, month_date, 'normal')
            plant                     = getReceivedCratesSum(user.id, month_date, 'normal')

            month_lists['dispatch']   = dispatch
            month_lists['plant']      = plant

            crate_lists.append(month_lists)    
        user.crates =  crate_lists

    total_crate_lists = []
    for months in month_list:
        total_month_lists               = {} 
        month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
        dispatch                  = getTotalDispatchedCratesSum(month_date, 'normal')
        plant                     = getTotalReceivedCratesSum(month_date, 'normal')

        total_month_lists['dispatch']   = dispatch
        total_month_lists['plant']      = plant

        total_crate_lists.append(total_month_lists)
    
    context = {}
    context['user_list']                    = user_list
    context['today_date']                   = today.strftime("%m/%Y")
    context['month_list']                   = days_in_months(year,month)
    context['month_lists']                  = month_lists
    context['total_opening_crates']         = sum(total_opening_crates)
    context['total_dispatch_crate']         = sum(total_dispatch_crate)
    context['total_plant_crate']            = sum(total_plant_crate)
    context['total_short_for_month']        = sum(total_short_for_month)
    context['total_short_since_starting']   = sum(total_short_since_starting)
    context['total_crate_lists']            = total_crate_lists
    context['page_title']                   = "Crate Summary"

    template = 'crate/index.html'
    return render(request, template, context)

# ajax List View
@login_required
def ajaxCrateSummary(request):
    today   = date.today()  
    if request.GET['crate_date']:
        crate_date = request.GET['crate_date']
        crate_date = crate_date.split('/')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)
           
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)
    
    total_opening_crates        = []
    total_dispatch_crate        = []
    total_plant_crate           = []
    total_short_for_month       = []
    total_short_since_starting  = []
    for user in user_list:
        previous_month_last_date = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        if SpPlantCrateLedger.objects.filter(user_id=user.id,updated_datetime__lt=previous_month_last_date).exists():
            opening_crates              = getLastCrateBalance(user.id, previous_month_last_date, request.GET['crate_type'])
            user.opening_crates         = opening_crates
        else:
            user.opening_crates         = 0
        total_opening_crates.append(user.opening_crates)
        
        if getTotalCrates(user.id, year, month, request.GET['crate_type'], 'dispatch'):
            user.total_dispatch_crates  = getTotalCrates(user.id, year, month, request.GET['crate_type'], 'dispatch')
        else:
            user.total_dispatch_crates  = 0
        total_dispatch_crate.append(user.total_dispatch_crates)
        
        if getTotalCrates(user.id, year, month, request.GET['crate_type'], 'plant'):        
            user.total_plant_crates     = getTotalCrates(user.id, year, month, request.GET['crate_type'], 'plant')
        else:
            user.total_plant_crates     = 0      
        total_plant_crate.append(user.total_plant_crates)
        
        user.short_for_month        =  int(user.total_plant_crates)-int(user.total_dispatch_crates)
        user.short_since_starting   =  int(user.opening_crates)+int(user.short_for_month)

        total_short_for_month.append(user.short_for_month)
        total_short_since_starting.append(user.short_since_starting)

        crate_lists = []
        for months in month_list:
            month_lists               = {} 
            month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['dispatch']   = getDispatchedCratesSum(user.id, month_date, request.GET['crate_type'])
            month_lists['plant']      = getReceivedCratesSum(user.id, month_date, request.GET['crate_type'])
            crate_lists.append(month_lists)    
        user.crates =  crate_lists

    total_crate_lists = []
    for months in month_list:
        total_month_lists               = {} 
        month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
        dispatch                  = getTotalDispatchedCratesSum(month_date, request.GET['crate_type'])
        plant                     = getTotalReceivedCratesSum(month_date, request.GET['crate_type'])

        total_month_lists['dispatch']   = dispatch
        total_month_lists['plant']      = plant

        total_crate_lists.append(total_month_lists)    
 
    context = {}
    context['user_list']                    = user_list
    context['today_date']                   = today.strftime("%m/%Y")
    context['month_list']                   = days_in_months(year,month)
    context['month_lists']                  = month_lists
    context['total_opening_crates']         = sum(total_opening_crates)
    context['total_dispatch_crate']         = sum(total_dispatch_crate)
    context['total_plant_crate']            = sum(total_plant_crate)
    context['total_short_for_month']        = sum(total_short_for_month)
    context['total_short_since_starting']   = sum(total_short_since_starting)
    context['total_crate_lists']            = total_crate_lists

    template = 'crate/ajax-crate-summary.html'
    return render(request, template, context)    

# export view
@login_required
def exportCrateSummary(request, crate_date, crate_type):
    today   = date.today()  
    # crates_details = SpUserCrates.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).order_by('id')
    if crate_date:
        crate_date = crate_date.split('-')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)
           
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)
    
    total_opening_crates        = []
    total_dispatch_crate        = []
    total_plant_crate           = []
    total_short_for_month       = []
    total_short_since_starting  = []
    for user in user_list:
        previous_month_last_date = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        if SpPlantCrateLedger.objects.filter(user_id=user.id,updated_datetime__lt=previous_month_last_date).exists():
            opening_crates              = getLastCrateBalance(user.id, previous_month_last_date, crate_type)
            user.opening_crates         = opening_crates
        else:
            user.opening_crates         = 0
        total_opening_crates.append(user.opening_crates)

        if getTotalCrates(user.id, year, month, crate_type, 'dispatch'):
            user.total_dispatch_crates  = getTotalCrates(user.id, year, month, crate_type, 'dispatch')
        else:
            user.total_dispatch_crates  = 0
        total_dispatch_crate.append(user.total_dispatch_crates)

        if getTotalCrates(user.id, year, month, crate_type, 'plant'):        
            user.total_plant_crates     = getTotalCrates(user.id, year, month, crate_type, 'plant')
        else:
            user.total_plant_crates     = 0      
        total_plant_crate.append(user.total_plant_crates)

        user.short_for_month        =  int(user.total_plant_crates)-int(user.total_dispatch_crates)
        user.short_since_starting   =  int(user.opening_crates)+int(user.short_for_month)

        total_short_for_month.append(user.short_for_month)
        total_short_since_starting.append(user.short_since_starting)

        crate_lists = []
        for months in month_list:
            month_lists               = {} 
            month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['dispatch']   = getDispatchedCratesSum(user.id, month_date, crate_type)
            month_lists['plant']      = getReceivedCratesSum(user.id, month_date, crate_type)
            crate_lists.append(month_lists)    
        user.crates =  crate_lists

    total_crate_lists = []
    for months in month_list:
        total_month_lists               = {} 
        month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
        dispatch                  = getTotalDispatchedCratesSum(month_date, crate_type)
        plant                     = getTotalReceivedCratesSum(month_date, crate_type)

        total_month_lists['dispatch']   = dispatch
        total_month_lists['plant']      = plant

        total_crate_lists.append(total_month_lists)

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=crate-summary-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'CRATE SUMMARY'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = len(month_list)+len(month_list)+7
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'SAAHAJ MILK PRODUCER COMPANY LIMITED, AGRA, UTTAR PRADESH'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=24, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    # color_codes = []
    columns = []
    columns += [ 'Customer Code' ]
    columns += [ 'Name of Distributor/SS' ]
    columns += [ 'Opening' ]
    if month_list:
        for months in month_list:
            columns += [ months ]
            columns += [ months ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    row_num = 2
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        if col_num >3:
            cell.font = Font(size=9, color='000000', bold=True)
        else:                  
            cell.font = Font(size=11, color='000000', bold=True) 
            
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num  == 3:
            column_dimensions.width = 10
        elif col_num >3:
            column_dimensions.width = 10
        else:
            column_dimensions.width = 25

    columns = []
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    if month_list:
        for months in month_list:
            columns += [ 'Dispatch' ]
            columns += [ 'Plant' ]
    columns += [ 'Plant' ]
    columns += [ 'Dispatch' ]
    columns += [ 'Short for Month' ]
    columns += [ 'Short Since starting' ]
    row_num = 3
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True)              
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num  == 3:
            column_dimensions.width = 10
        elif col_num >3:
            column_dimensions.width = 10
        else:
            column_dimensions.width = 25

    for id, user in enumerate(user_list):
        row_num += 1
        # Define the data for each cell in the row    
        row = []
        row += [ user.emp_sap_id ]
        row += [ user.first_name +' '+ user.middle_name +' '+ user.last_name ]
        row += [ user.opening_crates ]
        if user.crates:
            for crate in user.crates:
                row += [ crate['dispatch'] ]
                row += [ crate['plant'] ]
        row += [ user.total_plant_crates ]        
        row += [ user.total_dispatch_crates ]
        row += [ user.short_for_month ]
        row += [ user.short_since_starting ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            cell.fill = PatternFill()
            cell.font = Font(size=12, color='000000')
    

    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ ' ' ]
    bottom_columns += [ 'Total Crates Per Day' ]
    bottom_columns += [ sum(total_opening_crates) ]

    if total_crate_lists:
        for crate_list in total_crate_lists:
            bottom_columns += [ crate_list['dispatch'] ]
            bottom_columns += [ crate_list['plant'] ]

    bottom_columns += [ sum(total_plant_crate) ]
    bottom_columns += [ sum(total_dispatch_crate) ]
    bottom_columns += [ sum(total_short_for_month) ]
    bottom_columns += [ sum(total_short_since_starting) ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True)              
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num  == 3:
            column_dimensions.width = 10
        elif col_num >3:
            column_dimensions.width = 10
        else:
            column_dimensions.width = 25

    row_num += 1
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By SAAHAJ MILK PRODUCER COMPANY LIMITED'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response


#user crate summary
@login_required
def userCrateSummary(request):
    today   = date.today()  
    year  = today.year
    month = today.month

    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    total_dispatch                  = []
    total_return_to_plant           = []
    total_plant                     = []
    total_difference_with_vehicle   = []
    total_crate_difference          = []
    if user:
        previous_month_last_date    = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        opening_crates              = getLastCrateBalance(user['id'], previous_month_last_date, 'normal')

        opening_crates_list = [opening_crates]
        for id, months in enumerate(month_list):
            month_lists                         = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['dispatch']                 = getDispatchedCratesSum(user['id'], month_date, 'normal')
            month_lists['return_to_plant']          = getReturnCratesSum(user['id'], month_date, 'normal')
            month_lists['plant']                    = getReceivedCratesSum(user['id'], month_date, 'normal')
            
            month_lists['difference_with_vehicle']   = int(month_lists['plant'])-int(month_lists['return_to_plant'])
            month_lists['crate_difference']          = int(opening_crates_list[id])-(int(month_lists['dispatch'])+int(month_lists['plant']))
            opening_crates_list.append(month_lists['crate_difference'])
            crate_lists.append(month_lists)

            total_dispatch.append(month_lists['dispatch'])
            total_return_to_plant.append(month_lists['return_to_plant'])
            total_plant.append(month_lists['plant'])
            total_difference_with_vehicle.append(month_lists['difference_with_vehicle'])
            total_crate_difference.append(month_lists['crate_difference'])

    else:
        opening_crates              = 0            

    if len(crate_lists) > 0:
        last_opening_balance = crate_lists[-1]
        last_opening_balance = last_opening_balance['crate_difference']
    else:
        last_opening_balance = 0
           
    context = {}
    context['user_list']                        = user_list
    context['crate_lists']                      = crate_lists
    context['opening_crates']                   = opening_crates
    context['total_dispatch']                   = sum(total_dispatch)
    context['total_return_to_plant']            = sum(total_return_to_plant)
    context['total_plant']                      = sum(total_plant)
    context['total_difference_with_vehicle']    = sum(total_difference_with_vehicle)
    context['total_crate_difference']           = sum(total_crate_difference)
    context['last_opening_balance']             = last_opening_balance
    context['today_date']                       = today.strftime("%m/%Y")
    context['month_list']                       = days_in_months(year,month)
    context['page_title']                       = "User Crate Summary"

    template = 'crate/user-crate-summary.html'
    return render(request, template, context)

# ajax List View
@login_required
def ajaxUserCrateSummary(request):
    today   = date.today()  
    if request.GET['crate_date']:
        crate_date = request.GET['crate_date']
        crate_date = crate_date.split('/')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)
           
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(id=request.GET['user_id'],user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    total_dispatch                  = []
    total_return_to_plant           = []
    total_plant                     = []
    total_difference_with_vehicle   = []
    total_crate_difference          = []
    if user:
        previous_month_last_date    = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        opening_crates              = getLastCrateBalance(request.GET['user_id'], previous_month_last_date, request.GET['crate_type'])

        opening_crates_list = [opening_crates]
        for id, months in enumerate(month_list):
            month_lists                             = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['dispatch']                 = getDispatchedCratesSum(request.GET['user_id'], month_date, request.GET['crate_type'])
            month_lists['return_to_plant']          = getReturnCratesSum(request.GET['user_id'], month_date, request.GET['crate_type'])
            month_lists['plant']                    = getReceivedCratesSum(request.GET['user_id'], month_date, request.GET['crate_type'])
            
            month_lists['difference_with_vehicle']   = int(month_lists['plant'])-int(month_lists['return_to_plant'])
            month_lists['crate_difference']          = int(opening_crates_list[id])-(int(month_lists['dispatch'])+int(month_lists['plant']))
            opening_crates_list.append(month_lists['crate_difference'])
            crate_lists.append(month_lists)

            total_dispatch.append(month_lists['dispatch'])
            total_return_to_plant.append(month_lists['return_to_plant'])
            total_plant.append(month_lists['plant'])
            total_difference_with_vehicle.append(month_lists['difference_with_vehicle'])
            total_crate_difference.append(month_lists['crate_difference'])

    else:
        opening_crates              = 0    
    
    if len(crate_lists) > 0:
        last_opening_balance = crate_lists[-1]
        last_opening_balance = last_opening_balance['crate_difference']
    else:
        last_opening_balance = 0

    context = {}
    context['user_list']                        = user_list
    context['crate_lists']                      = crate_lists
    context['opening_crates']                   = opening_crates
    context['total_dispatch']                   = sum(total_dispatch)
    context['total_return_to_plant']            = sum(total_return_to_plant)
    context['total_plant']                      = sum(total_plant)
    context['total_difference_with_vehicle']    = sum(total_difference_with_vehicle)
    context['total_crate_difference']           = sum(total_crate_difference)
    context['last_opening_balance']             = last_opening_balance
    context['today_date']                       = today.strftime("%m/%Y")
    context['month_list']                       = days_in_months(year,month)

    template = 'crate/ajax-user-crate-summary.html'
    return render(request, template, context) 


# export view
@login_required
def exportUserCrateSummary(request, crate_date, crate_type, user_id):
    today   = date.today()  
    # crates_details = SpUserCrates.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).order_by('id')
    if crate_date:
        crate_date = crate_date.split('-')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)

    user_id = int(user_id)       
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(id=user_id, user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name', 'emp_sap_id', 'store_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    total_dispatch                  = []
    total_return_to_plant           = []
    total_plant                     = []
    total_difference_with_vehicle   = []
    total_crate_difference          = []
    if user:
        user_town                   = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user_id, 'town_name')
        previous_month_last_date    = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        opening_crates              = getLastCrateBalance(user_id, previous_month_last_date, crate_type)

        opening_crates_list = [opening_crates]
        for id, months in enumerate(month_list):
            month_lists                         = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['dispatch']                 = getDispatchedCratesSum(user_id, month_date, crate_type)
            month_lists['return_to_plant']          = getReturnCratesSum(user_id, month_date, crate_type)
            month_lists['plant']                    = getReceivedCratesSum(user_id, month_date, crate_type)
            
            month_lists['difference_with_vehicle']   = int(month_lists['plant'])-int(month_lists['return_to_plant'])
            month_lists['crate_difference']          = int(opening_crates_list[id])-(int(month_lists['dispatch'])+int(month_lists['plant']))
            opening_crates_list.append(month_lists['crate_difference'])
            crate_lists.append(month_lists)

            total_dispatch.append(month_lists['dispatch'])
            total_return_to_plant.append(month_lists['return_to_plant'])
            total_plant.append(month_lists['plant'])
            total_difference_with_vehicle.append(month_lists['difference_with_vehicle'])
            total_crate_difference.append(month_lists['crate_difference'])
    else:
        opening_crates              = 0

    if len(crate_lists) > 0:
        last_opening_balance = crate_lists[-1]
        last_opening_balance = last_opening_balance['crate_difference']
    else:
        last_opening_balance = 0

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=user-crate-summary-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'USER CRATE SUMMARY'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    # cell = worksheet.cell(row=1, column=2)  
    # cell.value = str(year)+'/'+str(month)
    # cell.font = header_font
    # cell.alignment = wrapped_alignment
    # cell.border = black_border
    # cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    # cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = 6
    
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'SAAHAJ MILK PRODUCER COMPANY LIMITED, AGRA, UTTAR PRADESH'
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(size=15, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    # color_codes = []
    columns = []
    columns += [ 'Vendor Name' ]
    columns += [ str(user['first_name'])+' '+str(user['first_name'])+' '+str(user['first_name'])+'/'+str(user['store_name']) ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'Location' ]
    columns += [ user_town ]
    row_num = 2
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        if col_num == 2 or col_num == 6:
            cell.font = Font(size=11, color='000000')
        else:
            cell.font = Font(size=11, color='000000', bold=True)  
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    month_name = datetime(year,month,1).strftime( '%B' )
    
    # color_codes = []
    columns = []
    columns += [ 'Vendor Code' ]
    columns += [ user['emp_sap_id'] ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'Month' ]
    columns += [ str(month_name)+'-'+str(year) ]
    row_num = 3
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        if col_num == 2 or col_num == 6:
            cell.font = Font(size=11, color='000000')
        else:
            cell.font = Font(size=11, color='000000', bold=True)     
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # color_codes = []
    columns = []
    columns += [ 'Date' ]
    columns += [ 'Dispatched from Plant' ]
    columns += [ 'Return to Vehicle' ]
    columns += [ 'Deposit in Plant' ]
    columns += [ 'Difference with Vehicle' ]
    columns += [ 'Crate Difference' ]
    row_num = 4
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20


    # color_codes = []
    columns = []
    columns += [ 'Opening' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ opening_crates ]
    row_num += 1
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for id, crate in enumerate(crate_lists):
        row_num += 1
        # Define the data for each cell in the row    
        row = []
        row += [ crate['month_date'] ]
        row += [ crate['dispatch'] ]
        row += [ crate['return_to_plant'] ]
        row += [ crate['plant'] ]
        row += [ crate['difference_with_vehicle'] ]
        row += [ crate['crate_difference'] ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            cell.fill = PatternFill()
            cell.font = Font(size=12, color='000000')

    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ 'Total' ]
    bottom_columns += [ sum(total_dispatch) ]
    bottom_columns += [ sum(total_return_to_plant) ]
    bottom_columns += [ sum(total_plant) ]
    bottom_columns += [ sum(total_difference_with_vehicle) ]
    bottom_columns += [ last_opening_balance ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    row_num += 1
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By SAAHAJ MILK PRODUCER COMPANY LIMITED'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=9, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response