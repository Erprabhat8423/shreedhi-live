import sys
import os
import json
from django.core import serializers
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import getConfigurationResult,getModelColumnById,clean_data
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from django.conf import settings
from ..decorators import *
from datetime import datetime
from calendar import monthrange
from datetime import date
import calendar
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.utils.html import strip_tags




@login_required
def index(request):
    page = request.GET.get('page')
    context = {}
    months = [1,2,3,4,5,6,7,8,9,10,11,12]
    month_holidays = []
    for month in months :
        temp = {}
        temp['month'] = month
        temp['month_name'] = calendar.month_name[month]
        holidays = SpHolidays.objects.filter(start_date__month=month)
        if len(holidays):
            temp['month_holidays'] = holidays
            month_holidays.append(temp)
        else:
            holidays = SpHolidays.objects.filter(end_date__month=month)
            if len(holidays):
                temp['month_holidays'] = holidays
                month_holidays.append(temp)



    
    context['total_holidays'] = SpHolidays.objects.all().count()
    current_date = date.today() 
    context['current_year'] = current_year = current_date.year
    context['current_month'] = current_month = current_date.month
    
    if int(current_month) == 12:
        context['next_month'] = 1
    else:
        context['next_month'] = int(current_month) + 1
    if int(current_month) == 1:
        context['previous_month'] = 12
    else:
        context['previous_month'] = int(current_month) - 1
    
    context['current_month_name'] = calendar.month_name[current_month]

    last_holiday = SpHolidays.objects.filter().order_by('-id').first()
    if last_holiday:
        holiday_dates = []
        start_date = last_holiday.start_date
        delta = last_holiday.end_date - last_holiday.start_date
        for i in range(delta.days + 1):
            holiday_date = start_date + timedelta(days=i)
            holiday_date = holiday_date.strftime('%Y-%m-%d')
            holiday_dates.append(holiday_date)

        calendarObj = calendar.Calendar()
        calendar_dates = []
        for week in calendarObj.monthdatescalendar(current_year, current_month):
            dates = []
            for week_date in week:
                calendar_datass = []
                tmp = {}
                tmp['full_date'] = week_date
                tmp['day'] = week_date.strftime('%A')
                tmp['short_date'] = week_date.strftime('%d')
                tmp['month'] = week_date.strftime('%m')
                                  

                if str(week_date) in holiday_dates:
                     tmp['is_holiday'] = 1
                     tmp['holiday'] = last_holiday.holiday
                else:
                    tmp['is_holiday'] = 0
                
                dates.append(tmp)               
            calendar_dates.append(dates)

        context['calendar_dates'] = calendar_dates

    context['holiday'] = last_holiday
    context['holidays'] = month_holidays
    context['holiday_types'] = SpHolidayTypes.objects.all()
    context['page_limit'] = getConfigurationResult('page_limit')
    context['page_title'] = "Manage Holidays"
    template = 'holiday/index.html'
    return render(request, template, context)


@login_required
def filterHoliday(request,filter_status):
    if request.method == 'POST':
        context = {}
        months = [1,2,3,4,5,6,7,8,9,10,11,12]
        month_holidays = []
        for month in months :
            temp = {}
            temp['month'] = month
            temp['month_name'] = calendar.month_name[month]
            if int(filter_status) < 0:
                holidays = SpHolidays.objects.filter(start_date__month=month)
                if len(holidays):
                    temp['month_holidays'] = holidays
                    month_holidays.append(temp)
                else:
                    holidays = SpHolidays.objects.filter(end_date__month=month)
                    if len(holidays):
                        temp['month_holidays'] = holidays
                        month_holidays.append(temp)
            else:
                holidays = SpHolidays.objects.filter(holiday_status=filter_status,start_date__month=month)
                if len(holidays):
                    temp['month_holidays'] = holidays
                    month_holidays.append(temp)
                else:
                    holidays = SpHolidays.objects.filter(holiday_status=filter_status,end_date__month=month)
                    if len(holidays):
                        temp['month_holidays'] = holidays
                        month_holidays.append(temp)


        if int(filter_status) < 0:
            context['total_holidays'] = SpHolidays.objects.all().count()
        else:
            context['total_holidays'] = SpHolidays.objects.filter(holiday_status=filter_status).count()
        context['holidays'] = month_holidays       
        template = 'holiday/ajax-holiday-filter.html'
        return render(request, template, context)


@login_required
def holidayCalendar(request,holiday_id):
    context = {}
    calendarObj = calendar.Calendar()
    current_date = date.today() 

    if 'year' in request.GET and int(request.GET.get('year')) != "":
        context['current_year'] = current_year = int(request.GET.get('year'))
    else:
        context['current_year'] = current_year = current_date.year
    
    if 'month' in request.GET and int(request.GET.get('month')) != "":
        context['current_month'] = current_month = int(request.GET.get('month'))
    else:
        context['current_month'] = current_month = current_date.month

    if int(current_month) == 12:
        context['next_month'] = 1
    else:
        context['next_month'] = int(current_month) + 1
    if int(current_month) == 1:
        context['previous_month'] = 12
    else:
        context['previous_month'] = int(current_month) - 1
    
    context['current_month_name'] = calendar.month_name[current_month]

    last_holiday = SpHolidays.objects.raw(''' SELECT * FROM sp_holidays WHERE id = %s ''',[holiday_id])[0]
    holiday_dates = []

    start_date = last_holiday.start_date
    delta = last_holiday.end_date - last_holiday.start_date
    for i in range(delta.days + 1):
        holiday_date = start_date + timedelta(days=i)
        holiday_date = holiday_date.strftime('%Y-%m-%d')
        holiday_dates.append(holiday_date)

    calendarObj = calendar.Calendar()
    calendar_dates = []
    for week in calendarObj.monthdatescalendar(current_year, current_month):
        dates = []
        for week_date in week:
            calendar_datass = []
            tmp = {}
            tmp['full_date'] = week_date
            tmp['day'] = week_date.strftime('%A')
            tmp['short_date'] = week_date.strftime('%d')
            tmp['month'] = week_date.strftime('%m')
                                
            if str(week_date) in holiday_dates:
                    tmp['is_holiday'] = 1
                    tmp['holiday'] = last_holiday.holiday
            else:
                tmp['is_holiday'] = 0
            
            dates.append(tmp)               
        calendar_dates.append(dates)

    context['calendar_dates'] = calendar_dates
    context['holiday'] = last_holiday
    template = 'holiday/holiday-calendar.html'
    return render(request, template, context)

@login_required
def addHoliday(request):
    if request.method == "POST":
        context = {}
        holiday_dates = []

        holiday = SpHolidays()
        holiday.holiday = clean_data(request.POST['holidayName'])
        holiday.holiday_type_id = request.POST['holiday_type_id']
        holiday.holiday_type = getModelColumnById(SpHolidayTypes,request.POST['holiday_type_id'],'holiday_type')
        holiday.start_time = "00:00:00" 
        holiday.start_date = datetime.strptime(clean_data(request.POST['from_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
        holiday.end_date = datetime.strptime(clean_data(request.POST['to_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
        holiday.end_time = "00:00:00" 
        holiday.description = clean_data(request.POST['holiday_description'])
        holiday.status = 1 
        holiday.holiday_status = 1               
        holiday.save()

        if holiday.id:
            
            user_name   = getUserName(request.user.id)
            sendFocNotificationToUsers(holiday.id,'', 'add', 0, request.user.id, user_name, 'SpHolidays',request.user.role_id)
    
            heading     = 'New holiday has been created.'
            activity    = 'New holiday has been created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Leave Management', 'Leave Request', heading, activity, request.user.id, user_name, 'add.png', '2', 'app.png')

            context['flag'] = True
            context['message'] = "Record has been save successfully."
        else:
            context['flag'] = False
            context['message'] = "Failed to save record."

        return JsonResponse(context)
        
    else:
        context = {}
        context['holiday_types'] = SpHolidayTypes.objects.filter(status=1)
        template = 'holiday/add-holiday.html'
        return render(request, template, context)
    


@login_required
def editHoliday(request,holiday_id):
    if request.method == "POST":
        context = {}
        holiday_name       = clean_data(request.POST['holidayName'])
        holiday_id       = request.POST['holiday_id']
        if SpHolidays.objects.filter(holiday=holiday_name).exclude(id=holiday_id).exists():
            context['flag'] = False
            context['message'] = "Holiday already exists."
        else:
            holiday = SpHolidays.objects.get(id=holiday_id)
            holiday.holiday = clean_data(request.POST['holidayName'])
            holiday.holiday_type_id = request.POST['holiday_type_id']
            holiday.holiday_type = getModelColumnById(SpHolidayTypes,request.POST['holiday_type_id'],'holiday_type')
            holiday.start_time = "00:00:00" 
            holiday.start_date = datetime.strptime(clean_data(request.POST['from_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
            holiday.end_date = datetime.strptime(clean_data(request.POST['to_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
            holiday.end_time = "00:00:00" 
            holiday.description = clean_data(request.POST['holiday_description'])
            holiday.status = 1 
            holiday.holiday_status = 1 
            holiday.save()

            if holiday.id:
                context['flag'] = True
                context['message'] = "Record has been updated successfully."

            else:
                context['flag'] = False
                context['message'] = "Failed to update record."

        return JsonResponse(context)
    else:
        context = {}
        context['holiday'] = holiday = SpHolidays.objects.get(id=holiday_id)
        context['holiday_types'] = SpHolidayTypes.objects.filter(status=1)
        template = 'holiday/edit-holiday.html'
        return render(request, template, context)



@login_required
def updateHolidayStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpHolidays.objects.get(id=id)
            data.status = is_active
            data.save()
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)


def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None

@login_required
def exportToPDF(request, filter_status):
    if int(filter_status) < 0:
        holidays = SpHolidays.objects.all().values().order_by('-id')
    else:
        holidays = SpHolidays.objects.filter(holiday_status=filter_status).values().order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('holiday/holiday_pdf_template.html', {'holidays': holidays, 'url': baseurl})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'holidays.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

@login_required
def exportToXlsx(request, filter_status):

    if int(filter_status) < 0:
        holidays = SpHolidays.objects.all().order_by('-id')
    else:
        holidays = SpHolidays.objects.filter(holiday_status=filter_status).order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=holidays.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Holidays'
    
    # Define the titles for columns
    columns = []
    columns += [ 'Holiday' ]
    columns += [ 'Type' ]
    columns += [ 'Duration' ] 
    columns += [ 'Description' ]
    
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for holiday in holidays:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        row += [ holiday.holiday ]
        row += [ holiday.holiday_type ]
        duration = holiday.start_date.strftime('%d %b') +' '+holiday.start_date.strftime('%a')+' - '+holiday.end_date.strftime('%d %b') +' '+holiday.end_date.strftime('%a')+', '+holiday.end_date.strftime('%Y')
        row += [ duration ]
        row += [ strip_tags(holiday.description) ] 
       
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response