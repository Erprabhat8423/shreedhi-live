import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password,check_password
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
import time
import math 
from datetime import datetime, date
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage

# Create your views here.

@login_required
def attendanceReport(request):
    
    context = {}
    today_attendance_users = SpUserAttendance.objects.raw(''' SELECT id,user_id FROM sp_user_attendance WHERE date(attendance_date_time) = CURDATE()
     group by user_id ''')
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')
    users = []
    if len(today_attendance_users):
        for user in today_attendance_users:
            temp = {}
            start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
            sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id LIMIT 1 ''',[user.user_id])
            end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
            sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id])

            temp['name'] = start_attendance[0].name
            temp['emp_sap_id'] = start_attendance[0].emp_sap_id
            temp['profile_image'] = start_attendance[0].profile_image
            temp['dis_sap_id'] = start_attendance[0].dis_sap_id
            temp['store_name'] = start_attendance[0].store_name
            temp['store_image'] = start_attendance[0].store_image
            temp['start_time'] = start_attendance[0].start_time
            temp['latitude'] = start_attendance[0].latitude
            temp['longitude'] = start_attendance[0].longitude
            if end_attendance:
                temp['end_time'] = end_attendance[0].end_time
            else:
                temp['end_time'] = None

            now = datetime.now().strftime('%Y-%m-%d')
            start_datetime = now + ' '+start_attendance[0].start_time
            if temp['end_time'] is None:
                end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                end_datetime = now + ' '+end_attendance[0].end_time
            
            start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
            end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
            time_delta = (end_datetime - start_datetime)
            total_seconds = time_delta.total_seconds()

            hours = math.floor(total_seconds / 3600)
            mins = math.floor((total_seconds - (hours * 3600)) / 60)

            temp['working_hours'] = str(hours)+'.'+str(mins)

            users.append(temp)

    context['users'] = users
    template = 'reports/attendance-report.html'
    return render(request, template, context)

@login_required
def ajaxAttendanceReport(request):
    if 'attendance_date' in request.GET and request.GET['attendance_date'] != "" :
            today                   = request.GET['attendance_date']
            today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()
    
    context = {}

    today_attendance_users = SpUserAttendance.objects.raw(''' SELECT id,user_id FROM sp_user_attendance WHERE date(attendance_date_time) = %s
     group by user_id ''',[today])

    users = []
    if len(today_attendance_users):
        for user in today_attendance_users:
            temp = {}
            print(user.user_id)
            start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
            sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s 
            order by sp_user_attendance.id LIMIT 1 ''',[user.user_id,today])
            end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
            sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s 
            order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id,today])

            temp['name'] = start_attendance[0].name
            temp['emp_sap_id'] = start_attendance[0].emp_sap_id
            temp['profile_image'] = start_attendance[0].profile_image
            temp['dis_sap_id'] = start_attendance[0].dis_sap_id
            temp['store_name'] = start_attendance[0].store_name
            temp['store_image'] = start_attendance[0].store_image
            temp['start_time'] = start_attendance[0].start_time
            temp['latitude'] = start_attendance[0].latitude
            temp['longitude'] = start_attendance[0].longitude
            if end_attendance:
                temp['end_time'] = end_attendance[0].end_time
            else:
                temp['end_time'] = None

            now = today
            start_datetime = now + ' '+start_attendance[0].start_time
            if temp['end_time'] is None:
                end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                end_datetime = now + ' '+end_attendance[0].end_time
            
            start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
            end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
            time_delta = (end_datetime - start_datetime)
            total_seconds = time_delta.total_seconds()
            hours = (total_seconds/60)/60

            temp['working_hours'] = hours

        
            users.append(temp)


    context['users'] = users
    template = 'reports/ajax-attendance-report.html'
    return render(request, template, context)


@login_required
def leaveReport(request):
    today = date.today()
    
    if request.user.role_id == 0:
        leaveReport = SpUserLeaves.objects.all().order_by('-id')

    else:
        leaveReport = SpUserLeaves.objects.raw('''SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
    left join sp_users on sp_users.id = sp_user_leaves.user_id 
    where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' order by id desc ''',[request.user.id])
    

    user_type = SpPermissionWorkflowRoles.objects.filter(sub_module_id=38,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()  
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context = {}

    user_ids = SpUserLeaves.objects.all().distinct().values_list('user_id',flat=True)
    users = SpUsers.objects.filter(id__in=user_ids).values('id','first_name','middle_name','last_name','emp_sap_id')

    context['leaveReport'] = leaveReport
    context['role_id'] = request.user.role_id
    context['level_id'] = level_id
    context['today_date'] = today.strftime("%d/%m/%Y")
    context['page_title'] = "Manage Leaves"
    context['users'] = users

    template = 'reports/attendance-report/leave-report.html'
    return render(request, template, context)


# ajax order list
@login_required
def ajaxLeaveReportLists(request):
    context = {}
    today = date.today()
    user_id = request.GET['user_id']
    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            context['leave_status'] = leave_status
            leaveReport = leaveReport.filter(leave_status=leave_status).order_by('-id')

        # if 'leave_from_date' in request.GET and request.GET['leave_from_date'] and request.GET['leave_to_date']:
        #     leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
        #     leave_to_date = datetime.strptime(request.GET['leave_to_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
            
        #     leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        # if request.GET['leave_from_date']:
        #     if request.GET['leave_to_date'] in today.strftime("%d/%m/%Y"):
        #         leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
                
        #         leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        if user_id:
                leaveReport = leaveReport.filter(user_id=user_id).order_by('-id')
    else:
        condition = ''
        
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status
        if user_id:
            condition += ' and sp_user_leaves.user_id = "%s"' % user_id
            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)

        
    

    user_type = SpRoleWorkflowPermissions.objects.filter(sub_module_id=8, permission_slug='add',
                                                             workflow_level_role_id=request.user.role_id).exclude(
            role_id=request.user.role_id).values('level_id').order_by('-id').first()
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context['leaveReport'] = leaveReport
    context['level_id'] = level_id

    context['role_id'] = request.user.role_id
    template = 'reports/attendance-report/ajax-leave-report-lists.html'
    return render(request, template, context)


@login_required
def leaveStatusDetails(request):
    leave_id = request.GET.get('leave_id')
    initiate_leave_details = SpUserLeaves.objects.get(id=leave_id)
    leave_details = SpApprovalStatus.objects.filter(row_id=leave_id, model_name='SpUserLeaves', status=1).values(
        'final_status_user_id').distinct().values('final_status_user_name', 'final_update_date_time', 'level_id')

    context = {}
    context['initiate_leave_details'] = initiate_leave_details
    context['leave_details'] = leave_details
    template = 'reports/attendance-report/leave-status-details.html'

    return render(request, template, context)


@login_required
def leaveExportToXlsx(request, columns, userId,leave_status):
    column_list = columns.split(",")

    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and user_id = "%s"' % userId
        
        leaveReport = SpUserLeaves.objects.raw("""SELECT * FROM sp_user_leaves WHERE 1 {condition}  order by id desc """.format(condition=condition))
    else:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and sp_user_leaves.user_id = "%s"' % userId

            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)

        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=leave-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030")
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Leave-reports'

    # Define the titles for columns
    columns = []

    if 'user_name' in column_list:
        columns += ['Name']

    if 'leave_from_date' in column_list:
        columns += ['Leave Apply From Date']

    if 'leave_to_date' in column_list:
        columns += ['Leave Apply To Date']

    if 'status' in column_list:
        columns += ['Status']

    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    for results in leaveReport:
        row_num += 1
        # Define the data for each cell in the row
        row = []
        if 'user_name' in column_list:
            row += [results.user_name]

        if 'leave_from_date' in column_list:
            if None in [results.leave_from_date]:
                leave_from_date = ['']
            else:
                leave_from_date = [results.leave_from_date]
            row += leave_from_date

        if 'leave_to_date' in column_list:
            if None in [results.leave_to_date]:
                leave_to_date = ['']
            else:
                leave_to_date = [results.leave_to_date]
            row += leave_to_date

        if 'status' in column_list:
            if results.leave_status == 1:
                status = 'Pending'
                row += [status]
            elif results.leave_status == 2:
                status = 'Forwarded'
                row += [status]
            elif results.leave_status == 3:
                status = 'Approved'
                row += [status]
            elif results.leave_status == 4:
                status = 'Declined'
                row += [status]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
    workbook.save(response)

    return response


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


# Automaticly downloads to PDF file
@login_required
def leaveExportToPDF(request, columns, userId,leave_status):
    column_list = columns.split(",")

    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and user_id = "%s"' % userId
        
        leaveReport = SpUserLeaves.objects.raw("""SELECT * FROM sp_user_leaves WHERE 1 {condition}  order by id desc """.format(condition=condition))
    else:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and sp_user_leaves.user_id = "%s"' % userId

            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)
    
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('reports/attendance-report/leave_pdf_template.html',
                        {'leaveReport': leaveReport, 'url': baseurl, 'columns': column_list,
                         'columns_length': len(column_list)})
    
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'leave-report.pdf'
    content = "attachment; filename=%s" % filename
    response['Content-Disposition'] = content
    return response


@login_required
def editLeaveStatus(request):
    id = request.GET.get('id')
    context = {}
    context['leaveData'] = SpUserLeaves.objects.get(id=id)
    template = 'reports/attendance-report/edit-leave-status.html'
    return render(request, template, context)

@login_required
def updateLeaveRemark(request): 
    context = {}
    context['level_id']     = request.GET.get('level_id')
    context['leave_status'] = request.GET.get('leave_status')
    template = 'reports/attendance-report/update-leave-remark.html'
    return render(request, template, context)

#update order status
@login_required
def updateLeaveStatus(request):
    response = {}
    leave_id        = request.POST.getlist('leave_id[]')
    level_id        = request.POST['level_id']
    leave_status    = request.POST['leave_status']
    if request.user.role_id == 0:
        for leave in leave_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id)
            if approvals_request:
                for approval in approvals_request:
                    approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                    approval_data.level_id                  = leave_status
                    if leave_status == '2':
                        approval_data.level                    = 'Forward'
                    elif leave_status == '3':
                        approval_data.level                    = 'Approve'
                    elif leave_status == '4':
                        approval_data.level                    = 'Declined'         
                    approval_data.status                    = 1
                    approval_data.final_status_user_id      = request.user.id
                    approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    approval_data.save()

                user_level_approval_count = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id, status=0).count()
                if user_level_approval_count == 0:
                    leave                   = SpUserLeaves.objects.get(id=leave)   
                    leave.leave_status      = leave_status
                    if request.POST['remark']:
                        leave.remark      = request.POST['remark']
                    leave.save()
            else:
                leave                   = SpUserLeaves.objects.get(id=leave)   
                leave.leave_status      = leave_status
                if request.POST['remark']:
                    leave.remark      = request.POST['remark']
                leave.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                leave.save()
                today   = date.today()
                 
    
    else:    
        for leave in leave_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', role_id=request.user.role_id, level_id=level_id)
            for approval in approvals_request:
                approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                approval_data.status                    = 1
                approval_data.final_status_user_id      = request.user.id
                approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                approval_data.save()

            user_level_approval_count = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id, status=0).count()
            if user_level_approval_count == 0:
                leave                   = SpUserLeaves.objects.get(id=leave)   
                leave.leave_status      = leave_status
                if request.POST['remark']:
                    leave.remark      = request.POST['remark']
                leave.save()   

    
    if leave_status == '2':
        for leave in leave_id:
            approvals_requests = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', status=0)
            if approvals_requests:
                for approval in approvals_requests:
                    notification                        = SpUserNotifications()
                    notification.row_id                 = approval.row_id
                    notification.user_id                = approval.user_id
                    notification.model_name             = 'SpUserLeaves'
                    notification.notification           = 'leave '+approval.level+' Request has been sent.'
                    notification.is_read                = 0
                    notification.created_by_user_id     = request.user.id
                    notification.created_by_user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    notification.save()

    if leave_status == '2':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been forwarded'
            activity    = 'Leave Request has been forwarded by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'forwaord.png', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)
            
            message_title = "Leave request forwarded"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been forwarded  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request forwarded',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------# 
    elif leave_status == '3':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been approved'
            activity    = 'Leave Request has been approved by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'approved.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            #update user leave count
            user_basic_details              = SpBasicDetails.objects.get(user_id=user_id)
            leave_count                     = int(user_basic_details.leave_count)-1   
            user_basic_details.leave_count  = leave_count
            user_basic_details.save()

            message_title = "Leave request approved"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been approved  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------#    

    elif leave_status == '4':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been declined'
            activity    = 'Leave Request has been declined by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'declined.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            message_title = "Leave request declined"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been declined  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request declined',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------#

    response['error'] = False
    response['message'] = "Leave status has been updated successfully."
    return JsonResponse(response)
    
@login_required
def employeeReport(request):
    
    context = {}
    today = date.today()
    users = SpUsers.objects.raw(''' SELECT id, first_name, middle_name, last_name, role_name, is_tagged, tagged_by, tagged_date, created_by FROM sp_users WHERE id!=%s and user_type=%s and (role_id=%s or role_id=%s)  ''', [1, 1, 4, 5])
    for user in users:
        name = user.first_name
        if user.middle_name:
            name += ' '+user.middle_name
        if user.last_name:
            name += ' '+user.last_name    
        user.name = name
        user.added_employee_count = SpUsers.objects.filter(created_by=user.id, created_at__icontains=today.strftime("%Y-%m-%d")).count()
        user.tagged_employee_count = SpUsers.objects.filter(tagged_by=user.id, tagged_date__icontains=today.strftime("%Y-%m-%d")).count()
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')

    context['users'] = users
    context['page_title'] = "Employee Tagging Report"
    template = 'reports/attendance-report/employee-report.html'
    return render(request, template, context)


@login_required
def ajaxEmployeeReport(request):
    if 'search_date' in request.GET and request.GET['search_date'] != "" :
        today                   = request.GET['search_date']
        today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()
        today                   = today.strftime("%Y-%m-%d")
        
    context = {}
    users = SpUsers.objects.raw(''' SELECT id, first_name, middle_name, last_name, role_name, is_tagged, tagged_by, tagged_date, created_by FROM sp_users WHERE id!=%s and user_type=%s and (role_id=%s or role_id=%s)  ''', [1, 1, 4, 5])
    for user in users:
        name = user.first_name
        if user.middle_name:
            name += ' '+user.middle_name
        if user.last_name:
            name += ' '+user.last_name    
        user.name = name
        user.added_employee_count = SpUsers.objects.filter(created_by=user.id, created_at__icontains=today).count()
        user.tagged_employee_count = SpUsers.objects.filter(tagged_by=user.id, tagged_date__icontains=today).count()
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')

    context['users'] = users
    template = 'reports/attendance-report/ajax-employee-report.html'
    return render(request, template, context)

@login_required
def exportEmployeeReport(request, columns, search_date):
    column_list = columns.split (",")
    if search_date != "" :
        today                   = search_date
    else:
        today                   = date.today()
        today                   = today.strftime("%Y-%m-%d")
        
    context = {}
    users = SpUsers.objects.raw(''' SELECT id, first_name, middle_name, last_name, role_name, is_tagged, tagged_by, tagged_date, created_by FROM sp_users WHERE id!=%s and user_type=%s and (role_id=%s or role_id=%s)  ''', [1, 1, 4, 5])
    for user in users:
        name = user.first_name
        if user.middle_name:
            name += ' '+user.middle_name
        if user.last_name:
            name += ' '+user.last_name    
        user.name = name
        user.added_employee_count = SpUsers.objects.filter(created_by=user.id, created_at__icontains=today).count()
        user.tagged_employee_count = SpUsers.objects.filter(tagged_by=user.id, tagged_date__icontains=today).count()
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')

    users = users
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employee-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Employee Report'
    
    # Define the titles for columns
    columns = []

    if 'employee_name' in column_list:
        columns += [ 'Employee Name' ]

    if 'added_employee' in column_list:
        columns += [ 'Added Employee' ]
    
    if 'tagged_employee' in column_list:
        columns += [ 'Tagged Employee' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        if 'employee_name' in column_list:
            row += [ user.name ]

        if 'added_employee' in column_list:
            row += [ user.added_employee_count ]
        
        if 'tagged_employee' in column_list:
            row += [ user.tagged_employee_count ]  
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 
    
@login_required
def incentiveReport(request):
    context = {}
    import datetime
    today = datetime.date.today()
    first = today.replace(day=1)
    lastMonth = first - datetime.timedelta(days=1)
    date = lastMonth.strftime('%Y-%m')
    superstockists = SpUserIncentive.objects.filter(payment_cycle=2,created_at__icontains=date).all()
    user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=2,created_at__icontains=date).values_list('id',flat=True)
    superstockists_distinct_slab = SpUserIncentiveDetails.objects.filter(user_incentive_id__in=user_incentive_ids).all()
    for superstockist in superstockists:
        ss_details = SpUsers.objects.get(id=superstockist.user_id)
        superstockist.emp_sap_id = ss_details.emp_sap_id
        superstockist.name = ss_details.first_name+" "+ss_details.middle_name+" "+ss_details.last_name
        superstockist.store_name = ss_details.store_name
        
        slab_amount_list = []
        for slab in superstockists_distinct_slab:
            key = "slab_"+str(slab.master_slab_id)
            slab_incentive = SpUserIncentiveDetails.objects.filter(user_incentive_id=superstockist.id,master_slab_id=slab.master_slab_id).first()
            if slab_incentive:
                slab_amount = slab_incentive.slab_amount
            else:
                slab_amount = 0
            slab_amount_list.append(slab_amount)
        superstockist.slab_amount_list = slab_amount_list
               
    for slab in superstockists_distinct_slab:
        product_class_id = getModelColumnById(SpSlabMasterList,slab.master_slab_id,'product_class_id')
        slab.slab_detail = getModelColumnById(SpProductClass,product_class_id,'product_class')+" ( "+ str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'more_than_quantity'))+" - "+str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'upto_quantity'))+" )"
        
        
    context['current_date'] = datetime.date.today().strftime('%m/%Y')
    context['superstockists'] = superstockists
    context['superstockists_distinct_slab'] = superstockists_distinct_slab
    context['page_title'] = "Superstockist Incentive Report"
    template = 'reports/attendance-report/incentive-report.html'
    return render(request, template, context)


@login_required
def ajaxIncentiveReport(request):
    import datetime
    if 'search_date' in request.GET and request.GET['search_date'] != "" :
        today                   = request.GET['search_date']
        date                   = datetime.datetime.strptime(str(today), '%m/%Y').strftime('%Y-%m')
    else:
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        
    context = {}
    if 'fortnight' in request.GET and request.GET['fortnight'] == "" :
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        date = lastMonth.strftime('%Y-%m')
    
    if 'fortnight' in request.GET and request.GET['fortnight'] != "" :
        if int(datetime.datetime.strptime(str(today), '%m/%Y').strftime('%m')) == 2:
            mid_date=14
        else:
            mid_date=15
        
        if request.GET['fortnight'] == "1":
            start_date = str(date)+"-01"
            end_date = str(date)+"-"+str(mid_date)
            date =  str(date)+"-"+str(mid_date)
            
        elif request.GET['fortnight'] == "2":
            start_date = str(date)+"-"+str(mid_date+1)
            end_date = str(date)+"-"+str(mid_date)
            last_date = len(days_in_months(int(datetime.datetime.strptime(str(today), '%m/%Y').strftime('%Y')), int(datetime.datetime.strptime(str(today), '%m/%Y').strftime('%m'))))
            date =  str(date)+"-"+str(last_date)
            
        payment_cycle = 1
    else:
        payment_cycle = 2
        
        # superstockists = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).all()
        # user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).values_list('id',flat=True)
        
    superstockists = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).all()
    user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).values_list('id',flat=True)
    
    superstockists_distinct_slab = SpUserIncentiveDetails.objects.filter(user_incentive_id__in=user_incentive_ids).all()
        
    for superstockist in superstockists:
        ss_details = SpUsers.objects.get(id=superstockist.user_id)
        superstockist.emp_sap_id = ss_details.emp_sap_id
        superstockist.name = ss_details.first_name+" "+ss_details.middle_name+" "+ss_details.last_name
        superstockist.store_name = ss_details.store_name
        
        slab_amount_list = []
        for slab in superstockists_distinct_slab:
            key = "slab_"+str(slab.master_slab_id)
            slab_incentive = SpUserIncentiveDetails.objects.filter(user_incentive_id=superstockist.id,master_slab_id=slab.master_slab_id).first()
            if slab_incentive:
                slab_amount = slab_incentive.slab_amount
            else:
                slab_amount = 0
            slab_amount_list.append(slab_amount)
        superstockist.slab_amount_list = slab_amount_list
               
    for slab in superstockists_distinct_slab:
        product_class_id = getModelColumnById(SpSlabMasterList,slab.master_slab_id,'product_class_id')
        slab.slab_detail = getModelColumnById(SpProductClass,product_class_id,'product_class')+" ( "+ str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'more_than_quantity'))+" - "+str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'upto_quantity'))+" )"
        
        
    context['current_date'] = datetime.date.today().strftime('%m/%Y')
    context['superstockists'] = superstockists
    context['superstockists_distinct_slab'] = superstockists_distinct_slab
    template = 'reports/attendance-report/ajax-incentive-report.html'
    return render(request, template, context)


@login_required
def exportIncentiveReport(request, fortnight, search_date):
    
    import datetime
    if search_date != "" :
        today                   = search_date
        date                   = today
    else:
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        
    context = {}
    if fortnight == "NA" :
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        date = lastMonth.strftime('%Y-%m')
    
    if fortnight != "NA" :
        if int(datetime.datetime.strptime(str(today), '%Y-%m').strftime('%m')) == 2:
            mid_date=14
        else:
            mid_date=15
        
        if fortnight == "1":
            start_date = str(date)+"-01"
            end_date = str(date)+"-"+str(mid_date)
            date =  str(date)+"-"+str(mid_date)
            
        elif fortnight == "2":
            start_date = str(date)+"-"+str(mid_date+1)
            end_date = str(date)+"-"+str(mid_date)
            last_date = len(days_in_months(int(datetime.datetime.strptime(str(today), '%Y-%m').strftime('%Y')), int(datetime.datetime.strptime(str(today), '%Y-%m').strftime('%m'))))
            date =  str(date)+"-"+str(last_date)
            
        payment_cycle = 1
    else:
        payment_cycle = 2
        
        # superstockists = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).all()
        # user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).values_list('id',flat=True)
        
    superstockists = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).all()
    user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).values_list('id',flat=True)
    
    superstockists_distinct_slab = SpUserIncentiveDetails.objects.filter(user_incentive_id__in=user_incentive_ids).all()
        
    for superstockist in superstockists:
        ss_details = SpUsers.objects.get(id=superstockist.user_id)
        superstockist.emp_sap_id = ss_details.emp_sap_id
        superstockist.name = ss_details.first_name+" "+ss_details.middle_name+" "+ss_details.last_name
        superstockist.store_name = ss_details.store_name
        
        slab_amount_list = []
        for slab in superstockists_distinct_slab:
            key = "slab_"+str(slab.master_slab_id)
            slab_incentive = SpUserIncentiveDetails.objects.filter(user_incentive_id=superstockist.id,master_slab_id=slab.master_slab_id).first()
            if slab_incentive:
                slab_amount = slab_incentive.slab_amount
            else:
                slab_amount = 0
            slab_amount_list.append(slab_amount)
        superstockist.slab_amount_list = slab_amount_list
               
    for slab in superstockists_distinct_slab:
        product_class_id = getModelColumnById(SpSlabMasterList,slab.master_slab_id,'product_class_id')
        slab.slab_detail = getModelColumnById(SpProductClass,product_class_id,'product_class')+" ( "+ str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'more_than_quantity'))+" - "+str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'upto_quantity'))+" )"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=incentive-report.xlsx'.format(
        date=date
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Incentive Report'
    
    # Define the titles for columns
    columns = []
    
    
    columns += [ 'Sap Code' ]
    columns += [ 'Distributor Name' ]
    # columns += [ 'SS Incentive' ]
    columns += [ 'Distributor Incentive' ]
    # columns += [ 'Primary  TPT Amount' ]
    # columns += [ 'Distributor TPT Amount' ]
    for slab in superstockists_distinct_slab:
        columns += [ slab.slab_detail ]
    columns += [ 'NET Amount' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for superstockist in superstockists:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        
        row += [ superstockist.emp_sap_id ]
        row += [ superstockist.name+" ( "+superstockist.store_name+" )" ]
        # row += [ superstockist.ss_incentive ]
        row += [ superstockist.distributor_incentive ]
        # row += [ superstockist.primary_transporter_amount ]
        #row += [ superstockist.secondary_transporter_amount ]
        for slab_amount in superstockist.slab_amount_list:
            row += [ slab_amount ]
        row += [ superstockist.net_amount ]
        
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response        

    
