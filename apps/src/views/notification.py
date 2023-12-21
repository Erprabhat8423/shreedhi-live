import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password,check_password
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
import time
import math 
from datetime import datetime, date
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from ..decorators import has_par


@login_required
@has_par(sub_module_id=40,permission='list')
def index(request):
    today   = date.today()
    context = {}
    notifications = SpNotifications.objects.all().order_by('-id').filter(created_at__icontains=today.strftime("%Y-%m-%d"))
    for notification in notifications:
        notification.recipient_role = getModelColumnById(SpUsers,notification.to_user_id,'role_name')

    last_notification = SpNotifications.objects.all().order_by('-id').filter(created_at__icontains=today.strftime("%Y-%m-%d")).first()

    context['notifications']     = notifications
    context['last_notification']     = last_notification
    context['page_title']        = "Bulk Notifications"
    context['roles']       = SpRoles.objects.filter(status=1)
    template = 'notification/index.html'
    return render(request, template, context)
@login_required
def getNotificationDetails(request,notification_id):
    last_notification = SpNotifications.objects.get(id=notification_id)
    context = {}
    context['last_notification']     = last_notification
    template = 'notification/notification-details.html'
    return render(request, template, context)

@login_required
@has_par(sub_module_id=40,permission='list')
def ajaxNotificationReport(request):
    condition = ''

    if 'filter_role' in request.POST and request.POST['filter_role'] != "" :
        condition += ' AND to_user_id in (SELECT id FROM sp_users WHERE role_id='+str(request.POST['filter_role'])+')'

    if 'filter_notification_type' in request.POST and request.POST['filter_notification_type'] != "" :
        condition += ' AND notification_type='+request.POST['filter_notification_type']
    
    if 'filter_keyword' in request.POST and request.POST['filter_keyword'] != "" :
        condition += ' AND to_user_name LIKE "%%'+request.POST['filter_keyword']+'%%"' 

    filter_from_dateArr = request.POST['filter_from_date'].split('/')
    filter_from_date = str(filter_from_dateArr[2])+'-'+str(filter_from_dateArr[1])+'-'+str(filter_from_dateArr[0])

    filter_to_dateArr = request.POST['filter_to_date'].split('/')
    filter_to_date = str(filter_to_dateArr[2])+'-'+str(filter_to_dateArr[1])+'-'+str(filter_to_dateArr[0])

    condition += ' AND date(created_at) BETWEEN "'+ filter_from_date +'" AND "'+filter_to_date+'"'

    context = {}
    notifications = SpNotifications.objects.raw(''' SELECT * FROM sp_notifications WHERE 1 {condition} ORDER BY id DESC '''.format(condition=condition) )
    for notification in notifications:
        notification.recipient_role = getModelColumnById(SpUsers,notification.to_user_id,'role_name')

    context['notifications']     = notifications
    template = 'notification/ajax-notification-report.html'
    return render(request, template, context)


@login_required
def sendNotification(request):
    response = {}
    if request.method == "POST":
        contact_no_list = []
        emails = []

        activity_image = None
        user_name   = str(request.user.first_name)
        if request.user.middle_name:
            user_name += ' '+str(request.user.middle_name)
        if request.user.last_name:
            user_name += ' '+str(request.user.last_name)

        
        if bool(request.FILES.get('activity_image', False)) == True:
            uploaded_activity_image = request.FILES['activity_image']
            storage = FileSystemStorage()
            timestamp = int(time.time())
            activity_image_name = uploaded_activity_image.name
            temp = activity_image_name.split('.')
            activity_image_name = 'notification_'+str(timestamp)+"."+temp[(len(temp) - 1)]
            
            activity_image = storage.save(activity_image_name, uploaded_activity_image)
            activity_image = storage.url(activity_image)

            # create thumbnail
            image_file = str(settings.MEDIA_ROOT) + '/'+activity_image_name
            size = 300, 300
            im = Image.open(image_file)
            
            if im.mode in ("RGBA", "P"):
                im = im.convert("RGB")

            im.thumbnail(size, Image.ANTIALIAS)
            thumbnail_image = str(settings.MEDIA_ROOT) + '/notification_images/'+activity_image_name
            im.save(thumbnail_image, "JPEG",quality=100, rotated=False)

            activity_image = 'media/notification_images/'+activity_image_name 
            if os.path.isfile(image_file):
                os.remove(image_file)
            
            notification_image = baseurl+'/'+activity_image
        else:
            notification_image = ""

        if request.POST['user_type'] == '0':
            customers = request.POST.getlist('customer_id[]')
            for id, customer in enumerate(customers):
                print(customers[id])
                customer_name = getModelColumnById(SpUsers, customers[id], 'first_name')
                if getModelColumnById(SpUsers, customers[id], 'middle_name'):
                    customer_name += ' '+getModelColumnById(SpUsers, customers[id], 'middle_name')
                if getModelColumnById(SpUsers, customers[id], 'last_name'):
                    customer_name += ' '+getModelColumnById(SpUsers, customers[id], 'last_name')
                
                if getModelColumnById(SpUsers, customers[id], 'primary_contact_number'):
                    contact_no_list.append(getModelColumnById(SpUsers, customers[id], 'primary_contact_number'))
                
                notification                        = SpNotifications()
                notification.module                 = 'MIS Reports'
                notification.sub_module             = 'Notification'
                notification.heading                = request.POST['heading']
                notification.activity               = request.POST['message']
                notification.activity_image         = activity_image
                notification.from_user_id           = request.user.id
                notification.from_user_name         = user_name
                notification.to_user_id             = customers[id]
                notification.to_user_name           = customer_name
                notification.platform               = 1
                notification.platform_icon          = 'web.png'
                notification.read_status            = 1
                notification.notification_type      = request.POST['notification_type']
                if request.POST['notification_type'] == '1':
                    notification.icon                   = 'push.png'
                    #-----------------------------notify android block-------------------------------#
                    message_title = request.POST['heading']
                    message_body  = request.POST['message']
                    registration_ids = []
                    if getModelColumnById(SpUsers,customers[id],'firebase_token')  is not None:
                        registration_ids.append(getModelColumnById(SpUsers,customers[id],'firebase_token'))
                        data_message = {}
                        data_message['id']              = 1
                        data_message['status']          = 'notification'
                        data_message['click_action']    = 'FLUTTER_NOTIFICATION_CLICK'
                        data_message['image']           = notification_image
                        send_android_notification(message_title,message_body,data_message,registration_ids)

                elif request.POST['notification_type'] == '2':
                    notification.icon                   = 'email.png'
                    if getModelColumnById(SpUsers,customers[id],'official_email')  is not None:
                        emails.append('sortstringsolutions@gmail.com')
                        # emails.append(getModelColumnById(SpUsers,employees[id],'official_email'))
                        
                else:
                    notification.icon                   = 'sms.png'        
                notification.save()
                
        else:    
            employees = request.POST.getlist('employee_id[]')
            for id, employee in enumerate(employees):
                employee_name = getModelColumnById(SpUsers, employees[id], 'first_name')
                if getModelColumnById(SpUsers, employees[id], 'middle_name'):
                    employee_name += ' '+getModelColumnById(SpUsers, employees[id], 'middle_name')
                if getModelColumnById(SpUsers, employees[id], 'last_name'):
                    employee_name += ' '+getModelColumnById(SpUsers, employees[id], 'last_name')

                if getModelColumnById(SpUsers, employees[id], 'primary_contact_number'):
                    contact_no_list.append(getModelColumnById(SpUsers, employees[id], 'primary_contact_number'))

                notification                        = SpNotifications()
                notification.module                 = 'MIS Reports'
                notification.sub_module             = 'Notification'
                notification.heading                = request.POST['heading']
                notification.activity               = request.POST['message']
                notification.activity_image         = activity_image
                notification.from_user_id           = request.user.id
                notification.from_user_name         = user_name
                notification.to_user_id             = employees[id]
                notification.to_user_name           = employee_name   
                notification.platform               = 1
                notification.platform_icon          = 'web.png'
                notification.read_status            = 1
                notification.notification_type      = request.POST['notification_type']

                if request.POST['notification_type'] == '1':
                    notification.icon                   = 'push.png'
                    #-----------------------------notify android block-------------------------------#
                    message_title = request.POST['heading']
                    message_body  = request.POST['message']
                    registration_ids = []
                    if getModelColumnById(SpUsers,employees[id],'firebase_token')  is not None:
                        registration_ids.append(getModelColumnById(SpUsers,employees[id],'firebase_token'))
                        data_message = {}
                        data_message['id']              = 1
                        data_message['status']          = 'notification'
                        data_message['click_action']    = 'FLUTTER_NOTIFICATION_CLICK'
                        data_message['image']           = notification_image
                        send_android_notification(message_title,message_body,data_message,registration_ids)

                elif request.POST['notification_type'] == '2':
                    notification.icon                   = 'email.png'
                    if getModelColumnById(SpUsers,employees[id],'official_email')  is not None:
                        emails.append('sortstringsolutions@gmail.com')
                        # emails.append(getModelColumnById(SpUsers,employees[id],'official_email'))

                else:
                    notification.icon                   = 'sms.png'

                notification.save()

        separtor = ","
        contact_no = separtor.join(contact_no_list)
        
        if request.POST['notification_type'] == '0' and len(contact_no_list) > 0:
            if 'sender_id' in  request.POST and request.POST['sender_id'] != "" :
                # pass
                sendSMS(request.POST['sender_id'],contact_no,request.POST['message'])
        
        if request.POST['notification_type'] == '2' and len(emails) > 0:
            sendBulkEmail(request.POST['heading'],request.POST['message'],emails,activity_image)

        response['flag'] = True
        response['message'] = "Notification sent successfully."
        return JsonResponse(response)
    else:    
        context = {}
        context['roles']       = SpRoles.objects.filter(status=1).exclude(Q(id=8)|Q(id=9))
        context['routes']      = SpRoutes.objects.filter(status=1)
        context['towns']       =  SpTowns.objects.all()
        sms_sender_ids         = getConfigurationResult('sms_sender_ids')
        if sms_sender_ids is not None:
            context['sms_sender_ids']       =  sms_sender_ids.split(',')
        else:
            context['sms_sender_ids']       =  []

        template = 'notification/send-notification.html'
        return render(request, template, context)

@login_required
def getRouteTownsOptions(request):
    response = {}
    options = '<option value="">Select Locality</option>'
    towns = SpRoutesTown.objects.filter(route_id=request.POST['route_id']).order_by('order_index')
    for town in towns :
        options += "<option value="+str(town.town_id)+">"+town.town_name+"</option>"
    response['options'] = options
    return JsonResponse(response)

@login_required
def getUsersOptions(request):
    response = {}
    if 'type' not in request.POST and request.POST['type'] == "":
        response['flag']         = False
        response['message']  = "Type missing"
        return JsonResponse(response) 
    
    route_ids = request.POST['route_ids']
    town_ids = request.POST['town_ids']
    role_ids = request.POST['role_ids']
    employee_town_ids = request.POST['employee_town_ids']

    options = ''
    count = 0
    if int(request.POST['type']) == 0:
        if route_ids != "" and town_ids == "":
            route_ids = route_ids.split(',')
            user_ids = SpUserAreaAllocations.objects.filter(route_id__in=route_ids).distinct().values_list('user_id',flat=True)
            if len(user_ids):
                users = SpUsers.objects.filter(id__in=user_ids).filter(Q(user_type=2) | Q(user_type=3)).values('id','first_name','middle_name','last_name','emp_sap_id')
                count = len(users)
                for user in users:
                    user_name = user['first_name']+' '
                    if user['middle_name'] is not None:
                         user_name += user['middle_name']+' '
                    if user['last_name'] is not None:
                         user_name += user['last_name']+' '
                    options += "<option value="+str(user['id'])+" selected>"+user_name+"</option>"

        elif town_ids != "" and route_ids != "":
            town_ids = town_ids.split(',')
            user_ids = SpUserAreaAllocations.objects.filter(town_id__in=town_ids).distinct().values_list('user_id',flat=True)
            if len(user_ids):
                users = SpUsers.objects.filter(id__in=user_ids).filter(Q(user_type=2) | Q(user_type=3)).values('id','first_name','middle_name','last_name','emp_sap_id')
                count = len(users)
                for user in users:
                    user_name = user['first_name']+' '
                    if user['middle_name'] is not None:
                         user_name += user['middle_name']+' '
                    if user['last_name'] is not None:
                         user_name += user['last_name']+' '
                    options += "<option value="+str(user['id'])+" selected>"+user_name+"</option>"
    else:

        if role_ids != "" and employee_town_ids != "" :
            employee_town_ids = employee_town_ids.split(',')
            role_ids = role_ids.split(',')

            user_ids = SpUserAreaAllocations.objects.filter(town_id__in=employee_town_ids).distinct().values_list('user_id',flat=True)
            if len(user_ids):
                users = SpUsers.objects.filter(role_id__in=role_ids,user_type=1,id__in=user_ids).filter(Q(user_type=2) | Q(user_type=3)).values('id','first_name','middle_name','last_name','emp_sap_id')
                count = len(users)
                for user in users:
                    user_name = user['first_name']+' '
                    if user['middle_name'] is not None:
                         user_name += user['middle_name']+' '
                    if user['last_name'] is not None:
                         user_name += user['last_name']+' '
                    options += "<option value="+str(user['id'])+" selected>"+user_name+"</option>"

        if role_ids != "" and employee_town_ids == "":
            role_ids = role_ids.split(',')
            users = SpUsers.objects.filter(role_id__in=role_ids,user_type=1).values('id','first_name','middle_name','last_name','emp_sap_id')
            count = len(users)
            for user in users:
                user_name = user['first_name']+' '
                if user['middle_name'] is not None:
                        user_name += user['middle_name']+' '
                if user['last_name'] is not None:
                        user_name += user['last_name']+' '
                options += "<option value="+str(user['id'])+" selected>"+user_name+"</option>"
            

        if employee_town_ids != "" and role_ids == "":
            employee_town_ids = employee_town_ids.split(',')
            user_ids = SpUserAreaAllocations.objects.filter(town_id__in=employee_town_ids).distinct().values_list('user_id',flat=True)
            if len(user_ids):
                users = SpUsers.objects.filter(id__in=user_ids).filter(Q(user_type=2) | Q(user_type=3)).values('id','first_name','middle_name','last_name','emp_sap_id')
                count = len(users)
                for user in users:
                    user_name = user['first_name']+' '
                    if user['middle_name'] is not None:
                         user_name += user['middle_name']+' '
                    if user['last_name'] is not None:
                         user_name += user['last_name']+' '
                    options += "<option value="+str(user['id'])+" selected>"+user_name+"</option>"


    response['options']         = options
    response['customer_count']  = count
    return JsonResponse(response) 

@login_required
@has_par(sub_module_id=40,permission='export')
def exportNotificationReport(request, columns, filter_notification_type,role,from_date,to_date,keyword):
    column_list = columns.split (",")
    condition = ''

    if role != "" and str(role) != "-1" :
        condition += ' AND to_user_id in (SELECT id FROM sp_users WHERE role_id='+str(role)+')'

    if filter_notification_type != "" and int(filter_notification_type) != 4:
        condition += ' AND notification_type='+filter_notification_type
    
    if keyword != "" and keyword != "keyword":
        condition += ' AND to_user_name LIKE "%%'+keyword+'%%"' 

    condition += ' AND date(created_at) BETWEEN "'+ from_date +'" AND "'+to_date+'"'

    context = {}
    notifications = SpNotifications.objects.raw(''' SELECT * FROM sp_notifications WHERE 1 {condition} ORDER BY id DESC '''.format(condition=condition) )

    for notification in notifications:
        notification.recipient_role = getModelColumnById(SpUsers,notification.to_user_id,'role_name')

    notification = notifications

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=notification-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Notification Report'
    
    # Define the titles for columns
    columns = []

    if 'user_name' in column_list:
        columns += [ 'User Name' ]

    if 'notification_heading' in column_list:
        columns += [ 'Heading' ]
    
    if 'notification_message' in column_list:
        columns += [ 'Message' ] 

    if 'notification_type' in column_list:
        columns += [ 'Notification Type' ]
    
    if 'send_at' in column_list:
        columns += [ 'Send At' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for notification in notifications:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        if 'user_name' in column_list:
            row += [ notification.to_user_name+' ('+ notification.recipient_role +')' ]

        if 'notification_heading' in column_list:
            row += [ notification.heading ]
            
        if 'notification_message' in column_list:
            row += [ notification.activity ]   
        
        if 'notification_message' in column_list:
            if notification.notification_type:
                if notification.notification_type == 0:
                    row += [ 'SMS' ]
                else:
                    row += [ 'Push' ]    
            else:
                row += [ '-' ]
        
        if 'send_at' in column_list:
            row += [ notification.created_at ] 

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response    
