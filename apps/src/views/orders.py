import sys
import os
import openpyxl
import collections, functools, operator 
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from ..decorators import has_par
from django.db import transaction



# Create your views here.

# Order View
@login_required
@has_par(sub_module_id=8,permission='list')
def index(request):
    page    = request.GET.get('page')
    today   = date.today()
    if request.user.role_id == 0:
        orders  = SpOrders.objects.all().order_by('-updated_date').filter(order_date__icontains=today.strftime("%Y-%m-%d"))
        for order in orders:
            order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')
    else:
        orders = SpOrders.objects.raw('''SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id
    left join sp_users on sp_users.id = sp_orders.user_id 
    where DATE(sp_orders.order_date) = %s and sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpOrders' group by sp_approval_status.row_id order by updated_date desc ''',[today.strftime("%Y-%m-%d"), request.user.id])
    
    towns       = SpTowns.objects.all()
    routes      = SpRoutes.objects.all()
    user_list   = SpUsers.objects.filter(user_type=2)

    distributor_order_count     = SpOrders.objects.filter(user_type='Distributor').count()
    super_stockist_order_count  = SpOrders.objects.filter(user_type='SuperStockist').count()
    total_orders                = distributor_order_count+super_stockist_order_count
    user_type                   = SpPermissionWorkflowRoles.objects.filter(sub_module_id=8,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()    
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    
    order_ids = []
    user_ids  = []
    for order in orders:
        order_ids.append(order.id)
        user_ids.append(order.user_id)

    if len(user_ids) > 0:
        user_count = SpUsers.objects.filter(status=1, user_type=2).exclude(id__in=user_ids).exclude(emp_sap_id__exact='').count()
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(id__in=user_ids).exclude(emp_sap_id__exact='').values_list('id', flat=True)
    else:
        user_count = SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').count()
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').values_list('id', flat=True) 
    
    total_user_count =  SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').count()
    order_counts = []
    products_list = []
    product_class = SpProductClass.objects.filter(status=1).order_by('-id')
    for product in product_class:
        products = SpProducts.objects.filter(product_class_id=product.id).values_list('id', flat=True)
        if products and len(order_ids)>0:
            quantity  = SpOrderDetails.objects.filter(order_id__in=order_ids, product_id__in=products, order_date__icontains=today.strftime("%Y-%m-%d")).aggregate(Sum('quantity'))['quantity__sum']
            if quantity:
                product.product_count = quantity
            else:
                product.product_count = 0
        else:
            product.product_count = 0

    total_order_quantity = 0
    
    if len(order_ids)>0:
        first_order_id = order_ids[0]
    else:
        first_order_id = 0
          
    context = {}
    context['orders']                           = orders
    context['user_list']                        = user_list
    context['products_list']                    = products_list
    context['product_class']                    = product_class
    context['total_order_quantity']             = total_order_quantity
    context['level_id']                         = level_id
    context['role_id']                          = request.user.role_id
    context['user_count']                       = user_count
    context['total_user_count']                 = total_user_count
    context['user_id']                          = list(user_id)
    context['order_counts']                     = order_counts
    context['total_orders']                     = total_orders
    context['distributor_order_count']          = distributor_order_count
    context['super_stockist_order_count']       = super_stockist_order_count
    context['towns']                            = towns
    context['routes']                           = routes
    context['first_order_id']                   = first_order_id
    context['today_date']                       = today.strftime("%d/%m/%Y")
    context['total_pages']                      = 'all'
    context['page_limit']                       = getConfigurationResult('page_limit')
    context['page_title']                       = "Order Management"

    template = 'order-management/index.html'
    return render(request, template, context)

#search Order
@login_required
@has_par(sub_module_id=8,permission='list')
def searchOrder(request, order_status):
    page    = request.GET.get('page')
    today   = date.today()
    if request.user.role_id == 0:
        orders  = SpOrders.objects.all().order_by('-updated_date').filter(order_date__icontains=today.strftime("%Y-%m-%d"))
        if order_status:
            orders = orders.filter(order_status=order_status)
        for order in orders:
            order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')    
    else:
        condition = ''
        if order_status:
            condition += ' and sp_orders.order_status = %s' % order_status
            
        orders = SpOrders.objects.raw('''SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id
    left join sp_users on sp_users.id = sp_orders.user_id 
    where DATE(sp_orders.order_date) = %s and sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpOrders' %s group by sp_approval_status.row_id order by updated_date desc ''',[today.strftime("%Y-%m-%d"), request.user.id, condition])
    
    towns       = SpTowns.objects.all()
    routes      = SpRoutes.objects.all()
    user_list   = SpUsers.objects.filter(user_type=2)

    distributor_order_count     = SpOrders.objects.filter(user_type='Distributor').count()
    super_stockist_order_count  = SpOrders.objects.filter(user_type='SuperStockist').count()
    total_orders                = distributor_order_count+super_stockist_order_count
    user_type                  = SpPermissionWorkflowRoles.objects.filter(sub_module_id=8,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()        
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    
    order_ids = []
    user_ids  = []
    for order in orders:
        order_ids.append(order.id)
        user_ids.append(order.user_id)

    if len(user_ids) > 0:
        user_count = SpUsers.objects.filter(status=1, user_type=2).exclude(id__in=user_ids).exclude(emp_sap_id__exact='').count()
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(id__in=user_ids).exclude(emp_sap_id__exact='').values_list('id', flat=True)
    else:
        user_count = SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').count()
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').values_list('id', flat=True) 
    
    total_user_count =  SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').count()

    order_counts = []
    products_list = []
    product_class = SpProductClass.objects.filter(status=1).order_by('-id')
    for product in product_class:
        products = SpProducts.objects.filter(product_class_id=product.id).values_list('id', flat=True)
        if products and len(order_ids)>0:
            quantity  = SpOrderDetails.objects.filter(order_id__in=order_ids, product_id__in=products, order_date__icontains=today.strftime("%Y-%m-%d")).aggregate(Sum('quantity'))['quantity__sum']
            if quantity:
                product.product_count = quantity
            else:
                product.product_count = 0
        else:
            product.product_count = 0

    total_order_quantity = 0
    
    if len(order_ids)>0:
        first_order_id = order_ids[0]
    else:
        first_order_id = 0
            
    context = {}
    context['orders']                           = orders
    context['order_status']                     = order_status
    context['user_list']                        = user_list
    context['products_list']                    = products_list
    context['product_class']                    = product_class
    context['total_order_quantity']             = total_order_quantity
    context['level_id']                         = level_id
    context['role_id']                          = request.user.role_id
    context['order_counts']                     = order_counts
    context['total_orders']                     = total_orders
    context['total_user_count']                 = total_user_count
    context['user_id']                          = list(user_id)
    context['distributor_order_count']          = distributor_order_count
    context['super_stockist_order_count']       = super_stockist_order_count
    context['towns']                            = towns
    context['routes']                           = routes
    context['first_order_id']                   = first_order_id
    context['today_date']                       = today.strftime("%d/%m/%Y")
    context['total_pages']                      = 'all'
    context['page_limit']                       = getConfigurationResult('page_limit')
    context['page_title']                       = "Order Management"

    template = 'order-management/index.html'
    return render(request, template, context)

#ajax order list
@login_required
@has_par(sub_module_id=8,permission='list')
def ajaxOrdersLists(request):
    order_status    = request.GET['order_status']
    town_id         = request.GET['town_id']
    route_id        = request.GET['route_id']
    user_sap_id     = request.GET['user_sap_id']
    order_date      = request.GET['order_date']

    if request.user.role_id == 0:
        orders = SpOrders.objects.all().order_by('-updated_date')
        if order_status:
            orders = orders.filter(order_status=order_status)
        if town_id:
            orders = orders.filter(town_id=town_id)
        if route_id:
            orders = orders.filter(route_id=route_id)
        if user_sap_id:
            orders = orders.filter(user_sap_id=user_sap_id) 
        if order_date:
            order_date  = datetime.strptime(str(order_date), '%d/%m/%Y').strftime('%Y-%m-%d')
            orders      = orders.filter(order_date__icontains=order_date)
        for order in orders:
            order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')        
    else:
        
        condition = ''
        if order_status:
            condition += ' and sp_orders.order_status = %s' % order_status
        if town_id:
            condition += ' and sp_orders.town_id = %s' % town_id
        if route_id:
            condition += ' and sp_orders.route_id = %s' % route_id    
        if user_sap_id:
            user_sap_id = '"'+user_sap_id+'"'
            condition += ' and sp_orders.user_sap_id = %s' % user_sap_id    
        
        order_date  = '"'+datetime.strptime(str(order_date), '%d/%m/%Y').strftime('%Y-%m-%d')+'"'
        query = """ SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id
    left join sp_users on sp_users.id = sp_orders.user_id 
    where DATE(sp_orders.order_date) = %s and sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpOrders' %s group by sp_approval_status.row_id order by updated_date desc """ % (order_date, request.user.id, condition)
        
        orders = SpOrders.objects.raw(query)

    order_ids = []
    user_ids  = []
    for order in orders:
        order_ids.append(order.id)
        user_ids.append(order.user_id)

    if len(user_ids) > 0:
        user_count = SpUsers.objects.filter(status=1, user_type=2).exclude(id__in=user_ids).exclude(emp_sap_id__exact='').count()
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(id__in=user_ids).exclude(emp_sap_id__exact='').values_list('id', flat=True)
    else:
        user_count = SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').count()
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').values_list('id', flat=True) 
    
    total_user_count =  SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').count()

    products_list = []
    product_class = SpProductClass.objects.filter(status=1).order_by('-id')
    for product in product_class:
        products = SpProducts.objects.filter(product_class_id=product.id).values_list('id', flat=True)
        if products and len(order_ids)>0:
            order_date  = datetime.strptime(str(request.GET['order_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
            quantity  = SpOrderDetails.objects.filter(order_id__in=order_ids, product_id__in=products, order_date__icontains=order_date).aggregate(Sum('quantity'))['quantity__sum']
            if quantity:
                product.product_count = quantity
            else:
                product.product_count = 0
        else:
            product.product_count = 0
    
    if len(order_ids)>0:
        first_order_id = order_ids[0]
    else:
        first_order_id = 0
        
    context = {}
    context['orders']           = orders
    context['role_id']          = request.user.role_id
    context['order_status']     = order_status
    context['product_class']    = product_class
    context['total_user_count'] = total_user_count
    context['first_order_id']   = first_order_id
    context['user_id']          = list(user_id)
    template = 'order-management/ajax-order-lists.html'
    return render(request, template, context)

def getFreeSchemes(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'variant_name')+''
        if free_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'product_id')
            free += str(free_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if free_scheme.container_quantity>0:
            free += ' and '     
        if free_scheme.pouch_quantity>0:
            if free_scheme.pouch_quantity == 1:
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
            else:    
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
              
        free_scheme = free
    else:
        free_scheme = None  
    return free_scheme

def getlogFreeSchemes(product_variant_id, order_id, user_id,vehicle_id):
    try:
        free_scheme = SpLogisticOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id,vehicle_id=vehicle_id)
    except SpLogisticOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'variant_name')+''
        if free_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'product_id')
            free += str(free_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if free_scheme.container_quantity>0:
            free += ' and '     
        if free_scheme.pouch_quantity>0:
            if free_scheme.pouch_quantity == 1:
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
            else:    
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
              
        free_scheme = free
    else:
        free_scheme = None  
    return free_scheme

def getOrderFreeSchemes(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'variant_name')+'-'
        if free_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'product_id')
            free += str(free_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if free_scheme.container_quantity>0:
            free += ' and '     
        if free_scheme.pouch_quantity>0:
            if free_scheme.pouch_quantity == 1:
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
            else:    
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
              
        free_scheme = free
    else:
        free_scheme = None  
    return free_scheme
def getOrderlogisticFreeSchemes(product_variant_id, order_id, user_id,vehicle_id):
    try:
        free_scheme =SpLogisticOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id,vehicle_id=vehicle_id)
    except SpLogisticOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'variant_name')+'-'
        if free_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'product_id')
            free += str(free_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if free_scheme.container_quantity>0:
            free += ' and '     
        if free_scheme.pouch_quantity>0:
            if free_scheme.pouch_quantity == 1:
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
            else:    
                free += ' '+str(free_scheme.pouch_quantity)+' free '+str(free_scheme.free_variant_packaging_type)+'.'
              
        free_scheme = free
    else:
        free_scheme = None  
    return free_scheme




def getFlatSchemeByVariant(product_variant_id, order_id, user_id):
    try:
        flat_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='flat', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        flat = 'Discount of Rs. '+str(flat_scheme.incentive_amount)+' applied.'
        flat_scheme = flat
    else:
        flat_scheme = None  
    return flat_scheme

def getlogisticFlatSchemeByVariant(product_variant_id, order_id, user_id,vehicle_id):
    try:
        flat_scheme = SpLogisticOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='flat', user_id=user_id,vehicle_id=vehicle_id)
    except SpLogisticOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        flat = 'Discount of Rs. '+str(flat_scheme.incentive_amount)+' applied.'
        flat_scheme = flat
    else:
        flat_scheme = None  
    return flat_scheme

def getFlatSchemeByVariants(product_variant_id, order_id, user_id):
    try:
        flat_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='flat', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        flat = flat_scheme.incentive_amount
        flat_scheme = flat
    else:
        flat_scheme = None  
    return flat_scheme    
    
def getlogisticFlatSchemeByVariants(product_variant_id, order_id, user_id,vehicle_id):
    try:
        flat_scheme = SpLogisticOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='flat', user_id=user_id,vehicle_id=vehicle_id)
    except SpLogisticOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        flat = flat_scheme.incentive_amount
        flat_scheme = flat
    else:
        flat_scheme = None  
    return flat_scheme    
    

def getQuantitativeScheme(order_id, user_id):
    try:
        quantitative_scheme = SpOrderSchemes.objects.get(order_id=order_id, scheme_type='quantitative', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        quantitative_scheme = None
    if quantitative_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, quantitative_scheme.free_variant_id, 'variant_name')+'-'
        if quantitative_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, quantitative_scheme.free_variant_id, 'product_id')
            free += str(quantitative_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if quantitative_scheme.pouch_quantity>0:
            free += ' and '+str(quantitative_scheme.pouch_quantity)+' free Pouches'
        free += ' under the '+getModelColumnById(SpSchemes, quantitative_scheme.scheme_id, 'name')+' Scheme'      
        quantitative_scheme = free
    else:
        quantitative_scheme = None  
    return quantitative_scheme

def getlogisticQuantitativeScheme(order_id, user_id):
    try:
        quantitative_scheme = SpOrderSchemes.objects.get(order_id=order_id, scheme_type='quantitative', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        quantitative_scheme = None
    if quantitative_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, quantitative_scheme.free_variant_id, 'variant_name')+'-'
        if quantitative_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, quantitative_scheme.free_variant_id, 'product_id')
            free += str(quantitative_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if quantitative_scheme.pouch_quantity>0:
            free += ' and '+str(quantitative_scheme.pouch_quantity)+' free Pouches'
        free += ' under the '+getModelColumnById(SpSchemes, quantitative_scheme.scheme_id, 'name')+' Scheme'      
        quantitative_scheme = free
    else:
        quantitative_scheme = None  
    return quantitative_scheme


def getFlatScheme(order_id, user_id):
    try:
        flat_scheme = SpOrderSchemes.objects.get(order_id=order_id, scheme_type='flat', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        free = ''
        free += str(flat_scheme.incentive_amount)+' Incentive amount has been applied under the '+getModelColumnById(SpFlatSchemes, flat_scheme.scheme_id, 'name')+' Scheme'      
        flat_scheme = free
    else:
        flat_scheme = None  
    return flat_scheme    

def getBulkpackScheme(order_id, user_id):
    try:
        bulk_pack_scheme = SpOrderSchemes.objects.get(order_id=order_id, scheme_type='bulkpack', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        bulk_pack_scheme = None
    if bulk_pack_scheme:
        free = ''
        free += str(bulk_pack_scheme.incentive_amount)+' Incentive amount has been applied under the '+getModelColumnById(SpBulkpackSchemes, bulk_pack_scheme.scheme_id, 'name')+' Scheme'      
        bulk_pack_scheme = free
    else:
        bulk_pack_scheme = None  
    return bulk_pack_scheme

#get order details view
@login_required
@has_par(sub_module_id=8,permission='view')
def getOrderDetails(request):
    id                          = request.GET.get('id')

    if request.user.role_id == 0:
        order_details  = SpOrders.objects.get(id=id)
    else:
        order_details = SpOrders.objects.raw('''SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id 
    where sp_orders.id = %s order by id desc LIMIT 1 ''',[request.GET.get('id')])[0] 

    order_details.profile_image = getModelColumnById(SpUsers, order_details.user_id, 'profile_image')
    order_details.store_name    = getModelColumnById(SpUsers, order_details.user_id, 'store_name')
    order_details.emp_sap_id    = getModelColumnById(SpUsers, order_details.user_id, 'emp_sap_id')
    order_details.user_details  = SpBasicDetails.objects.get(user_id=order_details.user_id)
    order_details.user_address  = SpAddresses.objects.get(user_id=order_details.user_id, type='permanent')

    order_item_list = SpOrderDetails.objects.filter(order_id=id)
    quantitative_scheme           = getQuantitativeScheme(id, getModelColumnById(SpOrders, id, 'user_id'))      
    flat_scheme                   = 0
    bulk_scheme                   = getBulkpackScheme(id, getModelColumnById(SpOrders, id, 'user_id'))         
    for order_item in order_item_list:
        order_item.free_schemes        = getOrderFreeSchemes(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
        order_item.free_scheme         = getOrderFreeScheme(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
        order_item.flat_scheme         = getFlatSchemeByVariant(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
        order_item.flat_schemes         = getFlatSchemeByVariants(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
    crate_sum  = SpOrderDetails.objects.filter(order_id=id,product_container_type='Crate').aggregate(Sum('quantity'))['quantity__sum']
    matki_sum  = SpOrderDetails.objects.filter(order_id=id,product_container_type='Matki').aggregate(Sum('quantity'))['quantity__sum']
    
    total_incentive = getFlatBulkIncentive(id, getModelColumnById(SpOrders, id, 'user_id'))
    find_order_amount = round(float(order_details.order_total_amount)-float(total_incentive), 2)
    
    product_class = SpProductClass.objects.filter(status=1).order_by('-id')
    for product in product_class:
        products = SpProducts.objects.filter(product_class_id=product.id).values_list('id', flat=True)
        if products:
            quantity  = SpOrderDetails.objects.filter(order_id=id,product_id__in=products).aggregate(Sum('quantity'))['quantity__sum']
            if quantity:
                product.product_count = quantity
            else:
                product.product_count = 0
        else:
            product.product_count = 0
            
    context = {}
    context['order_details']             = order_details
    context['order_item_list']           = order_item_list
    context['quantitative_scheme']       = quantitative_scheme
    context['flat_scheme']               = flat_scheme
    context['bulk_scheme']               = bulk_scheme
    context['crate_sum']                 = crate_sum
    context['matki_sum']                 = matki_sum
    context['total_incentive']           = total_incentive
    context['find_order_amount']         = find_order_amount
    context['product_class']             = product_class
    context['role_id']                   = request.user.role_id

    template = 'order-management/get-order-details.html'
    return render(request, template, context)

#get order details view
@login_required
def getDefaulterUserList(request):
    users = request.GET['user_id'].replace('[', '')
    users = users.replace(']', '')
    users = list(users.split(", "))
 
    user_list = []
    for user in users:
        user_details = {}
        user_detail = SpUsers.objects.get(id=int(user))
        if SpUserAreaAllocations.objects.filter(user_id=user_detail.id).exists():
           user_details['id']            = user_detail.id
           user_details['first_name']    = user_detail.first_name
           user_details['middle_name']   = user_detail.middle_name
           user_details['last_name']     = user_detail.last_name
           user_details['store_name']    = user_detail.store_name
           user_details['contact_no']    = user_detail.primary_contact_number
           user_details['route_name']    = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user_detail.id, 'route_name')
           user_details['town_name']     = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user_detail.id, 'town_name')
           user_list.append(user_details)

    context = {}
    context['users'] = user_list 
    template = 'order-management/get-defaulter-user-list.html'
    return render(request, template, context)
#-------------------------------------------------------------------------------------------
def updateUserOrderStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpOrders.objects.get(id=id)
            data.block_unblock=is_active
            data.save()  
                         
            if is_active =='1':
                msg = 'Unblocked'
            else:
                msg = 'Blocked'
            response['error'] = False
            response['message'] = "Order has been "+msg+" successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('orders')
#-------------------------------------------------------------------------------------------    
#update order status
@login_required
@has_par(sub_module_id=8,permission='edit')
def updateOrderStatus(request):
    response = {}
    order_id = request.POST.getlist('order_id[]')
    level_id = request.POST['level_id']
    if request.user.role_id == 0:
        for order in order_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', level_id=level_id)
            if approvals_request:
                for approval in approvals_request:
                    approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                    approval_data.status                    = 1
                    approval_data.final_status_user_id      = request.user.id
                    approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    approval_data.save()

                user_level_approval_count = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', level_id=level_id, status=0).count()
                if user_level_approval_count == 0:
                    order                   = SpOrders.objects.get(id=order)   
                    order.order_status      = level_id
                    order.save()
            else:
                order                   = SpOrders.objects.get(id=order)   
                order.order_status      = level_id
                order.save()
    else:    
        for order in order_id:
            if level_id == '3':
                approvals_request = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', level_id=level_id)
            else:
                approvals_request = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', role_id=request.user.role_id, level_id=level_id)
                
            for approval in approvals_request:
                approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                approval_data.status                    = 1
                approval_data.final_status_user_id      = request.user.id
                approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                approval_data.save()

            user_level_approval_count = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', level_id=level_id, status=0).count()
            if user_level_approval_count == 0:
                order                   = SpOrders.objects.get(id=order)   
                order.order_status      = level_id
                order.save()   

    
    if level_id == '2':
        for order in order_id:
            approvals_requests = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', status=0)
            if approvals_requests:
                for approval in approvals_requests:
                    notification                        = SpUserNotifications()
                    notification.row_id                 = approval.row_id
                    notification.user_id                = approval.user_id
                    notification.model_name             = 'SpOrders'
                    notification.notification           = 'Order '+approval.level+' Request has been sent.'
                    notification.is_read                = 0
                    notification.created_by_user_id     = request.user.id
                    notification.created_by_user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    notification.save()


    if level_id == '2':
        for order in order_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been forwarded'
            activity    = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been forwarded by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            
            saveActivity('Order Management', 'Order Summary', heading, activity, request.user.id, user_name, 'forwaord.png', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpOrders,order,'user_id')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            message_title = "Order forwarded"
            message_body = "An order ("+getModelColumnById(SpOrders,order,'order_code')+") has been forwarded by "+user_name
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(order,'SpOrders','Order Management','Order forwarded',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            

    elif level_id == '3':
        for order in order_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been approved'
            activity    = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been approved by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            
            saveActivity('Order Management', 'Order Summary', heading, activity, request.user.id, user_name, 'approved.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpOrders,order,'user_id')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            message_title = "Order approved"
            message_body = "An order ("+getModelColumnById(SpOrders,order,'order_code')+") has been approved by "+user_name
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(order,'SpOrders','Order Management','Order approved',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#

    response['error'] = False
    response['message'] = "Order status has been updated successfully."
    return JsonResponse(response)

@login_required
@has_par(sub_module_id=8,permission='export')
def exportToXlsx(request, columns, town_id, route_id, user_sap_id, order_date):
    column_list = columns.split (",")
    orders = SpOrders.objects.all().order_by('-id')
    if town_id != 0:
        orders = orders.filter(town_id=town_id)
    if route_id != 0:
        orders = orders.filter(route_id=route_id)
    if user_sap_id != '0':
        orders = orders.filter(user_sap_id=user_sap_id) 
    if order_date:
        orders = orders.filter(order_date__icontains=order_date)
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Orders'
    
    # Define the titles for columns
    columns = []

    if 'order_id' in column_list:
        columns += [ 'Order ID' ]

    if 'distributor_ss' in column_list:
        columns += [ 'Distributor/Super-Stockist Name' ]
    
    if 'shift' in column_list:
        columns += [ 'Shift' ] 

    if 'route' in column_list:
        columns += [ 'Route' ]

    if 'amount' in column_list:
        columns += [ 'Amount' ] 

    if 'status' in column_list:
        columns += [ 'Status' ]       

    if 'order_date' in column_list:
        columns += [ 'Order Date' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(name='Arial Nova Cond Light',size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    
    for order in orders:
        row_num += 1
        # Define the data for each cell in the row 

        store_name = getModelColumnById(SpUsers, order.user_id, 'store_name') 
        row = []
        if 'order_id' in column_list:
            row += [ order.order_code ]

        if 'distributor_ss' in column_list:
            row += [ store_name  + '(' + order.user_name + '/' + order.user_sap_id + ')' ]
        
        if 'shift' in column_list:
            row += [ order.order_shift_name ] 

        if 'route' in column_list:
            row += [ order.route_name ]

        if 'amount' in column_list:
            row += [ order.order_total_amount ]

        if 'status' in column_list:
            if order.order_status == 1:
                order_status = 'Initiated'
            elif order.order_status == 2:
                order_status ='Forwarded'
            elif order.order_status == 3:
                order_status ='Approved'       
            else:
                order_status='Delivered'

            row += [ order_status ]

        if 'order_date' in column_list:
            order_date = str(order.order_date).replace('+00:00', '')
            row += [ datetime.strptime(str(order_date), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %I:%M:%p') ]                   
       
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
    workbook.save(response)

    return response

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#Automaticly downloads to PDF file
@login_required
@has_par(sub_module_id=8,permission='export')
def exportToPDF(request, columns, town_id, route_id, user_sap_id, order_date):
    column_list = columns.split (",")
    context = {}
    orders = SpOrders.objects.all().order_by('-id')
    if town_id != 0:
        orders = orders.filter(town_id=town_id)
    if route_id != 0:
        orders = orders.filter(route_id=route_id)
    if user_sap_id != '0':
        orders = orders.filter(user_sap_id=user_sap_id) 
    if order_date:
        orders = orders.filter(order_date__icontains=order_date)
    for order in orders:
        order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')    
     
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('order-management/order_pdf_template.html', {'orders': orders, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'orders.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

#get order details view
@login_required
@has_par(sub_module_id=8,permission='edit')
def editOrder(request):
    response = {}
    id                          = request.GET.get('id')
    order_details               = SpOrders.objects.get(id=id)
    order_details.profile_image = getModelColumnById(SpUsers, order_details.user_id, 'profile_image')
    order_details.user_details  = SpBasicDetails.objects.get(user_id=order_details.user_id)
    order_details.user_address  = SpAddresses.objects.get(user_id=order_details.user_id, type='permanent')
    order_item_list             = SpLogisticPlanDetail.objects.filter(order_id=id)
    if SpInvoices.objects.filter(user_id=order_details.user_id,route_id=order_details.route_id,created_date=order_details.order_date.strftime('%Y-%m-%d')).exists():
        response['error'] = True
        response['message'] = 'Invoice Allredy genrated of this order you can not edit order '
        return JsonResponse(response)

    
    quantitative_scheme           = getQuantitativeScheme(id, order_details.user_id)      
    flat_scheme                   = 0 
    bulk_scheme                   = getBulkpackScheme(id, order_details.user_id)
    for order_item in order_item_list:
        order_item.quantity                       = int(order_item.quantity)
        order_item.user_product_variants = SpUserProductVariants.objects.filter(user_id=order_details.user_id, product_id=order_item.product_id)
        order_item.free_schemes          = getOrderlogisticFreeSchemes(order_item.product_variant_id, order_item.order_id, order_details.user_id,vehicle_id=order_item.vehicle_id)
        order_item.free_scheme           = getlogisticOrderFreeScheme(order_item.product_variant_id, order_item.order_id, order_details.user_id,vehicle_id=order_item.vehicle_id)
        order_item.flat_scheme           = getlogisticFlatSchemeByVariant(order_item.product_variant_id, order_item.order_id, order_details.user_id,vehicle_id=order_item.vehicle_id)
        order_item.flat_schemes          = getlogisticFlatSchemeByVariants(order_item.product_variant_id, order_item.order_id, order_details.user_id,vehicle_id=order_item.vehicle_id)

    try:
        bulk_pack_scheme_amount = SpLogisticOrderSchemes.objects.get(order_id=id, scheme_type='bulkpack')
        bulk_pack_scheme_id     = bulk_pack_scheme_amount.scheme_id
        bulk_pack_unit_id       = bulk_pack_scheme_amount.unit_id
        bulk_pack_unit_name     = bulk_pack_scheme_amount.unit_name
        bulk_pack_scheme_amount = bulk_pack_scheme_amount.incentive_amount
        
    except SpLogisticOrderSchemes.DoesNotExist:
        bulk_pack_scheme_amount = ''
        bulk_pack_scheme_id     = ''
        bulk_pack_unit_id       = ''
        bulk_pack_unit_name     = ''

    try:
        flat_pack_scheme_amount = ''
        flat_pack_scheme_id     = ''
        flat_pack_unit_id       = ''
        flat_pack_unit_name     = ''
    except SpLogisticOrderSchemes.DoesNotExist:
        flat_pack_scheme_amount = ''
        flat_pack_scheme_id     = ''
        flat_pack_unit_id       = ''
        flat_pack_unit_name     = ''

    try:
        quantitative_pack_scheme_amount = SpLogisticOrderSchemes.objects.get(order_id=id, scheme_type='quantitative')
        quantitative_pack_scheme_id             = quantitative_pack_scheme_amount.scheme_id
        quantitative_pack_free_variant_id       = quantitative_pack_scheme_amount.free_variant_id
        quantitative_pack_container_quantity    = quantitative_pack_scheme_amount.container_quantity
        quantitative_pack_pouch_quantity        = quantitative_pack_scheme_amount.pouch_quantity
    except SpLogisticOrderSchemes.DoesNotExist:
        quantitative_pack_scheme_id             = ''
        quantitative_pack_free_variant_id       = ''
        quantitative_pack_container_quantity    = ''
        quantitative_pack_pouch_quantity        = ''
    SpChallans.objects.filter(user_id=order_details.user_id,created_date=order_details.order_date.strftime("%Y-%m-%d")).delete()
    products = SpProducts.objects.filter(status=1)
    vehicle = SpVehicles.objects.filter(status=1)
    context = {}
    context['order_details']             = order_details
    context['order_item_list']           = order_item_list
    context['vehicle']                   = vehicle
    context['products']                  = products
    context['quantitative_scheme']       = quantitative_scheme
    context['flat_scheme']               = flat_scheme
    context['bulk_scheme']               = bulk_scheme
    context['bulk_pack_scheme_amount']   = bulk_pack_scheme_amount
    context['bulk_pack_scheme_id']       = bulk_pack_scheme_id
    context['bulk_pack_unit_id']         = bulk_pack_unit_id
    context['bulk_pack_unit_name']       = bulk_pack_unit_name
    context['flat_pack_scheme_amount']   = flat_pack_scheme_amount
    context['flat_pack_scheme_id']       = flat_pack_scheme_id
    context['flat_pack_unit_id']         = flat_pack_unit_id
    context['flat_pack_unit_name']       = flat_pack_unit_name
    context['quantitative_pack_scheme_id']              = quantitative_pack_scheme_id
    context['quantitative_pack_free_variant_id']        = quantitative_pack_free_variant_id
    context['quantitative_pack_container_quantity']     = quantitative_pack_container_quantity
    context['quantitative_pack_pouch_quantity']         = quantitative_pack_pouch_quantity
    template = 'order-management/edit-order.html'

    return render(request, template, context)

#add order
@login_required
@has_par(sub_module_id=8,permission='add')
def addOrder(request):
    page    = request.GET.get('page')
    today   = date.today()
    products = SpProducts.objects.filter(status=1)
    paymentmode = Sp_Mode_Of_Payments.objects.all()
    if request.user.role_id == 0:
        orders  = SpOrders.objects.all().order_by('-updated_date').filter(order_date__icontains=today.strftime("%Y-%m-%d"))
        for order in orders:
            order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')
    else:
        orders = SpOrders.objects.raw('''SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id
    left join sp_users on sp_users.id = sp_orders.user_id 
    where DATE(sp_orders.order_date) = %s and sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpOrders' group by sp_approval_status.row_id order by updated_date desc ''',[today.strftime("%Y-%m-%d"), request.user.id])
    order_ids = []
    user_ids  = []
    for order in orders:
        order_ids.append(order.id)
        user_ids.append(order.user_id)

    if len(user_ids) > 0:
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(id__in=user_ids).exclude(emp_sap_id__exact='').all()
    else:
        user_id    = SpUsers.objects.filter(status=1, user_type=2).exclude(emp_sap_id__exact='').all()
    vehicle        =SpVehicles.objects.all()
    context = {}
    context['orders']           = orders
    context['vehicle']           = vehicle
    context['paymentmode']      = paymentmode
    context['products']         = products
    context['role_id']          = request.user.role_id
    context['user_id']          = list(user_id)
    return render(request,'order-management/add-order.html',context)

#save order
@login_required
@has_par(sub_module_id=8,permission='add')
@transaction.atomic
def saveOrder(request):
    today   = date.today()
    user_id             = request.POST['user_id']
    mode_of_payment     = request.POST['mode_of_payment']
    commit_amount       = request.POST['commit_amount']
    product_ids          = request.POST.getlist('product_id[]')
    product_variant_ids  = request.POST.getlist('product_variant_id[]')
    quantityss            = request.POST.getlist('quantity[]')
    ratess               = request.POST.getlist('rate[]')
    amounts              = request.POST.getlist('amount[]')
    vehicle_ids          = request.POST.getlist('vehicle_id[]')
    user                = SpBasicDetails.objects.get(user_id=user_id)
    try:
        if SpOrders.objects.filter(user_id=user_id,order_date__icontains=today).exists():
            heading     = 'Order already placed of this user.'
            messages.error(request, heading, extra_tags='success')
            return redirect('/orders')
    except SpOrders.DoesNotExist:
        heading = None
    product_id   =[]
    product_variant_id =[]
    amount =[]
    rate =[]
    vehicle_id = []
    quantity = []
    for id, product_variant in enumerate(product_variant_ids):

            if (product_ids[id] in product_id):
                if product_variant_ids[id] in product_variant_id:
                    index_value  = product_variant_id.index(product_variant_ids[id])
                    if vehicle_ids[id] == vehicle_id[index_value]:
                        total_amount = amount[index_value]+float(amounts[id])
                        amount[index_value] = total_amount
                        total_quantity = quantity[index_value]+float(quantityss[id])
                        quantity[index_value] = total_quantity
                    else:
                        product_id.append(product_ids[id])
                        product_variant_id.append(product_variant_ids[id])
                        amount.append(float(amounts[id]))
                        rate.append(ratess[id])
                        quantity.append(float(quantityss[id]))
                        vehicle_id.append(vehicle_ids[id])
                else:
                    product_id.append(product_ids[id])
                    product_variant_id.append(product_variant_ids[id])
                    amount.append(float(amounts[id]))
                    rate.append(ratess[id])
                    quantity.append(float(quantityss[id]))
                    vehicle_id.append(vehicle_ids[id])          
            else:
                product_id.append(product_ids[id])
                product_variant_id.append(product_variant_ids[id])
                amount.append(float(amounts[id]))
                rate.append(ratess[id])
                quantity.append(float(quantityss[id]))
                vehicle_id.append(vehicle_ids[id])


    if len(product_variant_id) > 0:
        if SpOrders.objects.count() == 0:
            last_order_id = 1
        else:
            last_order_id = SpOrders.objects.order_by('-id').first()
            last_order_id = last_order_id.id+1 

        user_area_details = SpUserAreaAllocations.objects.get(user_id=user_id)
        
        if user_area_details.route_id:
            transporter_details = SpVehicles.objects.filter(route_id=user_area_details.route_id).order_by('-id').first()    
            if transporter_details:
                transporter_name = ''
                if transporter_details.owner_name:
                    transporter_name += transporter_details.owner_name
                if transporter_details.owner_address:
                    transporter_name += ', '+str(transporter_details.owner_address)
                transporter_description = ''
                if transporter_details.driver_name:
                    transporter_description += 'Driver Name: '+str(transporter_details.driver_name)
                vehicle_no = ''
                if transporter_details.registration_number:
                    vehicle_no += transporter_details.registration_number            
            else:
                transporter_name = ''
                transporter_description = ''
                vehicle_no = ''

        else:
            transporter_name = ''
            transporter_description = ''
            vehicle_no = ''

        #save order
        order                       = SpOrders()
        order.order_code            =  getConfigurationResult('org_code')+str(last_order_id)
        order.user_id               =  user_id
        order.user_sap_id           =  getModelColumnById(SpUsers, user_id, 'emp_sap_id')
        order.user_name             =  getModelColumnById(SpUsers, user_id, 'first_name')+' '+getModelColumnById(SpUsers, user_id, 'middle_name')+' '+getModelColumnById(SpUsers, user_id, 'last_name')
        order.user_type             =  getUserRole(user_id)
        order.route_id              =  user_area_details.route_id
        order.route_name            =  user_area_details.route_name
        order.transporter_name      =  transporter_name
        order.transporter_details   =  transporter_description
        order.vehicle_no            =  vehicle_no
        order.town_id               =  user_area_details.town_id
        order.town_name             =  user_area_details.town_name
        order.mode_of_payment       =  mode_of_payment
        order.amount_to_be_paid     =  int(commit_amount)
        order.order_date            =  datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        order.updated_date          =  datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        order.order_status          =  1
        order.order_shift_id        =  1
        order.block_unblock         =  1
        order.order_shift_name      =  'Morning'
        order.platfrom_created      =  'WEB'
        order.order_total_amount    =  request.POST['order_amount']
        order.order_items_count     =  len(product_variant_id)
        order.outstanding_amount    =  user.outstanding_amount
        production_unit_id          =  getModelColumnByColumnId(SpBasicDetails, 'user_id', user_id, 'production_unit_id')
        order.production_unit_id    =  production_unit_id
        order.production_unit_name  =  getModelColumnById(SpProductionUnit, production_unit_id, 'production_unit_name')
        if getModelColumnByColumnId(SpBasicDetails, 'user_id', user_id, 'tcs_applicable') == 1:
            order.tcs_value      =  getModelColumnByColumnId(SpBasicDetails, 'user_id', user_id, 'tcs_value')
        order.save()

        total_order_amount = []
        total_order_quantity = []
        order =  SpOrders.objects.get(id=order.id)
        product_id_list = []
        product_variant_id_list     = []
        amount_list             = []
        quantity_list              = []
        rate_list  = []

        for id, product_variant in enumerate(product_variant_id):
            index = int(id)+1
            packaging_type = 'packaging_type_'+str(index)
            total_order_amount.append(float(amount[id]))
            total_order_quantity.append(float(quantity[id]))


            if product_id[id] in product_id_list and product_variant_id[id] in product_variant_id_list:
                index_value  = product_variant_id_list.index(product_variant_id[id])
                total_amount = amount_list[index_value]+float(amount[id])
                amount_list[index_value] = total_amount
                total_quantity = quantity_list[index_value]+float(quantity[id])
                quantity_list[index_value] = total_quantity
            else:
                product_id_list.append(product_id[id])
                product_variant_id_list.append(product_variant_id[id])
                amount_list.append(float(amount[id]))
                rate_list.append(rate[id])
                quantity_list.append(float(quantity[id]))
            
            if request.POST[packaging_type] == '0':
                pouch_quantity     = int(float(quantity[id])*float(getModelColumnById(SpProductVariants, product_variant_id[id], 'no_of_pouch')))
                quantity_in_ltr    = float(pouch_quantity)*float(getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_size'))
            else:
                pouch_quantity     = int(quantity[id])
                quantity_in_ltr    = float(quantity[id])*float(getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_size'))
                
            vehicle                            = SpLogisticPlanDetail()    
            vehicle.user_id                    = request.POST['user_id']
            vehicle.order_id                   = order.id
            vehicle.product_id                 = product_id[id]
            vehicle.product_name               = getModelColumnById(SpProducts, product_id[id], 'product_name') 
            vehicle.product_variant_id         = product_variant_id[id]
            vehicle.product_variant_name       = getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_name') 
            vehicle.product_variant_size       = getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_size')
            vehicle.product_no_of_pouch        = getModelColumnById(SpProductVariants, product_variant_id[id], 'no_of_pouch')
            vehicle.product_container_size     = getModelColumnById(SpProductVariants, product_variant_id[id], 'container_size')
            vehicle.product_container_type     = getModelColumnById(SpProductVariants, product_variant_id[id], 'container_name')
            vehicle.quantity_in_pouch          = pouch_quantity
            vehicle.quantity_in_ltr            = round(quantity_in_ltr,2)
            vehicle.quantity                   = quantity[id]
            vehicle.rate                       = rate[id]
            vehicle.amount                     = amount[id]
            vehicle.packaging_type             = request.POST[packaging_type]
            vehicle.gst                        = getModelColumnById(SpProductVariants, product_variant_id[id], 'gst')
            vehicle.product_packaging_type_name= getModelColumnById(SpProductVariants, product_variant_id[id], 'packaging_type_name')
            vehicle.is_allow                   = getModelColumnById(SpProductVariants, product_variant_id[id], 'is_allow_in_packaging')
            vehicle.order_date                 = order.order_date    
            vehicle.route_id                   = getModelColumnById(SpVehicles, vehicle_id[id], 'route_id')    
            vehicle.route_name                 = getModelColumnById(SpVehicles, vehicle_id[id], 'route_name')      
            vehicle.vehicle_id                 = getModelColumnById(SpVehicles, vehicle_id[id], 'id')  
            vehicle.vehicle_number             = getModelColumnById(SpVehicles, vehicle_id[id], 'registration_number')  
            vehicle.save()

            try:
                free_scheme = SpUserSchemes.objects.get(applied_on_variant_id=product_variant_id[id], user_id=request.POST['user_id'], scheme_type=1, scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), status=1)
            except SpUserSchemes.DoesNotExist:
                free_scheme = None
            
            if free_scheme:
                container_quantity = 0
                if free_scheme.pouch_quantity > 0:
                    if request.POST[packaging_type] == '0':
                        if free_scheme.packaging_type == 0:
                            # order_value = int(quantity[id])/int(free_scheme.minimum_order_quantity)
                            order_value = float(quantity[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0    
                        elif free_scheme.packaging_type ==1:
                            # pouch_value = int(quantity[id])*int(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            pouch_value = float(quantity[id])*float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            order_value = int(pouch_value)/int(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0        
                    else:
                        if free_scheme.packaging_type == 0:
                            # crate_value = int(quantity[id])/int(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            # order_value = int(crate_value)/int(free_scheme.minimum_order_quantity)
                            # order_value = str(order_value).split('.')
                            # order_value = int(order_value[0])
                            crate_value = float(quantity[id])/float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            if int(crate_value)>0:
                                order_value = int(crate_value)/int(free_scheme.minimum_order_quantity)
                                order_value = str(order_value).split('.')
                                order_value = int(order_value[0])
                            else:
                                order_value = 0 
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0
                        elif free_scheme.packaging_type == 1:
                            # order_value = int(quantity[id])/int(free_scheme.minimum_order_quantity)
                            order_value = float(quantity[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0          
                else:    
                    pouch_quantity     = 0
                    
                if int(pouch_quantity) > 0:
                    logistic_order_scheme                             = SpLogisticOrderSchemes()
                    logistic_order_scheme.order_id                    = order.id
                    logistic_order_scheme.user_id                     = request.POST['user_id']
                    logistic_order_scheme.scheme_id                   = free_scheme.scheme_id
                    logistic_order_scheme.scheme_type                 = "free"
                    logistic_order_scheme.variant_id                  = free_scheme.applied_on_variant_id
                    logistic_order_scheme.on_order_of                 = free_scheme.minimum_order_quantity
                    logistic_order_scheme.free_variant_id             = free_scheme.free_variant_id
                    logistic_order_scheme.free_variant_container_type = free_scheme.order_container_name
                    logistic_order_scheme.container_quantity          = container_quantity
                    logistic_order_scheme.free_variant_packaging_type = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'packaging_type_name')
                    logistic_order_scheme.free_variant_container_size = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'container_size')
                    logistic_order_scheme.pouch_quantity              = pouch_quantity
                    logistic_order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'variant_size'))
                    logistic_order_scheme.product_id                  = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_id')
                    logistic_order_scheme.product_class_id            = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_class_id')
                    logistic_order_scheme.route_id                    = getModelColumnById(SpVehicles, vehicle_id[id], 'route_id')
                    logistic_order_scheme.route_name                  = getModelColumnById(SpVehicles, vehicle_id[id], 'route_name')
                    logistic_order_scheme.vehicle_id                  = getModelColumnById(SpVehicles, vehicle_id[id], 'id')
                    logistic_order_scheme.vehicle_number              = getModelColumnById(SpVehicles, vehicle_id[id], 'registration_number')
                    logistic_order_scheme.created_at                  = order.order_date
                    logistic_order_scheme.save()
            try:
                flat_scheme = SpUserFlatSchemes.objects.get(applied_on_variant_id=product_variant_id[id], user_id=request.POST['user_id'], scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), incentive_amount__gt=0, status=1)
            except SpUserFlatSchemes.DoesNotExist:
                flat_scheme = None
            
            if flat_scheme:
                if int(quantity[id]) > 0:
                    if request.POST[packaging_type] == '0':
                        pouch_quantity     = float(quantity[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        flat_discount = (float(quantity[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch')))*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    else:
                        flat_discount = float(quantity[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                        pouch_quantity = float(quantity[id])/float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        pouch_quantity = str(pouch_quantity).split('.')
                        pouch_quantity = int(pouch_quantity[0])
                        if pouch_quantity > 0: 
                            pouch_quantity     = int(pouch_quantity)
                        else:
                            pouch_quantity     = 0
                    # if request.POST[packaging_type] == '0':
                    #     pouch_quantity     = int(quantity[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                    #     flat_discount = (float(quantity[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch')))*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                    #     flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    # else:
                    #     flat_discount = float(quantity[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                    #     flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    #     pouch_quantity = int(quantity[id])/int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                    #     pouch_quantity = str(pouch_quantity).split('.')
                    #     pouch_quantity = int(pouch_quantity[0])
                    #     if pouch_quantity > 0: 
                    #         pouch_quantity     = int(pouch_quantity)
                    #     else:
                    #         pouch_quantity     = 0

                    logistic_order_scheme                             = SpLogisticOrderSchemes()
                    logistic_order_scheme.order_id                    = order.id
                    logistic_order_scheme.user_id                     = request.POST['user_id']
                    logistic_order_scheme.scheme_id                   = flat_scheme.scheme_id
                    logistic_order_scheme.scheme_type                 = "flat"
                    logistic_order_scheme.incentive_amount            = flat_discount
                    logistic_order_scheme.variant_id                  = flat_scheme.applied_on_variant_id
                    logistic_order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'variant_size'))
                    logistic_order_scheme.product_id                  = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_id')
                    logistic_order_scheme.product_class_id            = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_class_id')
                    logistic_order_scheme.route_id                    = getModelColumnById(SpVehicles, vehicle_id[id], 'route_id')
                    logistic_order_scheme.route_name                  = getModelColumnById(SpVehicles, vehicle_id[id], 'route_name')
                    logistic_order_scheme.vehicle_id                  = getModelColumnById(SpVehicles, vehicle_id[id], 'id')
                    logistic_order_scheme.vehicle_number              = getModelColumnById(SpVehicles, vehicle_id[id], 'registration_number')
                    logistic_order_scheme.created_at                  = order.order_date
                    logistic_order_scheme.save()  

        
        for id, product_variant in enumerate(product_variant_id_list):
            index = int(id)+1
            packaging_type = 'packaging_type_'+str(index)
            try:
                free_scheme = SpUserSchemes.objects.get(applied_on_variant_id=product_variant_id_list[id], user_id=request.POST['user_id'], scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), status=1)
            except SpUserSchemes.DoesNotExist:
                free_scheme = None
            
            if free_scheme:
                container_quantity = 0
                if free_scheme.pouch_quantity>0:
                    if request.POST[packaging_type] == '0':
                        if free_scheme.packaging_type == 0:
                            order_value = float(quantity_list[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0    
                        elif free_scheme.packaging_type == 1:
                            pouch_value = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            order_value = int(pouch_value)/int(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0        
                    else:
                        if free_scheme.packaging_type == 0:
                            # crate_value = float(quantity_list[id])/float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            # order_value = int(crate_value)/int(free_scheme.minimum_order_quantity)
                            # order_value = str(order_value).split('.')
                            # order_value = int(order_value[0])
                            crate_value = float(quantity[id])/float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            if int(crate_value)>0:
                                order_value = int(crate_value)/int(free_scheme.minimum_order_quantity)
                                order_value = str(order_value).split('.')
                                order_value = int(order_value[0])
                            else:
                                order_value = 0 
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0
                        elif free_scheme.packaging_type == 1:
                            order_value = float(quantity_list[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0          
                else:    
                    pouch_quantity     = 0
                # if free_scheme.pouch_quantity > 0:
                #     if request.POST[packaging_type] == '0':
                #         if free_scheme.packaging_type == 0:
                #             order_value = int(quantity_list[id])/int(free_scheme.minimum_order_quantity)
                #             order_value = str(order_value).split('.')
                #             order_value = int(order_value[0])
                #             if int(order_value) > 0:
                #                 pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                #             else:
                #                 pouch_quantity     = 0    
                #         elif free_scheme.packaging_type ==1:
                #             pouch_value = int(quantity[id])*int(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                #             order_value = int(pouch_value)/int(free_scheme.minimum_order_quantity)
                #             order_value = str(order_value).split('.')
                #             order_value = int(order_value[0])
                #             if int(order_value) > 0:
                #                 pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                #             else:
                #                 pouch_quantity     = 0        
                #     else:
                #         if free_scheme.packaging_type == 0:
                #             crate_value = int(quantity_list[id])/int(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                #             order_value = int(crate_value)/int(free_scheme.minimum_order_quantity)
                #             order_value = str(order_value).split('.')
                #             order_value = int(order_value[0])
                #             if int(order_value) > 0:
                #                 pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                #             else:
                #                 pouch_quantity     = 0
                #         elif free_scheme.packaging_type == 1:
                #             order_value = int(quantity_list[id])/int(free_scheme.minimum_order_quantity)
                #             order_value = str(order_value).split('.')
                #             order_value = int(order_value[0])
                #             if int(order_value) > 0:
                #                 pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                #             else:
                #                 pouch_quantity     = 0          
                # else:    
                #     pouch_quantity     = 0
                    
                if int(pouch_quantity) > 0:
                    order_scheme                             = SpOrderSchemes()
                    order_scheme.order_id                    = order.id
                    order_scheme.user_id                     = request.POST['user_id']
                    order_scheme.scheme_id                   = free_scheme.scheme_id
                    order_scheme.scheme_type                 = "free"
                    order_scheme.variant_id                  = free_scheme.applied_on_variant_id
                    order_scheme.on_order_of                 = free_scheme.minimum_order_quantity
                    order_scheme.free_variant_id             = free_scheme.free_variant_id
                    order_scheme.free_variant_container_type = free_scheme.order_container_name
                    order_scheme.container_quantity          = container_quantity
                    order_scheme.free_variant_packaging_type = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'packaging_type_name')
                    order_scheme.free_variant_container_size = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'container_size')
                    order_scheme.pouch_quantity              = pouch_quantity
                    order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'variant_size'))
                    order_scheme.product_id                  = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_id')
                    order_scheme.product_class_id            = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_class_id')
                    order_scheme.created_at                  = order.order_date
                    order_scheme.save()
            try:
                flat_scheme = SpUserFlatSchemes.objects.get(applied_on_variant_id=product_variant_id_list[id], user_id=request.POST['user_id'], scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), incentive_amount__gt=0, status=1)
            except SpUserFlatSchemes.DoesNotExist:
                flat_scheme = None
            
            if flat_scheme:
                if int(quantity_list[id]) > 0:
                    if request.POST[packaging_type] == '0':
                        pouch_quantity     = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        flat_discount = (float(quantity_list[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch')))*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    else:
                        flat_discount = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                        pouch_quantity = float(quantity_list[id])/float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        pouch_quantity = str(pouch_quantity).split('.')
                        pouch_quantity = int(pouch_quantity[0])
                        if pouch_quantity > 0: 
                            pouch_quantity     = int(pouch_quantity)
                        else:
                            pouch_quantity     = 0
                    # if request.POST[packaging_type] == '0':
                    #     pouch_quantity     = int(quantity_list[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                    #     flat_discount = (float(quantity_list[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch')))*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                    #     flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    # else:
                    #     flat_discount = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                    #     flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    #     pouch_quantity = int(quantity_list[id])/int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                    #     pouch_quantity = str(pouch_quantity).split('.')
                    #     pouch_quantity = int(pouch_quantity[0])
                    #     if pouch_quantity > 0: 
                    #         pouch_quantity     = int(pouch_quantity)
                    #     else:
                    #         pouch_quantity     = 0

                    order_scheme                             = SpOrderSchemes()
                    order_scheme.order_id                    = order.id
                    order_scheme.user_id                     = request.POST['user_id']
                    order_scheme.scheme_id                   = flat_scheme.scheme_id
                    order_scheme.scheme_type                 = "flat"
                    order_scheme.incentive_amount            = flat_discount
                    order_scheme.variant_id                  = flat_scheme.applied_on_variant_id
                    order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'variant_size'))
                    order_scheme.product_id                  = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_id')
                    order_scheme.product_class_id            = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_class_id')
                    order_scheme.created_at                  = order.order_date
                    order_scheme.save()  


            # if request.POST[packaging_type] == '0':
            #     pouch_quantity     = int(quantity_list[id])*int(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'no_of_pouch'))
            #     quantity_in_ltr    = float(pouch_quantity)*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size'))
            # else:
            #     pouch_quantity     = int(quantity_list[id])
            #     quantity_in_ltr    = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size'))
            if request.POST[packaging_type] == '0':
                pouch_quantity     = int(float(quantity_list[id])*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'no_of_pouch')))
                quantity_in_ltr    = float(pouch_quantity)*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size'))
            else:
                pouch_quantity     = int(quantity_list[id])
                quantity_in_ltr    = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size'))
            order_detail                            = SpOrderDetails()    
            order_detail.user_id                    = request.POST['user_id']
            order_detail.order_id                   = order.id
            order_detail.product_id                 = product_id_list[id]
            order_detail.product_name               = getModelColumnById(SpProducts, product_id_list[id], 'product_name') 
            order_detail.product_variant_id         = product_variant_id_list[id]
            order_detail.product_variant_name       = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_name') 
            order_detail.product_variant_size       = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size')
            order_detail.product_no_of_pouch        = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'no_of_pouch')
            order_detail.product_container_size     = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'container_size')
            order_detail.product_container_type     = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'container_name')
            order_detail.quantity_in_pouch          = pouch_quantity
            order_detail.quantity_in_ltr            = round(quantity_in_ltr,2)
            order_detail.quantity                   = quantity_list[id]
            order_detail.rate                       = rate_list[id]
            order_detail.amount                     = amount_list[id]
            order_detail.packaging_type             = request.POST[packaging_type]
            order_detail.gst                        = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'gst')
            order_detail.product_packaging_type_name= getModelColumnById(SpProductVariants, product_variant_id_list[id], 'packaging_type_name')
            order_detail.is_allow                   = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'is_allow_in_packaging')
            order_detail.order_date                 = order.order_date    
            order_detail.save()
        
        order.order_total_amount    =  sum(total_order_amount)
        newtoday                 = today.strftime("%Y-%m-%d")
        current                  = int((today.strftime("%Y")))
        prev                     = int(today.strftime("%Y"))-1
        naxt                     = int(today.strftime("%Y"))+1
        today_date               = datetime.now().strftime('%Y')+'-'+'04'+'-'+'01'
        if newtoday > today_date:
            financial_year_start_date   = datetime.now().strftime('%Y')+'-'+'04'+'-'+'01'
        else:
            financial_year_start_date   = datetime.now().strftime(str(prev))+'-'+'04'+'-'+'01'
        order_amounts = SpOrders.objects.filter(user_id=user_id,block_unblock=1,order_date__range=[financial_year_start_date,newtoday]).aggregate(Sum('order_total_amount'))['order_total_amount__sum']  
        if order_amounts:
            order_amounts = round((order_amounts+order.order_total_amount),2)
        else:
           order_amounts = order.order_total_amount
        tcsvalues = SpTcsMaster.objects.filter().first()
        if float(order_amounts) > tcsvalues.tcs_value:
            order.tcs_value      =  tcsvalues.tcs_percentage
        else:
            order.tcs_value = None
        order.order_items_count     =  len(product_variant_id)
        order.updated_date          =  datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        order.save()
        
    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
    heading     = 'New Order('+order.order_code+') has been initiated'
    activity    = 'New Order('+order.order_code+') has been initiated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
    
    saveActivity('Order Management', 'Order Summary', heading, activity, request.user.id, user_name, 'Orderplaced.png', '2', 'app.png')

    messages.success(request, heading, extra_tags='success')
    return redirect('/orders')
 
#update order
@login_required
@has_par(sub_module_id=8,permission='edit')
@transaction.atomic
def updateOrder(request):
    today               = date.today()
    product_ids          = request.POST.getlist('product_id[]')
    product_variant_ids  = request.POST.getlist('product_variant_id[]')
    quantityss            = request.POST.getlist('quantity[]')
    ratess               = request.POST.getlist('rate[]')
    amounts              = request.POST.getlist('amount[]')
    vehicle_ids          = request.POST.getlist('vehicle_id[]')
    # packaging_type_counter = SpOrderDetails.objects.filter(order_id=request.POST['order_id']).count()
    product_id   =[]
    product_variant_id =[]
    amount =[]
    rate =[]
    vehicle_id = []
    quantity = []
    for id, product_variant in enumerate(product_variant_ids):
        if (product_ids[id] in product_id):
            if product_variant_ids[id] in product_variant_id:
                index_value  = product_variant_id.index(product_variant_ids[id])
                if vehicle_ids[id] == vehicle_id[index_value]:
                    total_amount = amount[index_value]+float(amounts[id])
                    amount[index_value] = total_amount
                    total_quantity = quantity[index_value]+float(quantityss[id])
                    quantity[index_value] = total_quantity
                else:
                    product_id.append(product_ids[id])
                    product_variant_id.append(product_variant_ids[id])
                    amount.append(float(amounts[id]))
                    rate.append(ratess[id])
                    quantity.append(float(quantityss[id]))
                    vehicle_id.append(vehicle_ids[id])
            else:
                product_id.append(product_ids[id])
                product_variant_id.append(product_variant_ids[id])
                amount.append(float(amounts[id]))
                rate.append(ratess[id])
                quantity.append(float(quantityss[id]))
                vehicle_id.append(vehicle_ids[id])          
        else:
            product_id.append(product_ids[id])
            product_variant_id.append(product_variant_ids[id])
            amount.append(float(amounts[id]))
            rate.append(ratess[id])
            quantity.append(float(quantityss[id]))
            vehicle_id.append(vehicle_ids[id])

    
    if len(product_variant_id) > 0:
        SpOrderDetails.objects.filter(order_id=request.POST['order_id']).delete()
        SpOrderSchemes.objects.filter(order_id=request.POST['order_id']).delete()
        SpLogisticPlanDetail.objects.filter(order_id=request.POST['order_id']).delete()
        SpLogisticOrderSchemes.objects.filter(order_id=request.POST['order_id']).delete()
        total_order_amount          = []
        total_order_quantity        = []
        order =  SpOrders.objects.get(id=request.POST['order_id'])
       # SpChallans.objects.filter(user_id=request.POST['user_id'],created_date=order.order_date.strftime("%Y-%m-%d")).delete()
        product_id_list             = []
        product_variant_id_list     = []
        amount_list                 = []
        quantity_list               = []
        rate_list                   = []
        for id, product_variant in enumerate(product_variant_id):
            index = int(id)+1
            packaging_type = 'packaging_type_'+str(index)
            # if packaging_type_counter > 0:
            #     index = int(id)+1
            # else:
            #     index = int(id)+2
            # packaging_type = 'packaging_type_'+str(index)
            total_order_amount.append(float(amount[id]))
            total_order_quantity.append(float(quantity[id]))

            if product_id[id] in product_id_list and product_variant_id[id] in product_variant_id_list:
                index_value  = product_variant_id_list.index(product_variant_id[id])
                total_amount = amount_list[index_value]+float(amount[id])
                amount_list[index_value] = total_amount
                total_quantity = quantity_list[index_value]+float(quantity[id])
                quantity_list[index_value] = total_quantity
            else:
                product_id_list.append(product_id[id])
                product_variant_id_list.append(product_variant_id[id])
                amount_list.append(float(amount[id]))
                rate_list.append(rate[id])
                quantity_list.append(float(quantity[id]))
            if request.POST[packaging_type] == '0':
                pouch_quantity     = int(float(quantity[id])*float(getModelColumnById(SpProductVariants, product_variant_id[id], 'no_of_pouch')))
                quantity_in_ltr    = float(pouch_quantity)*float(getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_size'))
            else:
                pouch_quantity     = int(quantity[id])
                quantity_in_ltr    = float(quantity[id])*float(getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_size'))
                
            vehicle                            = SpLogisticPlanDetail()    
            vehicle.user_id                    = request.POST['user_id']
            vehicle.order_id                   = order.id
            vehicle.product_id                 = product_id[id]
            vehicle.product_name               = getModelColumnById(SpProducts, product_id[id], 'product_name') 
            vehicle.product_variant_id         = product_variant_id[id]
            vehicle.product_variant_name       = getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_name') 
            vehicle.product_variant_size       = getModelColumnById(SpProductVariants, product_variant_id[id], 'variant_size')
            vehicle.product_no_of_pouch        = getModelColumnById(SpProductVariants, product_variant_id[id], 'no_of_pouch')
            vehicle.product_container_size     = getModelColumnById(SpProductVariants, product_variant_id[id], 'container_size')
            vehicle.product_container_type     = getModelColumnById(SpProductVariants, product_variant_id[id], 'container_name')
            vehicle.quantity_in_pouch          = pouch_quantity
            vehicle.quantity_in_ltr            = round(quantity_in_ltr,2)
            vehicle.quantity                   = quantity[id]
            vehicle.rate                       = rate[id]
            vehicle.amount                     = amount[id]
            vehicle.packaging_type             = request.POST[packaging_type]
            vehicle.gst                        = getModelColumnById(SpProductVariants, product_variant_id[id], 'gst')
            vehicle.product_packaging_type_name= getModelColumnById(SpProductVariants, product_variant_id[id], 'packaging_type_name')
            vehicle.is_allow                   = getModelColumnById(SpProductVariants, product_variant_id[id], 'is_allow_in_packaging')
            vehicle.order_date                 = order.order_date    
            vehicle.route_id                   = getModelColumnById(SpVehicles, vehicle_id[id], 'route_id')    
            vehicle.route_name                 = getModelColumnById(SpVehicles, vehicle_id[id], 'route_name')      
            vehicle.vehicle_id                 = getModelColumnById(SpVehicles, vehicle_id[id], 'id')  
            vehicle.vehicle_number             = getModelColumnById(SpVehicles, vehicle_id[id], 'registration_number')  
            vehicle.save()
        
            try:
                free_scheme = SpUserSchemes.objects.get(applied_on_variant_id=product_variant_id[id], user_id=request.POST['user_id'], scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), status=1)
            except SpUserSchemes.DoesNotExist:
                free_scheme = None
            
            if free_scheme:
                container_quantity = 0
                
                if free_scheme.pouch_quantity>0:
                    if request.POST[packaging_type] == '0':
                        if free_scheme.packaging_type == 0:
                            order_value = float(quantity[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0    
                        elif free_scheme.packaging_type == 1:
                            pouch_value = float(quantity[id])*float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            order_value = int(pouch_value)/int(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0        
                    else:
                        if free_scheme.packaging_type == 0:
                            crate_value = float(quantity[id])/float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            if int(crate_value)>0:
                                order_value = int(crate_value)/int(free_scheme.minimum_order_quantity)
                                order_value = str(order_value).split('.')
                                order_value = int(order_value[0])
                            else:
                                order_value = 0 
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0
                        elif free_scheme.packaging_type == 1:
                            order_value = float(quantity[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0          
                else:    
                    pouch_quantity     = 0
                    
                if int(pouch_quantity) > 0:
                    logistic_order_scheme                             = SpLogisticOrderSchemes()
                    logistic_order_scheme.order_id                    = order.id
                    logistic_order_scheme.user_id                     = request.POST['user_id']
                    logistic_order_scheme.scheme_id                   = free_scheme.scheme_id
                    logistic_order_scheme.scheme_type                 = "free"
                    logistic_order_scheme.variant_id                  = free_scheme.applied_on_variant_id
                    logistic_order_scheme.on_order_of                 = free_scheme.minimum_order_quantity
                    logistic_order_scheme.free_variant_id             = free_scheme.free_variant_id
                    logistic_order_scheme.free_variant_container_type = free_scheme.order_container_name
                    logistic_order_scheme.container_quantity          = container_quantity
                    logistic_order_scheme.free_variant_packaging_type = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'packaging_type_name')
                    logistic_order_scheme.free_variant_container_size = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'container_size')
                    logistic_order_scheme.pouch_quantity              = pouch_quantity
                    logistic_order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'variant_size'))
                    logistic_order_scheme.product_id                  = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_id')
                    logistic_order_scheme.product_class_id            = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_class_id')
                    logistic_order_scheme.route_id                    = getModelColumnById(SpVehicles, vehicle_id[id], 'route_id')
                    logistic_order_scheme.route_name                  = getModelColumnById(SpVehicles, vehicle_id[id], 'route_name')
                    logistic_order_scheme.vehicle_id                  = getModelColumnById(SpVehicles, vehicle_id[id], 'id')
                    logistic_order_scheme.vehicle_number              = getModelColumnById(SpVehicles, vehicle_id[id], 'registration_number')
                    logistic_order_scheme.created_at                  = order.order_date
                    logistic_order_scheme.save()
            try:
                flat_scheme = SpUserFlatSchemes.objects.get(applied_on_variant_id=product_variant_id[id], user_id=request.POST['user_id'], scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), incentive_amount__gt=0, status=1)
            except SpUserFlatSchemes.DoesNotExist:
                flat_scheme = None
            
            if flat_scheme:
                if int(quantity[id]) > 0:
                    if request.POST[packaging_type] == '0':
                        pouch_quantity     = float(quantity[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        flat_discount = (float(quantity[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch')))*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    else:
                        flat_discount = float(quantity[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                        pouch_quantity = float(quantity[id])/float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        pouch_quantity = str(pouch_quantity).split('.')
                        pouch_quantity = int(pouch_quantity[0])
                        if pouch_quantity > 0: 
                            pouch_quantity     = int(pouch_quantity)
                        else:
                            pouch_quantity     = 0
                    logistic_order_scheme                             = SpLogisticOrderSchemes()
                    logistic_order_scheme.order_id                    = order.id
                    logistic_order_scheme.user_id                     = request.POST['user_id']
                    logistic_order_scheme.scheme_id                   = flat_scheme.scheme_id
                    logistic_order_scheme.scheme_type                 = "flat"
                    logistic_order_scheme.incentive_amount            = flat_discount
                    logistic_order_scheme.variant_id                  = flat_scheme.applied_on_variant_id
                    logistic_order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'variant_size'))
                    logistic_order_scheme.product_id                  = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_id')
                    logistic_order_scheme.product_class_id            = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_class_id')
                    logistic_order_scheme.route_id                    = getModelColumnById(SpVehicles, vehicle_id[id], 'route_id')
                    logistic_order_scheme.route_name                  = getModelColumnById(SpVehicles, vehicle_id[id], 'route_name')
                    logistic_order_scheme.vehicle_id                  = getModelColumnById(SpVehicles, vehicle_id[id], 'id')
                    logistic_order_scheme.vehicle_number              = getModelColumnById(SpVehicles, vehicle_id[id], 'registration_number')
                    logistic_order_scheme.created_at                  = order.order_date
                    logistic_order_scheme.save()

        for id, product_variant in enumerate(product_variant_id_list):
            # if packaging_type_counter > 0:
            index = int(id)+1
            # else:
            #     index = int(id)+2
            packaging_type = 'packaging_type_'+str(index)

            try:
                free_scheme = SpUserSchemes.objects.get(applied_on_variant_id=product_variant_id_list[id], user_id=request.POST['user_id'], scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), status=1)
            except SpUserSchemes.DoesNotExist:
                free_scheme = None
            
            if free_scheme:
                container_quantity = 0
                
                if free_scheme.pouch_quantity>0:
                    if request.POST[packaging_type] == '0':
                        if free_scheme.packaging_type == 0:
                            order_value = float(quantity_list[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0    
                        elif free_scheme.packaging_type == 1:
                            pouch_value = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            order_value = int(pouch_value)/int(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0        
                    else:
                        if free_scheme.packaging_type == 0:
                            crate_value = float(quantity_list[id])/float(getModelColumnById(SpProductVariants, free_scheme.applied_on_variant_id, 'no_of_pouch'))
                            if int(crate_value)>0:
                                order_value = int(crate_value)/int(free_scheme.minimum_order_quantity)
                                order_value = str(order_value).split('.')
                                order_value = int(order_value[0])
                            else:
                                order_value = 0
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0
                        elif free_scheme.packaging_type == 1:
                            order_value = float(quantity_list[id])/float(free_scheme.minimum_order_quantity)
                            order_value = str(order_value).split('.')
                            order_value = int(order_value[0])
                            if int(order_value) > 0:
                                pouch_quantity     = int(order_value)*int(free_scheme.pouch_quantity)
                            else:
                                pouch_quantity     = 0          
                else:    
                    pouch_quantity     = 0  
                if int(pouch_quantity) > 0:
                    order_scheme                             = SpOrderSchemes()
                    order_scheme.order_id                    = order.id
                    order_scheme.user_id                     = request.POST['user_id']
                    order_scheme.scheme_id                   = free_scheme.scheme_id
                    order_scheme.scheme_type                 = "free"
                    order_scheme.variant_id                  = free_scheme.applied_on_variant_id
                    order_scheme.on_order_of                 = free_scheme.minimum_order_quantity
                    order_scheme.free_variant_id             = free_scheme.free_variant_id
                    order_scheme.free_variant_container_type = free_scheme.order_container_name
                    order_scheme.container_quantity          = container_quantity
                    order_scheme.free_variant_packaging_type = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'packaging_type_name')
                    order_scheme.free_variant_container_size = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'container_size')
                    order_scheme.pouch_quantity              = pouch_quantity
                    order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'variant_size'))
                    order_scheme.product_id                  = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_id')
                    order_scheme.product_class_id            = getModelColumnById(SpProductVariants,free_scheme.free_variant_id,'product_class_id')
                    order_scheme.created_at                  = order.order_date
                    order_scheme.save()
            try:
                flat_scheme = SpUserFlatSchemes.objects.get(applied_on_variant_id=product_variant_id_list[id], user_id=request.POST['user_id'], scheme_start_date__lte=today.strftime("%Y-%m-%d"), scheme_end_date__gte=today.strftime("%Y-%m-%d"), incentive_amount__gt=0, status=1)
            except SpUserFlatSchemes.DoesNotExist:
                flat_scheme = None
            
            if flat_scheme:
                if int(quantity_list[id]) > 0:
                    if request.POST[packaging_type] == '0':
                        pouch_quantity     = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        flat_discount = (float(quantity_list[id])*int(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch')))*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                    else:
                        flat_discount = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'variant_size'))
                        flat_discount = round((float(flat_scheme.incentive_amount)*float(flat_discount)),2)
                        pouch_quantity = float(quantity_list[id])/float(getModelColumnById(SpProductVariants, flat_scheme.applied_on_variant_id, 'no_of_pouch'))
                        pouch_quantity = str(pouch_quantity).split('.')
                        pouch_quantity = int(pouch_quantity[0])
                        if pouch_quantity > 0: 
                            pouch_quantity     = int(pouch_quantity)
                        else:
                            pouch_quantity     = 0

                    order_scheme                             = SpOrderSchemes()
                    order_scheme.order_id                    = order.id
                    order_scheme.user_id                     = request.POST['user_id']
                    order_scheme.scheme_id                   = flat_scheme.scheme_id
                    order_scheme.scheme_type                 = "flat"
                    order_scheme.incentive_amount            = flat_discount
                    order_scheme.variant_id                  = flat_scheme.applied_on_variant_id
                    order_scheme.quantity_in_ltr             = int(pouch_quantity)*float(getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'variant_size'))
                    order_scheme.product_id                  = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_id')
                    order_scheme.product_class_id            = getModelColumnById(SpProductVariants,flat_scheme.applied_on_variant_id,'product_class_id')
                    order_scheme.created_at                  = order.order_date
                    order_scheme.save()  
            if request.POST[packaging_type] == '0':
                pouch_quantity     = int(float(quantity_list[id])*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'no_of_pouch')))
                quantity_in_ltr    = float(pouch_quantity)*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size'))
            else:
                pouch_quantity     = int(quantity_list[id])
                quantity_in_ltr    = float(quantity_list[id])*float(getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size'))
            order_detail                            = SpOrderDetails()    
            order_detail.user_id                    = request.POST['user_id']
            order_detail.order_id                   = order.id
            order_detail.product_id                 = product_id_list[id]
            order_detail.product_name               = getModelColumnById(SpProducts, product_id_list[id], 'product_name') 
            order_detail.product_variant_id         = product_variant_id_list[id]
            order_detail.product_variant_name       = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_name') 
            order_detail.product_variant_size       = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'variant_size')
            order_detail.product_no_of_pouch        = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'no_of_pouch')
            order_detail.product_container_size     = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'container_size')
            order_detail.product_container_type     = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'container_name')
            order_detail.quantity_in_pouch          = pouch_quantity
            order_detail.quantity_in_ltr            = round(quantity_in_ltr,2)
            order_detail.quantity                   = quantity_list[id]
            order_detail.rate                       = rate_list[id]
            order_detail.amount                     = amount_list[id]
            order_detail.packaging_type             = request.POST[packaging_type]
            order_detail.gst                        = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'gst')
            order_detail.product_packaging_type_name= getModelColumnById(SpProductVariants, product_variant_id_list[id], 'packaging_type_name')
            order_detail.is_allow                   = getModelColumnById(SpProductVariants, product_variant_id_list[id], 'is_allow_in_packaging')
            order_detail.order_date                 = order.order_date    
            order_detail.save()
        
        order.order_total_amount                    =  sum(total_order_amount)
        newtoday                                    = today.strftime("%Y-%m-%d")
        current                                     = int((today.strftime("%Y")))
        prev                                        = int(today.strftime("%Y"))-1
        naxt                                        = int(today.strftime("%Y"))+1
        
        if newtoday == datetime.strptime(str(order.order_date), '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d'):
            today_date                                  = datetime.now().strftime('%Y')+'-'+'04'+'-'+'01'
            if newtoday > today_date:
                financial_year_start_date               = datetime.now().strftime('%Y')+'-'+'04'+'-'+'01'
            else:
                financial_year_start_date               = datetime.now().strftime(str(prev))+'-'+'04'+'-'+'01'
            order_amounts                               = SpOrders.objects.filter(user_id=request.POST['user_id'],block_unblock=1,order_date__range=[financial_year_start_date,newtoday]).aggregate(Sum('order_total_amount'))['order_total_amount__sum']  
            if order_amounts:
                order_amounts                               = round((order_amounts+order.order_total_amount),2)
            else:
                order_amounts                               = round((order.order_total_amount),2)
            tcsvalues                                   = SpTcsMaster.objects.filter().first()
            if float(order_amounts) > tcsvalues.tcs_value:
                order.tcs_value                         =  tcsvalues.tcs_percentage
            else:
                order.tcs_value                         = None
        order.order_items_count     =  len(product_variant_id)
        order.updated_date          =  datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        order.platform_updated      =  'WEB'
        order.save()
        
    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
    heading     = 'Order('+order.order_code+') has been updated.'
    activity    = 'Order('+order.order_code+') has been updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
    
    saveActivity('Order Management', 'Order Summary', heading, activity, request.user.id, user_name, 'Orderplaced.png', '2', 'app.png')

    messages.success(request, heading, extra_tags='success')
    return redirect('/orders')
       
#get order status view
@login_required
@has_par(sub_module_id=8,permission='view')
def orderStatusDetails(request):
    order_id                    = request.GET.get('order_id')
    initiate_order_details      = SpOrders.objects.get(id=order_id)  
    order_details               = SpApprovalStatus.objects.filter(row_id=order_id, model_name='SpOrders', status=1).values('final_status_user_id').distinct().values('final_status_user_name', 'final_update_date_time', 'level_id')
    
    context = {}
    context['initiate_order_details']   = initiate_order_details
    context['order_details']            = order_details
    template = 'order-management/order-status-details.html'

    return render(request, template, context)

def getFreeScheme(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.filter(order_id=order_id, free_variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        sum_free_scheme = 0
        for scheme in free_scheme:
            scheme = (int(getModelColumnById(SpProductVariants, scheme.free_variant_id, 'no_of_pouch'))*int(scheme.container_quantity))+int(scheme.pouch_quantity)  
            sum_free_scheme = sum_free_scheme+scheme
        free_scheme = sum_free_scheme
    else:
        free_scheme = 0
    return free_scheme

def getlogFreeScheme(product_variant_id, order_id, user_id,vehicle_id):
    try:
        free_scheme = SpLogisticOrderSchemes.objects.filter(order_id=order_id, free_variant_id=product_variant_id, scheme_type='free', user_id=user_id,vehicle_id=vehicle_id)
    except SpLogisticOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        sum_free_scheme = 0
        for scheme in free_scheme:
            scheme = (int(getModelColumnById(SpProductVariants, scheme.free_variant_id, 'no_of_pouch'))*int(scheme.container_quantity))+int(scheme.pouch_quantity)  
            sum_free_scheme = sum_free_scheme+scheme
        free_scheme = sum_free_scheme
    else:
        free_scheme = 0
    return free_scheme    

def getOrderFreeScheme(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free_scheme = (int(getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'no_of_pouch'))*int(free_scheme.container_quantity))+int(free_scheme.pouch_quantity)  
    else:
        free_scheme = 0
    return free_scheme    

def getlogisticOrderFreeScheme(product_variant_id, order_id, user_id,vehicle_id):
    try:
        free_scheme = SpLogisticOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id,vehicle_id=vehicle_id)
    except SpLogisticOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free_scheme = (int(getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'no_of_pouch'))*int(free_scheme.container_quantity))+int(free_scheme.pouch_quantity)  
    else:
        free_scheme = 0
    return free_scheme    


def getSummaryFreeSchemeInPoches(product_variant_id, order_date,unblock_order):
    try:
        free_scheme = SpOrderSchemes.objects.filter(free_variant_id=product_variant_id, scheme_type='free', created_at__icontains=order_date,order_id__in=unblock_order)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        sum_free_scheme = 0
        for scheme in free_scheme:
            scheme = (int(getModelColumnById(SpProductVariants, scheme.free_variant_id, 'no_of_pouch'))*int(scheme.container_quantity))+int(scheme.pouch_quantity)  
            sum_free_scheme = sum_free_scheme+scheme
        free_scheme = sum_free_scheme
    else:
        free_scheme = 0
    return free_scheme

def getcratetotal(product_variant_id, order_id, user_id,vehicle_id):
    try:
        free_scheme = SpLogisticOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id,vehicle_id=vehicle_id)
    except SpLogisticOrderSchemes.DoesNotExist:
        free_scheme = None
        
    if free_scheme:
        cratenumber=0    
        if free_scheme.pouch_quantity>0:
            if free_scheme.pouch_quantity == 1:
                pass
            else:    
                Noofpouch=getModelColumnById(SpProductVariants,free_scheme.free_variant_id, 'no_of_pouch')
                half_crate_pouch=int(Noofpouch/2)
                cratenumber=int(free_scheme.pouch_quantity//Noofpouch)
                remain=int(free_scheme.pouch_quantity%Noofpouch)
                if remain>=half_crate_pouch:
                    cratenumber=cratenumber+1
        return cratenumber


def getSummaryFreeSchemeInLiterKg(product_variant_id, order_date,unblock_order):
    try:
        free_scheme = SpOrderSchemes.objects.filter(free_variant_id=product_variant_id, scheme_type='free', created_at__icontains=order_date,order_id__in=unblock_order)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        sum_free_scheme = 0
        for scheme in free_scheme:
            liter_kg     = float(scheme.pouch_quantity)
            scheme  = (liter_kg*float(getModelColumnById(SpProductVariants, scheme.free_variant_id, 'variant_size')))
            sum_free_scheme = sum_free_scheme+scheme
        free_scheme = sum_free_scheme
    else:
        free_scheme = 0
    return free_scheme

def getSummaryFreeSchemeInCrate(product_variant_id, order_date,unblock_order):
    try:
        free_scheme = SpOrderSchemes.objects.filter(free_variant_id=product_variant_id, scheme_type='free', created_at__icontains=order_date,order_id__in=unblock_order)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        sum_free_scheme = 0
        for scheme in free_scheme:
            liter_kg     = float(scheme.quantity_in_ltr)
            sum_free_scheme = sum_free_scheme+liter_kg
        free_scheme = sum_free_scheme
    else:
        free_scheme = 0
    return free_scheme
    
def getSummaryBonusSchemeInPoches(product_variant_id, order_date,unblock_order):
    try:
        bonus_scheme = SpOrderSchemes.objects.filter(free_variant_id=product_variant_id, scheme_type='quantitative', created_at__icontains=order_date,order_id__in=unblock_order)
    except SpOrderSchemes.DoesNotExist:
        bonus_scheme = None
    if bonus_scheme:
        sum_bonus_scheme = []
        for scheme in bonus_scheme:
            scheme = float(scheme.quantity_in_ltr) 
            sum_bonus_scheme.append(scheme) 
        bonus_scheme = sum(sum_bonus_scheme)
    else:
        bonus_scheme = 0    
    return bonus_scheme

def getSummaryBonusSchemeInLiterKg(product_variant_id, order_date,unblock_order):
    try:
        bonus_scheme = SpOrderSchemes.objects.filter(free_variant_id=product_variant_id, scheme_type='quantitative', created_at__icontains=order_date,order_id__in=unblock_order)
    except SpOrderSchemes.DoesNotExist:
        bonus_scheme = None
    if bonus_scheme:
        sum_bonus_scheme = []
        for scheme in bonus_scheme:
            liter_kg     = float(scheme.quantity_in_ltr)
            sum_bonus_scheme.append(liter_kg) 
        bonus_scheme = sum(sum_bonus_scheme)
    else:
        bonus_scheme = 0
    return bonus_scheme

def getFlatBulkSchemeIncentive(order_id, user_id, scheme_type):
    try:
        free_scheme = SpOrderSchemes.objects.filter(order_id=order_id, scheme_type=scheme_type, user_id=user_id).aggregate(Sum('incentive_amount'))['incentive_amount__sum']
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free_scheme = free_scheme
    else:
        free_scheme = 0
    return free_scheme

def getlogFlatBulkSchemeIncentive(order_id, user_id, scheme_type,vehicle_id):
    try:
        free_scheme = SpLogisticOrderSchemes.objects.filter(order_id=order_id, scheme_type=scheme_type, user_id=user_id,vehicle_id=vehicle_id).aggregate(Sum('incentive_amount'))['incentive_amount__sum']
    except SpLogisticOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free_scheme = free_scheme
    else:
        free_scheme = 0
    return free_scheme


def getFlatBulkSchemeIncentive(order_id, user_id, scheme_type):
    try:
        free_scheme = SpOrderSchemes.objects.filter(order_id=order_id, scheme_type=scheme_type, user_id=user_id).aggregate(Sum('incentive_amount'))['incentive_amount__sum']
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free_scheme = free_scheme
    else:
        free_scheme = 0
    return free_scheme

def getEmployeesTotalQunatity(product_variant_id):
    try:
        quantity = SpHoReport.objects.filter(product_variant_id=product_variant_id).aggregate(Sum('quantity'))['quantity__sum']
    except SpOrderSchemes.DoesNotExist:
        quantity = None
    if quantity:
        quantity = quantity
    else:
        quantity = 0
    return quantity

def getEmployeesTotalQunatityInLtr(product_variant_id):
    try:
        quantity = SpHoReport.objects.filter(product_variant_id=product_variant_id, quantity__gt=0)
    except SpOrderSchemes.DoesNotExist:
        quantity = None
    if quantity:
        qty_in_ltr = []
        for product_quantity in quantity:
            liter_kg     = int(product_quantity.quantity)
            total_quantity_in_ltr  = (liter_kg*float(getModelColumnById(SpProductVariants, product_quantity.product_variant_id, 'variant_size')))
            qty_in_ltr.append(total_quantity_in_ltr) 
        quantity = sum(qty_in_ltr)
    else:
        quantity = 0
    return quantity

def getFocTotalQunatity(product_variant_id):
    try:
        quantity = SpHoReport.objects.filter(product_variant_id=product_variant_id).aggregate(Sum('foc_pouch'))['foc_pouch__sum']
    except SpOrderSchemes.DoesNotExist:
        quantity = None
    if quantity:
        quantity = quantity
    else:
        quantity = 0
    return quantity

def getFocTotalQunatityInLtr(product_variant_id):
    try:
        quantity = SpHoReport.objects.filter(product_variant_id=product_variant_id, foc_pouch__gt=0)
    except SpOrderSchemes.DoesNotExist:
        quantity = None
    if quantity:
        qty_in_ltr = []
        for product_quantity in quantity:
            liter_kg     = int(product_quantity.foc_pouch)
            total_quantity_in_ltr  = (liter_kg*float(getModelColumnById(SpProductVariants, product_quantity.product_variant_id, 'variant_size')))
            qty_in_ltr.append(total_quantity_in_ltr) 
        quantity = sum(qty_in_ltr)
    else:
        quantity = 0
    return quantity
        
#get order indent Report
@login_required
@has_par(sub_module_id=9,permission='list')
def indentReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today.strftime("%Y-%m-%d")).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0,block_unblock=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()

    user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment,sp_orders.block_unblock, sp_orders.amount_to_be_paid
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_orders on sp_orders.user_id = sp_users.id
    where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s and sp_orders.block_unblock = %s''',[today.strftime("%Y-%m-%d"), 3,1])

    product_milk_variant_list           =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='1' and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''')
    product_without_milk_variant_list   =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='1' and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''')
    milk_product_class_name             = SpProductClass.objects.get(id=1)

    indent_lists                                = []
    total_crate_sum                             = []
    total_free_milk_pouches_sum                 = []
    total_free_without_milk_pouches_sum         = []
    total_without_milk_crate_sum                = []
    total_milk_crates_quantity                  = []
    total_free_milk_pouches_quantity            = []
    total_free_without_milk_pouches_quantity    = []
    total_flat_scheme_incentive                 = []
    total_bulk_scheme_incentive                 = []
    total_milk_ltr_list_sum                     = []
    total_without_milk_ltr_sum                  = []
    for user in user_list:
        flat_scheme_incentive = getFlatBulkSchemeIncentive(user.order_id, user.id, 'flat')
        total_flat_scheme_incentive.append(flat_scheme_incentive)

        bulk_scheme_incentive = getFlatBulkSchemeIncentive(user.order_id, user.id, 'bulkpack')
        total_bulk_scheme_incentive.append(bulk_scheme_incentive)

        indent = {}
        product_milk_variants_list      = []
        total_milk_crates               = []
        total_free_scheme_milk_pouches  = []
        total_milk_ltr_list             = []
        for id, product_variant_milk in enumerate(product_milk_variant_list):
            total = 0
            total_free_scheme = 0
            total_milk_in_ltr = 0 
            product_variant_milk_lists = {}
            total_crates = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                if order_details.packaging_type == '0':
                    Item_in_liters    = float(order_details.quantity)
                else:
                    Item_in_liters    = round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                total             = total + Item_in_liters
                total_milk_in_ltr     = order_details.quantity_in_ltr
            else:
                Item_in_liters    = 0
                total             = total + Item_in_liters

            free_scheme       = getFreeScheme(product_variant_milk.id, user.order_id, user.id)
            total_free_scheme = total_free_scheme + free_scheme

            product_variant_milk_lists['id']                    = product_variant_milk.id    
            product_variant_milk_lists['milk_items']            = Item_in_liters
            product_variant_milk_lists['product_color_code']    = product_variant_milk.product_color_code
            product_variant_milk_lists['free_scheme']           = free_scheme
            
            total_milk_crates.append(total)
            total_crates = total_milk_crates
            total_free_scheme_milk_pouches.append(total_free_scheme)
            total_free_scheme_milk = total_free_scheme_milk_pouches  
            product_milk_variants_list.append(product_variant_milk_lists)
            total_milk_ltr_list.append(total_milk_in_ltr)
       
        total_milk_ltr_list_sum.append(total_milk_ltr_list)
        total_crate_sum.append(total_crates)
        sum_total_crates = sum(total_milk_crates)
        total_milk_crates_quantity.append(sum_total_crates)

        total_free_milk_pouches_sum.append(total_free_scheme_milk)
        sum_total_free_milk_pouches_sum = sum(total_free_scheme_milk_pouches)
        total_free_milk_pouches_quantity.append(sum_total_free_milk_pouches_sum)
        

        product_without_milk_variants_list = []
        total_without_milk_crates = []
        total_without_milk_ltr = []
        total_free_scheme_without_milk_pouches = []
        for product_variant_without_milk in product_without_milk_variant_list:
            totals = 0
            total_free_scheme = 0
            total_without_milk_in_ltr = 0
            product_variant_without_milk_lists = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_without_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                if order_details.packaging_type == '0':
                    Item_in_liters    = float(order_details.quantity)
                else:
                    Item_in_liters    = round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                totals              = totals + Item_in_liters
                total_without_milk_in_ltr     = order_details.quantity_in_ltr
            else:
                Item_in_liters = 0
                totals = totals + Item_in_liters
            
            free_scheme         = getFreeScheme(product_variant_without_milk.id, user.order_id, user.id)
            total_free_scheme   = total_free_scheme + free_scheme

            product_variant_without_milk_lists['id']                 = product_variant_without_milk.id    
            product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
            product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
            product_variant_without_milk_lists['free_scheme']        = free_scheme

            total_without_milk_crates.append(totals)
            total_without_milk_crates = total_without_milk_crates

            total_free_scheme_without_milk_pouches.append(total_free_scheme)
            total_free_scheme_without_milk = total_free_scheme_without_milk_pouches  

            product_without_milk_variants_list.append(product_variant_without_milk_lists) 

            total_without_milk_ltr.append(total_without_milk_in_ltr)    
     
          
        total_without_milk_crate_sum.append(total_without_milk_crates)
        total_without_milk_ltr_sum.append(total_without_milk_ltr) 
        total_free_without_milk_pouches_sum.append(total_free_scheme_without_milk)
        sum_total_free_without_milk_pouches_sum = sum(total_free_scheme_without_milk_pouches)
        total_free_without_milk_pouches_quantity.append(sum_total_free_without_milk_pouches_sum)

        balance_after_deposit = (float(user.order_total_amount)+float(user.outstanding_amount))-float(user.amount_to_be_paid)
        balance_security      = float(user.security_amount)-balance_after_deposit
        
        indent['id']                    = user.id
        indent['first_name']            = user.first_name
        indent['middle_name']           = user.middle_name
        indent['last_name']             = user.last_name
        indent['emp_sap_id']            = user.emp_sap_id
        indent['store_name']            = user.store_name
        indent['town_name']             = user.town_name
        indent['order_total_amount']    = user.order_total_amount
        indent['outstanding_amount']    = user.outstanding_amount
        indent['security_amount']       = user.security_amount
        indent['mode_of_payment']       = user.mode_of_payment
        indent['amount_to_be_paid']     = user.amount_to_be_paid
        indent['balance_after_deposit'] = balance_after_deposit
        indent['balance_security']      = balance_security
        indent['milk_items']            = product_milk_variants_list
        indent['without_milk_items']    = product_without_milk_variants_list
        indent_lists.append(indent)
    
    total_milk_qty_ltr_all = []
    #sum of milk product ltr
    if total_milk_ltr_list_sum:
        for column in enumerate(total_milk_ltr_list_sum[0]):
            count = round((sum([x[column[0]] for x in total_milk_ltr_list_sum])),2)
            total_milk_qty_ltr_all.append(count)

    total_withoutmilk_qty_ltr_all = []
    #sum of withoutmilk product ltr
    if total_without_milk_ltr_sum:
        for column in enumerate(total_without_milk_ltr_sum[0]):
            count = round((sum([x[column[0]] for x in total_without_milk_ltr_sum])),2)
            total_withoutmilk_qty_ltr_all.append(count) 

        
    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = round((sum([x[column[0]] for x in total_crate_sum])),2)
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = round((sum([x[column[0]] for x in total_without_milk_crate_sum])),2)
            total_without_milk_crates_qty.append(count)
    
    #sum of milk product pouches
    total_free_milk_pouches_qty = []
    if total_free_milk_pouches_sum:
        for column in enumerate(total_free_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_milk_pouches_sum])
            total_free_milk_pouches_qty.append(count)

    #sum of without milk product pouches
    total_free_without_milk_pouches_qty = []
    if total_free_without_milk_pouches_sum:
        for column in enumerate(total_free_without_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_without_milk_pouches_sum])
            total_free_without_milk_pouches_qty.append(count)        
    #print(total_free_without_milk_pouches_qty)
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today.strftime("%Y-%m-%d")).count()

    total_incentive_amount = [x + y for x, y in zip(total_flat_scheme_incentive, total_bulk_scheme_incentive)]
    
    organizations=SpOrganizations.objects.all()
    
    context = {}
    context["organizations"]                        = organizations
    context['user_list']                            = user_list
    context['indent_lists']                         = indent_lists
    context['total_milk_crates']                    = total_milk_crates_quantity
    context['total_free_milk_pouches']              = total_free_milk_pouches_qty
    context['total_milk_crates_qty']                = total_milk_crates_qty
    context['total_without_milk_crates_qty']        = total_without_milk_crates_qty
    context['total_milk_qty_ltr_all']               = total_milk_qty_ltr_all
    context['total_withoutmilk_qty_ltr_all']        = total_withoutmilk_qty_ltr_all
    context['total_free_without_milk_pouches_qty']  = total_free_without_milk_pouches_qty
    context['product_milk_variant_list']            = product_milk_variant_list
    context['product_without_milk_variant_list']    = product_without_milk_variant_list
    context['total_flat_scheme_incentive']          = total_flat_scheme_incentive
    context['total_bulk_scheme_incentive']          = total_bulk_scheme_incentive
    context['total_incentive_amount']               = total_incentive_amount
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['milk_product_class_name']              = milk_product_class_name
    context['page_title']                           = "Indent Report"
    template                                        = 'order-management/indent-report.html'
    
    return render(request, template, context)    


#get order indent Report
@login_required
@has_par(sub_module_id=9,permission='list')
def ajaxIndentReport(request):
    organization_id         = request.GET['organization_id']
    today                   = request.GET['order_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()

    if request.GET['organization_id']:
        user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        left join sp_orders on sp_orders.user_id = sp_users.id
        where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s and sp_orders.block_unblock = %s and sp_users.organization_id=%s''',[today,3,1,organization_id])
    else:   
        user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        left join sp_orders on sp_orders.user_id = sp_users.id
        where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s  and sp_orders.block_unblock = %s ''',[today, 3,1])

    product_milk_variant_list           =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='1' and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''')
    product_without_milk_variant_list   =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='1' and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''')
    milk_product_class_name             = SpProductClass.objects.get(id=1)

    indent_lists                                = []
    total_crate_sum                             = []
    total_free_milk_pouches_sum                 = []
    total_free_without_milk_pouches_sum         = []
    total_without_milk_crate_sum                = []
    total_milk_crates_quantity                  = []
    total_free_milk_pouches_quantity            = []
    total_free_without_milk_pouches_quantity    = []
    total_flat_scheme_incentive                 = []
    total_bulk_scheme_incentive                 = []
    total_milk_ltr_list_sum                     = []
    total_without_milk_ltr_sum                  = []
    for user in user_list:
        flat_scheme_incentive = getFlatBulkSchemeIncentive(user.order_id, user.id, 'flat')
        total_flat_scheme_incentive.append(flat_scheme_incentive)

        bulk_scheme_incentive = getFlatBulkSchemeIncentive(user.order_id, user.id, 'bulkpack')
        total_bulk_scheme_incentive.append(bulk_scheme_incentive)

        indent = {}
        product_milk_variants_list      = []
        total_milk_crates               = []
        total_free_scheme_milk_pouches  = []
        total_milk_ltr_list             = []
        for id, product_variant_milk in enumerate(product_milk_variant_list):
            total = 0
            total_free_scheme = 0
            total_milk_in_ltr=0
            product_variant_milk_lists = {}
            total_crates = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                if order_details.packaging_type == '0':
                    Item_in_liters    = float(order_details.quantity)
                else:
                    Item_in_liters = round(
                        (float(order_details.quantity)/float(order_details.product_no_of_pouch)), 2)
                total                 = total + Item_in_liters
                total_milk_in_ltr     = order_details.quantity_in_ltr
            else:
                Item_in_liters    = 0
                total             = total + Item_in_liters

            free_scheme       = getFreeScheme(product_variant_milk.id, user.order_id, user.id)
            total_free_scheme = total_free_scheme + free_scheme

            product_variant_milk_lists['id']                    = product_variant_milk.id    
            product_variant_milk_lists['milk_items']            = Item_in_liters
            product_variant_milk_lists['product_color_code']    = product_variant_milk.product_color_code
            product_variant_milk_lists['free_scheme']           = free_scheme
            
            total_milk_crates.append(total)
            total_crates = total_milk_crates
            total_free_scheme_milk_pouches.append(total_free_scheme)
            total_free_scheme_milk = total_free_scheme_milk_pouches  
            product_milk_variants_list.append(product_variant_milk_lists)
            total_milk_ltr_list.append(total_milk_in_ltr)
        total_crate_sum.append(total_crates)
        total_milk_ltr_list_sum.append(total_milk_ltr_list)
        sum_total_crates = sum(total_milk_crates)
        total_milk_crates_quantity.append(sum_total_crates)

        total_free_milk_pouches_sum.append(total_free_scheme_milk)
        sum_total_free_milk_pouches_sum = sum(total_free_scheme_milk_pouches)
        total_free_milk_pouches_quantity.append(sum_total_free_milk_pouches_sum)
        

        product_without_milk_variants_list = []
        total_without_milk_crates = []
        total_without_milk_ltr = []
        total_free_scheme_without_milk_pouches = []
        
        for product_variant_without_milk in product_without_milk_variant_list:
            totals = 0
            total_free_scheme = 0
            total_without_milk_in_ltr=0
            product_variant_without_milk_lists = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_without_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                if order_details.packaging_type == '0':
                    Item_in_liters    = float(order_details.quantity)
                else:
                    Item_in_liters = round(
                        (float(order_details.quantity)/float(order_details.product_no_of_pouch)), 2)
                totals              = totals + Item_in_liters
                total_without_milk_in_ltr     = order_details.quantity_in_ltr
            else:
                Item_in_liters = 0
                totals = totals + Item_in_liters
            
            free_scheme         = getFreeScheme(product_variant_without_milk.id, user.order_id, user.id)
            total_free_scheme   = total_free_scheme + free_scheme

            product_variant_without_milk_lists['id']                 = product_variant_without_milk.id    
            product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
            product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
            product_variant_without_milk_lists['free_scheme']        = free_scheme

            total_without_milk_crates.append(totals)
            total_without_milk_crates = total_without_milk_crates

            total_free_scheme_without_milk_pouches.append(total_free_scheme)
            total_free_scheme_without_milk = total_free_scheme_without_milk_pouches  

            product_without_milk_variants_list.append(product_variant_without_milk_lists)
            total_without_milk_ltr.append(total_without_milk_in_ltr)    
        total_without_milk_crate_sum.append(total_without_milk_crates)
        total_without_milk_ltr_sum.append(total_without_milk_ltr)
        total_free_without_milk_pouches_sum.append(total_free_scheme_without_milk)
        sum_total_free_without_milk_pouches_sum = sum(total_free_scheme_without_milk_pouches)
        total_free_without_milk_pouches_quantity.append(sum_total_free_without_milk_pouches_sum)
        
        balance_after_deposit = (float(user.order_total_amount)+float(user.outstanding_amount))-float(user.amount_to_be_paid)
        balance_security      = float(user.security_amount)-balance_after_deposit
        
        indent['id']                    = user.id
        indent['first_name']            = user.first_name
        indent['middle_name']           = user.middle_name
        indent['last_name']             = user.last_name
        indent['emp_sap_id']            = user.emp_sap_id
        indent['store_name']            = user.store_name
        indent['town_name']             = user.town_name
        indent['order_total_amount']    = user.order_total_amount
        indent['outstanding_amount']    = user.outstanding_amount
        indent['security_amount']       = user.security_amount
        indent['mode_of_payment']       = user.mode_of_payment
        indent['amount_to_be_paid']     = user.amount_to_be_paid
        indent['balance_after_deposit'] = balance_after_deposit
        indent['balance_security']      = balance_security
        indent['milk_items']            = product_milk_variants_list
        indent['without_milk_items']    = product_without_milk_variants_list
        indent_lists.append(indent)
    
   
    total_milk_qty_ltr_all = []
    #sum of milk product ltr
    if total_milk_ltr_list_sum:
        for column in enumerate(total_milk_ltr_list_sum[0]):
            count = round((sum([x[column[0]] for x in total_milk_ltr_list_sum])),2)
            total_milk_qty_ltr_all.append(count)

    total_withoutmilk_qty_ltr_all = []
    #sum of withoutmilk product ltr
    if total_without_milk_ltr_sum:
        for column in enumerate(total_without_milk_ltr_sum[0]):
            count = round((sum([x[column[0]] for x in total_without_milk_ltr_sum])),2)
            total_withoutmilk_qty_ltr_all.append(count) 

        
    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = round((sum([x[column[0]] for x in total_crate_sum])),2)
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = round((sum([x[column[0]] for x in total_without_milk_crate_sum])),2)
            total_without_milk_crates_qty.append(count)
    
    #sum of milk product pouches
    total_free_milk_pouches_qty = []
    if total_free_milk_pouches_sum:
        for column in enumerate(total_free_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_milk_pouches_sum])
            total_free_milk_pouches_qty.append(count)

    #sum of without milk product pouches
    total_free_without_milk_pouches_qty = []
    if total_free_without_milk_pouches_sum:
        for column in enumerate(total_free_without_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_without_milk_pouches_sum])
            total_free_without_milk_pouches_qty.append(count)        
    #print(total_free_without_milk_pouches_qty)
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    total_incentive_amount = [x + y for x, y in zip(total_flat_scheme_incentive, total_bulk_scheme_incentive)]
    
    today_date   = date.today()
    context = {}
    context['user_list']                            = user_list
    context['indent_lists']                         = indent_lists
    context['total_milk_crates']                    = total_milk_crates_quantity
    context['total_free_milk_pouches']              = total_free_milk_pouches_qty
    context['total_milk_crates_qty']                = total_milk_crates_qty
    context['total_milk_qty_ltr_all']               = total_milk_qty_ltr_all
    context['total_withoutmilk_qty_ltr_all']        = total_withoutmilk_qty_ltr_all
    context['total_without_milk_crates_qty']        = total_without_milk_crates_qty
    context['total_free_without_milk_pouches_qty']  = total_free_without_milk_pouches_qty
    context['product_milk_variant_list']            = product_milk_variant_list
    context['product_without_milk_variant_list']    = product_without_milk_variant_list
    context['total_flat_scheme_incentive']          = total_flat_scheme_incentive
    context['total_bulk_scheme_incentive']          = total_bulk_scheme_incentive
    context['total_incentive_amount']               = total_incentive_amount
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")
    context['milk_product_class_name']              = milk_product_class_name
    context['page_title']                           = "Indent Report"
    template                                        = 'order-management/ajax-indent-report.html'
    
    return render(request, template, context)

#get order indent Report
@login_required
@has_par(sub_module_id=9,permission='edit')
def generateIndentReport(request):
    response = {}
    today   = date.today()
    if request.GET['id'] == '0':
        msg = 'generating'
    else:
        msg = 're-generating'      
    try:
        approved_order = SpOrders.objects.filter(order_status__gte=3,block_unblock=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
        today_order = SpOrders.objects.filter(block_unblock=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
        if today_order == 0:
            response['error']       = True
            response['error_type']  = 'danger'
            response['message']     = "No Order Found."
        elif approved_order == today_order:
            orders   = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d"))
            for order in orders:
                data = SpOrders.objects.get(id=order.id)
                data.indent_status = 1
                data.save()

            response['error']       = False
            response['error_type']  = 'success'
            response['message']     = "Indent Report is "+msg+". Please wait..."
        else:
            response['error']       = True
            response['error_type']  = 'danger'
            response['message']     = "All the orders for the date("+today.strftime("%d/%m/%Y")+") are not approved. Kindly approved the orders first."   
    except ObjectDoesNotExist:
        response['error']       = True
        response['error_type']  = 'danger'
        response['message']     = "Method not allowed"
    except Exception as e:
        response['error']       = True
        response['error_type']  = 'danger'
        response['message']     = e
    return JsonResponse(response)


#get order indent Report
@login_required
@has_par(sub_module_id=9,permission='export')
def exportIndentReport(request, order_date, organization_id):
    organization_id         = organization_id
    today                   = order_date
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()

    if organization_id!='0':
        user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        left join sp_orders on sp_orders.user_id = sp_users.id
        where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s and sp_orders.block_unblock = %s and sp_users.organization_id=%s''',[today,3,1,organization_id])
    else:   
        user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        left join sp_orders on sp_orders.user_id = sp_users.id
        where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s and sp_orders.block_unblock = %s ''',[today, 3,1])

    product_milk_variant_list           =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='1' and sp_product_variants.status='1' order by sp_product_variants.order_of asc ''')
    product_without_milk_variant_list   =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='1' and sp_product_variants.status='1' order by sp_product_variants.order_of asc ''')
    milk_product_class_name             = SpProductClass.objects.get(id=1)

    indent_lists                                = []
    total_crate_sum                             = []
    total_free_milk_pouches_sum                 = []
    total_free_without_milk_pouches_sum         = []
    total_without_milk_crate_sum                = []
    total_milk_crates_quantity                  = []
    total_free_milk_pouches_quantity            = []
    total_free_without_milk_pouches_quantity    = []
    total_flat_scheme_incentive                 = []
    total_bulk_scheme_incentive                 = []
    total_milk_ltr_list_sum                     = []
    total_without_milk_ltr_sum                  = []
    for user in user_list:
        flat_scheme_incentive = getFlatBulkSchemeIncentive(user.order_id, user.id, 'flat')
        total_flat_scheme_incentive.append(flat_scheme_incentive)

        bulk_scheme_incentive = getFlatBulkSchemeIncentive(user.order_id, user.id, 'bulkpack')
        total_bulk_scheme_incentive.append(bulk_scheme_incentive)

        indent = {}
        product_milk_variants_list      = []
        total_milk_crates               = []
        total_free_scheme_milk_pouches  = []
        total_milk_ltr_list             = []
        
        for id, product_variant_milk in enumerate(product_milk_variant_list):
            total = 0
            total_free_scheme = 0
            total_milk_in_ltr = 0 
            product_variant_milk_lists = {}
            total_crates = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                if order_details.packaging_type == '0':
                    Item_in_liters    = float(order_details.quantity)
                else:
                    Item_in_liters    = round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                total             = total + Item_in_liters
                total_milk_in_ltr     = order_details.quantity_in_ltr
            else:
                Item_in_liters    = 0
                total             = total + Item_in_liters

            free_scheme       = getFreeScheme(product_variant_milk.id, user.order_id, user.id)
            total_free_scheme = total_free_scheme + free_scheme

            product_variant_milk_lists['id']                    = product_variant_milk.id    
            product_variant_milk_lists['milk_items']            = Item_in_liters
            product_variant_milk_lists['product_color_code']    = product_variant_milk.product_color_code
            product_variant_milk_lists['free_scheme']           = free_scheme
            
            total_milk_crates.append(total)
            total_crates = total_milk_crates
            total_free_scheme_milk_pouches.append(total_free_scheme)
            total_free_scheme_milk = total_free_scheme_milk_pouches  
            product_milk_variants_list.append(product_variant_milk_lists)
            total_milk_ltr_list.append(total_milk_in_ltr)
        
        total_milk_ltr_list_sum.append(total_milk_ltr_list)
        total_crate_sum.append(total_crates)
        sum_total_crates = sum(total_milk_crates)
        total_milk_crates_quantity.append(sum_total_crates)

        total_free_milk_pouches_sum.append(total_free_scheme_milk)
        sum_total_free_milk_pouches_sum = sum(total_free_scheme_milk_pouches)
        total_free_milk_pouches_quantity.append(sum_total_free_milk_pouches_sum)
        

        product_without_milk_variants_list = []
        total_without_milk_crates = []
        total_without_milk_ltr = []
        total_free_scheme_without_milk_pouches = []
        for product_variant_without_milk in product_without_milk_variant_list:
            totals = 0
            total_free_scheme = 0
            total_without_milk_in_ltr=0
            product_variant_without_milk_lists = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_without_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                if order_details.packaging_type == '0':
                    Item_in_liters    = float(order_details.quantity)
                else:
                    Item_in_liters = round(
                        (float(order_details.quantity)/float(order_details.product_no_of_pouch)), 2)
                totals              = totals + Item_in_liters
                total_without_milk_in_ltr     = order_details.quantity_in_ltr
            else:
                Item_in_liters = 0
                totals = totals + Item_in_liters
            
            free_scheme         = getFreeScheme(product_variant_without_milk.id, user.order_id, user.id)
            total_free_scheme   = total_free_scheme + free_scheme

            product_variant_without_milk_lists['id']                 = product_variant_without_milk.id    
            product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
            product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
            product_variant_without_milk_lists['free_scheme']        = free_scheme

            total_without_milk_crates.append(totals)
            total_without_milk_crates = total_without_milk_crates

            total_free_scheme_without_milk_pouches.append(total_free_scheme)
            total_free_scheme_without_milk = total_free_scheme_without_milk_pouches  
            product_without_milk_variants_list.append(product_variant_without_milk_lists)  
            total_without_milk_ltr.append(total_without_milk_in_ltr)    
        
          
        total_without_milk_crate_sum.append(total_without_milk_crates)
        total_without_milk_ltr_sum.append(total_without_milk_ltr)
        total_free_without_milk_pouches_sum.append(total_free_scheme_without_milk)
        sum_total_free_without_milk_pouches_sum = sum(total_free_scheme_without_milk_pouches)
        total_free_without_milk_pouches_quantity.append(sum_total_free_without_milk_pouches_sum)

        balance_after_deposit = (float(user.order_total_amount)+float(user.outstanding_amount))-float(user.amount_to_be_paid)
        balance_security      = float(user.security_amount)-balance_after_deposit
        indent['id']                    = user.id
        indent['first_name']            = user.first_name
        indent['middle_name']           = user.middle_name
        indent['last_name']             = user.last_name
        indent['emp_sap_id']            = user.emp_sap_id
        indent['store_name']            = user.store_name
        indent['town_name']             = user.town_name
        indent['order_total_amount']    = user.order_total_amount
        indent['outstanding_amount']    = user.outstanding_amount
        indent['security_amount']       = user.security_amount
        indent['mode_of_payment']       = user.mode_of_payment
        indent['amount_to_be_paid']     = user.amount_to_be_paid
        indent['balance_after_deposit'] = balance_after_deposit
        indent['balance_security']      = balance_security
        indent['milk_items']            = product_milk_variants_list
        indent['without_milk_items']    = product_without_milk_variants_list
        indent_lists.append(indent)
    
    total_milk_qty_ltr_all = []
    #sum of milk product ltr
    if total_milk_ltr_list_sum:
        for column in enumerate(total_milk_ltr_list_sum[0]):
            count = round((sum([x[column[0]] for x in total_milk_ltr_list_sum])),2)
            total_milk_qty_ltr_all.append(count)

    total_withoutmilk_qty_ltr_all = []
    #sum of withoutmilk product ltr
    if total_without_milk_ltr_sum:
        for column in enumerate(total_without_milk_ltr_sum[0]):
            count = round((sum([x[column[0]] for x in total_without_milk_ltr_sum])),2)
            total_withoutmilk_qty_ltr_all.append(count) 

        
    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = round((sum([x[column[0]] for x in total_crate_sum])),2)
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = round((sum([x[column[0]] for x in total_without_milk_crate_sum])),2)
            total_without_milk_crates_qty.append(count)
    
    #sum of milk product pouches
    total_free_milk_pouches_qty = []
    if total_free_milk_pouches_sum:
        for column in enumerate(total_free_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_milk_pouches_sum])
            total_free_milk_pouches_qty.append(count)

    #sum of without milk product pouches
    total_free_without_milk_pouches_qty = []
    if total_free_without_milk_pouches_sum:
        for column in enumerate(total_free_without_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_without_milk_pouches_sum])
            total_free_without_milk_pouches_qty.append(count)        
    
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    total_incentive_amount = [x + y for x, y in zip(total_flat_scheme_incentive, total_bulk_scheme_incentive)]

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=indent-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'INDENT REPORT'
    worksheet.merge_cells('A1:B1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 8.5
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=3)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(name='Calibri',size=10, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+11
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
    worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=5).value = 'Shreedhi'
    worksheet.cell(row=1, column=5).font = header_font
    worksheet.cell(row=1, column=5).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=5).font = Font(size=24, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=5).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    color_codes = []

    columns = []
    columns += [ 'Customer Code' ]
    columns += [ 'Name of Distributor/SS' ]
    columns += [ 'Town' ]

    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            color_code = str(product_variant.product_color_code).replace('#', '')
            color_codes.append(color_code)
            columns += [ product_variant.variant_name ]

    columns += [ 'TOTAL '+milk_product_class_name.product_class+' CRATES' ]
    color_code = 'FFFFFFFF'
    color_codes.append(color_code)
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            color_code = str(product_variant.product_color_code).replace('#', '')
            color_codes.append(color_code)
            columns += [ product_variant.variant_name ]
    # columns += [ 'FLAT INCENTIVE' ]
    # columns += [ 'BULK PACK INCENTIVE' ]
    columns += [ 'TOTAL DISCOUNT AMOUNT' ]
    columns += [ 'INVOICE AMOUNT' ]
    columns += [ 'PAYMENT DETAILS' ]
    columns += [ 'OUTSTANDING DETAILS' ]
    columns += [ 'SECURITY' ]
    columns += [ 'BALANCE AFTER DEPOSIT' ]
    columns += [ 'BALANCE SECURITY' ]

    row_num = 2
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num <=3:
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            
        elif col_num <= len(product_milk_variant_list)+3:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=10, bold=True)
                cell.fill = PatternFill()
        elif col_num == len(product_milk_variant_list)+4:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=10, color='000000', bold=True)
                cell.fill = PatternFill()        
        elif col_num <= len(product_milk_variant_list)+len(product_without_milk_variant_list)+4:
            for product_variant in product_without_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=10, bold=True)
                cell.fill = PatternFill()        
        else:
            cell.fill = PatternFill()
            cell.font = Font(size=10, color='000000', bold=True)              
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num<=3:
            column_dimensions.width = 15  
        else:
            column_dimensions.width = 6

    
    for id, indent in enumerate(indent_lists):
        row_num += 1
        # Define the data for each cell in the row 
        total_incetive = float(total_flat_scheme_incentive[id]) + float(total_bulk_scheme_incentive[id])    
        row = []
        row += [ indent['emp_sap_id'] ]
        # += [  indent['first_name'] +" "+indent['middle_name']+" "+indent['last_name']+" ("+indent['store_name']+")"  ]
        row += [  indent['store_name']+"("+indent['first_name'] +" "+indent['middle_name']+" "+indent['last_name']+")"  ]
        row += [ indent['town_name'] ]
        if indent['milk_items']:
            for item in indent['milk_items']:
                row += [ item['milk_items'] ]
        row += [ total_milk_crates_quantity[id] ]
        if indent['without_milk_items']:
            for item in indent['without_milk_items']:
                row += [ item['without_milk_items'] ]
        # row += [ total_flat_scheme_incentive[id] ]
        # row += [ total_bulk_scheme_incentive[id] ]
        row += [ total_incetive ]
        row += [ indent['order_total_amount']-total_incetive ]
        row += [ indent['amount_to_be_paid'] ]
        row += [ indent['outstanding_amount'] ]
        row += [ indent['security_amount'] ]
        row += [ indent['balance_after_deposit']-total_incetive ]
        row += [ indent['balance_security']+total_incetive ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num <=3:
                cell.font = Font(name='Calibri',size=12, color='000000')
            elif col_num <= len(product_milk_variant_list)+3:
                for product_variant in product_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(size=10, color='000000', bold=True)
                    cell.fill = PatternFill()
            elif col_num == len(product_milk_variant_list)+4:
                for product_variant in product_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(name='Calibri',size=12, color='000000')
                    cell.fill = PatternFill()        
            elif col_num <= len(product_milk_variant_list)+len(product_without_milk_variant_list)+4:
                for product_variant in product_without_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(size=10, color='000000', bold=True)
                    cell.fill = PatternFill()        
            else:
                cell.fill = PatternFill()
                cell.font = Font(size=12, color='000000')
            if col_num<=3:
                column_dimensions.width = 15 
            else:
                column_dimensions.width = 6
    total_columns = len(color_codes)+3           
    for rows in worksheet.iter_rows(min_row=2, max_row=len(indent_lists)+2, min_col=4, max_col=total_columns):
        for id, cell in enumerate(rows):
            cell.fill = PatternFill(start_color=color_codes[id], end_color=color_codes[id], fill_type = "solid")            
    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ 'GRAND TOTAL(CRATES)' ]

    if total_milk_crates_qty:
        for total_milk_crates in total_milk_crates_qty:
            bottom_columns += [ total_milk_crates ]

    bottom_columns += [ ' ' ]
    if total_without_milk_crates_qty:
        for total_without_milk_crates in total_without_milk_crates_qty:
            bottom_columns += [ total_without_milk_crates ]
    # bottom_columns += [ ' ' ]
    # bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = wrapped_alignment
        cell.font = Font(size=10, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        # column_dimensions.width = 8
        if col_num<=3:
            column_dimensions.width = 15  
        else:
            column_dimensions.width = 6
#------------------------------------------------------------------------------------
    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ 'TOTAL QUANTITY IN LTR' ]

    if total_milk_qty_ltr_all:
        for total_milk_ltr in total_milk_qty_ltr_all:
            bottom_columns += [ total_milk_ltr ]

    
    bottom_columns += [ ' ' ]
    if total_withoutmilk_qty_ltr_all:
        for total_without_milk in total_withoutmilk_qty_ltr_all:
            bottom_columns += [ total_without_milk ]
    # bottom_columns += [ ' ' ]
    # bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        # worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = wrapped_alignment
        cell.font = Font(size=10, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        # column_dimensions.width = 8
        if col_num<=3:
            column_dimensions.width = 15 
        else:
            column_dimensions.width = 6
#------------------------------------------------------------------------------------




    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' TRADE SCHEME POUCHES ' ]

    if total_free_milk_pouches_qty:
        for total_milk_pouches in total_free_milk_pouches_qty:
            bottom_columns += [ total_milk_pouches ]

    bottom_columns += [ ' ' ]
    if total_free_without_milk_pouches_qty:
        for total_without_milk_pouches in total_free_without_milk_pouches_qty:
            bottom_columns += [ total_without_milk_pouches ]
    # bottom_columns += [ ' ' ]
    # bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = wrapped_alignment
        cell.font = Font(size=10, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        # column_dimensions.width = 20
        if col_num<=3:
            column_dimensions.width = 15  
        else:
            column_dimensions.width = 6
    row_num += 1
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)
    # img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.png')
    # img.height = 50
    # img.alignment = wrapped_alignment
    # img.anchor = 'W'+str(last_row)
    # worksheet.add_image(img)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=10, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response
    
#get summary Report
@login_required
@has_par(sub_module_id=26,permission='list')
def summaryReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    unblock_order      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today).values_list('id', flat=True)
    product_classes = SpProductClass.objects.all().order_by('order_of')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_product_quantity_with_scheme = []
        total_quantity_wo_scheme_crate_sum = []
        total_scheme_quantity_crate_sum = []
        total_bonus_scheme_quantity_crate_sum = []
        total_employee_sale_crate_sum = []
        total_foc_quantity_crate_sum = []

        total_pouch_product_quantity_with_scheme = []
        total_pouch_wo_scheme_crate_sum = []
        total_pouch_scheme_crate_sum = []
        total_pouch_bonus_scheme_crate_sum = []
        total_pouch_employee_sale_crate_sum = []
        total_pouch_foc_crate_sum = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_quantity_wo_scheme = 0
                total_scheme = 0
                total_bonus_scheme = 0
                total_employee_sale = 0
                total_foc_quantity = 0

                total_pouch_wo_scheme = 0
                total_pouch_scheme = 0
                total_pouch_bonus_scheme = 0
                total_pouch_employee_sale = 0
                total_pouch_foc = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today.strftime("%Y-%m-%d"),order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today.strftime("%Y-%m-%d"),order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0

                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None

                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                                  = float(product_total_quantity)
                    product_variant.product_total_quantity_wo_scheme    = quantity_wo_scheme
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + quantity_wo_scheme

                    pouch_wo_scheme                                     = int(product_pouch_quantity)
                    product_variant.product_total_pouch_wo_scheme       = pouch_wo_scheme
                    total_pouch_wo_scheme                               = total_pouch_wo_scheme + pouch_wo_scheme
                else:
                    product_variant.product_total_quantity_wo_scheme    = 0
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + 0
                    
                    product_variant.product_total_pouch_wo_scheme        = 0
                    total_pouch_wo_scheme                                = total_pouch_wo_scheme + 0
                
                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)
                total_scheme                                        = total_scheme + getSummaryFreeSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)
                total_bonus_scheme                                  = total_bonus_scheme + getSummaryBonusSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)

                product_variant.product_total_employee_sale_quantity = round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)
                total_employee_sale                                  = total_employee_sale + round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)

                product_variant.product_total_foc_quantity           = round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)
                total_foc_quantity                                   = total_foc_quantity + round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)

                product_variant.product_total_pouch_scheme           = getSummaryFreeSchemeInPoches(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)
                total_pouch_scheme                                   = total_pouch_scheme + getSummaryFreeSchemeInPoches(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)

                product_variant.product_total_pouch_bonus_scheme      = getSummaryBonusSchemeInPoches(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)   
                total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + getSummaryBonusSchemeInPoches(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)    

                product_variant.product_total_pouch_employee_sale     = int(getEmployeesTotalQunatity(product_variant.id))
                total_pouch_employee_sale                             = total_pouch_employee_sale + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_pouch_foc               = int(getFocTotalQunatity(product_variant.id))
                total_pouch_foc                                       = total_pouch_foc + int(getFocTotalQunatity(product_variant.id))

                product_variant.total_product_quantity_with_scheme        = round((product_variant.product_total_quantity_wo_scheme+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity),2)

                total_product_quantity_with_scheme.append(product_variant.total_product_quantity_with_scheme)
                total_quantity_wo_scheme_crate_sum.append(total_quantity_wo_scheme) 
                total_scheme_quantity_crate_sum.append(total_scheme)
                total_bonus_scheme_quantity_crate_sum.append(total_bonus_scheme)
                total_employee_sale_crate_sum.append(total_employee_sale)
                total_foc_quantity_crate_sum.append(total_foc_quantity)

                product_variant.total_pouch_product_quantity_with_scheme   =  round((product_variant.product_total_pouch_wo_scheme+product_variant.product_total_pouch_scheme+product_variant.product_total_pouch_bonus_scheme+product_variant.product_total_pouch_employee_sale+product_variant.product_total_pouch_foc),2)

                total_pouch_product_quantity_with_scheme.append(product_variant.total_pouch_product_quantity_with_scheme)
                total_pouch_wo_scheme_crate_sum.append(total_pouch_wo_scheme)
                total_pouch_scheme_crate_sum.append(total_pouch_scheme)
                total_pouch_bonus_scheme_crate_sum.append(total_pouch_bonus_scheme)
                total_pouch_employee_sale_crate_sum.append(total_pouch_employee_sale)
                total_pouch_foc_crate_sum.append(total_pouch_foc)
        
            product_class.total_product_quantity_with_scheme             = sum(total_product_quantity_with_scheme)
            product_class.total_quantity_wo_scheme_crate_sum             = sum(total_quantity_wo_scheme_crate_sum) 
            product_class.total_scheme_quantity_crate_sum                = sum(total_scheme_quantity_crate_sum)
            product_class.total_bonus_scheme_quantity_crate_sum          = sum(total_bonus_scheme_quantity_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

            product_class.total_pouch_product_quantity_with_scheme       = sum(total_pouch_product_quantity_with_scheme)
            product_class.total_pouch_wo_scheme_crate_sum                = sum(total_pouch_wo_scheme_crate_sum)
            product_class.total_pouch_scheme_crate_sum                   = sum(total_pouch_scheme_crate_sum)
            product_class.total_pouch_bonus_scheme_crate_sum             = sum(total_pouch_bonus_scheme_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

    products = SpProducts.objects.filter(status=1).order_by('id')
    products_list = []
    for product in products:
        total_quantity_wo_scheme_crate_sums = []
        total_scheme_quantity_crate_sums = []
        total_bonus_scheme_quantity_crate_sums = []
        total_employee_sale_crate_sums = []
        total_foc_quantity_crate_sums = []

        product_variant_list    =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_id=%s and sp_product_variants.status='1' order by sp_product_variants.product_id desc ''', [product.id])
        
        product_list = {}
        for product_variant in product_variant_list:
            total_quantity_wo_schemes   = 0
            total_schemes               = 0
            total_bonus_schemes         = 0
            total_employee_sales        = 0
            total_foc_quantitys         = 0

            try:
                product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today.strftime("%Y-%m-%d"),order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                product_total_quantity = product_total_quantity['quantity_in_ltr']
            except SpOrderDetails.DoesNotExist:
                product_total_quantity = 0

                
            try:
                order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
            except SpOrderDetails.DoesNotExist:
                order_details = None

            if order_details:
                variant_size    = order_details.product_variant_size
                no_of_pouch     = order_details.product_no_of_pouch
            else:
                variant_size    = 0
                no_of_pouch     = 0

            if product_total_quantity:
                quantity_wo_schemes  = float(product_total_quantity)
                product_variant.product_total_quantity_wo_scheme    = quantity_wo_schemes
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + quantity_wo_schemes
            else:
                product_variant.product_total_quantity_wo_scheme    = 0
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + 0
            
            product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)
            total_schemes                                       = total_schemes + getSummaryFreeSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)

            product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)
            total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInLiterKg(product_variant.id, today.strftime("%Y-%m-%d"),unblock_order)

            product_variant.product_total_employee_sale_quantity = round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)
            total_employee_sales                                 = total_employee_sales + round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)

            product_variant.product_total_foc_quantity           = round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)
            total_foc_quantitys                                  = total_foc_quantitys + round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)

            total_quantity_wo_scheme_crate_sums.append(total_quantity_wo_schemes)
            total_scheme_quantity_crate_sums.append(total_schemes)
            total_bonus_scheme_quantity_crate_sums.append(total_bonus_schemes)
            total_employee_sale_crate_sums.append(total_employee_sales) 
            total_foc_quantity_crate_sums.append(total_foc_quantitys)
            
        product_list['product_name'] = product.product_name
        product_list['total_quantity'] = round((sum(total_quantity_wo_scheme_crate_sums)+sum(total_scheme_quantity_crate_sums)+sum(total_bonus_scheme_quantity_crate_sums)+sum(total_employee_sale_crate_sums)+sum(total_foc_quantity_crate_sums)),2)
        products_list.append(product_list)

    b=1
    condition = ""
    # conditions += " and sp_order_details.order_id in ("+str(order_id)+")"
    condition += " and sp_orders.order_date LIKE '%%"+str(today)+"%%'" 
    condition += " and sp_orders.block_unblock = %s" % b
    # condition = " and sp_orders.order_date LIKE '%%"+str(today.strftime("%Y-%m-%d"))+"%%'" 
    town_wise_orders = SpOrders.objects.raw(''' SELECT sp_orders.id, sp_orders.town_id, sp_orders.town_name  FROM sp_orders WHERE 1 {condition} group by sp_orders.town_id '''.format(condition=condition))    
    
    town_list = []
    for town in town_wise_orders:
        order_id = SpOrders.objects.filter(town_id = town.town_id,block_unblock=1, order_date__icontains=today.strftime("%Y-%m-%d")).values_list('id', flat=True)
        towns = {}
        towns['town_id']     = town.town_id
        towns['town_name']   = town.town_name

        product_class = SpProductClass.objects.filter(status=1).order_by('-id')
        for  id, product in enumerate(product_class):
            towns[product.id]   = getTownWiseVariantInLiters(product.id, order_id)

        town_list.append(towns)


    approved_order = SpOrders.objects.filter(order_status = 3,block_unblock=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    today_order = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today.strftime("%Y-%m-%d")).count()

    context = {}
    context['product_classes']                      = product_classes
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['products_list']                        = products_list
    context['town_list']                            = town_list
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['page_title']                           = "Summary Report"
    template                                        = 'order-management/summary-report.html'

    return render(request, template, context) 
             
#get summary Report
@login_required
@has_par(sub_module_id=26,permission='list')
def ajaxSummaryReport(request):
    today                   = request.GET['order_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    unblock_order      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today).values_list('id', flat=True)
    product_classes = SpProductClass.objects.all().order_by('order_of')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_product_quantity_with_scheme = []
        total_quantity_wo_scheme_crate_sum = []
        total_scheme_quantity_crate_sum = []
        total_bonus_scheme_quantity_crate_sum = []
        total_employee_sale_crate_sum = []
        total_foc_quantity_crate_sum = []

        total_pouch_product_quantity_with_scheme = []
        total_pouch_wo_scheme_crate_sum = []
        total_pouch_scheme_crate_sum = []
        total_pouch_bonus_scheme_crate_sum = []
        total_pouch_employee_sale_crate_sum = []
        total_pouch_foc_crate_sum = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_quantity_wo_scheme = 0
                total_scheme = 0
                total_bonus_scheme = 0
                total_employee_sale = 0
                total_foc_quantity = 0

                total_pouch_wo_scheme = 0
                total_pouch_scheme = 0
                total_pouch_bonus_scheme = 0
                total_pouch_employee_sale = 0
                total_pouch_foc = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0
                    
                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None
                
                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                                  = float(product_total_quantity)
                    product_variant.product_total_quantity_wo_scheme    = quantity_wo_scheme
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + quantity_wo_scheme

                    pouch_wo_scheme                                     = int(product_pouch_quantity)
                    product_variant.product_total_pouch_wo_scheme       = pouch_wo_scheme
                    total_pouch_wo_scheme                               = total_pouch_wo_scheme + pouch_wo_scheme
                else:
                    product_variant.product_total_quantity_wo_scheme    = 0
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + 0
                    
                    product_variant.product_total_pouch_wo_scheme        = 0
                    total_pouch_wo_scheme                                = total_pouch_wo_scheme + 0
                
                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)
                total_scheme                                        = total_scheme + getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)
                total_bonus_scheme                                  = total_bonus_scheme + getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)

                product_variant.product_total_employee_sale_quantity = round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)
                total_employee_sale                                  = total_employee_sale + round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)

                product_variant.product_total_foc_quantity           = round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)
                total_foc_quantity                                   = total_foc_quantity + round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)

                product_variant.product_total_pouch_scheme           = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_pouch_scheme                                   = total_pouch_scheme + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_pouch_bonus_scheme      = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)   
                total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)    

                product_variant.product_total_pouch_employee_sale     = int(getEmployeesTotalQunatity(product_variant.id))
                total_pouch_employee_sale                             = total_pouch_employee_sale + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_pouch_foc               = int(getFocTotalQunatity(product_variant.id))
                total_pouch_foc                                       = total_pouch_foc + int(getFocTotalQunatity(product_variant.id))

                product_variant.total_product_quantity_with_scheme        = round((product_variant.product_total_quantity_wo_scheme+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity),2)

                total_product_quantity_with_scheme.append(product_variant.total_product_quantity_with_scheme)
                total_quantity_wo_scheme_crate_sum.append(total_quantity_wo_scheme) 
                total_scheme_quantity_crate_sum.append(total_scheme)
                total_bonus_scheme_quantity_crate_sum.append(total_bonus_scheme)
                total_employee_sale_crate_sum.append(total_employee_sale)
                total_foc_quantity_crate_sum.append(total_foc_quantity)

                product_variant.total_pouch_product_quantity_with_scheme   =  round((product_variant.product_total_pouch_wo_scheme+product_variant.product_total_pouch_scheme+product_variant.product_total_pouch_bonus_scheme+product_variant.product_total_pouch_employee_sale+product_variant.product_total_pouch_foc),2)

                total_pouch_product_quantity_with_scheme.append(product_variant.total_pouch_product_quantity_with_scheme)
                total_pouch_wo_scheme_crate_sum.append(total_pouch_wo_scheme)
                total_pouch_scheme_crate_sum.append(total_pouch_scheme)
                total_pouch_bonus_scheme_crate_sum.append(total_pouch_bonus_scheme)
                total_pouch_employee_sale_crate_sum.append(total_pouch_employee_sale)
                total_pouch_foc_crate_sum.append(total_pouch_foc)
        
            product_class.total_product_quantity_with_scheme             = sum(total_product_quantity_with_scheme)
            product_class.total_quantity_wo_scheme_crate_sum             = sum(total_quantity_wo_scheme_crate_sum) 
            product_class.total_scheme_quantity_crate_sum                = sum(total_scheme_quantity_crate_sum)
            product_class.total_bonus_scheme_quantity_crate_sum          = sum(total_bonus_scheme_quantity_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

            product_class.total_pouch_product_quantity_with_scheme       = sum(total_pouch_product_quantity_with_scheme)
            product_class.total_pouch_wo_scheme_crate_sum                = sum(total_pouch_wo_scheme_crate_sum)
            product_class.total_pouch_scheme_crate_sum                   = sum(total_pouch_scheme_crate_sum)
            product_class.total_pouch_bonus_scheme_crate_sum             = sum(total_pouch_bonus_scheme_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

    products = SpProducts.objects.filter(status=1).order_by('id')
    products_list = []
    for product in products:
        total_quantity_wo_scheme_crate_sums = []
        total_scheme_quantity_crate_sums = []
        total_bonus_scheme_quantity_crate_sums = []
        total_employee_sale_crate_sums = []
        total_foc_quantity_crate_sums = []

        product_variant_list    =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_id=%s and sp_product_variants.status='1'  order by sp_product_variants.product_id desc ''', [product.id])
        
        product_list = {}
        for product_variant in product_variant_list:
            total_quantity_wo_schemes   = 0
            total_schemes               = 0
            total_bonus_schemes         = 0
            total_employee_sales        = 0
            total_foc_quantitys         = 0

            try:
                product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                product_total_quantity = product_total_quantity['quantity_in_ltr']
            except SpOrderDetails.DoesNotExist:
                product_total_quantity = 0

            try:
                order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
            except SpOrderDetails.DoesNotExist:
                order_details = None

            if order_details:
                variant_size    = order_details.product_variant_size
                no_of_pouch     = order_details.product_no_of_pouch
            else:
                variant_size    = 0
                no_of_pouch     = 0

            if product_total_quantity:
                quantity_wo_schemes  = float(product_total_quantity)
                product_variant.product_total_quantity_wo_scheme    = quantity_wo_schemes
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + quantity_wo_schemes
            else:
                product_variant.product_total_quantity_wo_scheme    = 0
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + 0
            
            product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)
            total_schemes                                       = total_schemes + getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)

            product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)
            total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)

            product_variant.product_total_employee_sale_quantity = round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)
            total_employee_sales                                 = total_employee_sales + round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)

            product_variant.product_total_foc_quantity           = round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)
            total_foc_quantitys                                  = total_foc_quantitys + round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)

            total_quantity_wo_scheme_crate_sums.append(total_quantity_wo_schemes)
            total_scheme_quantity_crate_sums.append(total_schemes)
            total_bonus_scheme_quantity_crate_sums.append(total_bonus_schemes)
            total_employee_sale_crate_sums.append(total_employee_sales) 
            total_foc_quantity_crate_sums.append(total_foc_quantitys)
            
        product_list['product_name'] = product.product_name
        product_list['total_quantity'] = round((sum(total_quantity_wo_scheme_crate_sums)+sum(total_scheme_quantity_crate_sums)+sum(total_bonus_scheme_quantity_crate_sums)+sum(total_employee_sale_crate_sums)+sum(total_foc_quantity_crate_sums)),2)
        products_list.append(product_list)

    b=1
    condition = ""
    # conditions += " and sp_order_details.order_id in ("+str(order_id)+")"
    condition += " and sp_orders.order_date LIKE '%%"+str(today)+"%%'" 
    condition += " and sp_orders.block_unblock = %s" % b
    town_wise_orders = SpOrders.objects.raw(''' SELECT sp_orders.id, sp_orders.town_id, sp_orders.town_name  FROM sp_orders WHERE 1 {condition} group by sp_orders.town_id '''.format(condition=condition))    
    
    town_list = []
    for town in town_wise_orders:
        order_id = SpOrders.objects.filter(town_id = town.town_id, order_date__icontains=today,block_unblock=1).values_list('id', flat=True)
        towns = {}
        towns['town_id']     = town.town_id
        towns['town_name']   = town.town_name

        product_class = SpProductClass.objects.filter(status=1).order_by('-id')
        for  id, product in enumerate(product_class):
            towns[product.id]   = getTownWiseVariantInLiters(product.id, order_id)

        town_list.append(towns)

    
    approved_order = SpOrders.objects.filter(order_status = 3,block_unblock=1, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today).count()

    today_date   = date.today()
    context = {}
    context['product_classes']                      = product_classes
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['products_list']                        = products_list
    context['town_list']                            = town_list
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")
    context['page_title']                           = "Summary Report"
    template                                        = 'order-management/ajax-summary-report.html'

    return render(request, template, context) 

def getTownFreeSchemeInLiterKg(product_variant_id, order_id):
    try:
        free_scheme = SpOrderSchemes.objects.filter(order_id=order_id, free_variant_id=product_variant_id, scheme_type='free')
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        sum_free_scheme = 0
        for scheme in free_scheme:
            liter_kg     = float(scheme.quantity_in_ltr)
            sum_free_scheme = sum_free_scheme+liter_kg
        free_scheme = sum_free_scheme
    else:
        free_scheme = 0
    return free_scheme

def getTownBonusSchemeInLiterKg(product_variant_id, order_id):
    try:
        bonus_scheme = SpOrderSchemes.objects.filter(order_id=order_id, free_variant_id=product_variant_id, scheme_type='quantitative')
    except SpOrderSchemes.DoesNotExist:
        bonus_scheme = None
    if bonus_scheme:
        sum_bonus_scheme = []
        for scheme in bonus_scheme:
            liter_kg     = float(scheme.quantity_in_ltr)
            sum_bonus_scheme.append(liter_kg) 
        bonus_scheme = sum(sum_bonus_scheme)
    else:
        bonus_scheme = 0
    return bonus_scheme

def getTownWiseVariantInLiters(product_class_id, order_id):
    order_id = list(order_id)
    order_id = str(order_id).replace('[', '')
    order_id = str(order_id).replace(']', '')
    
    conditions = ""
    conditions += " and sp_order_details.order_id in ("+str(order_id)+")"
    conditions += " and sp_product_variants.product_class_id = %s" % product_class_id

    product_details = SpOrderDetails.objects.raw(''' SELECT sp_order_details.id, sp_order_details.order_id, sp_order_details.product_variant_id, sp_order_details.product_variant_name, sp_order_details.product_no_of_pouch, sp_order_details.product_variant_size, sp_order_details.quantity, sp_order_details.quantity_in_ltr
    FROM sp_order_details left join sp_product_variants on sp_product_variants.id = sp_order_details.product_variant_id 
    WHERE 1 {conditions} '''.format(conditions=conditions))
   
    total_qty_in_ltr = []
    for products in product_details:
        product_quantity_ltr_kg             = products.quantity_in_ltr 
        free_scheme_product_qty_in_ltr      = getTownFreeSchemeInLiterKg(products.product_variant_id, products.order_id)
        bonus_scheme_product_qty_in_ltr     = getTownBonusSchemeInLiterKg(products.product_variant_id, products.order_id)

        total_qty_in_ltrs = product_quantity_ltr_kg+free_scheme_product_qty_in_ltr+bonus_scheme_product_qty_in_ltr
        total_qty_in_ltr.append(total_qty_in_ltrs)

    return round(sum(total_qty_in_ltr),2)

#get order summary Report
@login_required
@has_par(sub_module_id=26,permission='export')
def exportSummaryReport(request, order_date):
    today                   = order_date
    today_order_status      = SpOrders.objects.filter(block_unblock=1,indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    unblock_order      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today).values_list('id', flat=True)
    product_classes = SpProductClass.objects.all().order_by('order_of')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s and sp_product_variants.status='1' order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_product_quantity_with_scheme = []
        total_quantity_wo_scheme_crate_sum = []
        total_scheme_quantity_crate_sum = []
        total_bonus_scheme_quantity_crate_sum = []
        total_employee_sale_crate_sum = []
        total_foc_quantity_crate_sum = []

        total_pouch_product_quantity_with_scheme = []
        total_pouch_wo_scheme_crate_sum = []
        total_pouch_scheme_crate_sum = []
        total_pouch_bonus_scheme_crate_sum = []
        total_pouch_employee_sale_crate_sum = []
        total_pouch_foc_crate_sum = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_quantity_wo_scheme = 0
                total_scheme = 0
                total_bonus_scheme = 0
                total_employee_sale = 0
                total_foc_quantity = 0

                total_pouch_wo_scheme = 0
                total_pouch_scheme = 0
                total_pouch_bonus_scheme = 0
                total_pouch_employee_sale = 0
                total_pouch_foc = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0

                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None

                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                                  = float(product_total_quantity)
                    product_variant.product_total_quantity_wo_scheme    = quantity_wo_scheme
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + quantity_wo_scheme

                    pouch_wo_scheme                                     = int(product_pouch_quantity)
                    product_variant.product_total_pouch_wo_scheme       = pouch_wo_scheme
                    total_pouch_wo_scheme                               = total_pouch_wo_scheme + pouch_wo_scheme
                else:
                    product_variant.product_total_quantity_wo_scheme    = 0
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + 0
                    
                    product_variant.product_total_pouch_wo_scheme        = 0
                    total_pouch_wo_scheme                                = total_pouch_wo_scheme + 0
                
                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)
                total_scheme                                        = total_scheme + getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)
                total_bonus_scheme                                  = total_bonus_scheme + getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)

                product_variant.product_total_employee_sale_quantity = round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)
                total_employee_sale                                  = total_employee_sale + round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)

                product_variant.product_total_foc_quantity           = round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)
                total_foc_quantity                                   = total_foc_quantity + round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)

                product_variant.product_total_pouch_scheme           = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_pouch_scheme                                   = total_pouch_scheme + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_pouch_bonus_scheme      = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)   
                total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)    

                product_variant.product_total_pouch_employee_sale     = int(getEmployeesTotalQunatity(product_variant.id))
                total_pouch_employee_sale                             = total_pouch_employee_sale + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_pouch_foc               = int(getFocTotalQunatity(product_variant.id))
                total_pouch_foc                                       = total_pouch_foc + int(getFocTotalQunatity(product_variant.id))

                product_variant.total_product_quantity_with_scheme        = round((product_variant.product_total_quantity_wo_scheme+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity),2)

                total_product_quantity_with_scheme.append(product_variant.total_product_quantity_with_scheme)
                total_quantity_wo_scheme_crate_sum.append(total_quantity_wo_scheme) 
                total_scheme_quantity_crate_sum.append(total_scheme)
                total_bonus_scheme_quantity_crate_sum.append(total_bonus_scheme)
                total_employee_sale_crate_sum.append(total_employee_sale)
                total_foc_quantity_crate_sum.append(total_foc_quantity)

                product_variant.total_pouch_product_quantity_with_scheme   =  round((product_variant.product_total_pouch_wo_scheme+product_variant.product_total_pouch_scheme+product_variant.product_total_pouch_bonus_scheme+product_variant.product_total_pouch_employee_sale+product_variant.product_total_pouch_foc),2)

                total_pouch_product_quantity_with_scheme.append(product_variant.total_pouch_product_quantity_with_scheme)
                total_pouch_wo_scheme_crate_sum.append(total_pouch_wo_scheme)
                total_pouch_scheme_crate_sum.append(total_pouch_scheme)
                total_pouch_bonus_scheme_crate_sum.append(total_pouch_bonus_scheme)
                total_pouch_employee_sale_crate_sum.append(total_pouch_employee_sale)
                total_pouch_foc_crate_sum.append(total_pouch_foc)
        
            product_class.total_product_quantity_with_scheme             = sum(total_product_quantity_with_scheme)
            product_class.total_quantity_wo_scheme_crate_sum             = sum(total_quantity_wo_scheme_crate_sum) 
            product_class.total_scheme_quantity_crate_sum                = sum(total_scheme_quantity_crate_sum)
            product_class.total_bonus_scheme_quantity_crate_sum          = sum(total_bonus_scheme_quantity_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

            product_class.total_pouch_product_quantity_with_scheme       = sum(total_pouch_product_quantity_with_scheme)
            product_class.total_pouch_wo_scheme_crate_sum                = sum(total_pouch_wo_scheme_crate_sum)
            product_class.total_pouch_scheme_crate_sum                   = sum(total_pouch_scheme_crate_sum)
            product_class.total_pouch_bonus_scheme_crate_sum             = sum(total_pouch_bonus_scheme_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

    products = SpProducts.objects.filter(status=1).order_by('id')
    products_list = []
    for product in products:
        total_quantity_wo_scheme_crate_sums = []
        total_scheme_quantity_crate_sums = []
        total_bonus_scheme_quantity_crate_sums = []
        total_employee_sale_crate_sums = []
        total_foc_quantity_crate_sums = []

        product_variant_list    =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_id=%s and sp_product_variants.status='1'  order by sp_product_variants.product_id desc ''', [product.id])
        
        product_list = {}
        for product_variant in product_variant_list:
            total_quantity_wo_schemes   = 0
            total_schemes               = 0
            total_bonus_schemes         = 0
            total_employee_sales        = 0
            total_foc_quantitys         = 0

            try:
                product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                product_total_quantity = product_total_quantity['quantity_in_ltr']
            except SpOrderDetails.DoesNotExist:
                product_total_quantity = 0

            try:
                order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
            except SpOrderDetails.DoesNotExist:
                order_details = None

            if order_details:
                variant_size    = order_details.product_variant_size
                no_of_pouch     = order_details.product_no_of_pouch
            else:
                variant_size    = 0
                no_of_pouch     = 0

            if product_total_quantity:
                quantity_wo_schemes  = float(product_total_quantity)
                product_variant.product_total_quantity_wo_scheme    = quantity_wo_schemes
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + quantity_wo_schemes
            else:
                product_variant.product_total_quantity_wo_scheme    = 0
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + 0
            
            product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)
            total_schemes                                       = total_schemes + getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)

            product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)
            total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInLiterKg(product_variant.id, today,unblock_order)

            product_variant.product_total_employee_sale_quantity = round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)
            total_employee_sales                                 = total_employee_sales + round(float(int(getEmployeesTotalQunatity(product_variant.id))*float(variant_size)), 2)

            product_variant.product_total_foc_quantity           = round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)
            total_foc_quantitys                                  = total_foc_quantitys + round(float(int(getFocTotalQunatity(product_variant.id))*float(variant_size)), 2)

            total_quantity_wo_scheme_crate_sums.append(total_quantity_wo_schemes)
            total_scheme_quantity_crate_sums.append(total_schemes)
            total_bonus_scheme_quantity_crate_sums.append(total_bonus_schemes)
            total_employee_sale_crate_sums.append(total_employee_sales) 
            total_foc_quantity_crate_sums.append(total_foc_quantitys)
            
        product_list['product_name'] = product.product_name
        product_list['total_quantity'] = round((sum(total_quantity_wo_scheme_crate_sums)+sum(total_scheme_quantity_crate_sums)+sum(total_bonus_scheme_quantity_crate_sums)+sum(total_employee_sale_crate_sums)+sum(total_foc_quantity_crate_sums)),2)
        products_list.append(product_list)
    b=1
    condition = ""
    # conditions += " and sp_order_details.order_id in ("+str(order_id)+")"
    condition += " and sp_orders.order_date LIKE '%%"+str(today)+"%%'" 
    condition += " and sp_orders.block_unblock = %s" % b
    # condition = " and sp_orders.order_date LIKE '%%"+str(today)+"%%'" 
    town_wise_orders = SpOrders.objects.raw(''' SELECT sp_orders.id, sp_orders.town_id, sp_orders.town_name  FROM sp_orders WHERE 1 {condition} group by sp_orders.town_id '''.format(condition=condition))    
    
    town_list = []
    for town in town_wise_orders:
        order_id = SpOrders.objects.filter(town_id = town.town_id, order_date__icontains=today,block_unblock=1).values_list('id', flat=True)
        towns = {}
        towns['town_id']     = town.town_id
        towns['town_name']   = town.town_name

        product_class = SpProductClass.objects.filter(status=1).order_by('-id')
        for  id, product in enumerate(product_class):
            towns[product.id]   = getTownWiseVariantInLiters(product.id, order_id)

        town_list.append(towns)

    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today,block_unblock=1).count()
    today_order = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today).count()

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=summary-report.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'SUMMARY REPORT'
    worksheet.merge_cells('A1:A1')

    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True 
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    length = []
    for product_variant in product_classes:
            if product_variant.product_variant_list:
                length.append(len(product_variant.product_variant_list))            

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(name='Arial Nova Cond Light',size=12, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = len(length)+sum(length)+1
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'Shreedhi'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=24, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    columns = []
    columns += [ 'PARTICULARS' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.variant_name ]
                columns += [ product_variant.product_variant_total ]
    
        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = 50
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

                cell.font = Font(size=14, color='000000', bold=True)
                cell.fill = PatternFill()                
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = 8

    columns = []
    columns += [ 'TOTAL QUANTITY W/O SCHEME (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_quantity_wo_scheme ]
                columns += [ product_variant.total_quantity_wo_scheme_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
        
    
    columns = []
    columns += [ 'TOTAL SCHEME QUANTITY (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_scheme_quantity ]
                columns += [ product_variant.total_scheme_quantity_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')

    columns = []
    columns += [ 'TOTAL BONUS SCHEME QUANTITY (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_bonus_scheme_quantity ]
                columns += [ product_variant.total_bonus_scheme_quantity_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
             

    columns = []
    columns += [ 'TOTAL EMPLOYEE SALE (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_employee_sale_quantity ]
                columns += [ product_variant.total_employee_sale_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
         

    columns = []
    columns += [ 'FOC QTY (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_foc_quantity ]
                columns += [ product_variant.total_foc_quantity_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'TOTAL QUANTITY WITH SCHEME (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.total_product_quantity_with_scheme ]
                columns += [ product_variant.total_product_quantity_with_scheme ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000', bold=True)
        

    columns = []
    columns += [ ' ' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ ' ' ]
                columns += [ ' ' ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        color_code = str(product.product_color_code).replace('#', '')
                    
            

    columns = []
    columns += [ 'TOTAL POUCH W/O SCHEME' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_pouch_wo_scheme ]
                columns += [ product_variant.total_pouch_wo_scheme_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'TOTAL SCHEME POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_pouch_scheme ]
                columns += [ product_variant.total_pouch_scheme_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'BONUS SCHEME POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_pouch_bonus_scheme ]
                columns += [ product_variant.total_pouch_bonus_scheme_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
            
    
    columns = []
    columns += [ 'EMPLOYEE SALE POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_pouch_employee_sale ]
                columns += [ product_variant.total_employee_sale_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'FOC POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.product_total_pouch_foc ]
                columns += [ product_variant.total_foc_quantity_crate_sum ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000')
      

    columns = []
    columns += [ 'TOTAL VARIANT WISE POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    columns += [ product.total_pouch_product_quantity_with_scheme ]
                columns += [ product_variant.total_pouch_product_quantity_with_scheme ]
    
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                for cols, product_variant in enumerate(product_classes):
                    for product in product_variant.product_variant_list:
                        cell.font = Font(size=14, color='000000', bold=True)
        
    columns = []
    columns += [ 'Zone' ]
    for product_class in product_classes:
        columns += [ 'TOTAL '+product_class.product_class+' QTY' ]
    row_num += 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.font = Font(size=14, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
    for town in town_list:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        row += [ town['town_name'] ]
        for id, product_class in enumerate(product_classes):
            row += [ town[product_class.id] ]                 
       
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(size=14, color='000000') 

    # sum the values with same keys 
    row_num += 1
    row = []
    row += [ 'Total Sales' ]
    if product_classes:
        for product_variant in product_classes:
            if product_variant.product_variant_list:
                row += [ product_variant.total_product_quantity_with_scheme ]

    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = wrapped_alignment
        cell.border = black_border 
        if col_num == 1:
            cell.font = Font(size=16, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            cell.font = Font(size=16, color='000000', bold=True)
                

    columns = []
    columns += [ 'PARTICULARS' ]
    columns += [ 'QUANTITY' ]
    row_num += 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.font = Font(size=14, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

    for products in products_list:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        row += [ products['product_name'] ]
        row += [ products['total_quantity'] ]                 
       
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(size=14, color='000000')      

    row_num += 1
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    workbook.save(response)

    return response

#get packing station Report
@login_required
@has_par(sub_module_id=28,permission='list')
def packingStationReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(indent_status=1, block_unblock=1,order_date__icontains=today.strftime("%Y-%m-%d")).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    unblock_order      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today).values_list('id', flat=True)
    product_classes = SpProductClass.objects.filter(id=1).order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0
                total_schemes = 0
                total_bonus_schemes = 0
                total_employee_sales = 0
                total_foc_quantitys = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0
                
                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None

                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_pouch_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_employee_sale_quantity = int(getEmployeesTotalQunatity(product_variant.id))
                total_employee_sales                                 = total_employee_sales + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_foc_quantity           = int(getFocTotalQunatity(product_variant.id))
                total_foc_quantitys                                  = total_foc_quantitys + int(getFocTotalQunatity(product_variant.id))


                product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                if no_of_pouch == 0:
                    product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today,unblock_order))
                else:
                    product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)+getFocTotalQunatityInLtr(product_variant.id)
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)
           
            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)
    
    product_classess = SpProductClass.objects.filter().exclude(id=1).order_by('-id')
    for product_class in product_classess:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0
                total_schemes = 0
                total_bonus_schemes = 0
                total_employee_sales = 0
                total_foc_quantitys = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0

                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None

                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_pouch_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_employee_sale_quantity = int(getEmployeesTotalQunatity(product_variant.id))
                total_employee_sales                                 = total_employee_sales + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_foc_quantity           = int(getFocTotalQunatity(product_variant.id))
                total_foc_quantitys                                  = total_foc_quantitys + int(getFocTotalQunatity(product_variant.id))


                product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                if no_of_pouch == 0:
                    product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today,unblock_order))
                else:
                    product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)+getFocTotalQunatityInLtr(product_variant.id)
               
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)
           
            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)

    approved_order = SpOrders.objects.filter(order_status = 3,  block_unblock=1,order_date__icontains=today).count()
    today_order = SpOrders.objects.filter( block_unblock=1,order_date__icontains=today).count()

    last_update_report_time = getConfigurationResult('last_update_report_time')
    if last_update_report_time:
        last_update_report_time     = str(last_update_report_time).replace('+00:00', '')
        last_update_report_time     = datetime.strptime(str(last_update_report_time), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y | %I:%M:%p')
        last_update_report_time     = last_update_report_time
    else:
        last_update_report_time = 'N/A'

    context = {}
    context['product_classes']                      = product_classes
    context['product_classess']                     = product_classess
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['page_title']                           = "Packing Station Report"
    context['last_update_report_time']              = last_update_report_time
    context['refresh_time']                         = getConfigurationResult('refresh_time')
    template                                        = 'order-management/packing-station-report.html'

    return render(request, template, context)   
    
#get packing station Report
@login_required
@has_par(sub_module_id=28,permission='list')
def ajaxPackingStationReport(request):
    today                   = request.GET['order_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(block_unblock=1,indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    unblock_order      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today).values_list('id', flat=True)
    
    product_classes = SpProductClass.objects.filter(id=1).order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0
                total_schemes = 0
                total_bonus_schemes = 0
                total_employee_sales = 0
                total_foc_quantitys = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0
                
                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None

                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_pouch_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_employee_sale_quantity = int(getEmployeesTotalQunatity(product_variant.id))
                total_employee_sales                                 = total_employee_sales + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_foc_quantity           = int(getFocTotalQunatity(product_variant.id))
                total_foc_quantitys                                  = total_foc_quantitys + int(getFocTotalQunatity(product_variant.id))


                product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                if no_of_pouch == 0:
                    product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today,unblock_order))
                else:
                    product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)+getFocTotalQunatityInLtr(product_variant.id)
                               
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)
           
            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)
    
    product_classess = SpProductClass.objects.filter().exclude(id=1).order_by('-id')
    for product_class in product_classess:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0
                total_schemes = 0
                total_bonus_schemes = 0
                total_employee_sales = 0
                total_foc_quantitys = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0

                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None

                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_pouch_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_employee_sale_quantity = int(getEmployeesTotalQunatity(product_variant.id))
                total_employee_sales                                 = total_employee_sales + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_foc_quantity           = int(getFocTotalQunatity(product_variant.id))
                total_foc_quantitys                                  = total_foc_quantitys + int(getFocTotalQunatity(product_variant.id))


                product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                if no_of_pouch == 0:
                    product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today,unblock_order))
                else:
                    product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)+getFocTotalQunatityInLtr(product_variant.id)
               
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)
           
            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)

    approved_order = SpOrders.objects.filter(order_status = 3, block_unblock=1, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter( block_unblock=1, order_date__icontains=today).count()

    today_date   = date.today()
    context = {}
    context['product_classes']                      = product_classes
    context['product_classess']                     = product_classess
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")
    context['page_title']                           = "Packing Station Report"
    template                                        = 'order-management/ajax-packing-station-report.html'

    return render(request, template, context)      

#update packing station Report Refresh Time
@login_required
def updatePackingStationReportRefreshTime(request):
    response = {}
    try:
        config_id = Configuration.objects.filter().first()
        data                            = Configuration.objects.get(id=config_id.id)
        data.last_update_report_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        data.save()
        
        response['error'] = False
        response['message'] = "Record has been updated successfully."
        return JsonResponse(response)
    except ObjectDoesNotExist:
        response['error'] = True
        response['message'] = "Method not allowed"
        return JsonResponse(response)
    except Exception as e:
        response['error'] = True
        response['message'] = e
        return JsonResponse(response)
        
#export packing station Report
@login_required
@has_par(sub_module_id=28,permission='export')
def exportPackingStationReport(request, order_date):
    today                   = order_date
    # today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today).count()
    # order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    today_order_status      = SpOrders.objects.filter(indent_status=1, block_unblock=1,order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    unblock_order      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today).values_list('id', flat=True)
    
    product_classes = SpProductClass.objects.filter(id=1).order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0
                total_schemes = 0
                total_bonus_schemes = 0
                total_employee_sales = 0
                total_foc_quantitys = 0

                # try:
                #     product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                #     product_total_quantity = product_total_quantity['quantity_in_ltr']
                # except SpOrderDetails.DoesNotExist:
                #     product_total_quantity = 0

                # try:
                #     product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                #     product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                # except SpOrderDetails.DoesNotExist:
                #     product_pouch_quantity = 0    

                # try:
                #     order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).order_by('id').first()
                # except SpOrderDetails.DoesNotExist:
                #     order_details = None
                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0
                
                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None

                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_pouch_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                # product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                # total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                # product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                # total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)


                product_variant.product_total_employee_sale_quantity = int(getEmployeesTotalQunatity(product_variant.id))
                total_employee_sales                                 = total_employee_sales + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_foc_quantity           = int(getFocTotalQunatity(product_variant.id))
                total_foc_quantitys                                  = total_foc_quantitys + int(getFocTotalQunatity(product_variant.id))


                # product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                # product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                # if no_of_pouch == 0:
                #     product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today))
                # else:
                #     product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                # product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                # product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today)+getFocTotalQunatityInLtr(product_variant.id)
                product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                if no_of_pouch == 0:
                    product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today,unblock_order))
                else:
                    product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)+getFocTotalQunatityInLtr(product_variant.id)
                
                
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)
           
            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)
    
    product_classess = SpProductClass.objects.filter().exclude(id=1).order_by('-id')
    for product_class in product_classess:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.order_of asc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0
                total_schemes = 0
                total_bonus_schemes = 0
                total_employee_sales = 0
                total_foc_quantitys = 0

                # try:
                #     product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                #     product_total_quantity = product_total_quantity['quantity_in_ltr']
                # except SpOrderDetails.DoesNotExist:
                #     product_total_quantity = 0

                # try:
                #     product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                #     product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                # except SpOrderDetails.DoesNotExist:
                #     product_pouch_quantity = 0

                # try:
                #     order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).order_by('id').first()
                # except SpOrderDetails.DoesNotExist:
                #     order_details = None
                
                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_ltr = Sum('quantity_in_ltr'))
                    product_total_quantity = product_total_quantity['quantity_in_ltr']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                try:
                    product_pouch_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).aggregate(quantity_in_pouch = Sum('quantity_in_pouch'))
                    product_pouch_quantity = product_pouch_quantity['quantity_in_pouch']
                except SpOrderDetails.DoesNotExist:
                    product_pouch_quantity = 0

                try:
                    order_details = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today,order_id__in=unblock_order).order_by('id').first()
                except SpOrderDetails.DoesNotExist:
                    order_details = None
                
                if order_details:
                    variant_size    = order_details.product_variant_size
                    no_of_pouch     = order_details.product_no_of_pouch
                else:
                    variant_size    = 0
                    no_of_pouch     = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_pouch_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                # product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today)
                # total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today)

                # product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today)
                # total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today)

                # product_variant.product_total_employee_sale_quantity = int(getEmployeesTotalQunatity(product_variant.id))
                # total_employee_sales                                 = total_employee_sales + int(getEmployeesTotalQunatity(product_variant.id))

                # product_variant.product_total_foc_quantity           = int(getFocTotalQunatity(product_variant.id))
                # total_foc_quantitys                                  = total_foc_quantitys + int(getFocTotalQunatity(product_variant.id))


                # product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                # product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                # if no_of_pouch == 0:
                #     product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today))
                # else:
                #     product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                # product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                # product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today)+getFocTotalQunatityInLtr(product_variant.id)
                
                product_variant.product_total_scheme_quantity       = getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)
                total_schemes                                       = total_schemes + getSummaryFreeSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_bonus_scheme_quantity = getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)
                total_bonus_schemes                                 = total_bonus_schemes + getSummaryBonusSchemeInPoches(product_variant.id, today,unblock_order)

                product_variant.product_total_employee_sale_quantity = int(getEmployeesTotalQunatity(product_variant.id))
                total_employee_sales                                 = total_employee_sales + int(getEmployeesTotalQunatity(product_variant.id))

                product_variant.product_total_foc_quantity           = int(getFocTotalQunatity(product_variant.id))
                total_foc_quantitys                                  = total_foc_quantitys + int(getFocTotalQunatity(product_variant.id))


                product_variant_total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_bonus_scheme_quantity
                product_variant.total_quantity              = product_variant.product_total_pouches_packed+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                if no_of_pouch == 0:
                    product_variant.total_crates_packed         = 0+float(getSummaryFreeSchemeInCrate(product_variant.id, today,unblock_order))
                else:
                    product_variant.total_crates_packed         = round((int(product_variant.total_quantity)/int(no_of_pouch)),2) 
                product_quantity_ltr_kg                     = (float(variant_size))*int(product_variant_total_quantity) 
                product_variant.total_quantity_in_ltr_kg    = product_quantity_ltr_kg+getEmployeesTotalQunatityInLtr(product_variant.id)+getSummaryFreeSchemeInLiterKg(product_variant.id, today,unblock_order)+getFocTotalQunatityInLtr(product_variant.id)

                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)
           
            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)
    
    # approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    # today_order = SpOrders.objects.filter(order_date__icontains=today).count()
    approved_order = SpOrders.objects.filter(order_status = 3, block_unblock=1, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter( block_unblock=1, order_date__icontains=today).count()

    today_date   = date.today()
    context = {}
    context['product_classes']                      = product_classes
    context['product_classess']                     = product_classess
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")

#export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=packing-station-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Packing Station Report'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    product_classes_length = []
    for product_variant in product_classes:
                for product in product_variant.product_variant_list:
                    product_classes_length.append(len(product_variant.product_variant_list))
                product_classes_length.append(len(product_variant.product_variant_list))    
    product_classess_length = []
    for product_variant in product_classess:
                for product in product_variant.product_variant_list:
                    product_classess_length.append(len(product_variant.product_variant_list))
                product_classess_length.append(len(product_variant.product_variant_list)) 

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(name='Arial Nova Cond Light',size=12, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")
    
    column_length = len(product_classes_length)+1
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'Shreedhi'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=16, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    columns = []
    columns += [ 'PARTICULARS' ]

    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.variant_name ]
            columns += [ product_variant.product_variant_total ]    

    row_num = 2
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:    
            cell.font = Font(size=14, color='000000', bold=True)
            cell.fill = PatternFill()             
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    
    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'TOTAL POUCH TO BE PACKED' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.total_quantity ]
            row += [ product_variant.total_quantity ]

            # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000')     
    
    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'POUCH PER CRATE' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.no_of_pouch ]
            row += [ ' ' ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000')   

    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'TOTAL CRATES TO BE PACKED' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.total_crates_packed ]
            row += [ product_variant.total_crates_packed ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000')

    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'TOTAL QTY (LTR / KG)' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.total_quantity_in_ltr_kg ]
            row += [ product_variant.total_quantity_in_ltr_kg ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000')    

    row_num += 2
    bottom_columns = []
    bottom_columns += [ 'PARTICULARS' ]

    if product_classess:
        for product_variant in product_classess:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    bottom_columns += [ product.variant_name ]
                bottom_columns += [ product_variant.product_variant_total ]    

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(bottom_columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                cell.border = black_border
                if col_num == 1:
                    cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
                else:    
                    cell.font = Font(size=14, color='000000', bold=True)
                    cell.fill = PatternFill()             
                
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = 20

    bottom_columns = []
    bottom_columns += [ 'TOTAL POUCH TO BE PACKED' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    bottom_columns += [ product.total_quantity ]
                bottom_columns += [ product_variant.total_quantity ]    

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(bottom_columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                cell.border = black_border
                if col_num == 1:
                    cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
                else:    
                    cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000')
                    cell.fill = PatternFill() 

    bottom_columns = []
    bottom_columns += [ 'POUCH PER CRATE' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    bottom_columns += [ product.no_of_pouch ]
                bottom_columns += [ ' ' ]    

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(bottom_columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                cell.border = black_border
                if col_num == 1:
                    cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
                else:    
                    cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000')
                    cell.fill = PatternFill()    

    bottom_columns = []
    bottom_columns += [ 'TOTAL CRATES TO BE PACKED' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    bottom_columns += [ product.total_crates_packed ]
                bottom_columns += [ product_variant.total_crates_packed ]    

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(bottom_columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                cell.border = black_border
                if col_num == 1:
                    cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
                else:    
                    cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000')
                    cell.fill = PatternFill()

    bottom_columns = []
    bottom_columns += [ 'TOTAL QTY (LTR / KG)' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            if product_variant.product_variant_list:
                for product in product_variant.product_variant_list:
                    bottom_columns += [ product.total_quantity_in_ltr_kg ]
                bottom_columns += [ product_variant.total_quantity_in_ltr_kg ]    

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(bottom_columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                cell.border = black_border
                if col_num == 1:
                    cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
                else:    
                    cell.font = Font(size=12, color='000000')
                    cell.fill = PatternFill()

    row_num += 1
    last_row = row_num
    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
                                
    workbook.save(response)
    return response




#get order HO Report
@login_required
@has_par(sub_module_id=29,permission='list')
def hoReport(request):
    context = {}
    today                   = date.today()
    user_list = SpUsers.objects.raw('''SELECT id, first_name, middle_name, last_name FROM sp_users
    WHERE purchase_milk_from_org = 1 ''')

    product_milk_variant_list = SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code  
    FROM sp_product_variants LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id
    WHERE sp_product_variants.product_class_id = '3' 
    order by sp_product_variants.order_of asc 
    ''')
    context['product_milk_variant_list'] =  product_milk_variant_list

    product_without_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code FROM sp_product_variants 
    LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id 
    WHERE sp_product_variants.product_class_id != '3' 
    order by sp_product_variants.order_of asc 
    ''')

    context['product_without_milk_variant_list'] =  product_without_milk_variant_list
    context['page_title']                       = "Order Management > HO Report"
    context['product_variants'] = SpProductUnits.objects.raw(''' SELECT id FROM sp_product_variants ''')
    context['users'] = user_list
    context['last_row'] = last_row = SpHoReport.objects.first()
    template                                    = 'order-management/ho-report.html'
    return render(request, template, context)

#update HO Report
@login_required
@has_par(sub_module_id=29,permission='edit')
def updateHoReport(request):
    response = {}
    if request.method == "POST":

        # update history and delete data
        ho_old_records =  SpHoReport.objects.all()
        if(len(ho_old_records)):
            for ho_old_record in ho_old_records:
                ho_report_history = SpHoReportHistory()
                if ho_old_record.user_id is not None :
                    ho_report_history.user_id = ho_old_record.user_id
                    
                ho_report_history.product_variant_id = ho_old_record.product_variant_id

                if ho_old_record.user_id is not None :
                    ho_report_history.quantity = ho_old_record.quantity
                if ho_old_record.user_id is not None :
                    ho_report_history.foc_pouch = ho_old_record.foc_pouch

                ho_report_history.save()

            SpHoReport.objects.all().delete()


        #insert new data
        user_list = SpUsers.objects.raw('''SELECT id, first_name, middle_name, last_name FROM sp_users
            WHERE purchase_milk_from_org = 1 ''')
        for user in user_list:
            product_variant_list = SpProductUnits.objects.raw(''' SELECT id FROM sp_product_variants ''')
            for product_variant in product_variant_list :
                quantity_var = "quantity_"+str(user.id)+"_"+str(product_variant.id)
                if quantity_var in request.POST :
                    ho_report = SpHoReport()
                    ho_report.user_id = user.id
                    ho_report.product_variant_id = product_variant.id
                    ho_report.quantity = request.POST[quantity_var]
                    ho_report.save() 

        product_variant_list = SpProductUnits.objects.raw(''' SELECT id FROM sp_product_variants ''')
        for product_variant in product_variant_list :
            quantity_var = "foc_quantity_"+str(product_variant.id)
            if quantity_var in request.POST :
                ho_report = SpHoReport()
                ho_report.user_id = None
                ho_report.product_variant_id = product_variant.id
                ho_report.foc_pouch = request.POST[quantity_var]
                ho_report.save()     

        response['flag'] = True
        response['message'] = "Report has been updated successfully."
    else:
        response['flag'] = False
        response['message'] = "Method not allowed"
    return JsonResponse(response)

@login_required
@has_par(sub_module_id=29,permission='export')
def exportHoReport(request):

    user_list = SpUsers.objects.raw('''SELECT id, first_name, middle_name, last_name FROM sp_users
    WHERE purchase_milk_from_org = 1 ''')

    product_milk_variant_list = SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code  
    FROM sp_product_variants LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id
    WHERE sp_product_variants.product_class_id = '3' 
    order by sp_product_variants.order_of asc 
    ''')

    product_without_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code FROM sp_product_variants 
    LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id 
    WHERE sp_product_variants.product_class_id != '3' 
    order by sp_product_variants.order_of asc 
    ''')

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Ho-report.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'HO REPORT'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True

    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.now().strftime('%d-%m-%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+2
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'Shreedhi'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=20, bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    columns = []
    columns += [ 'EMPLOYEE NAME' ]

    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            columns += [ product_variant.variant_name ]

    
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            columns += [ product_variant.variant_name ]
    
    columns += [ 'AMOUNT' ]

    row_num = 2


    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=10, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 8

# price/unit row
    row_num += 1
    row = []
    row +=['PRICE/UNIT']
    total_unit_price = 0
    for product_variant in product_milk_variant_list:
        row += [ product_variant.sp_employee ]
        total_unit_price = total_unit_price + product_variant.sp_employee
    
    for product_variant in product_without_milk_variant_list:
        row += [ product_variant.sp_employee ]
        total_unit_price = total_unit_price + product_variant.sp_employee

    row += [ total_unit_price ]
    for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 

# employees row
    total_qty_amount = 0
    for user in user_list:
        row_num += 1
        row = []
        row += [ user.first_name+" "+user.middle_name+" "+user.last_name ]
        variant_total_amount = 0
        if product_milk_variant_list:
            for product_variant in product_milk_variant_list:
                row += [ get_ho_report_quantity(user.id,product_variant.id) ]
               # print(float(get_ho_report_quantity(user.id,product_variant.id) * product_variant.sp_employee))
                variant_total_amount = float(variant_total_amount + float(get_ho_report_quantity(user.id,product_variant.id) * product_variant.sp_employee))

        if product_without_milk_variant_list:
            for product_variant_wm in product_without_milk_variant_list:
                row += [ get_ho_report_quantity(user.id,product_variant_wm.id) ]
                variant_total_amount = float(variant_total_amount + float(get_ho_report_quantity(user.id,product_variant_wm.id) * product_variant_wm.sp_employee))

        row += [ variant_total_amount ]
        total_qty_amount = float(total_qty_amount + variant_total_amount)
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
    
    # total row
    row_num += 1
    row = []
    row +=['TOTAL NO. OF POUCHES / UNITS']
    total_qty = 0
    for product_variant in product_milk_variant_list:
        row += [ get_ho_report_total_variant_qty(product_variant.id) ]
    for product_variant in product_without_milk_variant_list:
        row += [ get_ho_report_total_variant_qty(product_variant.id) ]
        
    row += [ total_qty_amount ]

    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = wrapped_alignment
        cell.border = black_border

# FOC row
    row_num += 1
    row = []
    row +=['FOC POUCHES']
    total_foc_qty = 0
    for product_variant in product_milk_variant_list:
        row += [ get_ho_report_foc(product_variant.id) ]
        total_foc_qty = total_foc_qty + get_ho_report_foc(product_variant.id)
    
    for product_variant in product_without_milk_variant_list:
        row += [ get_ho_report_foc(product_variant.id) ]
        total_foc_qty = total_foc_qty + get_ho_report_foc(product_variant.id)
        
    row += [ total_foc_qty ]

    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = wrapped_alignment
        cell.border = black_border 

    row_num += 1
    last_row = row_num
    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response

@login_required
def truckSheetReport(request):
    if 'order_date' in request.GET and request.GET['order_date'] != "" :
        today                   = request.GET['order_date']
        today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today() 
        
    
    if 'user_id' in request.GET and request.GET['user_id'] != "" :
        user_id = request.GET['user_id']
        if SpOrders.objects.filter(user_id=user_id,block_unblock=1,order_date__icontains=today).exists() :
            first_order = SpOrders.objects.get(user_id=user_id,block_unblock=1, order_date__icontains=today)
            user_id = first_order.user_id
        else:
            first_order = []
    else:
        if SpOrders.objects.filter(order_date__icontains=today,block_unblock=1).exists() :
            first_order = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today).order_by('id').first()
            user_id = first_order.user_id
        else:
            first_order = []
            user_id = 0
    
    condition = " and order_date LIKE '%%"+str(today)+"%%'" 
    users = SpUsers.objects.raw(''' SELECT id,emp_sap_id, CONCAT(first_name," ",middle_name," ", last_name) as name 
                        FROM sp_users WHERE id in (SELECT user_id FROM sp_orders WHERE 1 {condition})
                        '''.format(condition=condition))
    product_milk_variant_list = SpProductVariants.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code,
    sp_product_variants.no_of_pouch as pouch_per_crate  
    FROM sp_product_variants LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id
    WHERE sp_product_variants.product_class_id = '3' 
    order by sp_product_variants.order_of asc 
    ''')

    product_without_milk_variant_list =  SpProductVariants.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code,
     sp_product_variants.no_of_pouch as pouch_per_crate
    FROM sp_product_variants 
    LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id 
    WHERE sp_product_variants.product_class_id != '3' 
    order by sp_product_variants.order_of asc 
    ''')

    
    product_milk_total_crates = 0
    product_milk_total_bonus_scheme = 0
    product_milk_total_quantity = 0
    product_without_milk_total_crates = 0
    product_without_milk_total_bonus_scheme = 0
    product_without_milk_total_quantity = 0

    if first_order :
        
        for product_milk_variant in product_milk_variant_list :
            if SpOrderDetails.objects.filter(order_id=first_order.id,product_variant_id=product_milk_variant.id).exists() :
                details = SpOrderDetails.objects.filter(order_id=first_order.id,product_variant_id=product_milk_variant.id).aggregate(Sum('quantity'))
                product_milk_variant.demand = details['quantity__sum']
                product_milk_total_crates = int(details['quantity__sum']) + int(product_milk_total_crates)
                bonus_scheme = 0
                product_milk_total_bonus_scheme = int(bonus_scheme) + int(product_milk_total_bonus_scheme)
                product_milk_variant.bonus_scheme = int(bonus_scheme)
                variant_wise_total_crates = int(details['quantity__sum']) + int(bonus_scheme)
                product_milk_variant.variant_wise_total_crates = variant_wise_total_crates

                if SpOrderSchemes.objects.filter(scheme_type='free',user_id=user_id,order_id=first_order.id,free_variant_id=product_milk_variant.id).exists():
                    records = SpOrderSchemes.objects.filter(scheme_type='free',user_id=user_id,order_id=first_order.id,free_variant_id=product_milk_variant.id)
                    total_free_pouch_per_crate = []
                    for record in records:
                        free_variant = SpUserProductVariants.objects.get(id=record.free_variant_id) 
                        free_pouch_per_crates = int(record.container_quantity * free_variant.no_of_pouch) + int(record.pouch_quantity)
                        total_free_pouch_per_crate.append(free_pouch_per_crates)
                    free_pouch_per_crate = sum(total_free_pouch_per_crate)        
                else:
                    free_pouch_per_crate = 0
                free_pouch_per_crate       = getFreeScheme(product_milk_variant.id, order_id=first_order.id, user_id=user_id)

                product_milk_variant.free_pouch_per_crate = int(free_pouch_per_crate)
                total_pouch_per_crate = int(free_pouch_per_crate) * int(product_milk_variant.pouch_per_crate)
                product_milk_variant.total_pouch_per_crate = total_pouch_per_crate
                total_no_of_pouch = (int(product_milk_variant.demand) * int(product_milk_variant.pouch_per_crate)) + int(free_pouch_per_crate)
                product_milk_variant.total_no_of_pouch = total_no_of_pouch
                product_milk_variant.sku_size = product_milk_variant.variant_size
                product_milk_variant.total_quantity = (float(product_milk_variant.variant_size) * int(total_no_of_pouch))
                product_milk_total_quantity = float(product_milk_total_quantity) + (float(product_milk_variant.variant_size) * int(total_no_of_pouch))
            else:
                product_milk_variant.demand = 0
                product_milk_total_crates = 0 + int(product_milk_total_crates)
                bonus_scheme = 0
                product_milk_total_bonus_scheme = int(bonus_scheme) + int(product_milk_total_bonus_scheme)
                product_milk_variant.bonus_scheme = int(bonus_scheme)
                variant_wise_total_crates = 0 + int(bonus_scheme)
                product_milk_variant.variant_wise_total_crates = variant_wise_total_crates
                free_pouch_per_crate       = getFreeScheme(product_milk_variant.id, order_id=first_order.id, user_id=user_id)
                product_milk_variant.free_pouch_per_crate = int(free_pouch_per_crate)
                total_pouch_per_crate = int(free_pouch_per_crate) * int(product_milk_variant.pouch_per_crate)
                product_milk_variant.total_pouch_per_crate = total_pouch_per_crate
                total_no_of_pouch = (int(product_milk_variant.demand) * int(product_milk_variant.pouch_per_crate)) + int(free_pouch_per_crate)
                product_milk_variant.total_no_of_pouch = total_no_of_pouch
                product_milk_variant.sku_size = product_milk_variant.variant_size
                product_milk_variant.total_quantity = (float(product_milk_variant.variant_size) * int(total_no_of_pouch))
                product_milk_total_quantity = float(product_milk_total_quantity) + (float(product_milk_variant.variant_size) * int(total_no_of_pouch))

        
        for product_without_milk_variant in product_without_milk_variant_list :
            if SpOrderDetails.objects.filter(order_id=first_order.id,product_variant_id=product_without_milk_variant.id).exists() :
                details = SpOrderDetails.objects.filter(order_id=first_order.id,product_variant_id=product_without_milk_variant.id).aggregate(Sum('quantity'))
                product_without_milk_variant.demand = details['quantity__sum']
                product_without_milk_total_crates = int(details['quantity__sum']) + int(product_without_milk_total_crates)
                bonus_scheme = 0
                product_without_milk_total_bonus_scheme = int(bonus_scheme) + int(product_without_milk_total_bonus_scheme)
                product_without_milk_variant.bonus_scheme = int(bonus_scheme)
                variant_wise_total_crates = int(details['quantity__sum']) + int(bonus_scheme)
                product_without_milk_variant.variant_wise_total_crates = variant_wise_total_crates
                free_pouch_per_crate       = getFreeScheme(product_without_milk_variant.id, order_id=first_order.id, user_id=user_id)
                product_without_milk_variant.free_pouch_per_crate = int(free_pouch_per_crate)
                total_pouch_per_crate = int(free_pouch_per_crate) * int(product_without_milk_variant.pouch_per_crate)
                product_without_milk_variant.total_pouch_per_crate = total_pouch_per_crate
                total_no_of_pouch = (int(product_without_milk_variant.demand) * int(product_without_milk_variant.pouch_per_crate)) + int(free_pouch_per_crate)
                product_without_milk_variant.total_no_of_pouch = total_no_of_pouch
                product_without_milk_variant.sku_size = product_without_milk_variant.variant_size
                product_without_milk_variant.total_quantity = (float(product_without_milk_variant.variant_size) * int(total_no_of_pouch))
                product_without_milk_total_quantity = float(product_without_milk_total_quantity) + (float(product_without_milk_variant.variant_size) * int(total_no_of_pouch))
            else:
                product_without_milk_variant.demand = 0
                product_without_milk_total_crates = 0 + int(product_without_milk_total_crates)
                bonus_scheme = 0
                product_without_milk_total_bonus_scheme = int(bonus_scheme) + int(product_without_milk_total_bonus_scheme)
                product_without_milk_variant.bonus_scheme = int(bonus_scheme)
                variant_wise_total_crates = 0 + int(bonus_scheme)
                product_without_milk_variant.variant_wise_total_crates = variant_wise_total_crates
                free_pouch_per_crate       = getFreeScheme(product_without_milk_variant.id, order_id=first_order.id, user_id=user_id)
                product_without_milk_variant.free_pouch_per_crate = int(free_pouch_per_crate)
                total_pouch_per_crate = int(free_pouch_per_crate) * int(product_without_milk_variant.pouch_per_crate)
                product_without_milk_variant.total_pouch_per_crate = total_pouch_per_crate
                total_no_of_pouch = (int(product_without_milk_variant.demand) * int(product_without_milk_variant.pouch_per_crate)) + int(free_pouch_per_crate)
                product_without_milk_variant.total_no_of_pouch = total_no_of_pouch
                product_without_milk_variant.sku_size = product_without_milk_variant.variant_size
                product_without_milk_variant.total_quantity = (float(product_without_milk_variant.variant_size) * int(total_no_of_pouch))
                product_without_milk_total_quantity = float(product_without_milk_total_quantity) + (float(product_without_milk_variant.variant_size) * int(total_no_of_pouch))

    context = {}
    context['product_milk_variant_list'] = product_milk_variant_list
    context['product_without_milk_variant_list'] = product_without_milk_variant_list
    context['users']                                = users
    user = SpUsers.objects.raw(''' SELECT sp_users.id,sp_users.emp_sap_id,sp_users.store_name, CONCAT(sp_users.first_name," ",sp_users.middle_name," ", sp_users.last_name) 
                    as name ,sp_user_area_allocations.route_name
                    FROM sp_users LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_users.id 
                    WHERE sp_users.id = %s ''',[user_id])
                
    if user:
        context['current_user']                              = user[0]
    else:
        context['current_user']                              = []
    context['page_title']                           = "Truck Sheet Report"
    context['first_order']                              = first_order
    context['today_date']                   = date.today().strftime("%d/%m/%Y") 
    context['product_milk_total_crates'] = product_milk_total_crates
    context['product_milk_total_quantity'] = product_milk_total_quantity
    context['product_milk_total_bonus_scheme'] = product_milk_total_bonus_scheme
    context['product_without_milk_total_crates'] = product_without_milk_total_crates
    context['product_without_milk_total_bonus_scheme'] = product_without_milk_total_bonus_scheme
    context['product_without_milk_total_quantity'] = product_without_milk_total_quantity

    is_ajax = request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'
    if(is_ajax):
        template                                        = 'order-management/ajax-truck-sheet-report.html'
        return render(request, template, context) 
    else:
        if 'export_action' in request.GET and request.GET['export_action'] == "excel" :
            # export excel
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename=Truck Sheet Report-'+ user[0].name+'-'+user[0].store_name +'.xlsx'.format(
                date=datetime.now().strftime('%d-%m-%Y'),
            )
            workbook = Workbook()

            # Define some styles and formatting that will be later used for cells
            header_font = Font(name='Arial Nova Cond Light', bold=True)
            left_alignment = Alignment(horizontal='left')
            centered_alignment = Alignment(horizontal='center')
            thin = Side(border_style="thin", color="303030") 
            black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wrapped_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            
            # Get active worksheet/tab
            worksheet = workbook.active
            worksheet.title = 'TRUCK SHEET REPORT'

            worksheet.page_setup.orientation = 'landscape'
            worksheet.page_setup.paperSize = 9
            worksheet.page_setup.fitToPage = True

           
            row_num = 1
            columns = []
            columns += [ 'NAME OF THE DISTRIBUTOR / SUPER STOCKIST' ]
            columns += [ user[0].name+'/'+user[0].store_name ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ '' ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                cell.alignment = wrapped_alignment
                if i > 1:
                    column_dimensions.width = 8
                else:
                    column_dimensions.width = 8

                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                cell.border = black_border
                i = int(i)+1

            row_num += 1
            columns = []
            columns += [ 'ROUTE NAME' ]
            columns += [ '' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ '' ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                cell.alignment = wrapped_alignment
                if i > 1:
                    column_dimensions.width = 8
                else:
                    column_dimensions.width = 8
                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                cell.border = black_border
                i = int(i)+1

            row_num += 1
            columns = []
            columns += [ user[0].route_name ]
            columns += [ '' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ '' ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                cell.alignment = wrapped_alignment
                if i > 1:
                    column_dimensions.width = 8
                else:
                    column_dimensions.width = 8
                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                cell.border = black_border
                i = int(i) + 1

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ '' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ '' ]
                # columns += [ '' ]
            
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                cell.alignment = wrapped_alignment
                column_dimensions.width = 20
                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                cell.border = black_border

                i = int(i) + 1

            # row_num += 1
            # worksheet.row_dimensions[1].height = 40

            # cell = worksheet.cell(row=1, column=2)  
            # cell.value = 'DATE('+datetime.now().strftime('%d-%m-%Y')+')'
            # cell.font = header_font
            # cell.alignment = wrapped_alignment
            # cell.border = black_border
            # cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
            # cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

            # column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+2
            
            # worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
            # worksheet.cell(row=1, column=3).value = 'Shreedhi'
            # worksheet.cell(row=1, column=3).font = header_font
            # worksheet.cell(row=1, column=3).alignment = wrapped_alignment
            # worksheet.cell(row=1, column=column_length).border = black_border
            # worksheet.cell(row=1, column=3).font = Font(size=20, bold=True)
            # worksheet.cell(row=1, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

            row_num += 1
            columns = []
            columns += [ 'NAME OF THE DISTRIBUTOR / SUPER STOCKIST' ]
            columns += [ 'PARTICULARS' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ product_variant.variant_name ]
                # columns += [ 'TOTAL CRATES' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = wrapped_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20

                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                cell.border = black_border

                i = int(i) + 1
            
            

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'DEMAND' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ product_variant.demand ]
                # columns += [ product_milk_total_crates ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 8
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1
            
            # row_num += 1
            # columns = []
            # columns += [ '' ]
            # columns += [ 'BONUS SCHEME' ]
            # if product_milk_variant_list:
            #     for product_variant in product_milk_variant_list:
            #         columns += [ product_variant.bonus_scheme ]
            #     columns += [ product_milk_total_bonus_scheme ]
            # i = 0
            # for col_num, column_title in enumerate(columns, 1):
            #     cell = worksheet.cell(row=row_num, column=col_num)
            #     cell.value = column_title
            #     cell.font = header_font
            #     column_letter = get_column_letter(col_num)
            #     column_dimensions = worksheet.column_dimensions[column_letter]
            #     if i > 1 :
            #         cell.alignment = centered_alignment
            #         column_dimensions.width = 8
            #     else:
            #         cell.alignment = left_alignment
            #         column_dimensions.width = 20
            #     if i == 1 :
            #         column_dimensions.width = 30
            #         cell.font = Font(name='Arial Nova Cond Light',size=12)
            #     else:
            #         cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

            #     cell.border = black_border

            #     i = int(i) + 1
            
            # row_num += 1
            # columns = []
            # columns += [ '' ]
            # columns += [ 'VARIANT WISE TOTAL CRATES' ]
            # if product_milk_variant_list:
            #     for product_variant in product_milk_variant_list:
            #         columns += [ product_variant.variant_wise_total_crates ]
            #     columns += [ int(product_milk_total_crates) + int(product_milk_total_bonus_scheme) ]
            # i = 0
            # for col_num, column_title in enumerate(columns, 1):
            #     cell = worksheet.cell(row=row_num, column=col_num)
            #     cell.value = column_title
            #     cell.font = header_font
            #     column_letter = get_column_letter(col_num)
            #     column_dimensions = worksheet.column_dimensions[column_letter]
            #     if i > 1 :
            #         cell.alignment = centered_alignment
            #         column_dimensions.width = 8
            #     else:
            #         cell.alignment = left_alignment
            #         column_dimensions.width = 20
            #     if i == 1 :
            #         column_dimensions.width = 30
            #         cell.font = Font(name='Arial Nova Cond Light',size=12)
            #     else:
            #         cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

            #     cell.border = black_border

            #     i = int(i) + 1
            
            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'POUCH PER CRATE' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ product_variant.pouch_per_crate ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'FREE POUCHES' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ product_variant.free_pouch_per_crate ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1
            
            # row_num += 1
            # columns = []
            # columns += [ '' ]
            # columns += [ 'TOTAL POUCH PER CRATE' ]
            # if product_milk_variant_list:
            #     for product_variant in product_milk_variant_list:
            #         columns += [ product_variant.total_pouch_per_crate ]
            #     columns += [ '' ]
            # i = 0
            # for col_num, column_title in enumerate(columns, 1):
            #     cell = worksheet.cell(row=row_num, column=col_num)
            #     cell.value = column_title
            #     cell.font = header_font
            #     column_letter = get_column_letter(col_num)
            #     column_dimensions = worksheet.column_dimensions[column_letter]
            #     if i > 1 :
            #         cell.alignment = centered_alignment
            #         column_dimensions.width = 8
            #     else:
            #         cell.alignment = left_alignment
            #         column_dimensions.width = 20
            #     if i == 1 :
            #         column_dimensions.width = 30
            #         cell.font = Font(name='Arial Nova Cond Light',size=12)
            #     else:
            #         cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

            #     cell.border = black_border

            #     i = int(i) + 1
            
            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'TOTAL NO. OF POUCHES' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ product_variant.total_no_of_pouch ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1
            
            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'SKU SIZE' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ product_variant.sku_size ]
                # columns += [ '' ]
            
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'TOTAL QTY (LTR / KG)' ]
            if product_milk_variant_list:
                for product_variant in product_milk_variant_list:
                    columns += [ product_variant.total_quantity ]
                # columns += [ product_milk_total_quantity ]
            
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1

            # Without milk
            row_num += 3
            columns = []
            columns += [ 'NAME OF THE DISTRIBUTOR / SUPER STOCKIST' ]
            columns += [ 'PARTICULARS' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ product_variant.variant_name ]
                # columns += [ 'TOTAL CRATES' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = wrapped_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20

                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                cell.border = black_border

                i = int(i) + 1
            
            

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'DEMAND' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ product_variant.demand ]
                # columns += [ product_without_milk_total_crates ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1
            
            # row_num += 1
            # columns = []
            # columns += [ '' ]
            # columns += [ 'BONUS SCHEME' ]
            # if product_without_milk_variant_list:
            #     for product_variant in product_without_milk_variant_list:
            #         columns += [ product_variant.bonus_scheme ]
            #     columns += [ product_without_milk_total_bonus_scheme ]
            # i = 0
            # for col_num, column_title in enumerate(columns, 1):
            #     cell = worksheet.cell(row=row_num, column=col_num)
            #     cell.value = column_title
            #     cell.font = header_font
            #     column_letter = get_column_letter(col_num)
            #     column_dimensions = worksheet.column_dimensions[column_letter]
            #     if i > 1 :
            #         cell.alignment = centered_alignment
            #         column_dimensions.width = 8
            #     else:
            #         cell.alignment = left_alignment
            #         column_dimensions.width = 20
            #     if i == 1 :
            #         column_dimensions.width = 30
            #         cell.font = Font(name='Arial Nova Cond Light',size=12)
            #     else:
            #         cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

            #     cell.border = black_border

            #     i = int(i) + 1
            
            # row_num += 1
            # columns = []
            # columns += [ '' ]
            # columns += [ 'VARIANT WISE TOTAL CRATES' ]
            # if product_without_milk_variant_list:
            #     for product_variant in product_without_milk_variant_list:
            #         columns += [ product_variant.variant_wise_total_crates ]
            #     columns += [ int(product_without_milk_total_crates) + int(product_without_milk_total_bonus_scheme) ]
            # i = 0
            # for col_num, column_title in enumerate(columns, 1):
            #     cell = worksheet.cell(row=row_num, column=col_num)
            #     cell.value = column_title
            #     cell.font = header_font
            #     column_letter = get_column_letter(col_num)
            #     column_dimensions = worksheet.column_dimensions[column_letter]
            #     if i > 1 :
            #         cell.alignment = centered_alignment
            #         column_dimensions.width = 8
            #     else:
            #         cell.alignment = left_alignment
            #         column_dimensions.width = 20
            #     if i == 1 :
            #         column_dimensions.width = 30
            #         cell.font = Font(name='Arial Nova Cond Light',size=12)
            #     else:
            #         cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

            #     cell.border = black_border

            #     i = int(i) + 1
            
            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'POUCH PER CRATE' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ product_variant.pouch_per_crate ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'FREE POUCHES' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ product_variant.free_pouch_per_crate ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1
            
            # row_num += 1
            # columns = []
            # columns += [ '' ]
            # columns += [ 'TOTAL POUCH PER CRATE' ]
            # if product_without_milk_variant_list:
            #     for product_variant in product_without_milk_variant_list:
            #         columns += [ product_variant.total_pouch_per_crate ]
            #     columns += [ '' ]
            # i = 0
            # for col_num, column_title in enumerate(columns, 1):
            #     cell = worksheet.cell(row=row_num, column=col_num)
            #     cell.value = column_title
            #     cell.font = header_font
            #     column_letter = get_column_letter(col_num)
            #     column_dimensions = worksheet.column_dimensions[column_letter]
            #     if i > 1 :
            #         cell.alignment = centered_alignment
            #         column_dimensions.width = 8
            #     else:
            #         cell.alignment = left_alignment
            #         column_dimensions.width = 20
            #     if i == 1 :
            #         column_dimensions.width = 30
            #         cell.font = Font(name='Arial Nova Cond Light',size=12)
            #     else:
            #         cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

            #     cell.border = black_border

            #     i = int(i) + 1
            
            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'TOTAL NO. OF POUCHES' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ product_variant.total_no_of_pouch ]
                # columns += [ '' ]
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1
            
            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'SKU SIZE' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ product_variant.sku_size ]
                # columns += [ '' ]
            
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'TOTAL QTY (LTR / KG)' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ product_variant.total_quantity ]
                # columns += [ product_without_milk_total_quantity ]
            
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                    cell.font = Font(name='Arial Nova Cond Light',size=12)
                else:
                    cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)

                cell.border = black_border

                i = int(i) + 1

            worksheet.merge_cells('B2:B4') 
            worksheet.merge_cells('A6:A11')
            worksheet.merge_cells('A15:A20') 
            worksheet.merge_cells('C2:D2') 
            worksheet.merge_cells('C3:D4') 
            worksheet.merge_cells('C1:U1') 
            worksheet.merge_cells('E2:U4') 
            # worksheet.merge_cells('E2:H'+str(len(product_milk_variant_list)-1)) 
            worksheet.merge_cells('C1:H1') 
            
            img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
            img.height = 63
            img.width = 175
            img.alignment = centered_alignment
            img.anchor = 'B2'
            worksheet.add_image(img)
            worksheet['B2'].alignment = wrapped_alignment


            worksheet['A6'] = user[0].name+'/'+user[0].store_name
            worksheet['A6'].alignment = wrapped_alignment
            worksheet['A15'] = user[0].name+'/'+user[0].store_name
            worksheet['A15'].alignment = wrapped_alignment

            worksheet['A5'].alignment = wrapped_alignment
            worksheet['A14'].alignment = wrapped_alignment
            
            worksheet['C1'] = ''
            worksheet['C2'] = 'DATE'
            worksheet['C3'] = datetime.now().strftime('%d-%m-%Y')
            worksheet['E2'] = 'Shreedhi'
            worksheet['E2'].alignment = wrapped_alignment
            worksheet['E2'].font = Font(name='Arial Nova Cond Light',size=14, bold=True)
            
            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'Total Normal Crate Issued' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ '' ]
                # columns += [ product_without_milk_total_quantity ]
            
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30

                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                column_dimensions.height = 25
                cell.border = black_border
                i = int(i) + 1

            row_num += 1
            columns = []
            columns += [ '' ]
            columns += [ 'Total Jumbo Crate Issued' ]
            if product_without_milk_variant_list:
                for product_variant in product_without_milk_variant_list:
                    columns += [ '' ]
                # columns += [ product_without_milk_total_quantity ]
            
            i = 0
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                if i > 1 :
                    cell.alignment = centered_alignment
                    column_dimensions.width = 8
                else:
                    cell.alignment = left_alignment
                    column_dimensions.width = 20
                if i == 1 :
                    column_dimensions.width = 30
                cell.font = Font(name='Arial Nova Cond Light',size=12, bold=True)
                column_dimensions.height = 25    
                cell.border = black_border

                i = int(i) + 1

            worksheet.merge_cells('C21:Y21') 
            worksheet.merge_cells('C22:Y22') 

            row_num += 1
            last_row = row_num
            column_length = len(product_without_milk_variant_list)+2
            worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

            worksheet.row_dimensions[last_row].height = 20
            worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
            worksheet.cell(row=last_row, column=1).font = header_font
            worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
            worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
            worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")


            workbook.save(response)
            return response

        else:
            template                                        = 'order-management/truck-sheet-report.html'
            return render(request, template, context)

#get amount to be paid Report
@login_required
@has_par(sub_module_id=12,permission='list')
def amountToBePaidReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d")).count()

    orders  = SpOrders.objects.all().order_by('-id').filter(block_unblock=1,order_date__icontains=today.strftime("%Y-%m-%d"))
    for order in orders:
        order.emp_sap_id = getModelColumnById(SpUsers, order.user_id, 'emp_sap_id')
        order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')
        security_amount = SpBasicDetails.objects.get(user_id=order.user_id)
        order.security_amount = security_amount.security_amount
        total_incentive = float(getFlatBulkSchemeIncentive(order.id, order.user_id, 'flat'))+float(getFlatBulkSchemeIncentive(order.id, order.user_id, 'bulkpack'))
        order_amount = SpOrders.objects.get(order_date__icontains=today.strftime("%Y-%m-%d"), user_id=order.user_id,block_unblock=1)
        order.invoice_amount = float(order_amount.order_total_amount)-total_incentive
        
    context = {}
    context['orders']                               = orders
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['page_title']                           = "Amount to be paid Report"
    template                                        = 'order-management/amount-to-be-paid-report.html'
    
    return render(request, template, context) 

#get ajax amount to be paid Report
@login_required
@has_par(sub_module_id=12,permission='list')
def ajaxAmountToBePaidReport(request):
    today                   = request.GET['order_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()

    orders  = SpOrders.objects.all().order_by('-id').filter(order_date__icontains=today,block_unblock=1)
    for order in orders:
        order.emp_sap_id = getModelColumnById(SpUsers, order.user_id, 'emp_sap_id')
        order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')
        security_amount = SpBasicDetails.objects.get(user_id=order.user_id)
        order.security_amount = security_amount.security_amount
        total_incentive = float(getFlatBulkSchemeIncentive(order.id, order.user_id, 'flat'))+float(getFlatBulkSchemeIncentive(order.id, order.user_id, 'bulkpack'))
        order_amount = SpOrders.objects.get(order_date__icontains=today, user_id=order.user_id,block_unblock=1)
        order.invoice_amount = float(order_amount.order_total_amount)-total_incentive
        
    today_date = date.today()
    context = {}
    context['orders']                               = orders
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")
    context['page_title']                           = "Amount to be paid Report"
    template                                        = 'order-management/ajax-amount-to-be-paid-report.html'
    
    return render(request, template, context)    
    
#get export amount to be paid Report
@login_required
@has_par(sub_module_id=12,permission='export')
def exportAmountToBePaidReport(request, order_date):
    today                   = order_date
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()

    orders  = SpOrders.objects.all().order_by('-id').filter(order_date__icontains=today,block_unblock=1)
    for order in orders:
        order.emp_sap_id = getModelColumnById(SpUsers, order.user_id, 'emp_sap_id')
        order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')
        security_amount = SpBasicDetails.objects.get(user_id=order.user_id)
        order.security_amount = security_amount.security_amount
        total_incentive = float(getFlatBulkSchemeIncentive(order.id, order.user_id, 'flat'))+float(getFlatBulkSchemeIncentive(order.id, order.user_id, 'bulkpack'))
        order_amount = SpOrders.objects.get(order_date__icontains=today, user_id=order.user_id,block_unblock=1,)
        order.invoice_amount = float(order_amount.order_total_amount)-total_incentive
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=amount_to_be_paid.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        horizontal='left',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Amount to be paid'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)
    
    column_length = 6
    
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'Invoice Amount to be paid as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill()

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    
    # Define the titles for columns
    columns = []

    columns += [ 'Customer Code' ]
    columns += [ 'Name of Distributor/SS' ]
    columns += [ 'Invoice Amount' ]
    columns += [ 'Amount to be paid' ]
    columns += [ 'Outstanding Amount' ]
    columns += [ 'Security Amount' ]

    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 32

    
    for order in orders:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        row += [ order.emp_sap_id ]
        row += [ order.store_name +'('+order.user_name+')' ]
        row += [ order.invoice_amount ]
        row += [ order.amount_to_be_paid ]
        row += [ order.outstanding_amount ]
        row += [ order.security_amount ]
       
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border  

    wrapped_alignment = Alignment(
        horizontal='center',
        wrap_text=True
    )

    row_num += 1
    last_row = row_num
    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)
    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

    workbook.save(response)

    return response


@login_required
@has_par(sub_module_id=39,permission='list')
def focRequests(request):
    context = {}
    page = request.GET.get('page')
    today   = date.today()
    # foc_requests = SpFocRequests.objects.all().order_by('-id')
    detailsOfUsersForFoc = SpUsers.objects.filter(status=1, user_type=1).exclude(role_id=0)
    if request.user.role_id == 0:
        foc_requests  = SpFocRequests.objects.all().order_by('-id').filter(foc_delivery_date__icontains=today.strftime("%Y-%m-%d"))
        # for order in foc_requests:
        #     order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')
    else:
        foc_requests = SpFocRequests.objects.raw('''SELECT sp_foc_requests.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_foc_requests left join sp_approval_status on sp_approval_status.row_id = sp_foc_requests.id
    left join sp_users on sp_users.id = sp_foc_requests.user_id 
    where DATE(sp_foc_requests.foc_delivery_date) = %s and sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpFocRequests' order by id desc ''',[today.strftime("%Y-%m-%d"), request.user.id])

    

    towns       = SpTowns.objects.all()
    user_type   = SpPermissionWorkflowRoles.objects.filter(sub_module_id=39,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()    
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'

    
    paginator = Paginator(foc_requests, getConfigurationResult('page_limit'))
    try:
        foc_requests = paginator.page(page)
    except PageNotAnInteger:
        foc_requests = paginator.page(1)
    except EmptyPage:
        foc_requests = paginator.page(paginator.num_pages)
    if page is not None:
        page = page
    else:
        page = 1

    total_pages = int(paginator.count / 100)
    if (paginator.count == 0):
        paginator.count = 1

    temp = int(total_pages) % paginator.count
    if (temp > 0 and 100 != paginator.count):
        total_pages = total_pages + 1
    else:
        total_pages = total_pages

    
    

    context['foc_requests'] = foc_requests 
    context['today_date'] = today.strftime("%d/%m/%Y")
    context['users_details'] = detailsOfUsersForFoc
    context['towns']        = towns
    context['role_id']      = request.user.role_id
    context['level_id']      = level_id
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')
    context['page_title'] = "Sampling Request Details"

    template = 'order-management/foc-requests/foc-requests.html'

    return render(request, template, context)


@login_required
def FocRequestDetails(request,foc_request_id):
    context = {}
    last_foc_request = SpFocRequests.objects.get(id=foc_request_id)
    if last_foc_request:
        context['last_foc_request'] = last_foc_request
        context['last_foc_request_details'] = SpFocRequestsDetails.objects.filter(foc_request_id=last_foc_request.id)
    template = 'order-management/foc-requests/foc-requests-details.html'
    return render(request, template, context)


@login_required
@has_par(sub_module_id=39,permission='list')
def ajaxFocRequestList(request):
    context = {}
    name = request.GET.get('name')
    today = request.GET.get('order_date')
    today = today.split('/')
    today = today[2]+'-'+today[1]+'-'+today[0]
    page = request.GET.get('page')
    foc_status    = request.GET['foc_status']
    

    # foc_requests = SpFocRequests.objects.all().order_by('-id')
    if request.user.role_id == 0:
        foc_requests = SpFocRequests.objects.all().order_by('-id')
        if foc_status:
            foc_requests = foc_requests.filter(foc_status=foc_status)
        if name:
                foc_requests = foc_requests.filter(user_name__icontains=name)
        if today:
                foc_requests = foc_requests.filter(foc_delivery_date__icontains=today)
        # for order in foc_requests:
        # order.store_name = getModelColumnById(SpUsers, order.user_id, 'store_name')
    else:
        condition = ''
        if foc_status:
            condition += ' and sp_foc_requests.foc_status = "%s"' % foc_status
        if name:
            condition += ' and sp_foc_requests.user_name = "%s"' % name
        if today:
            condition += ' and DATE(sp_foc_requests.foc_delivery_date)  = "%s"' % today 

        query = """SELECT sp_foc_requests.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_foc_requests left join sp_approval_status on sp_approval_status.row_id = sp_foc_requests.id
        left join sp_users on sp_users.id = sp_foc_requests.user_id 
        where sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpFocRequests' %s  order by id desc """ % (request.user.id,condition)

        

        #print(query)
            
        foc_requests = SpFocRequests.objects.raw(query)



    user_type    = SpRoleWorkflowPermissions.objects.filter(sub_module_id=8, permission_slug='add', workflow_level_role_id=request.user.role_id).exclude(role_id=request.user.role_id).values('level_id').order_by('-id').first()  
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'


    


    paginator = Paginator(foc_requests, getConfigurationResult('page_limit'))
    try:
        foc_requests = paginator.page(page)
    except PageNotAnInteger:
        foc_requests = paginator.page(1)
    except EmptyPage:
        foc_requests = paginator.page(paginator.num_pages)
    if page is not None:
        page = page
    else:
        page = 1

    total_pages = int(paginator.count / 100)
    if (paginator.count == 0):
        paginator.count = 1

    temp = int(total_pages) % paginator.count
    if (temp > 0 and 100 != paginator.count):
        total_pages = total_pages + 1
    else:
        total_pages = total_pages

    last_foc_request = SpFocRequests.objects.order_by('-id').first()
    if last_foc_request:
        context['last_foc_request'] = last_foc_request
        context['last_foc_request_details'] = SpFocRequestsDetails.objects.filter(foc_request_id=last_foc_request.id)

    context['foc_requests'] = foc_requests
    context['level_id']      = level_id
    context['foc_status']     = foc_status
    context['role_id']          = request.user.role_id
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')
    context['page_title'] = "Sample Requests Details"

    template = 'order-management/foc-requests/ajax-foc-requests-lists.html'

    return render(request, template, context)

#update order status
@login_required
@has_par(sub_module_id=39,permission='edit')
def updateFocStatus(request):
    response = {}
    order_id = request.POST.getlist('order_id[]')
    level_id = request.POST['level_id']
    if request.user.role_id == 0:
        for order in order_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=order, model_name='SpFocRequests', level_id=level_id)
            if approvals_request:
                for approval in approvals_request:
                    approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                    approval_data.status                    = 1
                    approval_data.final_status_user_id      = request.user.id
                    approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    approval_data.save()

                user_level_approval_count = SpApprovalStatus.objects.filter(row_id=order, model_name='SpFocRequests', level_id=level_id, status=0).count()
                if user_level_approval_count == 0:
                    order                   = SpFocRequests.objects.get(id=order)   
                    order.foc_status      = level_id
                    order.save()
            else:
                order                   = SpFocRequests.objects.get(id=order)   
                order.foc_status      = level_id
                order.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
               
                order.save()
                today   = date.today()
                 
    
    else:    
        for order in order_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=order, model_name='SpFocRequests', role_id=request.user.role_id, level_id=level_id)
            for approval in approvals_request:
                approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                approval_data.status                    = 1
                approval_data.final_status_user_id      = request.user.id
                approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                approval_data.save()

            user_level_approval_count = SpApprovalStatus.objects.filter(row_id=order, model_name='SpFocRequests', level_id=level_id, status=0).count()
            if user_level_approval_count == 0:
                order                   = SpFocRequests.objects.get(id=order)   
                order.foc_status      = level_id
                order.save()   

    
    if level_id == '2':
        for order in order_id:
            approvals_requests = SpApprovalStatus.objects.filter(row_id=order, model_name='SpFocRequests', status=0)
            if approvals_requests:
                for approval in approvals_requests:
                    notification                        = SpUserNotifications()
                    notification.row_id                 = approval.row_id
                    notification.user_id                = approval.user_id
                    notification.model_name             = 'SpFocRequests'
                    notification.notification           = 'Order '+approval.level+' Request has been sent.'
                    notification.is_read                = 0
                    notification.created_by_user_id     = request.user.id
                    notification.created_by_user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    notification.save()

    if level_id == '2':
        for order in order_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Sample Request has been forwarded'
            activity    = 'Sample Request has been forwarded by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            
            saveActivity('Order Management', 'Order Summary', heading, activity, request.user.id, user_name, 'forwaord.png', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpFocRequests,order,'request_by_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            message_title = "Sample request forwarded"
            message_body = "A sample request("+employee_name+" - "+user_role+") has been forwarded  by "+user_name
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(order,'SpFocRequests','Order Management','Sample request forwarded',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
             
    elif level_id == '3':
        for order in order_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Sample Request has been approved'
            activity    = 'Sample Request has been approved by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            
            saveActivity('Order Management', 'Sample Request Summary', heading, activity, request.user.id, user_name, 'approved.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpFocRequests,order,'request_by_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            message_title = "Sample request approved"
            message_body = "A sample request("+employee_name+" - "+user_role+") has been approved  by "+user_name
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(order,'SpFocRequests','Order Management','Sample request approved',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#

    response['error'] = False
    response['message'] = "Foc status has been updated successfully."
    return JsonResponse(response)


#get foc status view
@login_required
def focStatusDetails(request):
    foc_id                    = request.GET.get('order_id')
    initiate_foc_details      = SpFocRequests.objects.get(id=foc_id)  
    foc_details             = SpApprovalStatus.objects.filter(row_id=foc_id, model_name='SpFocRequests', status=1).values('final_status_user_id').distinct().values('final_status_user_name', 'final_update_date_time', 'level_id')
    
    context = {}
    context['initiate_foc_details']   = initiate_foc_details
    context['foc_details']            = foc_details
    template = 'order-management/foc-requests/foc-status-details.html'

    return render(request, template, context)

@login_required
def FocRecordByDateName(request):
    context = {}
    name = request.GET.get('name')
    today = request.GET.get('order_date')
    today = today.split('/')
    today = today[2]+'-'+today[1]+'-'+today[0]
    page = request.GET.get('page')
    foc_requests = SpFocRequests.objects.all().order_by('-id').filter(foc_delivery_date__icontains=today, user_name__icontains=name)
    paginator = Paginator(foc_requests, getConfigurationResult('page_limit'))
    try:
        foc_requests = paginator.page(page)
    except PageNotAnInteger:
        foc_requests = paginator.page(1)
    except EmptyPage:
        foc_requests = paginator.page(paginator.num_pages)
    if page is not None:
        page = page
    else:
        page = 1

    total_pages = int(paginator.count / 100)
    if (paginator.count == 0):
        paginator.count = 1

    temp = int(total_pages) % paginator.count
    if (temp > 0 and 100 != paginator.count):
        total_pages = total_pages + 1
    else:
        total_pages = total_pages

    last_foc_request = SpFocRequests.objects.order_by('-id').first()
    if last_foc_request:
        context['last_foc_request'] = last_foc_request
        context['last_foc_request_details'] = SpFocRequestsDetails.objects.filter(foc_request_id=last_foc_request.id)

    context['foc_requests'] = foc_requests
    context['today_date'] = today
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')

    template = 'order-management/foc-requests/ajax-foc-requests-lists.html'

    return render(request, template, context)



@login_required
def usersForFoc(request):
    context = {}
    detailsOfUsersForFoc = SpUsers.objects.all().filter(role_id__exclude=0, status=1, user_type=1)
    context['users_details'] = detailsOfUsersForFoc
    template = 'order-management/foc-requests/ajax-foc-requests-lists.html'
    print(context)
    return render(request, template, context)

    

@login_required
@has_par(sub_module_id=39,permission='export')
def focExportToXlsx(request, columns, userName, states, delvry_time):
    column_list = columns.split (",")
    foc_requests = SpFocRequests.objects.all().order_by('-id')

    if userName != '0':
        foc_requests = foc_requests.filter(user_name=userName) 
    if states != '0':
        foc_requests = foc_requests.filter(foc_status=states)
    if delvry_time != '0':
        foc_requests = foc_requests.filter(foc_delivery_date__icontains=delvry_time)
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Sampling Report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Foc-reports'
    
    # Define the titles for columns
    columns = []

    if 'user_name_foc' in column_list:
        columns += [ 'User Name' ]

    if 'request_by_name_foc' in column_list:
        columns += [ 'Employee Name' ]

    if 'foc_status_foc' in column_list:
        columns += [ 'Status' ]

    if 'foc_delivery_date_foc' in column_list:
        columns += [ 'Foc Delivery Date' ]

    if 'created_at_foc' in column_list:
        columns += [ 'Created At' ]

    

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    
    for result in foc_requests:
        row_num += 1
        # Define the data for each cell in the row 

        row = []

        if 'user_name_foc' in column_list:
            row += [ result.user_name ]

        if 'request_by_name_foc' in column_list:
            row += [ result.request_by_name ]

       

        if 'foc_status_foc' in column_list:
            if result.foc_status == 1:
                foc_status = 'Initiated'
            elif result.foc_status == 2:
                foc_status ='Forwarded'
            elif result.foc_status == 3:
                foc_status ='Approved'       
            else:
                foc_status='Delivered'

            row += [ foc_status ]

        if 'foc_delivery_date_foc' in column_list:
            foc_delivery_date = str(result.foc_delivery_date).replace('+00:00', '')
            row += [ datetime.strptime(str(foc_delivery_date), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %I:%M:%p') ]

        if 'created_at_foc' in column_list:
            created_at = str(result.created_at).replace('+00:00', '')
            row += [ datetime.strptime(str(created_at), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %I:%M:%p') ] 

         
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
    workbook.save(response)

    return response

    
def foc_render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None

#Automaticly downloads to PDF file
@login_required
@has_par(sub_module_id=39,permission='export')
def focExportToPDF(request, columns, userName, states, delvry_time):
    context = {}
    column_list = columns.split (",")
    foc_requests = SpFocRequests.objects.all().order_by('-id')

    if userName != '0':
        foc_requests = foc_requests.filter(user_name=userName) 
    if states != '0':
        foc_requests = foc_requests.filter(foc_status=states)
    if delvry_time != '0':
        foc_requests = foc_requests.filter(foc_delivery_date__icontains=delvry_time)
    
    baseurl = settings.BASE_URL
    pdf = foc_render_to_pdf('order-management/foc-requests/foc_pdf_template.html', {'foc_requests': foc_requests, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Sampling Report.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response

#get order sap Report
@login_required
@has_par(sub_module_id=9,permission='list')
def truckSheetDispatchReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today.strftime("%Y-%m-%d")).count()
    today_orders             = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today.strftime("%Y-%m-%d")).values_list('id',flat=True)
    order_regenerate_status = SpOrders.objects.filter(order_date__icontains=today.strftime("%Y-%m-%d")).count()
    distnict_route          = SpLogisticPlanDetail.objects.filter(order_date__icontains=today).values_list('route_id',flat=True).distinct()
    distnict_vehicle        = SpLogisticPlanDetail.objects.filter(order_date__icontains=today,order_id__in=today_orders).values_list('vehicle_id',flat=True).distinct()
    block_order_ids         = SpOrders.objects.filter(block_unblock=0, order_date__icontains=today).values_list('id',flat=True)
    clean_dist_order_ids = []
 
    if distnict_route:
    # vehicle = SpLogisticPlanDetail.objects.filter(route_id = route_id,order_date__icontains=today.strftime("%Y-%m-%d")).values_list('vehicle_id').distinct()
        vehicle_id              = SpLogisticPlanDetail.objects.filter(route_id= distnict_route[0],order_date__icontains=today.strftime("%Y-%m-%d")).values_list('vehicle_id', flat=True).distinct().first()
        dist_order_ids          = SpLogisticPlanDetail.objects.filter(order_date__icontains=today,vehicle_id=vehicle_id).values_list('order_id',flat=True).distinct()    
        vehicles = SpVehicles.objects.filter(route_id = distnict_route[0]).values('id','registration_number')
        if block_order_ids:
            clean_dist_order_ids = [e for e in dist_order_ids if e not in block_order_ids]
        else:
            clean_dist_order_ids = dist_order_ids
    else:
        vehicles=''
        vehicle_id=''
    user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_orders on sp_orders.user_id = sp_users.id
    where DATE(sp_orders.order_date) = %s and sp_orders.block_unblock=%s''',[today.strftime("%Y-%m-%d"),1])

    product_milk_variant_list           =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.item_sku_code, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='1' and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''')
    product_without_milk_variant_list   =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.item_sku_code, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='1' and sp_product_variants.status='1' order by sp_product_variants.order_of asc ''')
    milk_product_class_name             = SpProductClass.objects.get(id=1)
    indent_lists                                = []
    total_crate_sum                             = []
    total_free_milk_pouches_sum                 = []
    total_free_without_milk_pouches_sum         = []
    total_without_milk_crate_sum                = []
    total_milk_crates_quantity                  = []
    total_free_milk_pouches_quantity            = []
    total_free_without_milk_pouches_quantity    = []
    total_flat_scheme_incentive                 = []
    total_bulk_scheme_incentive                 = []

    for user in user_list:
        if user.order_id in clean_dist_order_ids:
            flat_scheme_incentive = getlogFlatBulkSchemeIncentive(user.order_id, user.id, 'flat',vehicle_id)
            total_flat_scheme_incentive.append(flat_scheme_incentive)

            bulk_scheme_incentive = getlogFlatBulkSchemeIncentive(user.order_id, user.id, 'bulkpack',vehicle_id)
            total_bulk_scheme_incentive.append(bulk_scheme_incentive)

            indent = {}
            product_milk_variants_list      = []
            total_milk_crates               = []
            total_free_scheme_milk_pouches  = []
            
            for id, product_variant_milk in enumerate(product_milk_variant_list):
                total = 0
                total_free_scheme = 0
                product_variant_milk_lists = {}
                total_crates = {}
                try:
                    order_detailss = SpLogisticPlanDetail.objects.filter(product_variant_id=product_variant_milk.id, order_id=user.order_id,vehicle_id = vehicle_id)
                except SpLogisticPlanDetail.DoesNotExist:
                    order_detailss = None
                if order_detailss:
                    Item_in_liters = 0
                    for order_details in order_detailss:
                        if order_details.packaging_type == '0':
                            Item_in_liters += float(order_details.quantity)
                        else:
                            Item_in_liters    += round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                        total             = total + Item_in_liters
                else:
                    Item_in_liters    = 0
                    total             = total + Item_in_liters

                free_scheme       = getlogFreeScheme(product_variant_milk.id, user.order_id, user.id,vehicle_id)
                total_free_scheme = total_free_scheme + free_scheme

                product_variant_milk_lists['id']                    = product_variant_milk.id    
                product_variant_milk_lists['milk_items']            = Item_in_liters
                product_variant_milk_lists['product_color_code']    = product_variant_milk.product_color_code
                product_variant_milk_lists['free_scheme']           = free_scheme
                
                total_milk_crates.append(total)
                total_crates = total_milk_crates
                total_free_scheme_milk_pouches.append(total_free_scheme)
                total_free_scheme_milk = total_free_scheme_milk_pouches  
                product_milk_variants_list.append(product_variant_milk_lists)

            total_crate_sum.append(total_crates)
            sum_total_crates = sum(total_milk_crates)
            total_milk_crates_quantity.append(sum_total_crates)

            total_free_milk_pouches_sum.append(total_free_scheme_milk)
            sum_total_free_milk_pouches_sum = sum(total_free_scheme_milk_pouches)
            total_free_milk_pouches_quantity.append(sum_total_free_milk_pouches_sum)
            

            product_without_milk_variants_list = []
            total_without_milk_crates = []
            total_free_scheme_without_milk_pouches = []
            for product_variant_without_milk in product_without_milk_variant_list:
                totals = 0
                total_free_scheme = 0
                product_variant_without_milk_lists = {}
                try:
                    order_detailss = SpLogisticPlanDetail.objects.filter(product_variant_id=product_variant_without_milk.id, order_id=user.order_id,vehicle_id = vehicle_id)
                except SpLogisticPlanDetail.DoesNotExist:
                    order_detailss = None
                if order_detailss:
                    Item_in_liters = 0
                    for order_details in order_detailss:
                        if order_details.packaging_type == '0':
                            Item_in_liters    += float(order_details.quantity)
                        else:
                            Item_in_liters += round(
                                (float(order_details.quantity)/float(order_details.product_no_of_pouch)), 2)
                        totals              = totals + Item_in_liters
                else:
                    Item_in_liters = 0
                    totals = totals + Item_in_liters
                
                free_scheme         = getlogFreeScheme(product_variant_milk.id, user.order_id, user.id,vehicle_id)
                
                total_free_scheme   = total_free_scheme + free_scheme

                product_variant_without_milk_lists['id']                 = product_variant_without_milk.id    
                product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
                product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
                product_variant_without_milk_lists['free_scheme']        = free_scheme

                total_without_milk_crates.append(totals)
                total_without_milk_crates = total_without_milk_crates

                total_free_scheme_without_milk_pouches.append(total_free_scheme)
                total_free_scheme_without_milk = total_free_scheme_without_milk_pouches  

                product_without_milk_variants_list.append(product_variant_without_milk_lists)    
            
            
            total_without_milk_crate_sum.append(total_without_milk_crates)

            total_free_without_milk_pouches_sum.append(total_free_scheme_without_milk)
            sum_total_free_without_milk_pouches_sum = sum(total_free_scheme_without_milk_pouches)
            total_free_without_milk_pouches_quantity.append(sum_total_free_without_milk_pouches_sum)

            contacts = SpContactNumbers.objects.filter(user_id=user.id).values_list("contact_number" , flat=True )
            if len(contacts)>0:
                indent['contact_number1']            = contacts[0]
            else:
                indent['contact_number1']            = ''
            if len(contacts)>1:
                indent['contact_number2']            = contacts[1]
            else:
                indent['contact_number2']            = ''
            
            indent['id']                    = user.id
            indent['first_name']            = user.first_name
            indent['middle_name']           = user.middle_name
            indent['last_name']             = user.last_name
            indent['emp_sap_id']            = user.emp_sap_id
            indent['store_name']            = user.store_name
            indent['town_name']             = user.town_name
            indent['milk_items']            = product_milk_variants_list
            indent['without_milk_items']    = product_without_milk_variants_list
            indent_lists.append(indent)
    
    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = sum([x[column[0]] for x in total_crate_sum])
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = sum([x[column[0]] for x in total_without_milk_crate_sum])
            total_without_milk_crates_qty.append(count)
    
    #sum of milk product pouches
    total_free_milk_pouches_qty = []
    if total_free_milk_pouches_sum:
        for column in enumerate(total_free_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_milk_pouches_sum])
            total_free_milk_pouches_qty.append(count)

    #sum of without milk product pouches
    total_free_without_milk_pouches_qty = []
    if total_free_without_milk_pouches_sum:
        for column in enumerate(total_free_without_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_without_milk_pouches_sum])
            total_free_without_milk_pouches_qty.append(count)        
    #print(total_free_without_milk_pouches_qty)
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today.strftime("%Y-%m-%d")).count()

    total_incentive_amount = [x + y for x, y in zip(total_flat_scheme_incentive, total_bulk_scheme_incentive)]
    
    route_id = SpVehicles.objects.filter(route_id__in=distnict_route).values('route_id').order_by("-id")
    routes = SpRoutes.objects.filter(id__in = route_id)
    # vehicles = SpVehicles.objects.filter(route_id__in=route_id)
    
    # routes=SpLogisticPlanDetail.objects.filter(route_id__in=distnict_route,order_date__icontains=today).values_list()
    context = {}
    context['user_list']                            = user_list
    context['indent_lists']                         = indent_lists
    context['total_milk_crates']                    = total_milk_crates_quantity
    context['total_free_milk_pouches']              = total_free_milk_pouches_qty
    context['total_milk_crates_qty']                = total_milk_crates_qty
    context['total_without_milk_crates_qty']        = total_without_milk_crates_qty
    context['total_free_without_milk_pouches_qty']  = total_free_without_milk_pouches_qty
    context['product_milk_variant_list']            = product_milk_variant_list
    context['product_without_milk_variant_list']    = product_without_milk_variant_list
    context['total_flat_scheme_incentive']          = total_flat_scheme_incentive
    context['total_bulk_scheme_incentive']          = total_bulk_scheme_incentive
    context['total_incentive_amount']               = total_incentive_amount
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['clean_dist_order_ids']                 = len(clean_dist_order_ids)
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['milk_product_class_name']              = milk_product_class_name
    context['page_title']                           = "Truck Sheet Report"
    context['routes']                               = routes
    context['vehicles']                             = vehicles
    context['vehicle_id']                           = vehicle_id
    if distnict_route:
        # context['route_code']                           = getModelColumnByColumnId(SpRoutes,'id',distnict_route[0],'route_code')
        context['route_name']                           = getModelColumnByColumnId(SpVehicles,'id',vehicle_id,'route_name')
        context['registration_number']                  = getModelColumnByColumnId(SpVehicles,'id',vehicle_id,'registration_number')
        context['select_route_id']                      = distnict_route[0]
    else:
        context['route_code']=''
        context['route_name'] =''
        context['registration_number']=''
        context['select_route_id'] =''
    template                                        = 'order-management/trucksheet-dispatch-report.html'
    
    return render(request, template, context)    



#get order SAP Report
@login_required
@has_par(sub_module_id=9,permission='list')
def ajaxTruckSheetDispatchReport(request):
    today                   = request.GET['order_date']
    route_code              = request.GET['route_code']
    route_id                = request.GET['route_id']
    vehicle_id              = request.GET['vehicle_id']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(block_unblock=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter( order_date__icontains=today).count()
    block_order_ids         = SpOrders.objects.filter(block_unblock=0, order_date__icontains=today).values_list('id',flat=True)
    try:
        routes = SpRoutes.objects.get(id=route_id)
    except SpRoutes.DoesNotExist:
        routes=None
    if routes:
        route_code = routes.route_code
        route_id   = routes.id
        route_name   = routes.route
    else:
        route_code = ''
        route_name=''
    dist_order_ids          = SpLogisticPlanDetail.objects.filter(order_date__icontains=today,vehicle_id=vehicle_id).values_list('order_id',flat=True).distinct() 
    clean_dist_order_ids = [e for e in dist_order_ids if e not in block_order_ids]
    user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_orders on sp_orders.user_id = sp_users.id
    where DATE(sp_orders.order_date) = %s and sp_orders.block_unblock=%s ''', [today,1])
    # order_list = SpLogisticPlanDetail.objects.raw('''SELECT sp_logistic_plan_detail.*,sp_orders.* FROM sp_logistic_plan_detail left join sp_orders on sp_logistic_plan_detail.order_id = sp_orders.id 
    # where DATE(sp_orders.order_date) =%s and sp_orders.block_unblock=%s and sp_logistic_plan_detail.vehicle_id=%s''',[today,1,vehicle_id])
    # distnict_order_id=SpLogisticPlanDetail.objects.filter(order_date=today,vehicle_id=vehicle_id).values_list('order_id',flat=True).distinct()

    product_milk_variant_list           =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.item_sku_code, sp_product_variants.variant_name, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id  WHERE sp_product_variants.product_class_id='1' and sp_product_variants.status='1'   order by sp_product_variants.order_of asc ''')
    product_without_milk_variant_list   =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.item_sku_code, sp_product_variants.variant_name, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id  WHERE sp_product_variants.product_class_id!='1' and sp_product_variants.status='1'   order by sp_product_variants.order_of asc ''')
    milk_product_class_name             = SpProductClass.objects.get(id=1)

    indent_lists                                = []
    total_crate_sum                             = []
    total_free_milk_pouches_sum                 = []
    total_free_without_milk_pouches_sum         = []
    total_without_milk_crate_sum                = []
    total_milk_crates_quantity                  = []
    total_free_milk_pouches_quantity            = []
    total_free_without_milk_pouches_quantity    = []
    total_flat_scheme_incentive                 = []
    total_bulk_scheme_incentive                 = []
    for user in user_list:
        if user.order_id in clean_dist_order_ids:
            flat_scheme_incentive = getlogFlatBulkSchemeIncentive(user.order_id, user.id, 'flat',vehicle_id)
            total_flat_scheme_incentive.append(flat_scheme_incentive)

            bulk_scheme_incentive = getlogFlatBulkSchemeIncentive(user.order_id, user.id, 'bulkpack',vehicle_id)
            total_bulk_scheme_incentive.append(bulk_scheme_incentive)

            indent = {}
            product_milk_variants_list      = []
            total_milk_crates               = []
            total_free_scheme_milk_pouches  = []
            
            for id, product_variant_milk in enumerate(product_milk_variant_list):
                total = 0
                total_free_scheme = 0
                product_variant_milk_lists = {}
                total_crates = {}
                try:
                    order_detailss = SpLogisticPlanDetail.objects.filter(product_variant_id=product_variant_milk.id, order_id=user.order_id,vehicle_id = vehicle_id)
                except SpLogisticPlanDetail.DoesNotExist:
                    order_detailss = None
                if order_detailss:
                    Item_in_liters = 0
                    for order_details in order_detailss:
                        if order_details.packaging_type == '0':
                            Item_in_liters += float(order_details.quantity)
                        else:
                            Item_in_liters    += round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                        total             = total + Item_in_liters
                else:
                    Item_in_liters    = 0
                    total             = total + Item_in_liters

                free_scheme       = getlogFreeScheme(product_variant_milk.id, user.order_id, user.id,vehicle_id)
                total_free_scheme = total_free_scheme + free_scheme

                product_variant_milk_lists['id']                    = product_variant_milk.id    
                product_variant_milk_lists['milk_items']            = Item_in_liters
                product_variant_milk_lists['product_color_code']    = product_variant_milk.product_color_code
                product_variant_milk_lists['free_scheme']           = free_scheme
                
                total_milk_crates.append(total)
                total_crates = total_milk_crates
                total_free_scheme_milk_pouches.append(total_free_scheme)
                total_free_scheme_milk = total_free_scheme_milk_pouches  
                product_milk_variants_list.append(product_variant_milk_lists)

            total_crate_sum.append(total_crates)
            sum_total_crates = sum(total_milk_crates)
            total_milk_crates_quantity.append(sum_total_crates)

            total_free_milk_pouches_sum.append(total_free_scheme_milk)
            sum_total_free_milk_pouches_sum = sum(total_free_scheme_milk_pouches)
            total_free_milk_pouches_quantity.append(sum_total_free_milk_pouches_sum)
            

            product_without_milk_variants_list = []
            total_without_milk_crates = []
            total_free_scheme_without_milk_pouches = []
            for product_variant_without_milk in product_without_milk_variant_list:
                totals = 0
                total_free_scheme = 0
                product_variant_without_milk_lists = {}
                try:
                    order_detailss = SpLogisticPlanDetail.objects.filter(product_variant_id=product_variant_without_milk.id, order_id=user.order_id,vehicle_id = vehicle_id)
                except SpLogisticPlanDetail.DoesNotExist:
                    order_detailss = None
                if order_detailss:
                    Item_in_liters = 0
                    for order_details in order_detailss:
                        if order_details.packaging_type == '0':
                            Item_in_liters    += float(order_details.quantity)
                        else:
                            Item_in_liters    += round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                        totals              = totals + Item_in_liters
                else:
                    Item_in_liters = 0
                    totals = totals + Item_in_liters
                
                free_scheme         = getlogFreeScheme(product_variant_milk.id, user.order_id, user.id,vehicle_id)
                total_free_scheme   = total_free_scheme + free_scheme

                product_variant_without_milk_lists['id']                 = product_variant_without_milk.id    
                product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
                product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
                product_variant_without_milk_lists['free_scheme']        = free_scheme

                total_without_milk_crates.append(totals)
                total_without_milk_crates = total_without_milk_crates

                total_free_scheme_without_milk_pouches.append(total_free_scheme)
                total_free_scheme_without_milk = total_free_scheme_without_milk_pouches  

                product_without_milk_variants_list.append(product_variant_without_milk_lists)    
            
            
            total_without_milk_crate_sum.append(total_without_milk_crates)

            total_free_without_milk_pouches_sum.append(total_free_scheme_without_milk)
            sum_total_free_without_milk_pouches_sum = sum(total_free_scheme_without_milk_pouches)
            total_free_without_milk_pouches_quantity.append(sum_total_free_without_milk_pouches_sum)
            
            contacts = SpContactNumbers.objects.filter(user_id=user.id).values_list("contact_number" , flat=True )
            if len(contacts)>0:
                indent['contact_number1']            = contacts[0]
            else:
                indent['contact_number1']            = ''
            if len(contacts)>1:
                indent['contact_number2']            = contacts[1]
            else:
                indent['contact_number2']            = ''
                
            indent['id']                    = user.id
            indent['first_name']            = user.first_name
            indent['middle_name']           = user.middle_name
            indent['last_name']             = user.last_name
            indent['emp_sap_id']            = user.emp_sap_id
            indent['store_name']            = user.store_name
            indent['town_name']             = user.town_name
            indent['order_total_amount']    = user.order_total_amount
            indent['outstanding_amount']    = user.outstanding_amount
            indent['security_amount']       = user.security_amount
            indent['mode_of_payment']       = user.mode_of_payment
            indent['amount_to_be_paid']     = user.amount_to_be_paid
            indent['milk_items']            = product_milk_variants_list
            indent['without_milk_items']    = product_without_milk_variants_list
            indent_lists.append(indent)
        
    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = sum([x[column[0]] for x in total_crate_sum])
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = sum([x[column[0]] for x in total_without_milk_crate_sum])
            total_without_milk_crates_qty.append(count)
    
    #sum of milk product pouches
    total_free_milk_pouches_qty = []
    if total_free_milk_pouches_sum:
        for column in enumerate(total_free_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_milk_pouches_sum])
            total_free_milk_pouches_qty.append(count)

    #sum of without milk product pouches
    total_free_without_milk_pouches_qty = []
    if total_free_without_milk_pouches_sum:
        for column in enumerate(total_free_without_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_without_milk_pouches_sum])
            total_free_without_milk_pouches_qty.append(count)        
    #print(total_free_without_milk_pouches_qty)
    approved_order = SpOrders.objects.filter(order_status = 3,block_unblock=1, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today).count()

    total_incentive_amount = [float(x) + float(y) for x, y in zip(total_flat_scheme_incentive, total_bulk_scheme_incentive)]
    vehicle = SpVehicles.objects.filter(route_id=route_id).first() 
    if vehicle:
        vehicle_number = vehicle.registration_number
        transporter = vehicle.dealer_name
    else:
        vehicle_number = ''
        transporter = ''
    routes=SpRoutes.objects.all()
    today_date   = date.today()
    today_date = today_date.strftime("%d/%m/%Y")
    context = {}
    context['user_list']                            = user_list
    context['indent_lists']                         = indent_lists
    context['total_milk_crates']                    = total_milk_crates_quantity
    context['total_free_milk_pouches']              = total_free_milk_pouches_qty
    context['total_milk_crates_qty']                = total_milk_crates_qty
    context['total_without_milk_crates_qty']        = total_without_milk_crates_qty
    context['total_free_without_milk_pouches_qty']  = total_free_without_milk_pouches_qty
    context['product_milk_variant_list']            = product_milk_variant_list
    context['product_without_milk_variant_list']    = product_without_milk_variant_list
    context['total_flat_scheme_incentive']          = total_flat_scheme_incentive
    context['total_bulk_scheme_incentive']          = total_bulk_scheme_incentive
    context['total_incentive_amount']               = total_incentive_amount
    context['today_order_status']                   = today_order_status
    context['clean_dist_order_ids']                 = len(clean_dist_order_ids)
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['order_date']                           = request.GET['order_date']
    context['today_date']                           = today_date
    context['milk_product_class_name']              = milk_product_class_name
    context['page_title']                           = "Truck Sheet Report"
    context['route_code']                           = route_code
    context['route_name']                           = route_name
    context['vehicle_number']                       = getModelColumnById(SpVehicles, vehicle_id, 'registration_number') 
    context['transporter']                          = transporter
    template                                        = 'order-management/ajax-trucksheet-dispatch-report.html'
    
    return render(request, template, context)



#get order SAP Report
@login_required
@has_par(sub_module_id=9,permission='export')
def exportTruckSheetDispatchReport(request,order_date,route_id,vehicle_id):
    today                   = order_date
    route_id                = route_id
    vehicle_id              = vehicle_id
    today_order_status      = SpOrders.objects.filter(block_unblock=1,order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(order_date__icontains=today).count()
    block_order_ids         = SpOrders.objects.filter(block_unblock=0, order_date__icontains=today).values_list('id',flat=True)
    try:
        routes = SpRoutes.objects.get(id=route_id)
    except SpRoutes.DoesNotExist:
        routes=None
    if routes:
        route_code = routes.route_code
        route_id   = routes.id
        route_name   = routes.route
    else:
        route_code = ''
        route_name=''

    dist_order_ids          = SpLogisticPlanDetail.objects.filter(order_date__icontains=today,vehicle_id=vehicle_id).values_list('order_id',flat=True).distinct()
    clean_dist_order_ids = [e for e in dist_order_ids if e not in block_order_ids]
    user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_users.store_name, sp_orders.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_orders on sp_orders.user_id = sp_users.id
    where DATE(sp_orders.order_date) = %s and sp_orders.block_unblock=%s''',[today,1])

    product_milk_variant_list           =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.item_sku_code, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='1'and sp_product_variants.status='1'   order by sp_product_variants.order_of asc ''')
    product_without_milk_variant_list   =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.item_sku_code, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='1' and sp_product_variants.status='1'  order by sp_product_variants.order_of asc ''')
    milk_product_class_name             = SpProductClass.objects.get(id=1)

    indent_lists                                = []
    total_crate_sum                             = []
    total_free_milk_pouches_sum                 = []
    total_free_without_milk_pouches_sum         = []
    total_without_milk_crate_sum                = []
    total_milk_crates_quantity                  = []
    total_free_milk_pouches_quantity            = []
    total_free_without_milk_pouches_quantity    = []
    total_flat_scheme_incentive                 = []
    total_bulk_scheme_incentive                 = []
    for user in user_list:
        if user.order_id in clean_dist_order_ids:
            flat_scheme_incentive = getlogFlatBulkSchemeIncentive(user.order_id, user.id, 'flat',vehicle_id)
            total_flat_scheme_incentive.append(flat_scheme_incentive)

            bulk_scheme_incentive = getlogFlatBulkSchemeIncentive(user.order_id, user.id, 'bulkpack',vehicle_id)
            total_bulk_scheme_incentive.append(bulk_scheme_incentive)

            indent = {}
            product_milk_variants_list      = []
            total_milk_crates               = []
            total_free_scheme_milk_pouches  = []
            
            for id, product_variant_milk in enumerate(product_milk_variant_list):
                total = 0
                total_free_scheme = 0
                product_variant_milk_lists = {}
                total_crates = {}
                try:
                    order_detailss = SpLogisticPlanDetail.objects.filter(product_variant_id=product_variant_milk.id, order_id=user.order_id,vehicle_id = vehicle_id)
                except SpLogisticPlanDetail.DoesNotExist:
                    order_detailss = None
                if order_detailss:
                    Item_in_liters = 0
                    for order_details in order_detailss:
                        if order_details.packaging_type == '0':
                            Item_in_liters    += float(order_details.quantity)
                        else:
                            Item_in_liters    += round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                        total             = total + Item_in_liters
                else:
                    Item_in_liters    = 0
                    total             = total + Item_in_liters


                free_scheme       = getlogFreeScheme(product_variant_milk.id, user.order_id, user.id,vehicle_id)
                total_free_scheme = total_free_scheme + free_scheme

                product_variant_milk_lists['id']                    = product_variant_milk.id    
                product_variant_milk_lists['milk_items']            = Item_in_liters
                product_variant_milk_lists['product_color_code']    = product_variant_milk.product_color_code
                product_variant_milk_lists['free_scheme']           = free_scheme
                
                total_milk_crates.append(total)
                total_crates = total_milk_crates
                total_free_scheme_milk_pouches.append(total_free_scheme)
                total_free_scheme_milk = total_free_scheme_milk_pouches  
                product_milk_variants_list.append(product_variant_milk_lists)

            total_crate_sum.append(total_crates)
            sum_total_crates = sum(total_milk_crates)
            total_milk_crates_quantity.append(sum_total_crates)

            total_free_milk_pouches_sum.append(total_free_scheme_milk)
            sum_total_free_milk_pouches_sum = sum(total_free_scheme_milk_pouches)
            total_free_milk_pouches_quantity.append(sum_total_free_milk_pouches_sum)
            

            product_without_milk_variants_list = []
            total_without_milk_crates = []
            total_free_scheme_without_milk_pouches = []
            for product_variant_without_milk in product_without_milk_variant_list:
                totals = 0
                total_free_scheme = 0
                product_variant_without_milk_lists = {}
                try:
                    order_detailss = SpLogisticPlanDetail.objects.filter(product_variant_id=product_variant_without_milk.id, order_id=user.order_id,vehicle_id = vehicle_id)
                except SpLogisticPlanDetail.DoesNotExist:
                    order_detailss = None
                if order_detailss:
                    Item_in_liters = 0
                    for order_details in order_detailss:
                        if order_details.packaging_type == '0':
                            Item_in_liters    += float(order_details.quantity)
                        else:
                            Item_in_liters    += round((float(order_details.quantity)/float(order_details.product_no_of_pouch)),2)
                        totals              = totals + Item_in_liters
                else:
                    Item_in_liters = 0
                    totals = totals + Item_in_liters
                
                free_scheme         = getlogFreeScheme(product_variant_milk.id, user.order_id, user.id,vehicle_id)
                total_free_scheme   = total_free_scheme + free_scheme

                product_variant_without_milk_lists['id']                 = product_variant_without_milk.id    
                product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
                product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
                product_variant_without_milk_lists['free_scheme']        = free_scheme

                total_without_milk_crates.append(totals)
                total_without_milk_crates = total_without_milk_crates

                total_free_scheme_without_milk_pouches.append(total_free_scheme)
                total_free_scheme_without_milk = total_free_scheme_without_milk_pouches  

                product_without_milk_variants_list.append(product_variant_without_milk_lists)    
            total_without_milk_crate_sum.append(total_without_milk_crates)

            total_free_without_milk_pouches_sum.append(total_free_scheme_without_milk)
            sum_total_free_without_milk_pouches_sum = sum(total_free_scheme_without_milk_pouches)
            total_free_without_milk_pouches_quantity.append(sum_total_free_without_milk_pouches_sum)

            contacts = SpContactNumbers.objects.filter(user_id=user.id).values_list("contact_number" , flat=True )
            if len(contacts)>0:
                indent['contact_number1']            = contacts[0]
            else:
                indent['contact_number1']            = ''
            if len(contacts)>1:
                indent['contact_number2']            = contacts[1]
            else:
                indent['contact_number2']            = ''
                
            indent['id']                    = user.id
            indent['first_name']            = user.first_name
            indent['middle_name']           = user.middle_name
            indent['last_name']             = user.last_name
            indent['emp_sap_id']            = user.emp_sap_id
            indent['store_name']            = user.store_name
            indent['town_name']             = user.town_name
            indent['order_total_amount']    = user.order_total_amount
            indent['outstanding_amount']    = user.outstanding_amount
            indent['security_amount']       = user.security_amount
            indent['mode_of_payment']       = user.mode_of_payment
            indent['amount_to_be_paid']     = user.amount_to_be_paid
            indent['milk_items']            = product_milk_variants_list
            indent['without_milk_items']    = product_without_milk_variants_list
            indent_lists.append(indent)
    
    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = sum([x[column[0]] for x in total_crate_sum])
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = sum([x[column[0]] for x in total_without_milk_crate_sum])
            total_without_milk_crates_qty.append(count)
    
    #sum of milk product pouches
    total_free_milk_pouches_qty = []
    if total_free_milk_pouches_sum:
        for column in enumerate(total_free_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_milk_pouches_sum])
            total_free_milk_pouches_qty.append(count)

    #sum of without milk product pouches
    total_free_without_milk_pouches_qty = []
    if total_free_without_milk_pouches_sum:
        for column in enumerate(total_free_without_milk_pouches_sum[0]):
            count = sum([x[column[0]] for x in total_free_without_milk_pouches_sum])
            total_free_without_milk_pouches_qty.append(count)        
    
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()
    vehicle = SpVehicles.objects.filter(route_id=route_id).first() 
    if vehicle:
        transporter = vehicle.dealer_name
    else:
        vehicle_number = ''
        transporter = ''
    vehicle_number                       = SpVehicles.objects.get(id=vehicle_id)  
    routes=SpRoutes.objects.all()
    total_incentive_amount = [x + y for x, y in zip(total_flat_scheme_incentive, total_bulk_scheme_incentive)]

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=trucksheet-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Truck Sheet Report'
    worksheet.merge_cells('A1:B1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 8.5
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.width = 100
    img.alignment = 'center'
    img.anchor = 'B1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=3)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(name='Arial Nova Cond Light',size=12, color='FFFFFFFF', bold=True)

    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+3
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'Truck Loading Sheet'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=24,  color='000000', bold=True)

    color_codes = []

    columns = []
    columns += [ 'Route Name' ]
    columns += [ route_name ]
    # columns += [ '' ]
    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            columns += [ " " ]
    columns += [ " " ]
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            columns += [ " " ]
    
    columns[len(columns)-5]='Date'
    columns[len(columns)-2]= datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')
    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+3
    
    row_num = 2
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        if col_num <3 or col_num >(column_length-5):
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20   
    worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=column_length-5)
    worksheet.merge_cells(start_row=2, start_column=column_length-4, end_row=2, end_column=column_length-2)
    worksheet.merge_cells(start_row=2, start_column=column_length-1, end_row=2, end_column=column_length)
    color_codes = []

    columns = []
    columns += [ 'Route Code' ]
    columns += [ route_code ]
    # columns += [ ' ' ]
    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            columns += [ " " ]
    columns += [ " " ]
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            columns += [ " " ]
    
    columns[len(columns)-5]='Transporter'
    if transporter is not None:
        columns[len(columns)-2]= transporter
    else:
        columns[len(columns)-2]='_'
    # columns += [ 'Transporter ' ]
    # if transporter is not None:
    #     columns += [ transporter ]
    # else:
    #     columns += [ '-' ]
    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+3
    
    row_num = 3
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        if col_num <3 or col_num >(column_length-5):
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20
    worksheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=column_length-5)
    worksheet.merge_cells(start_row=3, start_column=column_length-4, end_row=3, end_column=column_length-2)
    worksheet.merge_cells(start_row=3, start_column=column_length-1, end_row=3, end_column=column_length)
    color_codes = []

    columns = []
    columns += [ 'Type of sheet' ]
    columns += [ 'Normal/Free Samples' ]
    # columns += [ ' ' ]
    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            columns += [ " " ]
    columns += [ " " ]
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            columns += [ " " ]
    
    # columns += [ 'Vehicle Number' ]
    # if vehicle_number is not None:
    #     columns += [ vehicle_number.registration_number ]
    # else:
    #     columns += [ '-' ]
    columns[len(columns)-5]='Vehicle Number'
    if vehicle_number is not None:
        columns[len(columns)-2]= vehicle_number.registration_number
    else:
        columns[len(columns)-2]='_'
    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+3
    
    row_num = 4
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        if col_num <3 or col_num >(column_length-5):
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20   
        column_dimensions.height = 20   
    worksheet.merge_cells(start_row=4, start_column=3, end_row=4, end_column=column_length-5)
    worksheet.merge_cells(start_row=4, start_column=column_length-4, end_row=4, end_column=column_length-2)
    worksheet.merge_cells(start_row=4, start_column=column_length-1, end_row=4, end_column=column_length)
    
    columns = []
    columns += [ 'S.No.' ]
    columns += [ 'Store Name (Distributor/Sap Code)' ]
    # columns += [ 'Mobile Number 1' ]
    # columns += [ 'Mobile Number 2' ]
    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            columns += [ product_variant.variant_name ]
    columns += [ 'TOTAL '+milk_product_class_name.product_class+' CRATES' ]
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            columns += [ product_variant.variant_name ]
    # columns += [ 'Total Dispatch (Nos)' ]

    row_num = 5
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num>2:
            column_dimensions.width = 6
        else:
          column_dimensions.width = 15  
    for id, indent in enumerate(indent_lists):
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        row += [ id+1 ]
        # row += [  indent['first_name'] +" "+indent['middle_name']+" "+indent['last_name']+" ("+indent['store_name']+")"  ]
        row += [ indent['store_name']+" ( "+indent['first_name'] +" "+indent['middle_name']+" "+indent['last_name']+"/"+indent['emp_sap_id']+") "]
        # row += [ indent['contact_number1'] ]
        # row += [ indent['contact_number2'] ]
        if indent['milk_items']:
            for item in indent['milk_items']:
                row += [ item['milk_items'] ]
        row += [ total_milk_crates_quantity[id] ]        
        if indent['without_milk_items']:
            for item in indent['without_milk_items']:
                row += [ item['without_milk_items'] ]
        # row += [ '' ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num <=4:
                cell.font = Font(name='Arial Nova Cond Light',size=10, color='000000',bold=True)
            elif col_num <= len(product_milk_variant_list)+3:
                for product_variant in product_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(size=10, color='000000',bold=True)
                    cell.fill = PatternFill()
            elif col_num == len(product_milk_variant_list)+4:
                for product_variant in product_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(name='Arial Nova Cond Light',size=10, color='000000',bold=True)
                    cell.fill = PatternFill()        
            elif col_num <= len(product_milk_variant_list)+len(product_without_milk_variant_list)+4:
                for product_variant in product_without_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(size=10, color='000000',bold=True)
                    cell.fill = PatternFill()        
            else:
                cell.fill = PatternFill()
                cell.font = Font(size=10, color='000000',bold=True)
     
    row_num += 1
    # Define the data for each cell in the row 
    row = []
    row += [ "GRAND TOTAL(CRATES)" ]
    row += [ " " ]
    # row += [  ""  ]
    # row += [  ""  ]
    if total_milk_crates_qty:
        for total_milk_crates in total_milk_crates_qty:
            row += [ total_milk_crates ]
    row += [ '' ]        
    if total_without_milk_crates_qty:
        for total_without_milk_crates in total_without_milk_crates_qty:
            row += [ total_without_milk_crates ]
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        if col_num <=2:
            cell.alignment=Alignment(
            horizontal='right',
            vertical='center',
            wrap_text=True
            )
        else:
            cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num <=2:
            cell.font = Font(name='Arial Nova Cond Light',size=10, color='000000',bold=True)
        elif col_num <= len(product_milk_variant_list)+2:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=10, color='000000',bold=True)
                cell.fill = PatternFill()
        elif col_num == len(product_milk_variant_list)+3:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(name='Arial Nova Cond Light',size=10, color='000000',bold=True)
                cell.fill = PatternFill()        
        elif col_num <= len(product_milk_variant_list)+len(product_without_milk_variant_list)+3:
            for product_variant in product_without_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=10, color='000000',bold=True)
                cell.fill = PatternFill()        
        else:
            cell.fill = PatternFill()
            cell.font = Font(size=10, color='000000',bold=True)         
    worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
     
    row_num += 1
    # Define the data for each cell in the row 
    row = []
    row += [ "Scheme in Nos" ]
    row += [ "" ]
    # row += [  ""  ]
    # row += [  ""  ]
    if total_free_milk_pouches_qty:
        for total_milk_pouches in total_free_milk_pouches_qty:
            row += [ total_milk_pouches ]
    row += [ '' ]        
    if total_free_without_milk_pouches_qty:
        for total_without_milk_pouches in total_free_without_milk_pouches_qty:
            row += [ total_without_milk_pouches ]
  
    
    
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        if col_num <=2:
            cell.alignment=Alignment(
            horizontal='right',
            vertical='center',
            wrap_text=True
            )
        else:
            cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num <=2:
            cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000',bold=True)
        elif col_num <= len(product_milk_variant_list)+2:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=10, color='000000')
                cell.fill = PatternFill()
        elif col_num == len(product_milk_variant_list)+3:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(name='Arial Nova Cond Light',size=12, color='000000',bold=True)
                cell.fill = PatternFill()        
        elif col_num <= len(product_milk_variant_list)+len(product_without_milk_variant_list)+3:
            for product_variant in product_without_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=10, color='000000',bold=True)
                cell.fill = PatternFill()        
        else:
            cell.fill = PatternFill()
            cell.font = Font(size=10, color='000000',bold=True)       
    worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    
    row_num += 2
    # Define the data for each cell in the row 
    row = []
    row += [ "" ]
    row += [  ""  ]
    row += [ "Dispatched" ]
    row += [  ""  ]
    row += [  "Recevied"  ]
    row += [  ""  ]
    
    
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        if col_num >1:
            cell.alignment = wrapped_alignment
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
    worksheet.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
    
    row_num += 1
    # Define the data for each cell in the row 
    row = []
    row += [ "" ]
    row += [ "Total  Crates" ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        if col_num >1:
            cell.alignment = wrapped_alignment
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
    worksheet.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)   
    row_num += 1
    # Define the data for each cell in the row 
    row = []
    row += [ "" ]
    row += [ "Total Boxes " ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    
    
    
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        if col_num >1:
            cell.alignment = wrapped_alignment
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
    worksheet.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)    
    row_num += 1
    # Define the data for each cell in the row 
    row = []
    row += [ "" ]
    row += [ "Total Matka " ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    
    
    
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        if col_num >1:
            cell.alignment = wrapped_alignment
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
    worksheet.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)    
    row_num += 1
    # Define the data for each cell in the row 
    row = []
    row += [ "" ]
    row += [ "Total Scheme Nos. " ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  ""  ]
    
    
    
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        if col_num >1:
            cell.alignment = wrapped_alignment
            cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
    worksheet.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)    
    row_num += 5
    # Define the data for each cell in the row 
    row = []
    row += [ "" ]
    row += [ "Dispatch Supervisor" ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  "Dispatch Incharge"  ]
    row += [  ""  ]
    row += [  ""  ]
    row += [  "Security "  ]
    row += [  ""  ]
    row += [  "Official"  ]
    
    
    
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.fill = PatternFill()
        cell.font = Font(size=10, color='000000',bold=True)
            
                       
    total_columns = len(color_codes)+3           
    for rows in worksheet.iter_rows(min_row=2, max_row=len(indent_lists)+2, min_col=4, max_col=total_columns):
        for id, cell in enumerate(rows):
            cell.fill = PatternFill(start_color=color_codes[id], end_color=color_codes[id], fill_type = "solid")            
    # Define the titles for bottom_columns
    row_num += 2
    bottom_columns = []
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = wrapped_alignment
        cell.font = Font(size=11, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20
    
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)
    # img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.png')
    # img.height = 50
    # img.alignment = wrapped_alignment
    # img.anchor = 'W'+str(last_row)
    # worksheet.add_image(img)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response

# Sales Report List
@login_required
#@has_par(sub_module_id=8,permission='list')
def salesReport(request):
    today                = date.today() 
    today                  =today.strftime("%Y-%m-%d") 
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    orders  = SpOrders.objects.all().order_by('-id').filter(order_date__icontains=today,block_unblock=1,)
    orders_list=[]
    if orders:
        for order in orders:
            order_list={}
            order_list["user_sap_id"]=order.user_sap_id
            order_list["order_date"]=order.order_date.strftime("%Y-%m-%d")
            order_list['store_name'] = getModelColumnById(SpUsers, order.user_id, 'store_name') 
            order_list["user_name"]=order.user_name
            order_list["town_name"]=order.town_name
            if order.tcs_value:
                totalsum =round((order.order_total_amount*order.tcs_value/100),2)
                total= round(order.order_total_amount,2)+totalsum
                order_list["order_total_amount"]=total
                order_list["tcs_amount"]=totalsum
            else:
                order_list["order_total_amount"]=order.order_total_amount
                order_list["tcs_amount"]=0.0
            order_list["amount_to_be_paid"]=order.amount_to_be_paid
            try:
                user_ledger1 = SpUserLedger.objects.filter(user_id=order.user_id).order_by('-id').exclude(created_at__icontains=today).first()
            except SpChallans.DoesNotExist:
                user_ledger1 = None
            if user_ledger1:
                order_list["opening_amount"] = user_ledger1.balance
            else:
                order_list["opening_amount"] = 0.0
            try:
                user_ledger = SpUserLedger.objects.filter(user_id=order.user_id,created_at__icontains=today).order_by('-id').first()
            except SpChallans.DoesNotExist:
                user_ledger = None
            if user_ledger:
                order_list["closing_amount"] = user_ledger.balance
            else:
                order_list["closing_amount"] = 0.0
            # print(order_list)
            sp_basics=SpBasicDetails.objects.all().order_by('-id').filter(user_id=order.user_id)
            for basic in sp_basics:
                order_list["security_amount"] = basic.security_amount 
            order_details=SpOrderDetails.objects.all().order_by('-id').filter(order_id=order.id)
            total_milk_amount=0
            other_amount=0
            for order_detail in order_details:
                if order_detail.product_name=='MILK':
                    total_milk_amount+=order_detail.amount
                else:
                    other_amount+=order_detail.amount
            order_list["total_milk_amount"]=round((total_milk_amount),2)
            order_list["other_amount"]= round((other_amount),2)        
                # print(order_detail.product_name)
            if SpUserCrateLedger.objects.filter(user_id=order.user_id ,updated_datetime__icontains = today).exists():
                user_dispatch= SpUserCrateLedger.objects.filter(user_id=order.user_id , normal_debit__gt=0, updated_datetime__icontains=today).order_by('-id').values('normal_debit','jumbo_debit').first()
                user_receive= SpUserCrateLedger.objects.filter(user_id=order.user_id , normal_credit__gt=0, updated_datetime__icontains=today).order_by('-id').values('normal_credit','jumbo_credit').first()
                user_crate_balance= SpUserCrateLedger.objects.filter(user_id=order.user_id , updated_datetime__icontains=today).order_by('-id').values('normal_balance','jumbo_balance').first()
                if user_receive:
                    order_list["reciving_crate"]=user_receive['normal_credit'] + user_receive['jumbo_credit']
                else:
                    order_list["reciving_crate"]=0
                if user_dispatch:
                    order_list["dispatch_crate"]=user_dispatch['normal_debit'] + user_dispatch['jumbo_debit']
                else:
                    order_list["dispatch_crate"]=0
                order_list["Total_crate"]=user_crate_balance['jumbo_balance']+user_crate_balance['normal_balance']
                try:
                    user_ledger1 = SpUserCrateLedger.objects.filter(user_id=order.user_id).order_by('-id').exclude(updated_datetime__icontains=today).first()
                    # user_ledger1 = SpUserCrateLedger.objects.filter(user_id=order.user_id).order_by('-id')[1]
                    order_list["opening_balance_crt"]=user_ledger1.jumbo_balance + user_ledger1.normal_balance
                except:
                    order_list["opening_balance_crt"]=0        
            else:
                order_list["reciving_crate"]=0
                order_list["dispatch_crate"]=0
                order_list["jumbo_balance"]=0
                order_list["Total_crate"]=0  
                
            orders_list.append(order_list)
    

            
    
    context={}
    cdate = date.today().strftime("%d/%m/%Y")
    context['cdate']                              = cdate
    context['orders']                             = orders
    context['orders_list']                        = orders_list
    context['today_order_status']                 = today_order_status
    context['order_regenerate_status']            = order_regenerate_status
    context['page_title']                         = "Sales Report"
    template                                      = 'order-management/sales-report.html'
    return render(request, template, context)

@login_required
#@has_par(sub_module_id=8,permission='list')
def ajaxSalesReportList(request):
    today                   = request.GET['order_date']       
    
    today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1,order_date__icontains=today)
    if request.GET['id']:
        today_order_status      = today_order_status.filter(user_id=request.GET['id'])
    today_order_status      = today_order_status.count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today)
    if request.GET['id']:
        order_regenerate_status      = order_regenerate_status.filter(user_id=request.GET['id'])
    order_regenerate_status      = order_regenerate_status.count()
    orders  = SpOrders.objects.filter(order_date__icontains=today,block_unblock=1)
    if request.GET['id']:
        orders = orders.filter(user_id=request.GET['id'])
    orders = orders.order_by('-id') 
    orders_list=[]
    if orders:
        for order in orders:
            order_list={}
            order_list["user_sap_id"]=order.user_sap_id
            order_list["order_date"]=order.order_date.strftime("%Y-%m-%d")
            order_list['store_name'] = getModelColumnById(SpUsers, order.user_id, 'store_name') 
            order_list["user_name"]=order.user_name
            order_list["town_name"]=order.town_name
            if order.tcs_value:
                totalsum =round((order.order_total_amount*order.tcs_value/100),2)
                
                total= round(order.order_total_amount,2)+totalsum
                order_list["order_total_amount"]=total
                order_list["tcs_amount"]=totalsum
            else:
                order_list["order_total_amount"]=order.order_total_amount
                order_list["tcs_amount"]=0.0
            order_list["amount_to_be_paid"]=order.amount_to_be_paid
            try:
                user_ledger1 = SpUserLedger.objects.filter(user_id=order.user_id).order_by('-id').exclude(created_at__icontains=today).first()
            except SpChallans.DoesNotExist:
                user_ledger1 = None
            if user_ledger1:
                order_list["opening_amount"] = user_ledger1.balance
            else:
                order_list["opening_amount"] = 0.0
            try:
                user_ledger = SpUserLedger.objects.filter(user_id=order.user_id,created_at__icontains=today).order_by('-id').first()
            except SpChallans.DoesNotExist:
                user_ledger = None
            if user_ledger:
                order_list["closing_amount"] = user_ledger.balance
            else:
                order_list["closing_amount"] = 0.0
            # print(order_list)
            sp_basics=SpBasicDetails.objects.all().order_by('-id').filter(user_id=order.user_id)
            for basic in sp_basics:
                
                order_list["security_amount"] = basic.security_amount 
            order_details=SpOrderDetails.objects.all().order_by('-id').filter(order_id=order.id)
            total_milk_amount=0
            other_amount=0
            for order_detail in order_details:
                if order_detail.product_name=='MILK':
                    total_milk_amount+=order_detail.amount
                else:
                    other_amount+=order_detail.amount
            order_list["total_milk_amount"]=round((total_milk_amount),2)
            order_list["other_amount"]= round((other_amount),2)       
            if SpUserCrateLedger.objects.filter(user_id=order.user_id ,updated_datetime__icontains = today).exists():
                user_dispatch= SpUserCrateLedger.objects.filter(user_id=order.user_id , normal_debit__gt=0, updated_datetime__icontains=today).order_by('-id').values('normal_debit','jumbo_debit').first()
                user_receive= SpUserCrateLedger.objects.filter(user_id=order.user_id , normal_credit__gt=0, updated_datetime__icontains=today).order_by('-id').values('normal_credit','jumbo_credit').first()
                user_crate_balance= SpUserCrateLedger.objects.filter(user_id=order.user_id , updated_datetime__icontains=today).order_by('-id').values('normal_balance','jumbo_balance').first()
                if user_receive:
                    order_list["reciving_crate"]=user_receive['normal_credit'] + user_receive['jumbo_credit']
                else:
                    order_list["reciving_crate"]=0
                if user_dispatch:
                    order_list["dispatch_crate"]=user_dispatch['normal_debit'] + user_dispatch['jumbo_debit']
                else:
                    order_list["dispatch_crate"]=0
                order_list["Total_crate"]=user_crate_balance['jumbo_balance']+user_crate_balance['normal_balance']

                try:
                    user_ledger1 = SpUserCrateLedger.objects.filter(user_id=order.user_id).order_by('-id').exclude(updated_datetime__icontains=today).first()
                    order_list["opening_balance_crt"]=user_ledger1.jumbo_balance + user_ledger1.normal_balance
                except:
                    order_list["opening_balance_crt"]=0
            else:
                order_list["reciving_crate"]=0
                order_list["dispatch_crate"]=0
                # order_list["jumbo_balance"]=0
                order_list["Total_crate"]=0  
                order_list["opening_balance_crt"]=0
            orders_list.append(order_list)
    
    context={}
    context['cdate']                              = today 
    context['orders']                             = orders
    context['orders_list']                        = orders_list
    context['today_order_status']                 = today_order_status
    context['order_regenerate_status']            = order_regenerate_status
    template                                      = 'order-management/ajax-sales-report.html'
    return render(request, template, context)

#export Sales Report List
@login_required
def ajaxExportSaleReportList(request,order_date,id):
    if id == "0":
        today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=order_date).count()
        order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=order_date).count()
        orders  = SpOrders.objects.all().order_by('-id').filter(order_date__icontains=order_date, block_unblock=1)
    else:
        today_order_status      = SpOrders.objects.filter(indent_status=1,block_unblock=1, order_date__icontains=order_date,user_id=id).count()
        order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=order_date,user_id=id).count()
        orders  = SpOrders.objects.all().order_by('-id').filter(order_date__icontains=order_date,user_id=id ,block_unblock=1)
    if len(orders)==0:
        heading     = 'Order not found.'
        messages.error(request, heading, extra_tags='success')
        return redirect('/sales-report')
    if orders:
        orders_list=[]
        for order in orders:
            order_list={}
            order_list["user_sap_id"]=order.user_sap_id
            order_list["order_date"]=order.order_date.strftime("%Y-%m-%d")
            order_list['store_name'] = getModelColumnById(SpUsers, order.user_id, 'store_name') 
            order_list["user_name"]=order.user_name
            order_list["town_name"]=order.town_name
            if order.tcs_value:
                totalsum =round((order.order_total_amount*order.tcs_value/100),2)
                
                total= round(order.order_total_amount,2)+totalsum
                order_list["order_total_amount"]=total
                order_list["tcs_amount"]=totalsum
            else:
                order_list["order_total_amount"]=order.order_total_amount
                order_list["tcs_amount"]=0.0
            order_list["amount_to_be_paid"]=order.amount_to_be_paid
            try:
                user_ledger1 = SpUserLedger.objects.filter(user_id=order.user_id).order_by('-id').exclude(created_at__icontains=today).first()
            except SpChallans.DoesNotExist:
                user_ledger1 = None
            if user_ledger1:
                order_list["opening_amount"] = user_ledger1.balance
            else:
                order_list["opening_amount"] = 0.0
            try:
                user_ledger = SpUserLedger.objects.filter(user_id=order.user_id,created_at__icontains=today).order_by('-id').first()
            except SpChallans.DoesNotExist:
                user_ledger = None
            if user_ledger:
                order_list["closing_amount"] = user_ledger.balance
            else:
                order_list["closing_amount"] = 0.0
            # print(order_list)
            sp_basics=SpBasicDetails.objects.all().order_by('-id').filter(user_id=order.user_id)
            for basic in sp_basics:
                order_list["security_amount"] = basic.security_amount 
            order_details=SpOrderDetails.objects.all().order_by('-id').filter(order_id=order.id)
            total_milk_amount=0
            other_amount=0
            for order_detail in order_details:
                if order_detail.product_name=='MILK':
                    total_milk_amount+=order_detail.amount
                else:
                    other_amount+=order_detail.amount
            order_list["total_milk_amount"]=round((total_milk_amount),2)
            order_list["other_amount"]= round((other_amount),2)        
            if SpUserCrateLedger.objects.filter(user_id=order.user_id ,updated_datetime__icontains = order_date).exists():
                user_dispatch= SpUserCrateLedger.objects.filter(user_id=order.user_id , normal_debit__gt=0, updated_datetime__icontains=order_date).order_by('-id').values('normal_debit','jumbo_debit').first()
                user_receive= SpUserCrateLedger.objects.filter(user_id=order.user_id , normal_credit__gt=0, updated_datetime__icontains=order_date).order_by('-id').values('normal_credit','jumbo_credit').first()
                user_crate_balance= SpUserCrateLedger.objects.filter(user_id=order.user_id , updated_datetime__icontains=order_date).order_by('-id').values('normal_balance','jumbo_balance').first()
                if user_receive:
                    order_list["reciving_crate"]=user_receive['normal_credit'] + user_receive['jumbo_credit']
                else:
                    order_list["reciving_crate"]=0
                if user_dispatch:
                    order_list["dispatch_crate"]=user_dispatch['normal_debit'] + user_dispatch['jumbo_debit']
                else:
                    order_list["dispatch_crate"]=0
                # order_list["jumbo_balance"]=user_crate_balance.jumbo_balance
                order_list["Total_crate"]=user_crate_balance['jumbo_balance']+user_crate_balance['normal_balance']

                try:
                    user_ledger1 = SpUserCrateLedger.objects.filter(user_id=order.user_id).order_by('-id').exclude(updated_datetime__icontains=order_date).first()
                    order_list["opening_balance_crt"]=user_ledger1.jumbo_balance + user_ledger1.normal_balance
                except:
                    order_list["opening_balance_crt"]=0
            else:
                order_list["reciving_crate"]=0
                order_list["dispatch_crate"]=0
                # order_list["jumbo_balance"]=0
                order_list["Total_crate"]=0  
                order_list["opening_balance_crt"]=0
            orders_list.append(order_list)
        # print(orders_list)    

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=Sales Report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
            # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='left')
        thin = Side(border_style="thin", color="303030") 
        black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wrapped_alignment = Alignment(
            vertical='top',
            horizontal='left',
            wrap_text=True
        )
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Sales Report'
        worksheet.merge_cells('A1:A1') 
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = 9
        worksheet.page_setup.fitToPage = True
        column_length = 16

        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
       
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = wrapped_alignment
        worksheet.cell(row=1, column=column_length).border = black_border
        worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()
        worksheet.cell(row=1, column=2).fill = PatternFill()


            # Define the titles for columns
        # columns = []
        row_num = 1
        worksheet.row_dimensions[1].height = 0
        
        # Define the titles for columns
        columns = []

        columns += [ 'DATE' ]
        columns += [ 'S.No' ]
        columns += [ 'SAP ID' ]
        columns += [ 'Party Name' ]
        columns += [ 'Town' ]
        columns += [ 'Milk Opening Balance' ]
        columns += [ 'Milk Sale Amount' ]
        columns += [ 'Other sale Amount' ]
        columns += [ 'TCS Amount' ]
        columns += [ 'total sale' ]
        columns += [ 'Amt. Recd.' ]
        columns += [ 'Total closing outstading' ]
        columns += [ 'Opening Balance Crt.' ]
        columns += [ 'DISPATCH Crt.' ]
        columns += [ 'RECEIVING Crt.' ]
        columns += [ 'Total outstading Crt.' ]
        columns += [ 'SECURITY DIPOSIT' ]


        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32

        counter = 1
        for order in orders_list:
            row_num += 1
            # Define the data for each cell in the row 

            row = []
            row += [ order["order_date"]]
            row += [counter]
            row += [ order["user_sap_id"] ]
            row += [ order["store_name"]+"("+order["user_name"]+")" ]
            row += [ order["town_name"] ]
            row += [ order["opening_amount"] ]
            row += [ order["total_milk_amount"] ]
            row += [ order["other_amount"] ]
            row += [ order["tcs_amount"] ]
            row += [ order["order_total_amount"] ]
            row += [ order["amount_to_be_paid"] ]
            row += [ order["closing_amount"] ]
            row += [ order["opening_balance_crt"] ]
            row += [ order["dispatch_crate"] ]
            row += [ order["reciving_crate"] ]
            row += [ order["Total_crate"]]
            row += [ order["security_amount"] ]
            counter+=1
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border  
        wrapped_alignment = Alignment(
        horizontal='center',
        wrap_text=True
        )

        row_num += 1
        last_row = row_num
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)
        worksheet.row_dimensions[last_row].height = 20
        worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
        worksheet.cell(row=last_row, column=1).font = header_font
        worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
        worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
        worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
       

        workbook.save(response)

        return response

