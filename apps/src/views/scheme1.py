import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q,F
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
import time

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage

# Create your views here.

# products View
@login_required
def index(request):

    context = {}
    free_page = request.GET.get('free_page')
    free_schemes = SpSchemes.objects.filter(scheme_type=1).order_by('-id')
    for free_scheme in free_schemes:
        route_name = ''
        if free_scheme.route_id:
            routes = free_scheme.route_id.split(',')
            for route in routes:
                route_name = route_name + getModelColumnById(SpRoutes, route, 'route') + ", "
        if free_scheme.town_id:
            towns = free_scheme.town_id.split(',')
            for town in towns:
                route_name = route_name + getModelColumnById(SpTowns, town, 'town') + ", "        
            
        free_scheme.route_name  = route_name
    free_paginator = Paginator(free_schemes, 1000)

    flat_page = request.GET.get('flat_page')
    flat_incentive_schemes = SpFlatSchemes.objects.all().order_by('-id')
    flat_paginator = Paginator(flat_incentive_schemes, getConfigurationResult('page_limit'))

    bulkpack_page = request.GET.get('bulkpack_page')
    bulkpack_incentive_schemes = SpBulkpackSchemes.objects.all().order_by('-id')
    bulkpack_paginator = Paginator(bulkpack_incentive_schemes, getConfigurationResult('page_limit'))

    quantitative_page = request.GET.get('quantitative_page')
    quantitative_schemes = SpSchemes.objects.filter(scheme_type=2,).order_by('-id')
    quantitative_paginator = Paginator(quantitative_schemes, getConfigurationResult('page_limit'))

    try:
        free_schemes = free_paginator.page(free_page)
        flat_incentive_schemes = flat_paginator.page(flat_page)
    except PageNotAnInteger:
        free_schemes = free_paginator.page(1)
        flat_incentive_schemes = flat_paginator.page(1)
    except EmptyPage:
        free_schemes = free_paginator.page(free_paginator.num_pages)  
        flat_incentive_schemes = paginator.page(flat_paginator.num_pages) 

    if free_page is not None:
           free_page = free_page
    else:
           free_page = 1
    
    if flat_page is not None:
           flat_page = flat_page
    else:
           flat_page = 1
    
    total_free_pages = int(free_paginator.count/getConfigurationResult('page_limit')) 
    total_flat_incentive_pages = int(flat_paginator.count/getConfigurationResult('page_limit')) 
    total_bulkpack_incentive_pages = int(bulkpack_paginator.count/getConfigurationResult('page_limit')) 
    total_quantitative_scheme_pages = int(quantitative_paginator.count/getConfigurationResult('page_limit')) 
    
    if(free_paginator.count == 0):
        free_paginator.count = 1
    
    if(flat_paginator.count == 0):
        flat_paginator.count = 1
    
    if(bulkpack_paginator.count == 0):
        bulkpack_paginator.count = 1

    if(quantitative_paginator.count == 0):
        quantitative_paginator.count = 1

    free_temp = total_free_pages%free_paginator.count
    if(free_temp > 0 and getConfigurationResult('page_limit')!= free_paginator.count):
        total_free_pages = total_free_pages+1
    else:
        total_free_pages = total_free_pages

    flat_temp = total_flat_incentive_pages%flat_paginator.count
    if(flat_temp > 0 and getConfigurationResult('page_limit')!= flat_paginator.count):
        total_flat_incentive_pages = total_flat_incentive_pages+1
    else:
        total_flat_incentive_pages = total_flat_incentive_pages

    bulkpack_temp = total_bulkpack_incentive_pages%bulkpack_paginator.count
    if(bulkpack_temp > 0 and getConfigurationResult('page_limit')!= bulkpack_paginator.count):
        total_bulkpack_incentive_pages = total_bulkpack_incentive_pages+1
    else:
        total_flat_incentive_pages = total_flat_incentive_pages

    quantitative_temp = total_quantitative_scheme_pages%quantitative_paginator.count
    if(quantitative_temp > 0 and getConfigurationResult('page_limit')!= quantitative_paginator.count):
        total_quantitative_scheme_pages = total_quantitative_scheme_pages+1
    else:
        total_quantitative_scheme_pages = total_quantitative_scheme_pages

    
    last_free_scheme = SpSchemes.objects.filter(scheme_type=1).order_by('-id').first()
    if last_free_scheme:

        scheme_free_users = list(SpUserSchemes.objects.raw(''' SELECT sp_user_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_schemes.user_id
        WHERE sp_user_schemes.scheme_id = %s
        order by sp_user_schemes.id  ''',[last_free_scheme.id]))

        free_page = request.GET.get('free_page')
        free_user_paginator = Paginator(scheme_free_users, 50)
        try:
            scheme_free_users = free_user_paginator.page(free_page)
        except PageNotAnInteger:
            scheme_free_users = free_user_paginator.page(1)
        except EmptyPage:
            scheme_free_users = free_user_paginator.page(free_user_paginator.num_pages)  
        if free_page is not None:
            free_page = free_page
        else:
            free_page = 1
        total_scheme_free_users_pages = int(free_user_paginator.count/getConfigurationResult('page_limit')) 
        
        if(free_user_paginator.count == 0):
            free_user_paginator.count = 1

        temp = total_scheme_free_users_pages%free_user_paginator.count
        if(temp > 0 and getConfigurationResult('page_limit')!= free_user_paginator.count):
            total_scheme_free_users_pages = total_scheme_free_users_pages+1
        else:
            total_scheme_free_users_pages = total_scheme_free_users_pages
        context['total_scheme_free_users_pages']            = total_scheme_free_users_pages
    else:
        scheme_free_users = {}
        context['total_scheme_free_users_pages']            = 0

    
    last_flat_incentive_scheme = SpFlatSchemes.objects.order_by('-id').first()
    if last_flat_incentive_scheme:
        scheme_flat_incentive_users = list(SpUserFlatSchemes.objects.raw(''' SELECT sp_user_flat_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_flat_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_flat_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_flat_schemes.user_id
        WHERE sp_user_flat_schemes.scheme_id = %s
        order by sp_user_flat_schemes.id  ''',[last_flat_incentive_scheme.id]))
        flat_incentive_page = request.GET.get('flat_incentive_page')
        flat_incentive_user_paginator = Paginator(scheme_flat_incentive_users, 50)
        try:
            scheme_flat_incentive_users = flat_incentive_user_paginator.page(flat_incentive_page)
        except PageNotAnInteger:
            scheme_flat_incentive_users = flat_incentive_user_paginator.page(1)
        except EmptyPage:
            scheme_flat_incentive_users = flat_incentive_user_paginator.page(flat_incentive_user_paginator.num_pages)  
        if flat_incentive_page is not None:
            flat_incentive_page = flat_incentive_page
        else:
            flat_incentive_page = 1
        total_scheme_flat_incentive_users_pages = int(flat_incentive_user_paginator.count/getConfigurationResult('page_limit')) 
        
        if(flat_incentive_user_paginator.count == 0):
            flat_incentive_user_paginator.count = 1

        temp = total_scheme_flat_incentive_users_pages%flat_incentive_user_paginator.count
        if(temp > 0 and getConfigurationResult('page_limit')!= flat_incentive_user_paginator.count):
            total_scheme_flat_incentive_users_pages = total_scheme_flat_incentive_users_pages+1
        else:
            total_scheme_flat_incentive_users_pages = total_scheme_flat_incentive_users_pages
        context['total_scheme_flat_incentive_users_pages']            = total_scheme_flat_incentive_users_pages
    else:
        scheme_flat_incentive_users = {}
        context['total_scheme_flat_incentive_users_pages']            = 0


   
    context['free_schemes']                     = free_schemes
    context['flat_incentive_schemes']           = flat_incentive_schemes
    context['bulkpack_incentive_schemes']       = bulkpack_incentive_schemes
    context['quantitative_schemes']             = quantitative_schemes
    context['total_free_pages']                 = total_free_pages
    context['total_flat_incentive_pages']       = total_flat_incentive_pages
    context['page_limit']                       = getConfigurationResult('page_limit')
    context['scheme_free_users']                = scheme_free_users
    context['scheme_flat_incentive_users']      = scheme_flat_incentive_users
    context['page_title']                       = "Scheme Management"
    template                                    = 'schemes/schemes.html'
    return render(request, template, context)


@login_required
def latestFlatIncentiveUsers(request):
    context = {}
    latest_scheme = SpFlatSchemes.objects.raw(''' SELECT * FROM sp_flat_schemes order by id desc LIMIT 1  ''')

    if latest_scheme :
        latest_scheme = latest_scheme[0]
        
        users = SpUserFlatSchemes.objects.raw(''' SELECT sp_user_flat_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_flat_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_flat_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_flat_schemes.user_id
        WHERE sp_user_flat_schemes.scheme_id = %s
        order by sp_user_flat_schemes.id  ''',[latest_scheme.id])

        if users :
            context['users'] = users
        else : 
            context['users'] = []

    else : 
        context['users'] = []
    

    template = 'schemes/flat-incentive-users.html'
    return render(request, template, context)



@login_required
def freeSchemeUsers(request,scheme_id):
    context = {}
    scheme_free_users = list(SpUserSchemes.objects.raw(''' SELECT sp_user_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
    sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_schemes
    LEFT JOIN sp_users on sp_users.id = sp_user_schemes.user_id
    LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_schemes.user_id
    WHERE sp_user_schemes.scheme_id = %s
    order by sp_user_schemes.id  ''',[scheme_id]))

    free_page = request.GET.get('free_page')
    free_user_paginator = Paginator(scheme_free_users, 50)
    try:
        scheme_free_users = free_user_paginator.page(free_page)
    except PageNotAnInteger:
        scheme_free_users = free_user_paginator.page(1)
    except EmptyPage:
        scheme_free_users = free_user_paginator.page(free_user_paginator.num_pages)  
    if free_page is not None:
        free_page = free_page
    else:
        free_page = 1
    total_scheme_free_users_pages = int(free_user_paginator.count/getConfigurationResult('page_limit')) 
    
    if(free_user_paginator.count == 0):
        free_user_paginator.count = 1

    temp = total_scheme_free_users_pages%free_user_paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= free_user_paginator.count):
        total_scheme_free_users_pages = total_scheme_free_users_pages+1
    else:
        total_scheme_free_users_pages = total_scheme_free_users_pages
    context['total_scheme_free_users_pages']            = total_scheme_free_users_pages
    
    context['users'] = scheme_free_users

    template = 'schemes/free-scheme-users.html'
    return render(request, template, context)

@login_required
def userBulkpackIncentiveBifurcations(request,user_id, scheme_id):
    context = {}
    context['bifurcations'] = SpUserBulkpackSchemeBifurcation.objects.filter(user_id=user_id,bulkpack_scheme_id=scheme_id)
    template = 'schemes/user-bulkpack-incentive-bifurcations.html'
    return render(request, template, context)

@login_required
def userQuantitativeSchemeBifurcations(request,user_id, scheme_id):
    context = {}
    context['bifurcations'] = SpUserQuantitativeSchemeBifurcation.objects.filter(user_id=user_id,user_scheme_id=scheme_id)
    template = 'schemes/user-quantitative-scheme-bifurcations.html'
    return render(request, template, context)



@login_required
def flatIncentiveUsers(request,scheme_id):
    context = {}
    users = SpUserFlatSchemes.objects.raw(''' SELECT sp_user_flat_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_flat_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_flat_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_flat_schemes.user_id
        WHERE sp_user_flat_schemes.scheme_id = %s
        order by sp_user_flat_schemes.id  ''',[scheme_id])

    if users :
        context['users'] = users
    else : 
        context['users'] = []
    

    template = 'schemes/flat-incentive-users.html'
    return render(request, template, context)

@login_required
def latestBulkpackIncentiveUsers(request):
    context = {}
    latest_scheme = SpBulkpackSchemes.objects.raw(''' SELECT * FROM sp_bulkpack_schemes order by id desc LIMIT 1  ''')

    if latest_scheme :
        latest_scheme = latest_scheme[0]
        
        users = SpUserBulkpackSchemes.objects.raw(''' SELECT sp_user_bulkpack_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_bulkpack_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_bulkpack_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_bulkpack_schemes.user_id
        WHERE sp_user_bulkpack_schemes.scheme_id = %s
        order by sp_user_bulkpack_schemes.id  ''',[latest_scheme.id])

        if users :
            context['users'] = users
        else : 
            context['users'] = []

    else : 
        context['users'] = []
    template = 'schemes/bulkpack-incentive-users.html'
    return render(request, template, context)

@login_required
def bulkpackIncentiveUsers(request,scheme_id):
    context = {}
    users = SpUserBulkpackSchemes.objects.raw(''' SELECT sp_user_bulkpack_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_bulkpack_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_bulkpack_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_bulkpack_schemes.user_id
        WHERE sp_user_bulkpack_schemes.scheme_id = %s
        order by sp_user_bulkpack_schemes.id  ''',[scheme_id])

    if users :
        context['users'] = users
    else : 
        context['users'] = []

    template = 'schemes/bulkpack-incentive-users.html'
    return render(request, template, context)

@login_required
def latestQuantitativeIncentiveUsers(request):
    context = {}
    latest_scheme = SpBulkpackSchemes.objects.raw(''' SELECT * FROM sp_schemes WHERE scheme_type = 2 order by id desc LIMIT 1  ''')

    if latest_scheme :
        latest_scheme = latest_scheme[0]
        
        users = SpUserSchemes.objects.raw(''' SELECT sp_user_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_schemes.user_id
        WHERE sp_user_schemes.scheme_type = 2 AND sp_user_schemes.scheme_id = %s
        order by sp_user_schemes.id  ''',[latest_scheme.id])

        if users :
            context['users'] = users
        else : 
            context['users'] = []

    else : 
        context['users'] = []

    template = 'schemes/bonus-scheme-users.html'
    return render(request, template, context)

@login_required
def quantitativeIncentiveUsers(request,scheme_id):
    context = {}
    users = SpUserSchemes.objects.raw(''' SELECT sp_user_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,sp_users.store_name,sp_users.emp_sap_id,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_schemes.user_id
        WHERE sp_user_schemes.scheme_type = 2 AND sp_user_schemes.scheme_id = %s
        order by sp_user_schemes.id  ''',[scheme_id])

    if users :
        context['users'] = users
    else : 
        context['users'] = []

    template = 'schemes/bonus-scheme-users.html'
    return render(request, template, context)

@login_required
def getSchemeUsers(request,scheme_id):
    if SpProducts.objects.filter(id=product_id).exists() :
        product_variants = SpProductVariants.objects.filter(product_id=product_id).order_by('-id')
        page = request.GET.get('page')
        variant_paginator = Paginator(product_variants, getConfigurationResult('page_limit'))
        try:
            product_variants = variant_paginator.page(page)
        except PageNotAnInteger:
            product_variants = variant_paginator.page(1)
        except EmptyPage:
            product_variants = variant_paginator.page(variant_paginator.num_pages)  
        if page is not None:
            page = page
        else:
            page = 1
        total_variant_pages = int(variant_paginator.count/getConfigurationResult('page_limit')) 
        
        if(variant_paginator.count == 0):
            variant_paginator.count = 1

        temp = total_variant_pages%variant_paginator.count
        if(temp > 0 and getConfigurationResult('page_limit')!= variant_paginator.count):
            total_variant_pages = total_variant_pages+1
        else:
            total_variant_pages = total_variant_pages
        
        context = {}
        context['page_limit']             = getConfigurationResult('page_limit')
        context['total_variant_pages']     = total_variant_pages
        context['product_variants']   = product_variants
        template = 'schemes/product-variants.html'
        return render(request, template, context)
        
    else:
        return HttpResponse('Not found')




@login_required
def ajaxSchemeList(request):
    page = request.GET.get('page')
    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))   
    
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpOrganizations.objects.order_by('-id').first()
    
    context = {}
    context['organizations']          = organizations
    context['total_pages']            = total_pages
    context['organization_details']   = organization_details

    template = 'role-permission/ajax-schemes.html'
    return render(request, template, context)


@login_required
def ajaxSchemeUserLists(request,product_id):
    if SpProducts.objects.filter(id=product_id).exists() :

        product_variants = SpProductVariants.objects.filter(product_id=product_id).order_by('-id')
        page = request.GET.get('page')
        variant_paginator = Paginator(product_variants, getConfigurationResult('page_limit'))
        try:
            product_variants = variant_paginator.page(page)
        except PageNotAnInteger:
            product_variants = variant_paginator.page(1)
        except EmptyPage:
            product_variants = variant_paginator.page(variant_paginator.num_pages)  
        if page is not None:
            page = page
        else:
            page = 1
        total_variant_pages = int(variant_paginator.count/getConfigurationResult('page_limit')) 
        
        if(variant_paginator.count == 0):
            variant_paginator.count = 1

        temp = total_variant_pages%variant_paginator.count
        if(temp > 0 and getConfigurationResult('page_limit')!= variant_paginator.count):
            total_variant_pages = total_variant_pages+1
        else:
            total_variant_pages = total_variant_pages
        
        context = {}
        context['page_limit']             = getConfigurationResult('page_limit')
        context['total_variant_pages']     = total_variant_pages
        context['product_variants']   = product_variants
        template = 'schemes/ajax-product-variant-lists.html'
        return render(request, template, context)

    else:
        return HttpResponse('Not found')

    





@login_required
def addFreeScheme(request):
    if request.method == "POST":
        response = {}
        separator = ','
        
        route_id = separator.join(request.POST.getlist('route_id[]'))
        zone_id = separator.join(request.POST.getlist('zone_id[]'))
        town_id = separator.join(request.POST.getlist('town_id[]'))
        is_route = request.POST['is_route']
        user_id       = request.POST.getlist('user_id[]') 
        
        error = []
        for id, val in enumerate(user_id):
            if SpUserSchemes.objects.filter(applied_on_variant_id=request.POST['applied_on_variant_id'], user_id=user_id[id]).exists():
                error.append('1')
        if len(error) > 0:
            response['flag'] = False
            response['message'] = "Some User variant already mapped in different scheme, please check and try again."
        else:
            product_id = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'product_id')
            product = SpProducts.objects.get(id=product_id)
            scheme = SpSchemes()
            scheme.name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
            scheme.state_id = request.POST['state_id']
            scheme.is_route = request.POST['is_route']
            if is_route == '0':
                if len(request.POST.getlist('route_id[]'))>0:
                    scheme.route_id = route_id
            else:        
                scheme.zone_id = zone_id
                scheme.town_id = town_id
            scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['end_date'] != '' :
                scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            
            scheme.scheme_type = 1
            scheme.applied_on_variant_id = request.POST['applied_on_variant_id']
            scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
            scheme.minimum_order_quantity = request.POST['minimum_order_quantity']
            scheme.packaging_type = request.POST['packaging_type']
            scheme.order_container_id = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'container_id')
            scheme.order_container_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'container_name')
            scheme.order_packaging_id = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'packaging_type_id')
            scheme.order_packaging_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'packaging_type_name')
            scheme.free_variant_id = request.POST['free_variant_id']
            scheme.free_variant_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'variant_name')
            
            if request.POST['container_qty'] != "":
                scheme.container_quantity = request.POST['container_qty']
            else:
                scheme.container_quantity = 0

            if request.POST['pouch_qty'] != "":
                scheme.pouch_quantity = request.POST['pouch_qty']
            else:
                scheme.pouch_quantity = 0
                

            scheme.status = 1
            scheme.save()
            if scheme.id : 
                user_free_variants       = request.POST.getlist('user_free_variant_id[]')
                user_id       = request.POST.getlist('user_id[]') 
                user_minimum_order_quantity       = request.POST.getlist('user_minimum_order_quantity[]') 
                user_container_qty       = request.POST.getlist('user_container_qty[]') 
                user_pouch_qty       = request.POST.getlist('user_pouch_qty[]') 
                

                for id, val in enumerate(user_free_variants):
                    user_scheme =  SpUserSchemes()
                    user_scheme.user_id = user_id[id]
                    user_scheme.scheme_id = scheme.id
                    user_scheme.scheme_name = scheme.name
                    user_scheme.is_route = request.POST['is_route']
                    user_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                    if is_route == '0':
                        user_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                    else:    
                        user_scheme.zone_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'zone_id')
                        user_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                    user_scheme.scheme_start_date = scheme.scheme_start_date
                    if scheme.scheme_end_date is not None :
                        user_scheme.scheme_end_date = scheme.scheme_end_date
                    
                    user_scheme.scheme_type = 1
                    user_scheme.applied_on_variant_id = request.POST['applied_on_variant_id']
                    user_scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
                    user_scheme.minimum_order_quantity = user_minimum_order_quantity[id]
                    user_scheme.packaging_type = request.POST['packaging_type']
                    user_scheme.order_container_id = getModelColumnById(SpProductVariants,user_free_variants[id],'container_id')
                    user_scheme.order_container_name = getModelColumnById(SpProductVariants,user_free_variants[id],'container_name')
                    user_scheme.order_packaging_id = getModelColumnById(SpProductVariants,user_free_variants[id],'packaging_type_id')
                    user_scheme.order_packaging_name = getModelColumnById(SpProductVariants,user_free_variants[id],'packaging_type_name')
                    user_scheme.free_variant_id = user_free_variants[id]
                    user_scheme.free_variant_name = getModelColumnById(SpProductVariants,user_free_variants[id],'variant_name')
                    user_scheme.container_quantity = user_container_qty[id]
                    user_scheme.pouch_quantity = user_pouch_qty[id]
                    user_scheme.status = 1
                    user_scheme.save()

                #Save Activity
                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'Free scheme created'
                activity    = 'Free scheme created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Product & Variant Management', 'Free Scheme', heading, activity, request.user.id, user_name, 'productScheme.png', '1', 'web.png')

                response['flag'] = True
                response['message'] = "Record has been saved successfully."

        return JsonResponse(response)
    else:
        today  = date.today()
        template = 'schemes/add-free-scheme.html'
        context = {}
        context['states'] = SpStates.objects.all()
        # context['zones'] = SpZones.objects.all()
        context['routes'] = SpRoutes.objects.all()
        context['today_date'] = today.strftime("%d/%m/%Y")
        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products

        return render(request, template,context)


@login_required
def editFreeScheme(request,scheme_id):
    if request.method == "POST":
        response = {}
        separator = ','
       
        scheme_id = request.POST['scheme_id']
        route_id = separator.join(request.POST.getlist('route_id[]'))
        zone_id = separator.join(request.POST.getlist('zone_id[]'))
        town_id = separator.join(request.POST.getlist('town_id[]'))
        is_route = request.POST['is_route']
        user_id = request.POST.getlist('user_id[]')
        
        error = []
        for id, val in enumerate(user_id):
            if SpUserSchemes.objects.filter(applied_on_variant_id=request.POST['applied_on_variant_id'], user_id=user_id[id]).exclude(scheme_id=request.POST['scheme_id']).exists():
                error.append('1')
        if len(error) > 0:
            response['flag'] = False
            response['message'] = "Some User variant already mapped in different scheme, please check and try again."
        else:
            product_id = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'product_id')
            product = SpProducts.objects.get(id=product_id)
            scheme = SpSchemes.objects.get(id=request.POST['scheme_id'])
            scheme.name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
            scheme.state_id = request.POST['state_id']
            
            scheme.is_route = request.POST['is_route']
            if is_route == '0':
                if len(request.POST.getlist('route_id[]'))>0:
                    scheme.route_id = route_id
            else:        
                scheme.zone_id = zone_id
                scheme.town_id = town_id
            scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['end_date'] != '' :
                scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            
            scheme.scheme_type = 1
            scheme.applied_on_variant_id = request.POST['applied_on_variant_id']
            scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
            scheme.minimum_order_quantity = request.POST['minimum_order_quantity']
            scheme.packaging_type = request.POST['packaging_type']
            scheme.order_container_id = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'container_id')
            scheme.order_container_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'container_name')
            scheme.order_packaging_id = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'packaging_type_id')
            scheme.order_packaging_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'packaging_type_name')
            scheme.free_variant_id = request.POST['free_variant_id']
            scheme.free_variant_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'variant_name')
            
            if request.POST['container_qty'] != "":
                scheme.container_quantity = request.POST['container_qty']
            else:
                scheme.container_quantity = 0

            if request.POST['pouch_qty'] != "":
                scheme.pouch_quantity = request.POST['pouch_qty']
            else:
                scheme.pouch_quantity = 0
                

            scheme.status = 1
            scheme.save()
            if scheme.id : 
                SpUserSchemes.objects.filter(scheme_id=request.POST['scheme_id']).delete()

                user_free_variants       = request.POST.getlist('user_free_variant_id[]')
                user_id       = request.POST.getlist('user_id[]') 
                user_minimum_order_quantity       = request.POST.getlist('user_minimum_order_quantity[]') 
                user_container_qty       = request.POST.getlist('user_container_qty[]') 
                user_pouch_qty       = request.POST.getlist('user_pouch_qty[]') 
                

                for id, val in enumerate(user_free_variants):
                    user_scheme =  SpUserSchemes()
                    user_scheme.user_id = user_id[id]
                    user_scheme.scheme_id = scheme.id
                    user_scheme.scheme_name = scheme.name
                    user_scheme.is_route = request.POST['is_route']
                    user_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                    if is_route == '0':
                        user_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                    else:    
                        user_scheme.zone_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'zone_id')
                        user_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                    user_scheme.scheme_start_date = scheme.scheme_start_date
                    if scheme.scheme_end_date is not None :
                        user_scheme.scheme_end_date = scheme.scheme_end_date
                    
                    user_scheme.scheme_type = 1
                    user_scheme.applied_on_variant_id = request.POST['applied_on_variant_id']
                    user_scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
                    user_scheme.minimum_order_quantity = user_minimum_order_quantity[id]
                    user_scheme.packaging_type = request.POST['packaging_type']
                    user_scheme.order_container_id = getModelColumnById(SpProductVariants,user_free_variants[id],'container_id')
                    user_scheme.order_container_name = getModelColumnById(SpProductVariants,user_free_variants[id],'container_name')
                    user_scheme.order_packaging_id = getModelColumnById(SpProductVariants,user_free_variants[id],'packaging_type_id')
                    user_scheme.order_packaging_name = getModelColumnById(SpProductVariants,user_free_variants[id],'packaging_type_name')
                    user_scheme.free_variant_id = user_free_variants[id]
                    user_scheme.free_variant_name = getModelColumnById(SpProductVariants,user_free_variants[id],'variant_name')
                    user_scheme.container_quantity = user_container_qty[id]
                    user_scheme.pouch_quantity = user_pouch_qty[id]
                    user_scheme.status = 1
                    user_scheme.save()

                #Save Activity
                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'Flat scheme updated'
                activity    = 'Flat scheme updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Product & Variant Management', 'Flat Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

                response['flag'] = True
                response['message'] = "Record has been updated successfully."

        return JsonResponse(response)
    else:

        template = 'schemes/edit-free-scheme.html'
        context = {}
        context['scheme'] = scheme = SpSchemes.objects.get(id=scheme_id)
        context['users'] = SpUserSchemes.objects.raw(''' SELECT sp_user_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_schemes.user_id
        WHERE sp_user_schemes.scheme_id = %s
        order by sp_user_schemes.id  ''',[scheme_id])

        context['states'] = SpStates.objects.all()

        if scheme.is_route == 1:
            zone_id = scheme.zone_id
            zone_id = zone_id.split(',')
        else:
            zone_id = []

        zone_lists = SpZones.objects.filter(state_id=scheme.state_id)
        zone_list = []
        for zone in zone_lists:
            zones_dict = {}
            zones_dict['id'] =  zone.id
            zones_dict['zone'] =  zone.zone
            if str(zone.id) in zone_id:
                zones_dict['zone_value'] =  'selected' 
            zone_list.append(zones_dict)
        context['zones']  = zone_list


        routes_list = []
        if scheme.is_route == 0:
            route_id = scheme.route_id
            route_id = route_id.split(',')
        else:
            route_id = []
        routes = SpRoutes.objects.filter(state_id=scheme.state_id)
        for route in routes:
            routes_dict = {}
            routes_dict['id'] =  route.id
            routes_dict['route'] =  route.route
            if str(route.id) in route_id:
                routes_dict['route_value'] =  'selected'
            routes_list.append(routes_dict)  

        context['routes_list'] = routes_list

        town_list = []
        if scheme.is_route == 1:
            town_id = scheme.town_id
            town_id = town_id.split(',')
            
            towns  = SpTowns.objects.filter(zone_id__in=zone_id)
            for town in towns:
                towns_dict = {}
                towns_dict['id'] =  town.id
                towns_dict['town'] =  town.town
                if str(town.id) in town_id:
                    towns_dict['town_value'] =  'selected' 
                town_list.append(towns_dict)

        context['towns'] = town_list

        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products

        applied_product_variant = SpProductVariants.objects.get(id=scheme.applied_on_variant_id)
        context['container_name'] = applied_product_variant.container_name
        context['packaging_type_name'] = applied_product_variant.packaging_type_name
        return render(request, template,context)


@login_required
def addBonusScheme(request):
    if request.method == "POST":
        response = {}
        if SpSchemes.objects.filter(name=request.POST['scheme_name']).exists() :
            response['flag'] = False
            response['message'] = "Scheme name already exists."
        else:
            scheme = SpSchemes()
            scheme.name = request.POST['scheme_name']
            scheme.state_id = request.POST['state_id']
            separator = ','
            scheme.route_id = separator.join(request.POST.getlist('route_id[]'))
            scheme.town_id = separator.join(request.POST.getlist('town_id[]'))
            scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['end_date'] != '' :
                scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            
            scheme.scheme_type = 2
            scheme.applied_on_variant_id = None
            scheme.applied_on_variant_name = None
            scheme.minimum_order_quantity = None
            scheme.order_container_id = None
            scheme.order_container_name = None
            scheme.free_variant_id = request.POST['free_variant_id']
            scheme.free_variant_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'variant_name')
            
            if request.POST['container_qty'] != "":
                scheme.container_quantity = request.POST['container_qty']
            else:
                scheme.container_quantity = 0

            if request.POST['pouch_qty'] != "":
                scheme.pouch_quantity = request.POST['pouch_qty']
            else:
                scheme.pouch_quantity = 0

            scheme.status = 1
            scheme.save()
            if scheme.id : 

                minimum_order_quantities       = request.POST.getlist('minimum_order_quantity[]')
                order_container_ids       = request.POST.getlist('order_container_id[]') 

                for id, val in enumerate(minimum_order_quantities):
                    scheme_bifurcation =  SpQuantitativeSchemeBifurcation()
                    scheme_bifurcation.scheme_id = scheme.id
                    scheme_bifurcation.minimum_order_quantity = minimum_order_quantities[id]
                    scheme_bifurcation.order_container_id = order_container_ids[id]
                    scheme_bifurcation.order_container_name = getModelColumnById(SpContainers,order_container_ids[id],'container')
                    scheme_bifurcation.save()


                user_free_variants       = request.POST.getlist('user_free_variant_id[]')
                user_id       = request.POST.getlist('user_id[]') 
                user_minimum_order_quantity       = request.POST.getlist('user_minimum_order_quantity[]') 
                user_container_qty       = request.POST.getlist('user_container_qty[]') 
                user_pouch_qty       = request.POST.getlist('user_pouch_qty[]') 
                

                for id, val in enumerate(user_free_variants):
                    user_scheme =  SpUserSchemes()
                    user_scheme.user_id = user_id[id]
                    user_scheme.scheme_id = scheme.id
                    user_scheme.scheme_name = scheme.name
                    user_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                    user_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                    user_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                    user_scheme.scheme_start_date = scheme.scheme_start_date
                    if scheme.scheme_end_date is not None :
                        user_scheme.scheme_end_date = scheme.scheme_end_date
                    
                    user_scheme.scheme_type = 2
                    user_scheme.applied_on_variant_id = None
                    user_scheme.applied_on_variant_name = None
                    user_scheme.minimum_order_quantity = None
                    user_scheme.order_container_id = None
                    user_scheme.order_container_name = None
                    user_scheme.free_variant_id = user_free_variants[id]
                    user_scheme.free_variant_name = getModelColumnById(SpProductVariants,user_free_variants[id],'variant_name')
                    user_scheme.container_quantity = user_container_qty[id]
                    user_scheme.pouch_quantity = user_pouch_qty[id]
                    user_scheme.status = 1
                    user_scheme.save()

                    if user_scheme.id :
                        container_var = "user_order_container_id_"+str(user_id[id])
                        qty_var = "user_minimim_order_quantity_"+str(user_id[id])
                        user_order_containers       = request.POST.getlist(container_var)
                        user_minimum_order_quantity       = request.POST.getlist(qty_var)
                        total_minimum_order_quantity = 0
                        for next_id, val in enumerate(user_order_containers):
                            bifurcation = SpUserQuantitativeSchemeBifurcation()
                            bifurcation.user_id = user_id[id]
                            bifurcation.user_scheme_id = scheme.id
                            bifurcation.minimum_order_quantity = user_minimum_order_quantity[next_id]
                            bifurcation.order_container_id = user_order_containers[next_id]
                            bifurcation.order_container_name = getModelColumnById(SpContainers,user_order_containers[next_id],'container')
                            bifurcation.save()
                            total_minimum_order_quantity = int(total_minimum_order_quantity) + int(user_minimum_order_quantity[next_id])
                        
                        user_scheme.minimum_order_quantity = total_minimum_order_quantity
                        user_scheme.save()

                
                #Save Activity
                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'Quantitative scheme created'
                activity    = 'Quantitative scheme created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Product & Variant Management', 'Flat Scheme', heading, activity, request.user.id, user_name, 'productScheme.png', '1', 'web.png')

                response['flag'] = True
                response['message'] = "Record has been saved successfully."

        return JsonResponse(response)
        
    else:
        
        template = 'schemes/add-bonus-scheme.html'
        context = {}
        context['states'] = SpStates.objects.all()
        context['containers']          = SpContainers.objects.all()

        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products
       
        return render(request, template,context)



@login_required
def editBonusScheme(request,scheme_id):
    if request.method == "POST":
        response = {}
        if SpSchemes.objects.filter(name=request.POST['scheme_name']).exclude(id=scheme_id).exists() :
            response['flag'] = False
            response['message'] = "Scheme name already exists."
        else:
            scheme = SpSchemes.objects.get(id=request.POST['scheme_id'])
            scheme.name = request.POST['scheme_name']
            scheme.state_id = request.POST['state_id']
            separator = ','
            scheme.route_id = separator.join(request.POST.getlist('route_id[]'))
            scheme.town_id = separator.join(request.POST.getlist('town_id[]'))
            scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['end_date'] != '' :
                scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            
            scheme.scheme_type = 2
            scheme.applied_on_variant_id = None
            scheme.applied_on_variant_name = None
            scheme.minimum_order_quantity = None
            scheme.order_container_id = None
            scheme.order_container_name = None
            scheme.free_variant_id = request.POST['free_variant_id']
            scheme.free_variant_name = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'variant_name')
            
            if request.POST['container_qty'] != "":
                scheme.container_quantity = request.POST['container_qty']
            else:
                scheme.container_quantity = 0

            if request.POST['pouch_qty'] != "":
                scheme.pouch_quantity = request.POST['pouch_qty']
            else:
                scheme.pouch_quantity = 0

            scheme.status = 1
            scheme.save()
            if scheme.id : 
                SpQuantitativeSchemeBifurcation.objects.filter(scheme_id=request.POST['scheme_id']).delete()

                minimum_order_quantities       = request.POST.getlist('minimum_order_quantity[]')
                order_container_ids       = request.POST.getlist('order_container_id[]') 

                for id, val in enumerate(minimum_order_quantities):
                    scheme_bifurcation =  SpQuantitativeSchemeBifurcation()
                    scheme_bifurcation.scheme_id = scheme.id
                    scheme_bifurcation.minimum_order_quantity = minimum_order_quantities[id]
                    scheme_bifurcation.order_container_id = order_container_ids[id]
                    scheme_bifurcation.order_container_name = getModelColumnById(SpContainers,order_container_ids[id],'container')
                    scheme_bifurcation.save()


                SpUserSchemes.objects.filter(scheme_id=request.POST['scheme_id']).delete()
                SpUserQuantitativeSchemeBifurcation.objects.filter(user_scheme_id=request.POST['scheme_id']).delete()
                

                

                user_free_variants       = request.POST.getlist('user_free_variant_id[]')
                user_id       = request.POST.getlist('user_id[]') 
                user_minimum_order_quantity       = request.POST.getlist('user_minimum_order_quantity[]') 
                user_container_qty       = request.POST.getlist('user_container_qty[]') 
                user_pouch_qty       = request.POST.getlist('user_pouch_qty[]') 
                

                for id, val in enumerate(user_free_variants):
                    user_scheme =  SpUserSchemes()
                    user_scheme.user_id = user_id[id]
                    user_scheme.scheme_id = scheme.id
                    user_scheme.scheme_name = scheme.name
                    user_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                    user_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                    user_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                    user_scheme.scheme_start_date = scheme.scheme_start_date
                    if scheme.scheme_end_date is not None :
                        user_scheme.scheme_end_date = scheme.scheme_end_date
                    
                    user_scheme.scheme_type = 2
                    user_scheme.applied_on_variant_id = None
                    user_scheme.applied_on_variant_name = None
                    user_scheme.minimum_order_quantity = None
                    user_scheme.order_container_id = None
                    user_scheme.order_container_name = None
                    user_scheme.free_variant_id = user_free_variants[id]
                    user_scheme.free_variant_name = getModelColumnById(SpProductVariants,user_free_variants[id],'variant_name')
                    user_scheme.container_quantity = user_container_qty[id]
                    user_scheme.pouch_quantity = user_pouch_qty[id]
                    user_scheme.status = 1
                    user_scheme.save()

                    if user_scheme.id :
                        container_var = "user_order_container_id_"+str(user_id[id])
                        qty_var = "user_minimim_order_quantity_"+str(user_id[id])
                        user_order_containers       = request.POST.getlist(container_var)
                        user_minimum_order_quantity       = request.POST.getlist(qty_var)
                        total_minimum_order_quantity = 0
                        for next_id, val in enumerate(user_order_containers):
                            bifurcation = SpUserQuantitativeSchemeBifurcation()
                            bifurcation.user_id = user_id[id]
                            bifurcation.user_scheme_id = scheme.id
                            bifurcation.minimum_order_quantity = user_minimum_order_quantity[next_id]
                            bifurcation.order_container_id = user_order_containers[next_id]
                            bifurcation.order_container_name = getModelColumnById(SpContainers,user_order_containers[next_id],'container')
                            bifurcation.save()
                            total_minimum_order_quantity = int(total_minimum_order_quantity) + int(user_minimum_order_quantity[next_id])
                        
                        user_scheme.minimum_order_quantity = total_minimum_order_quantity
                        user_scheme.save()

                #Save Activity
                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'Quantitative scheme updated'
                activity    = 'Quantitative scheme updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Product & Variant Management', 'Flat Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

                response['flag'] = True
                response['message'] = "Record has been saved successfully."

        return JsonResponse(response)
        
    else:
        template = 'schemes/edit-bonus-scheme.html'
        context = {}
        context['states'] = SpStates.objects.all()
        context['containers']          = SpContainers.objects.all()

        context['scheme'] = scheme = SpSchemes.objects.get(id=scheme_id)
        context['scheme_bifurcations'] = SpQuantitativeSchemeBifurcation.objects.filter(scheme_id=scheme_id)
        
        users = SpUserSchemes.objects.raw(''' SELECT sp_user_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_schemes.user_id
        WHERE sp_user_schemes.scheme_id = %s
        order by sp_user_schemes.id  ''',[scheme_id])

        for user in users:
            user.scheme_bifurcations = SpUserQuantitativeSchemeBifurcation.objects.filter(user_id=user.user_id,user_scheme_id=scheme_id)

        context['users'] = users
        context['states'] = SpStates.objects.all()
        if scheme.route_id is not None:
            context['routes'] = SpRoutes.objects.raw(''' SELECT * FROM sp_routes WHERE id in (%s) ''',[scheme.route_id])

        if scheme.town_id is not None:
            context['towns'] = SpTowns.objects.raw(''' SELECT * FROM sp_towns WHERE id in (%s) ''',[scheme.town_id])


        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products
       
        return render(request, template,context)

@login_required
def addFlatIncentiveScheme(request):
    if request.method == "POST":
        today  = date.today()
        response = {}
        separator = ','
        
        route_id = separator.join(request.POST.getlist('route_id[]'))
        zone_id = separator.join(request.POST.getlist('zone_id[]'))
        town_id = separator.join(request.POST.getlist('town_id[]'))
        is_route = request.POST['is_route']
        if SpFlatSchemes.objects.filter(name=request.POST['scheme_name']).exists() :
            response['flag'] = False
            response['message'] = "Scheme name already exists."
        else:
            user_incentive_amount       = request.POST.getlist('user_incentive_amount[]')
            user_applied_on_variant_id  = request.POST.getlist('user_applied_on_variant_id[]')
            user_id       = request.POST.getlist('user_id[]') 
            is_disabled       = request.POST.getlist('is_disabled[]')
            
            error = []
            for id, val in enumerate(user_id):
                if SpUserFlatSchemes.objects.filter(applied_on_variant_id=user_applied_on_variant_id[id], user_id=user_id[id]).exists() and float(user_incentive_amount[id]) > 0:
                    error.append('1')
            if len(error) == 0:
                
                scheme = SpFlatSchemes()
                scheme.name = request.POST['scheme_name']
                scheme.state_id = request.POST['state_id']
                scheme.is_route = request.POST['is_route']
                if is_route == '0':
                    if len(request.POST.getlist('route_id[]'))>0:
                        scheme.route_id = route_id
                else:        
                    scheme.zone_id = zone_id
                    scheme.town_id = town_id
                scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['end_date'] != '' :
                    scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
               
                scheme.incentive_amount = request.POST['incentive_amount']
                scheme.applied_on_variant_id = request.POST['applied_on_variant_id']
                scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
                # scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))
    
                scheme.status = 1
                scheme.save()
                if scheme.id :  
    
                    for id, val in enumerate(user_id):
                        #if not int(is_disabled[id]) :
                        user_flat_scheme =  SpUserFlatSchemes()
                        user_flat_scheme.user_id = user_id[id]
                        user_flat_scheme.scheme_id = scheme.id
                        user_flat_scheme.scheme_name = scheme.name
                        user_flat_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                        if is_route == '0':
                            user_flat_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                        else:    
                            user_flat_scheme.zone_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'zone_id')
                            user_flat_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                        user_flat_scheme.scheme_start_date = scheme.scheme_start_date
                        if scheme.scheme_end_date is not None :
                            user_flat_scheme.scheme_end_date = scheme.scheme_end_date
                        
                        user_flat_scheme.incentive_amount = user_incentive_amount[id]
                        user_flat_scheme.applied_on_variant_id = user_applied_on_variant_id[id]
                        user_flat_scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,user_applied_on_variant_id[id],'variant_name')
                        #user_flat_scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))
                        user_flat_scheme.status = 1
                        user_flat_scheme.save()
                
                    #Save Activity
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = 'Flat scheme created'
                    activity    = 'Flat scheme created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    saveActivity('Product & Variant Management', 'Flat Scheme', heading, activity, request.user.id, user_name, 'productScheme.png', '1', 'web.png')
                    
                    response['flag'] = True
                    response['message'] = "Record has been saved successfully."
            else:
                response['flag'] = False
                response['message'] = "Some User variant already mapped in different flat scheme, please check and try again."

        return JsonResponse(response)
        
    else:
        template = 'schemes/add-flat-incentive-scheme.html'
        context = {}
        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products
        context['routes'] = SpRoutes.objects.all()
        context['states'] = SpStates.objects.all()
        context['units']          = SpProductUnits.objects.raw(''' SELECT id,largest_unit as unit FROM sp_product_units WHERE id in 
                        (SELECT DISTINCT variant_unit_id FROM sp_product_variants) GROUP BY largest_unit '''
                        )

        return render(request, template,context)


@login_required
def editFlatIncentiveScheme(request,scheme_id):
    if request.method == "POST":
        today  = date.today()
        response = {}
        separator = ','
        route_id = separator.join(request.POST.getlist('route_id[]'))
        zone_id = separator.join(request.POST.getlist('zone_id[]'))
        town_id = separator.join(request.POST.getlist('town_id[]'))
        is_route = request.POST['is_route']
        if SpFlatSchemes.objects.filter(name=request.POST['scheme_name']).exclude(id=request.POST['scheme_id']).exists() :
            response['flag'] = False
            response['message'] = "Scheme name already exists."
        else:
            user_incentive_amount       = request.POST.getlist('user_incentive_amount[]')
            user_applied_on_variant_id  = request.POST.getlist('user_applied_on_variant_id[]')
            user_id       = request.POST.getlist('user_id[]') 
            is_disabled       = request.POST.getlist('is_disabled[]')
            
            error = []
            for id, val in enumerate(user_id):
                if SpUserFlatSchemes.objects.filter(applied_on_variant_id=user_applied_on_variant_id[id], user_id=user_id[id]).exclude(scheme_id=request.POST['scheme_id']).exists() and float(user_incentive_amount[id]) > 0:
                    if SpUserFlatSchemes.objects.filter(applied_on_variant_id=user_applied_on_variant_id[id], user_id=user_id[id], incentive_amount__gt=0).exclude(scheme_id=request.POST['scheme_id']).exists():
                        error.append('1')
            if len(error) == 0:        
                scheme = SpFlatSchemes.objects.get(id=request.POST['scheme_id'])
                scheme.name = request.POST['scheme_name']
                scheme.state_id = request.POST['state_id']
                scheme.is_route = request.POST['is_route']
                if is_route == '0':
                    if len(request.POST.getlist('route_id[]'))>0:
                        scheme.route_id = route_id
                else:        
                    scheme.zone_id = zone_id
                    scheme.town_id = town_id
                scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['end_date'] != '' :
                    scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
               
                scheme.incentive_amount = request.POST['incentive_amount']
                scheme.applied_on_variant_id = request.POST['applied_on_variant_id']
                scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
                #scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))
    
                scheme.status = 1
                scheme.save()
                if scheme.id : 
                    SpUserFlatSchemes.objects.filter(scheme_id=request.POST['scheme_id']).delete() 
    
                    for id, val in enumerate(user_id):
                        if not SpUserFlatSchemes.objects.filter(scheme_id=request.POST['scheme_id'],user_id=user_id[id]).exists():
                            user_flat_scheme =  SpUserFlatSchemes()
                            user_flat_scheme.user_id = user_id[id]
                            user_flat_scheme.scheme_id = scheme.id
                            user_flat_scheme.scheme_name = scheme.name
                            user_flat_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                            if is_route == '0':
                                user_flat_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                            else:    
                                user_flat_scheme.zone_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'zone_id')
                                user_flat_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                            user_flat_scheme.scheme_start_date = scheme.scheme_start_date
                            if scheme.scheme_end_date is not None :
                                user_flat_scheme.scheme_end_date = scheme.scheme_end_date
                            
                            user_flat_scheme.incentive_amount = user_incentive_amount[id]
                            user_flat_scheme.applied_on_variant_id = user_applied_on_variant_id[id]
                            user_flat_scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,user_applied_on_variant_id[id],'variant_name')
                            #user_flat_scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))
                            user_flat_scheme.status = 1
                            user_flat_scheme.save()
    
                    #Save Activity
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = 'Flat scheme updated'
                    activity    = 'Flat scheme updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    saveActivity('Product & Variant Management', 'Flat Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')
                    
                    response['flag'] = True
                    response['message'] = "Record has been saved successfully."
            else:
                response['flag'] = False
                response['message'] = "Some User variant already mapped in different flat scheme, please check and try again."
                    

        return JsonResponse(response)
        
    else:
        template = 'schemes/edit-flat-incentive-scheme.html'
        context = {}
        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products

        context['scheme'] = scheme = SpFlatSchemes.objects.get(id=scheme_id)
        # class_list = scheme.product_class_id
        # class_list = class_list.split(',')
        # context['product_classes'] = product_classes =  SpProductClass.objects.filter(pk__in=class_list)
        context['states'] = SpStates.objects.all()

        if scheme.is_route == 1:
            zone_id = scheme.zone_id
            zone_id = zone_id.split(',')
        else:
            zone_id = []

        zone_lists = SpZones.objects.filter(state_id=scheme.state_id)
        zone_list = []
        for zone in zone_lists:
            zones_dict = {}
            zones_dict['id'] =  zone.id
            zones_dict['zone'] =  zone.zone
            if str(zone.id) in zone_id:
                zones_dict['zone_value'] =  'selected' 
            zone_list.append(zones_dict)
        context['zones']  = zone_list


        routes_list = []
        if scheme.is_route == 0:
            route_id = scheme.route_id
            route_id = route_id.split(',')
        else:
            route_id = []
        routes = SpRoutes.objects.filter(state_id=scheme.state_id)
        for route in routes:
            routes_dict = {}
            routes_dict['id'] =  route.id
            routes_dict['route'] =  route.route
            if str(route.id) in route_id:
                routes_dict['route_value'] =  'selected'
            routes_list.append(routes_dict)  

        context['routes_list'] = routes_list

        town_list = []
        if scheme.is_route == 1:
            town_id = scheme.town_id
            town_id = town_id.split(',')
            
            towns  = SpTowns.objects.filter(zone_id__in=zone_id)
            for town in towns:
                towns_dict = {}
                towns_dict['id'] =  town.id
                towns_dict['town'] =  town.town
                if str(town.id) in town_id:
                    towns_dict['town_value'] =  'selected' 
                town_list.append(towns_dict)

        context['towns'] = town_list


        context['units']          = SpProductUnits.objects.raw(''' SELECT id,largest_unit as unit FROM sp_product_units WHERE id in 
                        (SELECT DISTINCT variant_unit_id FROM sp_product_variants) GROUP BY largest_unit '''
                        )

        context['users'] = SpUserFlatSchemes.objects.raw(''' SELECT sp_user_flat_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_flat_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_flat_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_flat_schemes.user_id
        WHERE sp_user_flat_schemes.scheme_id = %s
        order by sp_user_flat_schemes.id  ''',[scheme_id])

        return render(request, template,context)

@login_required
def addBulkpackIncentiveScheme(request):
    if request.method == "POST":
        response = {}
        if SpBulkpackSchemes.objects.filter(name=request.POST['scheme_name']).exists() :
            response['flag'] = False
            response['message'] = "Scheme name already exists."
        else:
            scheme = SpBulkpackSchemes()
            scheme.name = request.POST['scheme_name']
            scheme.state_id = request.POST['state_id']
            separator = ','
            scheme.route_id = separator.join(request.POST.getlist('route_id[]'))
            scheme.town_id = separator.join(request.POST.getlist('town_id[]'))
            scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['end_date'] != '' :
                scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
           
            scheme.unit_id = request.POST['unit_id']
            scheme.unit_name = getModelColumnById(SpProductUnits,request.POST['unit_id'],'largest_unit')
            scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))

            scheme.status = 1
            scheme.save()
            if scheme.id : 
                above_upto_quantities       = request.POST.getlist('above_order_of[]')
                incentive_amounts       = request.POST.getlist('incentive_amount[]') 

                for id, val in enumerate(above_upto_quantities):
                    scheme_bifurcation =  SpBulkpackSchemeBifurcation()
                    scheme_bifurcation.bulkpack_scheme_id = scheme.id
                    scheme_bifurcation.above_upto_quantity = above_upto_quantities[id]
                    scheme_bifurcation.incentive_amount = incentive_amounts[id]
                    scheme_bifurcation.save()


                user_id       = request.POST.getlist('user_id[]') 
                is_disabled       = request.POST.getlist('is_disabled[]') 

                for id, val in enumerate(user_id):
                    if not int(is_disabled[id]) :
                        
                        user_bulkpack_scheme =  SpUserBulkpackSchemes()
                        user_bulkpack_scheme.user_id = user_id[id]
                        user_bulkpack_scheme.scheme_id = scheme.id
                        user_bulkpack_scheme.scheme_name = scheme.name
                        user_bulkpack_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                        user_bulkpack_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                        user_bulkpack_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                        user_bulkpack_scheme.scheme_start_date = scheme.scheme_start_date
                        if scheme.scheme_end_date is not None :
                            user_bulkpack_schemendeme_start_date = scheme.scheme_end_date
                        
                        user_bulkpack_scheme.unit_id = request.POST['unit_id']
                        user_bulkpack_scheme.unit_name = getModelColumnById(SpProductUnits,request.POST['unit_id'],'largest_unit')
                        user_bulkpack_scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))
                        user_bulkpack_scheme.status = 1
                        user_bulkpack_scheme.save()

                        if user_bulkpack_scheme.id :
                            incentive_var = "user_incentive_amount_"+str(user_id[id])
                            order_var = "user_above_order_of_"+str(user_id[id])
                            user_incentive_amount       = request.POST.getlist(incentive_var)
                            user_above_order       = request.POST.getlist(order_var)

                            for next_id, val in enumerate(user_incentive_amount):
                                bifurcation =  SpUserBulkpackSchemeBifurcation()
                                bifurcation.user_id = user_id[id]
                                bifurcation.bulkpack_scheme_id = scheme.id
                                bifurcation.above_upto_quantity = user_above_order[next_id]
                                bifurcation.incentive_amount = user_incentive_amount[next_id]
                                bifurcation.save()

                #Save Activity
                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'Bulkpack scheme created'
                activity    = 'Bulkpack scheme created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Product & Variant Management', 'Bulkpack Scheme', heading, activity, request.user.id, user_name, 'productScheme.png', '1', 'web.png')

                response['flag'] = True
                response['message'] = "Record has been saved successfully."

        return JsonResponse(response)
        
    else:
        template = 'schemes/add-bulkpack-incentive-scheme.html'
        context = {}
        context['states'] = SpStates.objects.all()
        context['units']          = SpProductUnits.objects.raw(''' SELECT id,largest_unit as unit FROM sp_product_units WHERE id in 
                        (SELECT DISTINCT variant_unit_id FROM sp_product_variants)  GROUP BY largest_unit'''
                        )

        return render(request, template,context)

@login_required
def editBulkpackIncentiveScheme(request,scheme_id):
    if request.method == "POST":
        response = {}
        if SpBulkpackSchemes.objects.filter(name=request.POST['scheme_name']).exclude(id=request.POST['scheme_id']).exists() :
            response['flag'] = False
            response['message'] = "Scheme name already exists."
        else:
            scheme = SpBulkpackSchemes.objects.get(id=request.POST['scheme_id'])
            scheme.name = request.POST['scheme_name']
            scheme.state_id = request.POST['state_id']
            separator = ','
            scheme.route_id = separator.join(request.POST.getlist('route_id[]'))
            scheme.town_id = separator.join(request.POST.getlist('town_id[]'))
            scheme.scheme_start_date = datetime.strptime(request.POST['start_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['end_date'] != '' :
                scheme.scheme_end_date = datetime.strptime(request.POST['end_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
           
            scheme.unit_id = request.POST['unit_id']
            scheme.unit_name = getModelColumnById(SpProductUnits,request.POST['unit_id'],'largest_unit')
            scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))

            scheme.status = 1
            scheme.save()
            if scheme.id : 
                above_upto_quantities       = request.POST.getlist('above_order_of[]')
                incentive_amounts       = request.POST.getlist('incentive_amount[]') 

                SpBulkpackSchemeBifurcation.objects.filter(bulkpack_scheme_id=request.POST['scheme_id']).delete()

                for id, val in enumerate(above_upto_quantities):
                    scheme_bifurcation =  SpBulkpackSchemeBifurcation()
                    scheme_bifurcation.bulkpack_scheme_id = scheme.id
                    scheme_bifurcation.above_upto_quantity = above_upto_quantities[id]
                    scheme_bifurcation.incentive_amount = incentive_amounts[id]
                    scheme_bifurcation.save()


                user_id       = request.POST.getlist('user_id[]') 
                is_disabled       = request.POST.getlist('is_disabled[]') 

                SpUserBulkpackSchemes.objects.filter(scheme_id=request.POST['scheme_id']).delete()
                SpUserBulkpackSchemeBifurcation.objects.filter(bulkpack_scheme_id=request.POST['scheme_id']).delete()

                for id, val in enumerate(user_id):
                        
                    user_bulkpack_scheme =  SpUserBulkpackSchemes()
                    user_bulkpack_scheme.user_id = user_id[id]
                    user_bulkpack_scheme.scheme_id = scheme.id
                    user_bulkpack_scheme.scheme_name = scheme.name
                    user_bulkpack_scheme.state_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'state_id')
                    user_bulkpack_scheme.route_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'route_id')
                    user_bulkpack_scheme.town_id = getModelColumnByColumnId(SpUserAreaAllocations,'user_id',user_id[id],'town_id')
                    user_bulkpack_scheme.scheme_start_date = scheme.scheme_start_date
                    if scheme.scheme_end_date is not None :
                        user_bulkpack_schemendeme_start_date = scheme.scheme_end_date
                    
                    user_bulkpack_scheme.unit_id = request.POST['unit_id']
                    user_bulkpack_scheme.unit_name = getModelColumnById(SpProductUnits,request.POST['unit_id'],'largest_unit')
                    user_bulkpack_scheme.product_class_id = ','.join(request.POST.getlist('product_class[]'))
                    user_bulkpack_scheme.status = 1
                    user_bulkpack_scheme.save()

                    if user_bulkpack_scheme.id :
                        incentive_var = "user_incentive_amount_"+str(user_id[id])
                        order_var = "user_above_order_of_"+str(user_id[id])
                        user_incentive_amount       = request.POST.getlist(incentive_var)
                        user_above_order       = request.POST.getlist(order_var)

                        for next_id, val in enumerate(user_incentive_amount):
                            bifurcation =  SpUserBulkpackSchemeBifurcation()
                            bifurcation.user_id = user_id[id]
                            bifurcation.bulkpack_scheme_id = scheme.id
                            bifurcation.above_upto_quantity = user_above_order[next_id]
                            bifurcation.incentive_amount = user_incentive_amount[next_id]
                            bifurcation.save()

                #Save Activity
                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'Bulkpack scheme updated'
                activity    = 'Bulkpack scheme updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Product & Variant Management', 'Bulkpack Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

                response['flag'] = True
                response['message'] = "Record has been saved successfully."

        return JsonResponse(response)
        
    else:
        template = 'schemes/edit-bulkpack-incentive-scheme.html'
        context = {}
        context['scheme'] = scheme = SpBulkpackSchemes.objects.get(id=scheme_id)
        context['scheme_bifurcations'] = SpBulkpackSchemeBifurcation.objects.filter(bulkpack_scheme_id=scheme_id)
        context['product_classes'] = SpProductClass.objects.raw(''' SELECT * FROM sp_product_class WHERE id in (%s) ''',[scheme.product_class_id])
        context['states'] = SpStates.objects.all()
        if scheme.route_id is not None:
            context['routes'] = SpRoutes.objects.raw(''' SELECT * FROM sp_routes WHERE id in (%s) ''',[scheme.route_id])

        if scheme.town_id is not None:
            context['towns'] = SpTowns.objects.raw(''' SELECT * FROM sp_towns WHERE id in (%s) ''',[scheme.town_id])


        context['units']          = SpProductUnits.objects.raw(''' SELECT id,largest_unit as unit FROM sp_product_units WHERE id in 
                        (SELECT DISTINCT variant_unit_id FROM sp_product_variants) GROUP BY largest_unit '''
                        )
        users = SpUserBulkpackSchemes.objects.raw(''' SELECT sp_user_bulkpack_schemes.*, sp_users.id as user_id,sp_users.first_name,sp_users.middle_name,sp_users.last_name,
        sp_user_area_allocations.town_name,sp_user_area_allocations.route_name,sp_user_area_allocations.state_name FROM sp_user_bulkpack_schemes
        LEFT JOIN sp_users on sp_users.id = sp_user_bulkpack_schemes.user_id
        LEFT JOIN sp_user_area_allocations on sp_user_area_allocations.user_id = sp_user_bulkpack_schemes.user_id
        WHERE sp_user_bulkpack_schemes.scheme_id = %s
        order by sp_user_bulkpack_schemes.id  ''',[scheme_id])

        for user in users:
            user.scheme_bifurcations = SpUserBulkpackSchemeBifurcation.objects.filter(user_id=user.user_id,bulkpack_scheme_id=scheme_id)
        context['users'] = users

       

        return render(request, template,context)

@login_required
def getIncentiveFilteredUsers(request):
    if request.method == "POST":
        
        context = {}
        state_id = request.POST['state_id']
        is_route = request.POST['is_route']
        route_ids = request.POST['route_ids']
        zone_ids = request.POST['zone_ids']
        town_ids = request.POST['town_ids']
        if int(request.POST['incentive_scheme_type']) == 1 :
    
            # unit_id = request.POST['unit_id']
            applied_on_variant_id = request.POST['applied_on_variant_id']
            
            template = "schemes/flat-filtered-distributors.html"
            if request.POST['applied_on_variant_id'] != '' :
                context['applied_on_variant_id'] = int(request.POST['applied_on_variant_id'])
                context['applied_on_variant_name'] = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
            else :
                context['applied_on_variant_id']   = ""
                context['applied_on_variant_name'] = ""
            
            if request.POST['incentive_amount'] != '' :
                context['incentive_amount'] = request.POST['incentive_amount']
            else :
                context['incentive_amount'] = 0
            
            
            if applied_on_variant_id != '' :
                if route_ids != "" and is_route == '0':
                    route_ids = request.POST['route_ids'].split(',')
                    route_users_list = []
                    for route_id in route_ids :
    
                        users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                        ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                        FROM sp_user_area_allocations 
                        LEFT JOIN sp_users on sp_users.id = sp_user_area_allocations.user_id 
                        where sp_users.user_type=2 and sp_users.emp_sap_id!='' and sp_users.status=1 and sp_user_area_allocations.route_id = %s ''',[route_id])
                        
                        for user in users: 
                            if SpUserSchemes.objects.filter(scheme_type=1,applied_on_variant_id=applied_on_variant_id,user_id=user.user_id).exists():
                                user.is_scheme_applied = 1
                            else:
                                user.is_scheme_applied = 0
                                
                            route_users_list.append(user)
    
                    context['users'] = route_users_list
    
                if zone_ids != ''  and is_route == '1':
                    zone_ids = request.POST['zone_ids'].split(',')
                    zone_users = []
                    for zone_id in zone_ids :
                        users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                        ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                        where sp_users.user_type=2 and sp_users.emp_sap_id!='' and sp_users.status=1 and sp_user_area_allocations.zone_id = %s ''',[zone_id])
    
                        for user in users:
                            if SpUserSchemes.objects.filter(scheme_type=1,applied_on_variant_id=applied_on_variant_id,user_id=user.user_id).exists():
                                user.is_scheme_applied = 1
                            else:
                                user.is_scheme_applied = 0
                            zone_users.append(user)
    
                    context['users'] = zone_users
    
                if town_ids != ''  and is_route == '1':
                    town_ids = request.POST['town_ids'].split(',')
                    town_users = []
                    for town_id in town_ids :
                        users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                        ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                        where sp_users.user_type=2 and sp_users.emp_sap_id!='' and sp_users.status=1 and sp_user_area_allocations.town_id = %s ''',[town_id])
    
                        for user in users:
                            if SpUserSchemes.objects.filter(scheme_type=1,applied_on_variant_id=applied_on_variant_id,user_id=user.user_id,town_id=town_id).exists():
                                user.is_scheme_applied = 1
                            else:
                                user.is_scheme_applied = 0
                            town_users.append(user)
    
                    context['users'] = town_users
            else :
                context['user'] = None

        else:
            template = "schemes/bulkpack-filtered-distributors.html"

            context['above_order_ofs'] = above_order_ofs = request.POST.getlist('above_order_of[]')

            context['incentive_amounts'] = request.POST.getlist('incentive_amount[]')

            route_ids = ','.join(request.POST.getlist('route_id[]'))
            town_ids = ','.join(request.POST.getlist('town_id[]'))
            unit_id = request.POST['unit_id']


            if request.POST['unit_id'] != '' :
                context['unit_id'] = request.POST['unit_id']
                context['unit_name'] = getModelColumnById(SpProductUnits,request.POST['unit_id'],'largest_unit')
            else :
                context['unit_id'] = ""
                context['unit_name'] = ""


            if request.POST['unit_id'] != '' :

                if state_id != '' :
                    users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                    ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                    FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                    where sp_users.user_type=2 and (sp_users.is_distributor = 1 or sp_users.is_super_stockist = 1) and sp_user_area_allocations.state_id = %s ''',[state_id])
                    for user in users:
                        if SpUserBulkpackSchemes.objects.filter(unit_id=unit_id,user_id=user.user_id,state_id=state_id).exists():
                            user.is_scheme_applied = 1
                        else:
                            user.is_scheme_applied = 0

                    context['users'] = users
        
                if route_ids != "" :
                    route_ids = route_ids.split(',')
                    route_users = []
                    for route_id in route_ids :

                        users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                        ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                        FROM sp_user_area_allocations 
                        LEFT JOIN sp_users on sp_users.id = sp_user_area_allocations.user_id 
                        where sp_users.user_type=2 and (sp_users.is_distributor = 1 or sp_users.is_super_stockist = 1) and sp_user_area_allocations.route_id = %s ''',[route_id])
                        
                        for user in users:

                            if SpUserBulkpackSchemes.objects.filter(unit_id=unit_id,user_id=user.user_id,route_id=route_id).exists():
                                user.is_scheme_applied = 1
                            else:
                                user.is_scheme_applied = 0

                            route_users.append(user)

                    context['users'] = route_users


                if town_ids != '':
                    town_ids = town_ids.split(',')
                    town_users = []
                    for town_id in town_ids :
                        users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                        ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                        where sp_users.user_type=2 and (sp_users.is_distributor = 1 or sp_users.is_super_stockist = 1) and sp_user_area_allocations.town_id = %s ''',[town_id])

                        for user in users:
                            if SpUserBulkpackSchemes.objects.filter(unit_id=unit_id,user_id=user.user_id,town_id=town_id).exists():
                                user.is_scheme_applied = 1
                            else:
                                user.is_scheme_applied = 0
                            town_users.append(user)

                    context['users'] = town_users

            else :
                context['user'] = None

        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products


        return render(request, template,context)

    else:
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed"
        return JsonResponse(response)


@login_required
def getFilteredUsers(request):
    if request.method == "POST":
        context = {}
        context['packaging_type'] = request.POST['packaging_type']
        context['container_name'] = request.POST['container_name']
        context['packaging_names'] = request.POST['packaging_name']
        state_id = request.POST['state_id']
        is_route = request.POST['is_route']
        route_ids = request.POST['route_ids']
        zone_ids = request.POST['zone_ids']
        town_ids = request.POST['town_ids']

        if int(request.POST['scheme_type']) == 1 :
            template = "schemes/free-filtered-distributors.html"
            if request.POST['applied_on_variant_id'] != '' :
                context['applied_on_variant_id'] = request.POST['applied_on_variant_id']
                context['applied_on_variant_name'] = getModelColumnById(SpProductVariants,request.POST['applied_on_variant_id'],'variant_name')
            else :
                context['applied_on_variant_id'] = ""
                context['applied_on_variant_name'] = ""
            
            if request.POST['minimum_order_quantity'] != '' :
                context['minimum_order_quantity'] = request.POST['minimum_order_quantity']
            else :
                context['minimum_order_quantity'] = ""

            applied_on_variant_id = request.POST['applied_on_variant_id']

            # if state_id != '' :
            #     users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
            #     ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
            #     FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
            #     where sp_users.user_type=2 and (sp_users.is_distributor = 1 or sp_users.is_super_stockist = 1) and sp_user_area_allocations.state_id = %s ''',[state_id])
            #     for user in users:
            #         if SpUserSchemes.objects.filter(scheme_type=1,applied_on_variant_id=applied_on_variant_id,user_id=user.user_id,state_id=state_id).exists():
            #             user.is_scheme_applied = 1
            #         else:
            #             user.is_scheme_applied = 0

            #     context['users'] = users
            

            if route_ids != "" and is_route == '0':
                route_ids = request.POST['route_ids'].split(',')
                route_users_list = []
                for route_id in route_ids :

                    users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                    ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                    FROM sp_user_area_allocations 
                    LEFT JOIN sp_users on sp_users.id = sp_user_area_allocations.user_id 
                    where sp_users.user_type=2 and sp_users.emp_sap_id!='' and sp_users.status=1 and sp_user_area_allocations.route_id = %s ''',[route_id])
                    
                    for user in users: 
                        if SpUserSchemes.objects.filter(scheme_type=1,applied_on_variant_id=applied_on_variant_id,user_id=user.user_id).exists():
                            user.is_scheme_applied = 1
                        else:
                            user.is_scheme_applied = 0
                            
                        route_users_list.append(user)

                context['users'] = route_users_list

            if zone_ids != ''  and is_route == '1':
                zone_ids = request.POST['zone_ids'].split(',')
                zone_users = []
                for zone_id in zone_ids :
                    users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                    ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                    FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                    where sp_users.user_type=2 and sp_users.emp_sap_id!='' and sp_users.status=1 and sp_user_area_allocations.zone_id = %s ''',[zone_id])

                    for user in users:
                        if SpUserSchemes.objects.filter(scheme_type=1,applied_on_variant_id=applied_on_variant_id,user_id=user.user_id).exists():
                            user.is_scheme_applied = 1
                        else:
                            user.is_scheme_applied = 0
                        zone_users.append(user)

                context['users'] = zone_users

            if town_ids != ''  and is_route == '1':
                town_ids = request.POST['town_ids'].split(',')
                town_users = []
                for town_id in town_ids :
                    users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                    ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                    FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                    where sp_users.user_type=2 and sp_users.emp_sap_id!='' and sp_users.status=1 and sp_user_area_allocations.town_id = %s ''',[town_id])

                    for user in users:
                        if SpUserSchemes.objects.filter(scheme_type=1,applied_on_variant_id=applied_on_variant_id,user_id=user.user_id,town_id=town_id).exists():
                            user.is_scheme_applied = 1
                        else:
                            user.is_scheme_applied = 0
                        town_users.append(user)

                context['users'] = town_users
        else:
            template = "schemes/bonus-filtered-distributors.html"
            context['order_containers'] = request.POST.getlist('order_container_id[]')
            context['minimum_order_quantities'] = request.POST.getlist('minimum_order_quantity[]')
            
            if state_id != '' :
                users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                where sp_users.user_type=2 and (sp_users.is_distributor = 1 or sp_users.is_super_stockist = 1) and sp_user_area_allocations.state_id = %s ''',[state_id])
                for user in users:
                    if SpUserSchemes.objects.filter(scheme_type=2,user_id=user.user_id,state_id=state_id).exists():
                        user.is_scheme_applied = 1
                    else:
                        user.is_scheme_applied = 0

                context['users'] = users
        
            if route_ids != "" :
                route_ids = request.POST['route_ids'].split(',')
                route_users = []
                for route_id in route_ids :

                    users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                    ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                    FROM sp_user_area_allocations 
                    LEFT JOIN sp_users on sp_users.id = sp_user_area_allocations.user_id 
                    where sp_users.user_type=2 and (sp_users.is_distributor = 1 or sp_users.is_super_stockist = 1) and sp_user_area_allocations.route_id = %s ''',[route_id])
                    
                    for user in users:

                        if SpUserSchemes.objects.filter(scheme_type=2,user_id=user.user_id,route_id=route_id).exists():
                            user.is_scheme_applied = 1
                        else:
                            user.is_scheme_applied = 0

                        route_users.append(user)

                context['users'] = route_users


            if town_ids != '':
                town_ids = request.POST['town_ids'].split(',')
                town_users = []
                for town_id in town_ids :
                    users = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.*, sp_users.id as user_id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                    ,sp_user_area_allocations.route_name,sp_user_area_allocations.town_name
                    FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
                    where sp_users.user_type=2 and (sp_users.is_distributor = 1 or sp_users.is_super_stockist = 1) and sp_user_area_allocations.town_id = %s ''',[town_id])

                    for user in users:
                        if SpUserSchemes.objects.filter(scheme_type=2,user_id=user.user_id,town_id=town_id).exists():
                            user.is_scheme_applied = 1
                        else:
                            user.is_scheme_applied = 0
                        town_users.append(user)

                context['users'] = town_users
        


        products = SpProducts.objects.filter(status=1)
        for product in products : 
            product.product_variants = SpProductVariants.objects.filter(status=1,product_id=product.id)

        context['products'] = products
        

        if request.POST['free_variant_id'] != '' :
            context['free_variant_id'] = int(request.POST['free_variant_id'])
            context['free_variant_name'] = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'variant_name')
            product_id = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'product_id')
            context['container'] = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'container_name')
            context['packaging_name'] = getModelColumnById(SpProductVariants,request.POST['free_variant_id'],'packaging_type_name')
        else:
            context['free_variant_id']  = int(0)
            context['free_variant_name'] = ''
            context['container'] = ''
            context['packaging_name'] = ''

        

        if request.POST['container_qty'] != '' :
            context['container_qty'] = request.POST['container_qty']
        else:
            context['container_qty']  = 0
        
        if request.POST['pouch_qty'] != '' :
            context['pouch_qty'] = request.POST['pouch_qty']
        else:
            context['pouch_qty']  = 0

        context['containers']     = SpContainers.objects.all()

        return render(request, template,context)

    else:
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed"
        return JsonResponse(response)


@login_required
def productUnitDetails(request,unit_id):
    response = {}
    if SpProductUnits.objects.filter(id=unit_id).exists():
        response['flag'] = True
        response['product_unit'] = model_to_dict(SpProductUnits.objects.get(id=unit_id))
    else:
        response['flag'] = False
        response['flag'] = "Unit not found"

    return JsonResponse(response)




@login_required
def updateFreeSchemeStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpSchemes.objects.get(id=id)
            data.status = is_active
            data.save()

            SpUserSchemes.objects.filter(scheme_type=1,scheme_id=id).update(status=is_active)

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Free Scheme status updated'
            activity    = 'Free Scheme status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Free Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')


            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')


@login_required
def updateQuantitativeSchemeStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpSchemes.objects.get(id=id)
            data.status = is_active
            data.save()
            SpUserSchemes.objects.filter(scheme_type=2,scheme_id=id).update(status=is_active)

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Quantitative Scheme status updated'
            activity    = 'Quantitative Scheme status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Quantitative Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')


            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')


@login_required
def updateFlatIncentiveSchemeStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpFlatSchemes.objects.get(id=id)
            data.status = is_active
            data.save()
            SpUserFlatSchemes.objects.filter(scheme_id=id).update(status=is_active)
            
            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Flat Scheme status updated'
            activity    = 'Flat Scheme status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Flat Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')


            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')


@login_required
def updateBulkpackIncentiveSchemeStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpBulkpackSchemes.objects.get(id=id)
            data.status = is_active
            data.save()
            SpUserBulkpackSchemes.objects.filter(scheme_id=id).update(status=is_active)
            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Bulkpack Scheme status updated'
            activity    = 'Bulkpack Scheme status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Bulkpack Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')




@login_required
def updateUserFreeSchemeStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpUserSchemes.objects.get(id=id)
            data.status = is_active
            data.save()

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'User Free Scheme status updated'
            activity    = 'User Free Scheme status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'User Free Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')


@login_required
def updateUserQuantitativeSchemeStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpUserSchemes.objects.get(id=id)
            data.status = is_active
            data.save()

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'User quantitative status updated'
            activity    = 'User quantitative status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'User quantitative', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')


@login_required
def updateUserFlatIncentiveSchemeStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpUserFlatSchemes.objects.get(id=id)
            data.status = is_active
            data.save()

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'User Flat Scheme status updated'
            activity    = 'User Flat Scheme status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'User Flat Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')


@login_required
def updateUserBulkpackIncentiveSchemeStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')
            data = SpUserBulkpackSchemes.objects.get(id=id)
            data.status = is_active
            data.save()

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Bulkpack Scheme status updated'
            activity    = 'Bulkpack Scheme status updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Bulkpack Scheme', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Object Does Not Exist"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/schemes')

@login_required
def importOutstanding(request):
    if request.method == 'POST' :
        outstanding_sheet = request.FILES["outstanding_sheet"]
        wb = openpyxl.load_workbook(outstanding_sheet)
        worksheet = wb.active
        excel_data = list()
        header_data = list()
        i = 0
        for row in worksheet.iter_rows():
            if(i > 0):
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            else:
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                header_data.append(row_data)    
            i = int(i) + 1
        
        header_list = ['SAPID', 'PARTY NAME', 'AMOUNT']
        for header in header_data[0]:
            if header not in header_list:
                messages.error(request, 'Invalid File Format', extra_tags='success')
                return redirect('/upload-outstanding')

        distributor_data = []
        for data in excel_data:
            temp = {}
            temp['sap_id'] = data[0]
            temp['outstanding_amount'] = data[2]
            if SpUsers.objects.filter(emp_sap_id=data[0]).exists():
                temp['distributor_name'] = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'first_name')
                temp['store_name'] = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'store_name')
            else:
                temp['distributor_name'] = "Invalid SAP ID"
                temp['store_name'] = ""
            
            distributor_data.append(temp)
        context = {}
        context['excel_data'] = distributor_data
        context['page_title']                       = "Upload outstanding"
        template = 'schemes/upload-outstanding.html'
        return render(request,template,context)

    else:
        context = {}
        context['page_title']                       = "Upload outstanding"
        template = 'schemes/upload-outstanding.html'
        return render(request,template,context)

@login_required
def updateOutstanding(request):
    if request.method == 'POST' :
        sap_id = request.POST.getlist('sap_id[]')
        outstanding_amount = request.POST.getlist('outstanding_amount[]')
        for id, val in enumerate(sap_id):
            if SpUsers.objects.filter(emp_sap_id=sap_id[id]).exists():
                user = SpUsers.objects.get(emp_sap_id=sap_id[id])
                SpBasicDetails.objects.filter(user_id=user.id).update(outstanding_amount=outstanding_amount[id])
        
        #Save Activity
        user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
        heading     = 'Outstanding amount updated'
        activity    = 'Outstanding amount updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
        saveActivity('Product & Variant Management', 'Update Outstanding amount', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

        response = {}
        response['flag'] = True
        response['message'] = "Records has been updated successfully."
        return JsonResponse(response)

    else:
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed"
        return JsonResponse(response)

@login_required
def exportOutstandingExcel(request):
    path = 'excel-templates/user_outstanding_amount.xlsx'
    print(os.path) 
    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=user_outstanding_amount.xlsx'

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Outstanding excel format exported'
            activity    = 'Outstanding excel format exported by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Export Outstanding', heading, activity, request.user.id, user_name, 'generateIndentReport.png', '1', 'web.png')
            
            return response
    else:
        raise Http404
        
@login_required
def importRateList(request):
    if request.method == 'POST' :
        rate_list_sheet = request.FILES["rate_list_sheet"]
        wb = openpyxl.load_workbook(rate_list_sheet)
        worksheet = wb.active
        excel_data = list()
        header_data = list()
        i = 0
        for row in worksheet.iter_rows():
            if(i > 0):
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            else:
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                header_data.append(row_data)    
            i = int(i) + 1
        
        header_list = ['SAPID', 'PARTY NAME', 'SKU', 'RATE']
        for header in header_data[0]:
            if header not in header_list:
                messages.error(request, 'Invalid File Format', extra_tags='success')
                return redirect('/upload-rate-list')

        distributor_data = []
        for data in excel_data:
            temp = {}
            temp['sap_id'] = data[0]
            
            if SpUsers.objects.filter(emp_sap_id=data[0]).exists():
                user_id = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'id')
                temp['distributor_name'] = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'first_name')
                temp['store_name'] = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'store_name')

                if SpUserProductVariants.objects.filter(user_id=user_id,item_sku_code=data[2]).exists():
                    temp['product_variant_name'] = getModelColumnByColumnId(SpUserProductVariants,'item_sku_code',data[2],'variant_name')
                    temp['sku_code'] = data[2]
                else:
                    temp['product_variant_name'] = "Invalid SKU"
                    temp['sku_code'] = data[2]

            else:
                temp['distributor_name'] = "Invalid SAP ID"
                temp['store_name'] = ""
                temp['product_variant_name'] = "Invalid SKU"
                temp['sku_code'] = data[2]
            if data[3]>'0':
                rate = data[3].replace(" ", "")
                temp['rate'] = round(float(rate),2)
            else:
                temp['rate'] = "Invalid Rate"
            
            distributor_data.append(temp)
        context = {}
        context['excel_data'] = distributor_data
        context['page_title']                       = "Upload Rate List"
        template = 'schemes/upload-rate-list.html'
        return render(request,template,context)

    else:
        context = {}
        context['page_title']                       = "Upload Rate List"
        template = 'schemes/upload-rate-list.html'
        return render(request,template,context)


@login_required
def updateRateList(request):
    if request.method == 'POST' :
        sap_id = request.POST.getlist('sap_id[]')
        variant_id = request.POST.getlist('variant_id[]')
        rate = request.POST.getlist('rate[]')
        for id, val in enumerate(sap_id):
            if SpUsers.objects.filter(emp_sap_id=sap_id[id]).exists():
                user = SpUsers.objects.get(emp_sap_id=sap_id[id])
                SpUserProductVariants.objects.filter(user_id=user.id,item_sku_code=variant_id[id]).update(sp_user=rate[id])
                SpUserProductVariants.objects.filter(user_id=user.id,item_sku_code=variant_id[id]).update(container_sp_user=F('sp_user')*F('no_of_pouch'))

        #Save Activity
        user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
        heading     = 'Rate list updated'
        activity    = 'Rate list updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
        saveActivity('Product & Variant Management', 'Rate list', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')
        
        response = {}
        response['flag'] = True
        response['message'] = "Records has been updated successfully."
        return JsonResponse(response)

    else:
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed"
        return JsonResponse(response)



@login_required
def exportRateListExcel(request):
    path = 'excel-templates/user_rate_list.xlsx'
    print(os.path) 
    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=user_rate_list.xlsx'

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Rate list excel format exported'
            activity    = 'Rate list excel format exported by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Export Rate list', heading, activity, request.user.id, user_name, 'generateIndentReport.png', '1', 'web.png')

            return response
    else:
        raise Http404

@login_required
def exportSchemeExcel(request):
    path = 'excel-templates/user_scheme.xlsx'
    print(os.path) 
    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=user_scheme.xlsx'

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Scheme excel format exported'
            activity    = 'Scheme excel format exported by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Export Scheme', heading, activity, request.user.id, user_name, 'generateIndentReport.png', '1', 'web.png')
            
            return response
    else:
        raise Http404

@login_required
def importScheme(request):
    if request.method == 'POST' :
        scheme_sheet = request.FILES["scheme_sheet"]
        wb = openpyxl.load_workbook(scheme_sheet)
        worksheet = wb.active
        excel_data = list()
        header_data = list()
        i = 0
        for row in worksheet.iter_rows():
            if(i > 0):
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            else:
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                header_data.append(row_data)    
            i = int(i) + 1
        
        header_list = ['SapId', 'State', 'Route', 'Town', 'Variant Material Code', 'On Order Of Quantity', 'Free Variant Material Code', 'Container Quantity', 'Pouch Quantity']
        for header in header_data[0]:
            if header not in header_list:
                messages.error(request, 'Invalid File Format', extra_tags='success')
                return redirect('/upload-scheme')

        scheme_data = []
        for data in excel_data:
            temp = {}
            temp['sap_id']                      = data[0].strip()
            temp['state']                       = data[1].strip()
            temp['route']                       = data[2].strip()
            temp['town']                        = data[3].strip()
            temp['variant_material_code']       = data[4].strip()
            temp['on_order_of_quantity']        = round(float(data[5]))
            temp['free_variant_material_code']  = data[6].strip()
            temp['container']                   = data[7]
            temp['pouch']                       = data[8]
            if SpStates.objects.filter(state=data[1]).exists():
                temp['state_temp'] = ''
            else:
                temp['state_temp'] = "Invalid state name"
            if SpRoutes.objects.filter(route=data[2]).exists():
                temp['route_temp'] = ''
            else:
                temp['route_temp'] = "Invalid route name" 
            if SpRoutesTown.objects.filter(town_name=data[3]).exists():
                temp['route_town_temp'] = ''
            else:
                temp['route_town_temp'] = "Invalid town name"  
            if SpUsers.objects.filter(emp_sap_id=data[0]).exists():
                temp['user_name']       = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'first_name')+' '+getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'middle_name')+' '+getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'last_name')
                temp['user_name_tmp']   = "1"
            else:
                temp['user_name_tmp']   = "0"
                temp['user_name']       = "Invalid SAP ID"
            if SpUserProductVariants.objects.filter(item_sku_code=data[4]).exists():
                temp['variant_material_code_temp'] = ''
            else:
                temp['variant_material_code_temp'] = "Invalid variant material code"    
            if SpUserProductVariants.objects.filter(item_sku_code=data[6]).exists():
                temp['free_variant_material_code_temp'] = ''
            else:
                temp['free_variant_material_code_temp'] = "Invalid free variant material code"
            scheme_data.append(temp)
        context = {}
        context['excel_data'] = scheme_data
        context['page_title']                       = "Upload Scheme"
        template = 'schemes/upload-scheme.html'
        return render(request,template,context)

    else:
        context = {}
        context['page_title']                       = "Upload Scheme"
        template = 'schemes/upload-scheme.html'
        return render(request,template,context)

@login_required
def updateScheme(request):
    if request.method == 'POST' :
        sap_id                      = request.POST.getlist('sap_id[]')
        state                       = request.POST.getlist('state[]')
        route                       = request.POST.getlist('route[]')
        town                        = request.POST.getlist('town[]')
        variant_material_code       = request.POST.getlist('variant_material_code[]')
        free_variant_material_code  = request.POST.getlist('free_variant_material_code[]')  
        on_order_of_quantity        = request.POST.getlist('on_order_of_quantity[]')
        container                   = request.POST.getlist('container[]')
        pouch                       = request.POST.getlist('pouch[]')
        
        SpSchemes.objects.filter(scheme_type=1).delete()
        SpUserSchemes.objects.filter(scheme_type=1).delete()

        for id, val in enumerate(sap_id):
            if SpUsers.objects.filter(emp_sap_id=sap_id[id]).exists():
                user = SpUsers.objects.get(emp_sap_id=sap_id[id])
                product_id = SpProductVariants.objects.get(item_sku_code=variant_material_code[id])
                product = SpProducts.objects.get(id=product_id.product_id)
                if int(container[id]) > 0 or int(pouch[id]) > 0:
                    scheme                          = SpSchemes()
                    scheme.name                     = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', variant_material_code[id], 'id'),'variant_name')
                    scheme.state_id                 = getModelColumnByColumnId(SpStates, 'state', state[id], 'id')
                    scheme.route_id                 = getModelColumnByColumnId(SpRoutes, 'route', route[id], 'id')
                    scheme.town_id                  = getModelColumnByColumnId(SpTowns, 'town', town[id], 'id')
                    scheme.scheme_start_date        = datetime.now().strftime('%Y-%m-%d')
                    scheme.scheme_type              = 1
                    scheme.applied_on_variant_id    = getModelColumnByColumnId(SpUserProductVariants, 'item_sku_code', variant_material_code[id], 'product_variant_id')
                    scheme.applied_on_variant_name  = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', variant_material_code[id], 'id'),'variant_name')
                    scheme.minimum_order_quantity   = on_order_of_quantity[id]
                    scheme.order_container_id       = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', free_variant_material_code[id], 'id'),'container_id')
                    scheme.order_container_name     = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', free_variant_material_code[id], 'id'),'container_name')
                    scheme.free_variant_id          = getModelColumnByColumnId(SpUserProductVariants, 'item_sku_code', free_variant_material_code[id], 'product_variant_id')
                    scheme.free_variant_name        = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', free_variant_material_code[id], 'id'),'variant_name')
                    scheme.container_quantity       = container[id]
                    scheme.pouch_quantity           = pouch[id]
                    scheme.status                   = 1
                    scheme.save()
                    if scheme.id : 
                        user_scheme                         =  SpUserSchemes()
                        user_scheme.user_id                 = user.id
                        user_scheme.scheme_id               = scheme.id
                        user_scheme.scheme_name             = scheme.name
                        user_scheme.state_id                = getModelColumnByColumnId(SpStates, 'state', state[id], 'id')
                        user_scheme.route_id                = getModelColumnByColumnId(SpRoutes, 'route', route[id], 'id')
                        user_scheme.town_id                 = getModelColumnByColumnId(SpTowns, 'town', town[id], 'id')
                        user_scheme.scheme_start_date       = datetime.now().strftime('%Y-%m-%d')
                        user_scheme.scheme_type             = 1
                        user_scheme.applied_on_variant_id   = getModelColumnByColumnId(SpUserProductVariants, 'item_sku_code', variant_material_code[id], 'product_variant_id')
                        user_scheme.applied_on_variant_name = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', variant_material_code[id], 'id'),'variant_name')
                        user_scheme.minimum_order_quantity  = on_order_of_quantity[id]
                        user_scheme.order_container_id      = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', free_variant_material_code[id], 'id'),'container_id')
                        user_scheme.order_container_name    = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', free_variant_material_code[id], 'id'),'container_name')
                        user_scheme.free_variant_id         = getModelColumnByColumnId(SpUserProductVariants, 'item_sku_code', free_variant_material_code[id], 'product_variant_id')
                        user_scheme.free_variant_name       = getModelColumnById(SpProductVariants,getModelColumnByColumnId(SpProductVariants, 'item_sku_code', free_variant_material_code[id], 'id'),'variant_name')
                        user_scheme.container_quantity      = container[id]
                        user_scheme.pouch_quantity          = pouch[id]
                        user_scheme.status                  = 1
                        user_scheme.save()
                                  
        #Save Activity
        user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
        heading     = 'Free Scheme list updated'
        activity    = 'Free Scheme list updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
        saveActivity('Product & Variant Management', 'Scheme list', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')
        
        response = {}
        response['flag'] = True
        response['message'] = "Records has been updated successfully."
        return JsonResponse(response)

    else:
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed"
        return JsonResponse(response)

#--------------------------------------------------------------------------------------------------------------
@login_required
def exportRevisedExcel(request):
    path = 'excel-templates/user_revised_amount.xlsx'
    print(os.path) 
    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=user_revised_amount.xlsx'

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'revise excel format exported'
            activity    = 'revise excel format exported by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Export revise', heading, activity, request.user.id, user_name, 'generateIndentReport.png', '1', 'web.png')
            
            return response
    else:
        raise Http404
#---------------------------------------------------------------------------------------------------------------

#--------------------------------------------------------------------------------------------------
@login_required
def importRevisedAmount(request):
    today   = date.today()
    error = []
    if request.method == 'POST' :
        outstanding_sheet = request.FILES["outstanding_sheet"]
        wb = openpyxl.load_workbook(outstanding_sheet)
        worksheet = wb.active
        excel_data = list()
        header_data = list()
        i = 0
        for row in worksheet.iter_rows():
            if(i > 0):
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            else:
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                header_data.append(row_data)    
            i = int(i) + 1
        
        header_list = ['SAPID', 'PARTY NAME', 'AMOUNT']
        for header in header_data[0]:
            if header not in header_list:
                messages.error(request, 'Invalid File Format', extra_tags='success')
                return redirect('/upload-revised-amount')

        distributor_data = []
        for data in excel_data:
            temp = {}
            temp['sap_id'] = data[0]
            temp['revised_amount'] = data[2]
            if SpUsers.objects.filter(emp_sap_id=data[0]).exists():
                temp['distributor_name'] = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'first_name')
                temp['store_name'] = getModelColumnByColumnId(SpUsers,'emp_sap_id',data[0],'store_name')
                if not SpOrders.objects.filter(user_sap_id =data[0],order_date__icontains=today.strftime("%Y-%m-%d")).exists():
                    error.append('1')
                    temp['distributor_name'] = "Order not found"
                    temp['error']="1"
                    temp['store_name'] = ""
            else:
                error.append('1')
                temp['distributor_name'] = "Invalid SAP ID"
                temp['error']="1"
                temp['store_name'] = ""
            
            distributor_data.append(temp)
        context = {}
       
        #context['error']=error
        context['error'] = len(error)  
        context['excel_data'] = distributor_data
        context['page_title']                       = "Upload Revised Amount"
        template = 'schemes/upload-revised-amount.html'
        return render(request,template,context)
    else:
        context = {}
        context['page_title']                       = "Upload Revised Amount"
        template = 'schemes/upload-revised-amount.html'
        return render(request,template,context)
#------------------------------------------------------------------------------------------------------------

def updateRevisedAmount(request):
    today   = date.today()
    if request.method == 'POST' :
        sap_id = request.POST.getlist('sap_id[]')
        revised_amounts=request.POST.getlist('revised_amount[]')
        for id, val in enumerate(sap_id):
            if SpUsers.objects.filter(emp_sap_id=sap_id[id]).exists():
                if SpOrders.objects.filter(user_sap_id =sap_id[id],order_date__icontains=today.strftime("%Y-%m-%d")).exists():
                    order_data = SpOrders.objects.get(user_sap_id =sap_id[id],order_date__icontains=today.strftime("%Y-%m-%d"))#.first()
                    user = SpUsers.objects.get(emp_sap_id=sap_id[id])
                    orderDate = datetime.now().strftime("%Y-%m-%d")
                    
                    order_data.revised_amount=order_data.amount_to_be_paid
                    order_data.amount_to_be_paid=revised_amounts[id]
                    order_data.save()
                    
                    basic_details = SpBasicDetails.objects.get(user_id=user.id)
                    
                    total_amount = round((float(basic_details.outstanding_amount)+float(order_data.order_total_amount)),2)
                    final_amount = round((float(total_amount)-float(revised_amounts[id])),2)
                    
                    basic_details.outstanding_amount = final_amount
                    basic_details.save()
                    
        #Save Activity
        user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
        heading     = 'revised amount updated'
        activity    = 'revised amount updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
        saveActivity('Product & Variant Management', 'Update revised amount', heading, activity, request.user.id, user_name, 'schemeEdit.png', '1', 'web.png')

        response = {}
        response['flag'] = True
        response['message'] = "Records has been updated successfully."
        return JsonResponse(response)
        # response.redirect('/upload-revised-amount.html')

    else:
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed"
        return JsonResponse(response)

#------------------------------------------------------------------------------------------------------------

