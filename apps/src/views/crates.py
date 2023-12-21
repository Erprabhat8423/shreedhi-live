import sys
import os
from django import dispatch
import openpyxl
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.db import transaction

# List View
@login_required
def index(request):
    today   = date.today()  
    year  = today.year
    month = today.month

    user_list = SpUsers.objects.filter(user_type=2, status=1).order_by('first_name')
    month_list = days_in_months(year,month)
    
    total_opening_crates        = []
    total_dispatch_crate        = []
    total_plant_crate           = []
    total_short_for_month       = []
    total_short_since_starting  = []
    for user in user_list:
        previous_month_last_date = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        if SpPlantCrateLedger.objects.filter(user_id=user.id,updated_datetime__lt=previous_month_last_date).exists():
            opening_crates              = getLastCrateBalance(user.id, previous_month_last_date, 'normal')
            user.opening_crates         = opening_crates
        else:
            user.opening_crates         = 0 
        total_opening_crates.append(user.opening_crates)

        if getTotalCrates(user.id, year, month, 'normal', 'dispatch'):
            user.total_dispatch_crates  = getTotalCrates(user.id, year, month, 'normal', 'dispatch')
        else:
            user.total_dispatch_crates  = 0    
        total_dispatch_crate.append(user.total_dispatch_crates)

        if getTotalCrates(user.id, year, month, 'normal', 'plant'):        
            user.total_plant_crates     = getTotalCrates(user.id, year, month, 'normal', 'plant')
        else:
            user.total_plant_crates     = 0    
        total_plant_crate.append(user.total_plant_crates)

        user.short_for_month        =  int(user.total_plant_crates)-int(user.total_dispatch_crates)
        user.short_since_starting   =  int(user.opening_crates)+int(user.short_for_month)

        total_short_for_month.append(user.short_for_month)
        total_short_since_starting.append(user.short_since_starting)

        crate_lists = []
        for months in month_list:
            month_lists               = {} 
            month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            dispatch                  = getDispatchedCratesSum(user.id, month_date, 'normal')
            plant                     = getReceivedCratesSum(user.id, month_date, 'normal')

            month_lists['dispatch']   = dispatch
            month_lists['plant']      = plant

            crate_lists.append(month_lists)    
        user.crates =  crate_lists

    total_crate_lists = []
    for months in month_list:
        total_month_lists               = {} 
        month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
        dispatch                  = getTotalDispatchedCratesSum(month_date, 'normal')
        plant                     = getTotalReceivedCratesSum(month_date, 'normal')

        total_month_lists['dispatch']   = dispatch
        total_month_lists['plant']      = plant

        total_crate_lists.append(total_month_lists)
    
    context = {}
    context['user_list']                    = user_list
    context['today_date']                   = today.strftime("%m/%Y")
    context['month_list']                   = days_in_months(year,month)
    context['month_list_count']             = len(month_list)+len(month_list)+8
    context['total_opening_crates']         = sum(total_opening_crates)
    context['total_dispatch_crate']         = sum(total_dispatch_crate)
    context['total_plant_crate']            = sum(total_plant_crate)
    context['total_short_for_month']        = sum(total_short_for_month)
    context['total_short_since_starting']   = sum(total_short_since_starting)
    context['total_crate_lists']            = total_crate_lists
    context['routes']                       = SpRoutes.objects.filter(status=1)
    context['page_title']                   = "Crate Summary"

    template = 'crate/index.html'
    return render(request, template, context)

# ajax List View
@login_required
def ajaxCrateSummary(request):
    today   = date.today()  
    if request.GET['crate_date']:
        crate_date = request.GET['crate_date']
        crate_date = crate_date.split('/')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)
           
    #user_list = SpUsers.objects.filter(user_type=2, status=1)
    condition = ''
    condition += ' and sp_users.user_type = %s' % 2
    condition += ' and sp_users.status = %s' % 1
    if request.GET['route_id']:
        condition += ' and sp_user_area_allocations.route_id = %s' % int(request.GET['route_id'])
        
    user_list = SpUsers.objects.raw(''' SELECT sp_users.* FROM sp_users left join sp_user_area_allocations on sp_user_area_allocations.user_id = sp_users.id WHERE 1 {condition} order by sp_users.first_name asc '''.format(condition=condition))
    
    month_list = days_in_months(year,month)
    
    total_opening_crates        = []
    total_dispatch_crate        = []
    total_plant_crate           = []
    total_short_for_month       = []
    total_short_since_starting  = []
    for user in user_list:
        previous_month_last_date = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        if SpUserCrateLedger.objects.filter(user_id=user.id,updated_datetime__lt=previous_month_last_date).exists():
            opening_crates              = getLastCrateBalance(user.id, previous_month_last_date, request.GET['crate_type'])
            user.opening_crates         = opening_crates
        else:
            user.opening_crates         = 0
        total_opening_crates.append(user.opening_crates)
        
        if getTotalCrates(user.id, year, month, request.GET['crate_type'], 'dispatch'):
            user.total_dispatch_crates  = getTotalCrates(user.id, year, month, request.GET['crate_type'], 'dispatch')
        else:
            user.total_dispatch_crates  = 0
        total_dispatch_crate.append(user.total_dispatch_crates)
        
        if getTotalCrates(user.id, year, month, request.GET['crate_type'], 'plant'):        
            user.total_plant_crates     = getTotalCrates(user.id, year, month, request.GET['crate_type'], 'plant')
        else:
            user.total_plant_crates     = 0      
        total_plant_crate.append(user.total_plant_crates)
        
        user.short_for_month        =  int(user.total_plant_crates)-int(user.total_dispatch_crates)
        user.short_since_starting   =  int(user.opening_crates)+int(user.short_for_month)

        total_short_for_month.append(user.short_for_month)
        total_short_since_starting.append(user.short_since_starting)

        crate_lists = []
        for months in month_list:
            month_lists               = {} 
            month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['dispatch']   = getDispatchedCratesSum(user.id, month_date, request.GET['crate_type'])
            month_lists['plant']      = getReceivedCratesSum(user.id, month_date, request.GET['crate_type'])
            crate_lists.append(month_lists)    
        user.crates =  crate_lists

    total_crate_lists = []
    for months in month_list:
        total_month_lists               = {} 
        month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
        dispatch                  = getTotalDispatchedCratesSum(month_date, request.GET['crate_type'])
        plant                     = getTotalReceivedCratesSum(month_date, request.GET['crate_type'])

        total_month_lists['dispatch']   = dispatch
        total_month_lists['plant']      = plant

        total_crate_lists.append(total_month_lists)    
 
    context = {}
    context['user_list']                    = user_list
    context['today_date']                   = today.strftime("%m/%Y")
    context['month_list']                   = days_in_months(year,month)
    context['month_list_count']             = len(month_list)+len(month_list)+8
    context['total_opening_crates']         = sum(total_opening_crates)
    context['total_dispatch_crate']         = sum(total_dispatch_crate)
    context['total_plant_crate']            = sum(total_plant_crate)
    context['total_short_for_month']        = sum(total_short_for_month)
    context['total_short_since_starting']   = sum(total_short_since_starting)
    context['total_crate_lists']            = total_crate_lists

    template = 'crate/ajax-crate-summary.html'
    return render(request, template, context)    

# export view
@login_required
def exportCrateSummary(request, crate_date, crate_type, route_id):
    today   = date.today()  
    # crates_details = SpUserCrates.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).order_by('id')
    if crate_date:
        crate_date = crate_date.split('-')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)
           
    condition = ''
    condition += ' and sp_users.user_type = %s' % 2
    condition += ' and sp_users.status = %s' % 1
    if route_id!='0':
        condition += ' and sp_user_area_allocations.route_id = %s' % int(route_id)
        
    user_list = SpUsers.objects.raw(''' SELECT sp_users.* FROM sp_users left join sp_user_area_allocations on sp_user_area_allocations.user_id = sp_users.id WHERE 1 {condition} order by sp_users.first_name asc '''.format(condition=condition))
    
    month_list = days_in_months(year,month)
    
    total_opening_crates        = []
    total_dispatch_crate        = []
    total_plant_crate           = []
    total_short_for_month       = []
    total_short_since_starting  = []
    for user in user_list:
        previous_month_last_date = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        if SpPlantCrateLedger.objects.filter(user_id=user.id,updated_datetime__lt=previous_month_last_date).exists():
            opening_crates              = getLastCrateBalance(user.id, previous_month_last_date, crate_type)
            user.opening_crates         = opening_crates
        else:
            user.opening_crates         = 0
        total_opening_crates.append(user.opening_crates)

        if getTotalCrates(user.id, year, month, crate_type, 'dispatch'):
            user.total_dispatch_crates  = getTotalCrates(user.id, year, month, crate_type, 'dispatch')
        else:
            user.total_dispatch_crates  = 0
        total_dispatch_crate.append(user.total_dispatch_crates)

        if getTotalCrates(user.id, year, month, crate_type, 'plant'):        
            user.total_plant_crates     = getTotalCrates(user.id, year, month, crate_type, 'plant')
        else:
            user.total_plant_crates     = 0      
        total_plant_crate.append(user.total_plant_crates)

        user.short_for_month        =  int(user.total_plant_crates)-int(user.total_dispatch_crates)
        user.short_since_starting   =  int(user.opening_crates)+int(user.short_for_month)

        total_short_for_month.append(user.short_for_month)
        total_short_since_starting.append(user.short_since_starting)

        crate_lists = []
        for months in month_list:
            month_lists               = {} 
            month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['dispatch']   = getDispatchedCratesSum(user.id, month_date, crate_type)
            month_lists['plant']      = getReceivedCratesSum(user.id, month_date, crate_type)
            crate_lists.append(month_lists)    
        user.crates =  crate_lists

    total_crate_lists = []
    for months in month_list:
        total_month_lists               = {} 
        month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
        dispatch                  = getTotalDispatchedCratesSum(month_date, crate_type)
        plant                     = getTotalReceivedCratesSum(month_date, crate_type)

        total_month_lists['dispatch']   = dispatch
        total_month_lists['plant']      = plant
        total_crate_lists.append(total_month_lists)

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=crate-summary-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'CRATE SUMMARY'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = len(month_list)+len(month_list)+7
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'Shreedhi'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=24, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    # color_codes = []
    columns = []
    columns += [ 'Customer Code' ]
    columns += [ 'Name of Distributor/SS' ]
    columns += [ 'Opening' ]
    if month_list:
        for months in month_list:
            columns += [ months ]
            columns += [ months ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    row_num = 2
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        if col_num >3:
            cell.font = Font(size=9, color='000000', bold=True)
        else:                  
            cell.font = Font(size=11, color='000000', bold=True) 
            
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num  == 3:
            column_dimensions.width = 10
        elif col_num >3:
            column_dimensions.width = 10
        else:
            column_dimensions.width = 25

    columns = []
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    if month_list:
        for months in month_list:
            columns += [ 'Dispatch' ]
            columns += [ 'Plant' ]
    columns += [ 'Plant' ]
    columns += [ 'Dispatch' ]
    columns += [ 'Short for Month' ]
    columns += [ 'Short Since starting' ]
    row_num = 3
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True)              
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num  == 3:
            column_dimensions.width = 10
        elif col_num >3:
            column_dimensions.width = 10
        else:
            column_dimensions.width = 25

    for id, user in enumerate(user_list):
        row_num += 1
        # Define the data for each cell in the row    
        row = []
        row += [ user.emp_sap_id ]
        row += [ user.store_name +"("+user.first_name +" "+user.middle_name+" "+user.last_name+" )"  ]
        row += [ user.opening_crates ]
        if user.crates:
            for crate in user.crates:
                row += [ crate['dispatch'] ]
                row += [ crate['plant'] ]
        row += [ user.total_plant_crates ]        
        row += [ user.total_dispatch_crates ]
        row += [ user.short_for_month ]
        row += [ user.short_since_starting ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            cell.fill = PatternFill()
            cell.font = Font(size=12, color='000000')
    

    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ ' ' ]
    bottom_columns += [ 'Total Crates Per Day' ]
    bottom_columns += [ sum(total_opening_crates) ]

    if total_crate_lists:
        for crate_list in total_crate_lists:
            bottom_columns += [ crate_list['dispatch'] ]
            bottom_columns += [ crate_list['plant'] ]

    bottom_columns += [ sum(total_plant_crate) ]
    bottom_columns += [ sum(total_dispatch_crate) ]
    bottom_columns += [ sum(total_short_for_month) ]
    bottom_columns += [ sum(total_short_since_starting) ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True)              
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if col_num  == 3:
            column_dimensions.width = 10
        elif col_num >3:
            column_dimensions.width = 10
        else:
            column_dimensions.width = 25

    row_num += 1
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response


#user crate summary
@login_required
def userCrateSummary(request):
    today   = date.today()  
    year  = today.year
    month = today.month

    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    total_dispatch                  = []
    total_return_to_plant           = []
    total_plant                     = []
    total_difference_with_vehicle   = []
    total_crate_difference          = []
    if user:
        previous_month_last_date    = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        opening_crates              = getLastCrateBalance(user['id'], previous_month_last_date, 'normal')

        opening_crates_list = [opening_crates]
        for id, months in enumerate(month_list):
            month_lists                         = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['dispatch']                 = getDispatchedCratesSum(user['id'], month_date, 'normal')
            month_lists['return_to_plant']          = getReturnCratesSum(user['id'], month_date, 'normal')
            month_lists['plant']                    = getReceivedCratesSum(user['id'], month_date, 'normal')
            
            month_lists['difference_with_vehicle']   = int(month_lists['plant'])-int(month_lists['return_to_plant'])
            month_lists['crate_difference']          = int(opening_crates_list[id])-(int(month_lists['dispatch'])+int(month_lists['plant']))
            opening_crates_list.append(month_lists['crate_difference'])
            crate_lists.append(month_lists)

            total_dispatch.append(month_lists['dispatch'])
            total_return_to_plant.append(month_lists['return_to_plant'])
            total_plant.append(month_lists['plant'])
            total_difference_with_vehicle.append(month_lists['difference_with_vehicle'])
            total_crate_difference.append(month_lists['crate_difference'])

    else:
        opening_crates              = 0            

    if len(crate_lists) > 0:
        last_opening_balance = crate_lists[-1]
        last_opening_balance = last_opening_balance['crate_difference']
    else:
        last_opening_balance = 0
           
    context = {}
    context['user_list']                        = user_list
    context['crate_lists']                      = crate_lists
    context['opening_crates']                   = opening_crates
    context['total_dispatch']                   = sum(total_dispatch)
    context['total_return_to_plant']            = sum(total_return_to_plant)
    context['total_plant']                      = sum(total_plant)
    context['total_difference_with_vehicle']    = sum(total_difference_with_vehicle)
    context['total_crate_difference']           = sum(total_crate_difference)
    context['last_opening_balance']             = last_opening_balance
    context['today_date']                       = today.strftime("%m/%Y")
    context['month_list']                       = days_in_months(year,month)
    context['page_title']                       = "User Crate Summary"

    template = 'crate/user-crate-summary.html'
    return render(request, template, context)

# ajax List View
@login_required
def ajaxUserCrateSummary(request):
    today   = date.today()  
    if request.GET['crate_date']:
        crate_date = request.GET['crate_date']
        crate_date = crate_date.split('/')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)
           
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(id=request.GET['user_id'],user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    total_dispatch                  = []
    total_return_to_plant           = []
    total_plant                     = []
    total_difference_with_vehicle   = []
    total_crate_difference          = []
    if user:
        previous_month_last_date    = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        opening_crates              = getLastCrateBalance(request.GET['user_id'], previous_month_last_date, request.GET['crate_type'])

        opening_crates_list = [opening_crates]
        for id, months in enumerate(month_list):
            month_lists                             = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['dispatch']                 = getDispatchedCratesSum(request.GET['user_id'], month_date, request.GET['crate_type'])
            month_lists['return_to_plant']          = getReturnCratesSum(request.GET['user_id'], month_date, request.GET['crate_type'])
            month_lists['plant']                    = getReceivedCratesSum(request.GET['user_id'], month_date, request.GET['crate_type'])
            
            month_lists['difference_with_vehicle']   = int(month_lists['plant'])-int(month_lists['return_to_plant'])
            month_lists['crate_difference']          = int(opening_crates_list[id])-(int(month_lists['dispatch'])+int(month_lists['plant']))
            opening_crates_list.append(month_lists['crate_difference'])
            crate_lists.append(month_lists)

            total_dispatch.append(month_lists['dispatch'])
            total_return_to_plant.append(month_lists['return_to_plant'])
            total_plant.append(month_lists['plant'])
            total_difference_with_vehicle.append(month_lists['difference_with_vehicle'])
            total_crate_difference.append(month_lists['crate_difference'])

    else:
        opening_crates              = 0    
    
    if len(crate_lists) > 0:
        last_opening_balance = crate_lists[-1]
        last_opening_balance = last_opening_balance['crate_difference']
    else:
        last_opening_balance = 0

    context = {}
    context['user_list']                        = user_list
    context['crate_lists']                      = crate_lists
    context['opening_crates']                   = opening_crates
    context['total_dispatch']                   = sum(total_dispatch)
    context['total_return_to_plant']            = sum(total_return_to_plant)
    context['total_plant']                      = sum(total_plant)
    context['total_difference_with_vehicle']    = sum(total_difference_with_vehicle)
    context['total_crate_difference']           = sum(total_crate_difference)
    context['last_opening_balance']             = last_opening_balance
    context['today_date']                       = today.strftime("%m/%Y")
    context['month_list']                       = days_in_months(year,month)

    template = 'crate/ajax-user-crate-summary.html'
    return render(request, template, context) 


# export view
@login_required
def exportUserCrateSummary(request, crate_date, crate_type, user_id):
    today   = date.today()  
    # crates_details = SpUserCrates.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).order_by('id')
    if crate_date:
        crate_date = crate_date.split('-')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month)> 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)

    user_id = int(user_id)       
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(id=user_id, user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name', 'emp_sap_id', 'store_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    total_dispatch                  = []
    total_return_to_plant           = []
    total_plant                     = []
    total_difference_with_vehicle   = []
    total_crate_difference          = []
    if user:
        user_town                   = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user_id, 'town_name')
        previous_month_last_date    = datetime.strptime(str(month_list[0]), '%d/%m/%Y').strftime('%Y-%m-%d')
        opening_crates              = getLastCrateBalance(user_id, previous_month_last_date, crate_type)

        opening_crates_list = [opening_crates]
        for id, months in enumerate(month_list):
            month_lists                         = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['dispatch']                 = getDispatchedCratesSum(user_id, month_date, crate_type)
            month_lists['return_to_plant']          = getReturnCratesSum(user_id, month_date, crate_type)
            month_lists['plant']                    = getReceivedCratesSum(user_id, month_date, crate_type)
            
            month_lists['difference_with_vehicle']   = int(month_lists['plant'])-int(month_lists['return_to_plant'])
            month_lists['crate_difference']          = int(opening_crates_list[id])-(int(month_lists['dispatch'])+int(month_lists['plant']))
            opening_crates_list.append(month_lists['crate_difference'])
            crate_lists.append(month_lists)

            total_dispatch.append(month_lists['dispatch'])
            total_return_to_plant.append(month_lists['return_to_plant'])
            total_plant.append(month_lists['plant'])
            total_difference_with_vehicle.append(month_lists['difference_with_vehicle'])
            total_crate_difference.append(month_lists['crate_difference'])
    else:
        opening_crates              = 0

    if len(crate_lists) > 0:
        last_opening_balance = crate_lists[-1]
        last_opening_balance = last_opening_balance['crate_difference']
    else:
        last_opening_balance = 0

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=user-crate-summary-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'USER CRATE SUMMARY'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    # cell = worksheet.cell(row=1, column=2)  
    # cell.value = str(year)+'/'+str(month)
    # cell.font = header_font
    # cell.alignment = wrapped_alignment
    # cell.border = black_border
    # cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    # cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = 6
    
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'Shreedhi'
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(size=15, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    # color_codes = []
    columns = []
    columns += [ 'Vendor Name' ]
    columns += [ str(user['first_name'])+' '+str(user['first_name'])+' '+str(user['first_name'])+'/'+str(user['store_name']) ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'Location' ]
    columns += [ user_town ]
    row_num = 2
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        if col_num == 2 or col_num == 6:
            cell.font = Font(size=11, color='000000')
        else:
            cell.font = Font(size=11, color='000000', bold=True)  
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    month_name = datetime(year,month,1).strftime( '%B' )
    
    # color_codes = []
    columns = []
    columns += [ 'Vendor Code' ]
    columns += [ user['emp_sap_id'] ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'Month' ]
    columns += [ str(month_name)+'-'+str(year) ]
    row_num = 3
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        if col_num == 2 or col_num == 6:
            cell.font = Font(size=11, color='000000')
        else:
            cell.font = Font(size=11, color='000000', bold=True)     
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # color_codes = []
    columns = []
    columns += [ 'Date' ]
    columns += [ 'Dispatched from Plant' ]
    columns += [ 'Return to Vehicle' ]
    columns += [ 'Deposit in Plant' ]
    columns += [ 'Difference with Vehicle' ]
    columns += [ 'Crate Difference' ]
    row_num = 4
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20


    # color_codes = []
    columns = []
    columns += [ 'Opening' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ opening_crates ]
    row_num += 1
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for id, crate in enumerate(crate_lists):
        row_num += 1
        # Define the data for each cell in the row    
        row = []
        row += [ crate['month_date'] ]
        row += [ crate['dispatch'] ]
        row += [ crate['return_to_plant'] ]
        row += [ crate['plant'] ]
        row += [ crate['difference_with_vehicle'] ]
        row += [ crate['crate_difference'] ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            cell.fill = PatternFill()
            cell.font = Font(size=12, color='000000')

    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ 'Total' ]
    bottom_columns += [ sum(total_dispatch) ]
    bottom_columns += [ sum(total_return_to_plant) ]
    bottom_columns += [ sum(total_plant) ]
    bottom_columns += [ sum(total_difference_with_vehicle) ]
    bottom_columns += [ last_opening_balance ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    row_num += 1
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=9, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response
    

def addCrate(request):
    today   = date.today() 
    user_list = [] 
    user = SpUsers.objects.filter(user_type=2, status=1)
    context = {}
    context['user_list']                    =user
    context['page_title']                    ="Crate Management"
    context['today_date']                   = today.strftime("%d/%m/%Y")
    template = 'crate/add-crates.html'
    return render(request,template, context)

@login_required
@transaction.atomic
def saveCreatesDetails(request):
    today                   = date.today() 
    response                = {}
    error_response          = {}
    if request.method == "POST":
        try:
            user_id                                     = request.POST.get('user_id')
            receive_check                               = request.POST.get('receive_check')
            dispatch_check                              = request.POST.get('dispatch_check')
            crate_date                                  = today.strftime("%Y-%m-%d")
            rnormal_crate                               = request.POST.get('rnormal_crate')
            rjumbo_crate                                = request.POST.get('rjumbo_crate') 
            dnormal_crate                               = request.POST.get('dnormal_crate')
            djumbo_crate                                = request.POST.get('djumbo_crate') 
            if receive_check!='0':  
                receive                                 = SpUserCrateLedger.objects.filter(user_id=user_id).order_by('-id').first()
                crate_recive                            = SpUserCrateLedger()
                crate_recive.user_id                    = user_id
                crate_recive.normal_credit              = int(rnormal_crate)                        
                crate_recive.normal_debit               = 0
                if receive:
                    crate_recive.normal_balance         = (receive.normal_balance)-int(rnormal_crate) 
                else:
                    crate_recive.normal_balance         = -int(rnormal_crate)
                crate_recive.jumbo_credit               = int(rjumbo_crate)
                crate_recive.jumbo_debit                = 0
                if receive:
                    crate_recive.jumbo_balance          = (receive.jumbo_balance)-int(rjumbo_crate) 
                else:
                    crate_recive.jumbo_balance          = -int(rjumbo_crate)
                crate_recive.updated_datetime           = crate_date
                crate_recive.updated_at                 = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                crate_recive.save()
                response['error']                       = False
                response['message']                     = "Record has been saved successfully."
 
            if dispatch_check!='0':
                dispatch_bal                            = SpUserCrateLedger.objects.filter(user_id=user_id).order_by('-id').first()  
                crate_dispatch                          = SpUserCrateLedger()
                crate_dispatch.user_id                  = user_id
                crate_dispatch.normal_credit            = 0
                crate_dispatch.normal_debit             = int(dnormal_crate)
                if dispatch_bal:
                    crate_dispatch.normal_balance       = (dispatch_bal.normal_balance)+int(dnormal_crate)
                else:
                    crate_dispatch.normal_balance       = int(dnormal_crate)               
                crate_dispatch.jumbo_credit             = 0
                crate_dispatch.jumbo_debit              = int(djumbo_crate)
                if dispatch_bal:
                    crate_dispatch.jumbo_balance        = (dispatch_bal.jumbo_balance)+int(djumbo_crate)
                else:
                    crate_dispatch.jumbo_balance        = int(djumbo_crate)
                crate_dispatch.updated_datetime         = crate_date
                crate_dispatch.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                crate_dispatch.save()
                response['error']                       = False
                response['message']                     = "Record has been saved successfully."                         
        except Exception as e:
            response['error'] = True
            response['message'] = str(e)
            error_response['message'] = str(e)
        return JsonResponse(response)
    return redirect('/add-crate')

@login_required
@transaction.atomic
def saveCrates(request):
    today   = date.today() 
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            user_id = request.POST.get('user_id')
            receive_check = request.POST.get('receive_check')
            dispatch_check = request.POST.get('dispatch_check')
            crate_date = today.strftime("%Y-%m-%d")
            rnormal_crate = request.POST.get('rnormal_crate')
            rjumbo_crate = request.POST.get('rjumbo_crate')  
            dnormal_crate = request.POST.get('dnormal_crate')
            djumbo_crate = request.POST.get('djumbo_crate') 
            error_count = 0 
            if rnormal_crate=="0" or  dnormal_crate=="0":
                error_count = 1
                error_response = "Please select a valid crate number."
            userlst = SpUserCrateLedger.objects.filter(user_id=user_id,normal_balance__gte=0).order_by('-id').exclude(updated_datetime__icontains=crate_date).first()
            if user_id =='':
                error_count = 1
                error_response = "Please select user"
            else:
                with transaction.atomic():
                    if receive_check!='0':
                        if rnormal_crate == "" and rjumbo_crate == "":
                            error_count = 1
                            error_response = "Please fill atleast normal or jumbo in receive crate."
                    if dispatch_check!='0':
                        if dnormal_crate == "" and djumbo_crate == "":
                            error_count = 1
                            error_response = "Please fill atleast normal or jumbo in dispatch crate."
                       
                    if(error_count > 0):
                        response['error'] = True
                        response['message'] = error_response
                    else:
                        current_crate=SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date)
                        if current_crate:
                            if len(current_crate)==2:
                                receive = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).values_list('id',flat=True)
                                dispatch = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).values_list('id',flat=True)
                                if dispatch_check!="0" and receive_check=="0":
                                    SpUserCrateLedger.objects.filter(id=dispatch[0]).delete()
                                    if dnormal_crate!=0 and djumbo_crate=="":
                                        receive                     = SpUserCrateLedger.objects.get(id = receive[0])
                                        temp                        =receive.normal_credit
                                        receive.delete()
                                        # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).delete()
                                        usercrate                   = SpUserCrateLedger()
                                        usercrate.user_id           = user_id
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = dnormal_crate 
                                        usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = 0
                                        usercrate.jumbo_balance     = userlst.jumbo_balance
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        
                                        usercraterecive                  = SpUserCrateLedger()
                                        usercraterecive.user_id          =user_id
                                        usercraterecive.normal_credit    =temp
                                        usercraterecive.normal_debit     =0
                                        if userlst:
                                            usercraterecive.normal_balance   =userlst.normal_balance+int(dnormal_crate)-temp
                                        else:
                                            usercraterecive.normal_balance   =int(dnormal_crate)-temp
                                        usercraterecive.jumbo_credit     =0
                                        usercraterecive.jumbo_debit      =0
                                        if userlst:
                                            usercraterecive.jumbo_balance    =userlst.jumbo_balance
                                        else:
                                            usercraterecive.jumbo_balance    =0
                                        usercraterecive.updated_datetime  = crate_date
                                        usercraterecive.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercraterecive.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    elif dnormal_crate=="" and djumbo_crate!="0":
                                        receive                     = SpUserCrateLedger.objects.get(id = receive[0])
                                        usercrate                   = SpUserCrateLedger()
                                        usercrate.user_id           =  user_id
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = 0  
                                        usercrate.normal_balance    = receive.normal_balance
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = djumbo_crate
                                        usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    else:
                                        receive                     = SpUserCrateLedger.objects.get(id = receive[0])
                                        usercrate                   = SpUserCrateLedger()
                                        usercrate.user_id           =  user_id
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = dnormal_crate  
                                        usercrate.normal_balance    = receive.normal_balance+int(dnormal_crate) 
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = djumbo_crate
                                        usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                elif receive_check!="0" and dispatch_check=="0":
                                    SpUserCrateLedger.objects.filter(id=receive[0]).delete()
                                    if rnormal_crate!=0 and rjumbo_crate=="":
                                        dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
                                        temp                            = dispatch.normal_debit
                                        dispatch.delete()
                                        # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).delete()
                                        usercrate                       = SpUserCrateLedger()
                                        usercrate.user_id               = user_id
                                        # if dispatch.normal_balance != 0:
                                        usercrate.normal_credit     = int(rnormal_crate) 
                                        usercrate.normal_debit      = 0
                                        usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = 0
                                        if userlst:
                                            usercrate.jumbo_balance     = userlst.jumbo_balance
                                        else:
                                            usercrate.jumbo_balance     = 0
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        usercratedispatch           =  SpUserCrateLedger()
                                        usercratedispatch.user_id           = user_id
                                        usercratedispatch.normal_credit     = 0
                                        usercratedispatch.normal_debit      = temp
                                        if userlst:
                                            usercratedispatch.normal_balance    = userlst.normal_balance-int(rnormal_crate)+temp
                                        else:
                                            usercratedispatch.normal_balance    = int(rnormal_crate)+temp
                                        usercratedispatch.jumbo_credit      = 0
                                        usercratedispatch.jumbo_debit       = 0
                                        if userlst:
                                            usercratedispatch.jumbo_balance     = userlst.jumbo_balance
                                        else:
                                            usercratedispatch.jumbo_balance     = 0
                                        usercratedispatch.updated_datetime  = crate_date
                                        usercratedispatch.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercratedispatch.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    # else:
                                        #     response['error'] = True
                                        #     response['message'] = "No opening balace of this create type."
                                        
                                    elif rjumbo_crate!=0 and rnormal_crate=="":
                                        dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
                                        usercrate                       = SpUserCrateLedger()
                                        usercrate.user_id               =  user_id
                                        if dispatch.jumbo_balance != 0:
                                            usercrate.normal_credit     = 0
                                            usercrate.normal_debit      = 0
                                            usercrate.normal_balance    = dispatch.normal_balance
                                            usercrate.jumbo_credit      = int(rjumbo_crate)
                                            usercrate.jumbo_debit       = 0
                                            usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
                                            usercrate.updated_datetime  = crate_date
                                            usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                        else:
                                            response['error'] = True
                                            response['message'] = "No opening balace of this create type."
                                    else:
                                        dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
                                        if dispatch.normal_balance != 0 and dispatch.jumbo_balance:
                                            usercrate                   = SpUserCrateLedger()
                                            usercrate.user_id           =  user_id
                                            usercrate.normal_credit     = int(rnormal_crate)
                                            usercrate.normal_debit      = 0
                                            usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
                                            usercrate.jumbo_credit      = int(rjumbo_crate)
                                            usercrate.jumbo_debit       = 0
                                            usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
                                            usercrate.updated_datetime  = crate_date
                                            usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                        else:
                                            response['error'] = True
                                            response['message'] = "No opening balace of this create type."
                                else:
                                    SpUserCrateLedger.objects.filter(id=dispatch[0]).delete()
                                    # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).delete()
                                    usercrate                          = SpUserCrateLedger()
                                    if dnormal_crate!='':
                                        usercrate.user_id              = user_id
                                        usercrate.normal_credit        = 0
                                        usercrate.normal_debit         = int(dnormal_crate)
                                        if userlst:  
                                            usercrate.normal_balance   = userlst.normal_balance+int(dnormal_crate) 
                                        else:
                                            usercrate.normal_balance   = int(dnormal_crate)
                                    else:
                                        usercrate.user_id              = user_id
                                        usercrate.normal_debit         = 0
                                        usercrate.normal_credit        = 0
                                        if userlst:  
                                            usercrate.normal_balance   = userlst.normal_balance
                                        else:
                                            usercrate.normal_balance   = 0
                                
                                    if djumbo_crate!="":
                                        usercrate.jumbo_credit             = int(dnormal_crate)
                                        usercrate.jumbo_debit              = 0 
                                        if userlst:
                                            usercrate.jumbo_balance    = userlst.jumbo_balance+int(djumbo_crate)
                                        else:
                                            usercrate.jumbo_balance    = int(djumbo_crate)
                                    else:
                                        usercrate.jumbo_debit               = 0
                                        usercrate.jumbo_credit             = 0
                                        if userlst:
                                            usercrate.jumbo_balance    = userlst.jumbo_balance
                                        else:
                                            usercrate.jumbo_balance    = 0
                                    usercrate.updated_datetime         = crate_date
                                    usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    usercrate.save()
                                    SpUserCrateLedger.objects.filter(id=receive[0]).delete()
                                    # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).delete()
                                    receive                       = SpUserCrateLedger()
                                    if rnormal_crate!="":
                                        receive.user_id           = user_id
                                        receive.normal_credit     = int(rnormal_crate)
                                        receive.normal_debit      = 0
                                        if userlst:
                                            receive.normal_balance    = int(userlst.normal_balance)+int(dnormal_crate)-int(rnormal_crate)
                                        else:
                                            receive.normal_balance    = int(dnormal_crate)-int(rnormal_crate)
                                    else:
                                        receive.normal_credit     = 0
                                        receive.normal_debit      = 0
                                        receive.normal_balance    = userlst.normal_balance
                                    
                                    if rjumbo_crate!="":
                                        receive.jumbo_credit      = int(rjumbo_crate)
                                        receive.jumbo_debit       = 0
                                        if userlst:
                                            receive.jumbo_balance     = int(userlst.jumbo_balance)+int(djumbo_crate)-int(rjumbo_crate)
                                    else:
                                        receive.jumbo_credit      = 0
                                        receive.jumbo_debit       = 0
                                        if userlst:
                                            receive.jumbo_balance    = userlst.jumbo_balance
                                        else:
                                            receive.jumbo_balance    = 0
                                    
                                    receive.updated_datetime      = crate_date
                                    usercrate.updated_at          = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    receive.save()
                                    response['error'] = False
                                    response['message'] = "Record has been saved successfully."
                                    
                            if len(current_crate)==1:
                                receive = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).values_list('id',flat=True)
                                dispatch = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).values_list('id',flat=True)
                                count = 0
                                if receive:
                                    if dispatch_check!="0" and receive_check=="0":
                                        if dnormal_crate!=0 and djumbo_crate=="":
                                            receive                     = SpUserCrateLedger.objects.get(id = receive[0])
                                            usercrate                   = SpUserCrateLedger()
                                            usercrate.user_id           =   user_id
                                            usercrate.normal_credit     = 0
                                            usercrate.normal_debit      = dnormal_crate 
                                            usercrate.normal_balance    = receive.normal_balance+int(dnormal_crate) 
                                            usercrate.jumbo_credit      = 0
                                            usercrate.jumbo_debit       = 0
                                            usercrate.jumbo_balance     = 0
                                            usercrate.updated_datetime  = crate_date
                                            usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                        elif dnormal_crate=="" and djumbo_crate!="0":
                                            receive                     = SpUserCrateLedger.objects.get(id = receive[0])
                                            usercrate                   = SpUserCrateLedger()
                                            usercrate.user_id           =   user_id
                                            usercrate.normal_credit     = 0
                                            usercrate.normal_debit      = 0  
                                            usercrate.normal_balance    = receive.normal_balance
                                            usercrate.jumbo_credit      = 0
                                            usercrate.jumbo_debit       = djumbo_crate
                                            usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
                                            usercrate.updated_datetime  = crate_date
                                            usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                        else:
                                            receive                            = SpUserCrateLedger.objects.get(id = receive[0])
                                            usercrate                          = SpUserCrateLedger()
                                            usercrate.normal_credit            = 0
                                            usercrate.user_id                  =   user_id
                                            usercrate.normal_debit             = dnormal_crate  
                                            usercrate.normal_balance           = receive.normal_balance+int(dnormal_crate) 
                                            usercrate.jumbo_credit             = 0
                                            usercrate.jumbo_debit              = djumbo_crate
                                            usercrate.jumbo_balance            = receive.jumbo_balance+int(djumbo_crate)
                                            usercrate.updated_datetime         = crate_date
                                            usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                    elif receive_check!="0" and dispatch_check=="0":
                                        if rnormal_crate!=0 and rjumbo_crate=="":
                                            usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
                                            if userlst.normal_balance !=0:
                                                usercrate.normal_credit            = int(rnormal_crate) 
                                                usercrate.normal_debit             = 0
                                                if userlst:
                                                    usercrate.normal_balance       = userlst.normal_balance-int(rnormal_crate)
                                                else:
                                                    usercrate.normal_balance       = int(rnormal_crate)
                                                usercrate.jumbo_credit             = 0
                                                usercrate.jumbo_debit              = 0
                                                if userlst:
                                                    usercrate.jumbo_balance        = userlst.jumbo_balance
                                                else:
                                                    usercrate.jumbo_balance        = 0 
                                                usercrate.updated_datetime         = crate_date
                                                usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                                usercrate.save()
                                                response['error'] = False
                                                response['message'] = "Record has been saved successfully."
                                            else:
                                                response['error'] = True
                                                response['message'] = "No opening balance of this create type.."
                                        elif rjumbo_crate!=0 and rnormal_crate=="":
                                            usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
                                            if userlst.jumbo_balance!=0:
                                                usercrate.normal_credit            = 0
                                                usercrate.normal_debit             = 0
                                                if userlst:
                                                    usercrate.normal_balance       = userlst.normal_balance
                                                else:
                                                    usercrate.normal_balance       = 0 
                                                usercrate.jumbo_credit             = int(rjumbo_crate)
                                                usercrate.jumbo_debit              = 0
                                                if userlst:
                                                    usercrate.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
                                                else:
                                                    usercrate.jumbo_balance        = int(rjumbo_crate)
                                                usercrate.updated_datetime         = crate_date
                                                usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                                usercrate.save()
                                                response['error'] = False
                                                response['message'] = "Record has been saved successfully."
                                            else:
                                                response['error'] = True
                                                response['message'] = "No opening balance of this create type.."
                                        else:
                                            usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
                                            if userlst.jumbo_balance!=0 and userlst.normal_balance!=0: 
                                                usercrate.normal_credit            = int(rnormal_crate)
                                                usercrate.normal_debit             = 0
                                                if userlst:
                                                    usercrate.normal_balance       = userlst.normal_balance-int(rnormal_crate)
                                                else:
                                                    usercrate.normal_balance       = int(rnormal_crate)
                                                usercrate.jumbo_credit             = int(rjumbo_crate)
                                                usercrate.jumbo_debit              = 0
                                                if userlst:
                                                    usercrate.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
                                                else:
                                                    usercrate.jumbo_balance        = int(rjumbo_crate)
                                                usercrate.updated_datetime         = crate_date
                                                usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                                usercrate.save()
                                                response['error'] = False
                                                response['message'] = "Record has been saved successfully."
                                            else:
                                                response['error'] = True
                                                response['message'] = "No opening balance of this create type.."
                                            
                                    else:
                                        usercrate1                              = SpUserCrateLedger.objects.get(id = receive[0])
                                        usercrate1.normal_debit                 = 0
                                        if rnormal_crate!='':
                                            usercrate1.normal_credit            = int(rnormal_crate)
                                            if userlst:
                                                usercrate1.normal_balance       = userlst.normal_balance-int(rnormal_crate)
                                            else:
                                                usercrate1.normal_balance       =  int(rnormal_crate)
                                        else:
                                            usercrate1.normal_credit            = 0
                                            if userlst:
                                                usercrate1.normal_balance       = userlst.normal_balance
                                            else:
                                                usercrate1.normal_balance       =  0
                                        
                                        if rjumbo_crate!='':
                                            usercrate1.jumbo_credit             = int(rjumbo_crate)
                                            usercrate1.jumbo_debit              = 0
                                            if userlst:
                                                usercrate1.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
                                            else:
                                                usercrate1.jumbo_balance        = int(rjumbo_crate)
                                        else:
                                            usercrate1.jumbo_credit             = 0
                                            usercrate1.jumbo_debit              = 0
                                            if userlst:
                                                usercrate1.jumbo_balance        = userlst.jumbo_balance
                                            else:
                                                usercrate1.jumbo_balance        = 0
                                        usercrate1.updated_datetime             = crate_date
                                        usercrate1.updated_at                   = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate1.save()
    
                                        receive                                 = SpUserCrateLedger.objects.get(id = receive[0])
                                        usercrate                               = SpUserCrateLedger()
                                        usercrate.user_id                       =   user_id
                                        if dnormal_crate!='':
                                            usercrate.normal_credit             = 0
                                            usercrate.normal_debit              = dnormal_crate  
                                            usercrate.normal_balance            = userlst.normal_balance+int(dnormal_crate) 
                                        else:
                                            usercrate.normal_credit             = 0
                                            usercrate.normal_debit              = 0
                                            usercrate.normal_balance            = userlst.normal_balance
    
                                        if rjumbo_crate!='':
                                            usercrate.jumbo_credit              = 0
                                            usercrate.jumbo_debit               = dnormal_crate
                                            usercrate.jumbo_balance             = userlst.jumbo_balance+int(dnormal_crate)
                                        else:
                                            usercrate.jumbo_credit              = 0
                                            usercrate.jumbo_debit               = 0
                                            usercrate.jumbo_balance             = userlst.jumbo_balance
                                        usercrate.updated_datetime              = crate_date
                                        usercrate.updated_at                    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        count+=1
                                        response['message'] = "Record has been saved successfully."
                                                                
                                if dispatch and count == 0:
                                    if dispatch_check!="0" and receive_check=="0":
                                        if dnormal_crate!=0 and djumbo_crate=="":
                                            usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
                                            usercrate.normal_credit            = 0
                                            usercrate.normal_debit             = dnormal_crate
                                            if userlst:
                                                usercrate.normal_balance       = userlst.normal_balance+int(dnormal_crate)
                                            else:
                                                usercrate.normal_balance       = int(dnormal_crate)
                                            usercrate.jumbo_credit             = 0
                                            usercrate.jumbo_debit              = 0
                                            if userlst:
                                                usercrate.jumbo_balance        = userlst.jumbo_balance
                                            else:
                                                usercrate.jumbo_balance        = 0
                                            usercrate.updated_datetime         = crate_date
                                            usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                        elif dnormal_crate=="" and djumbo_crate!="0":
                                            usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
                                            usercrate.normal_credit            = 0
                                            usercrate.normal_debit             = 0
                                            if userlst:
                                                usercrate.normal_balance       = userlst.normal_balance
                                            else:
                                                usercrate.normal_balance       = 0
                                            usercrate.jumbo_credit             = 0
                                            usercrate.jumbo_debit              = djumbo_crate
                                            if userlst:
                                                usercrate.jumbo_balance        = userlst.jumbo_balance+int(djumbo_crate)
                                            else:
                                                usercrate.jumbo_balance        = int(djumbo_crate)
                                            usercrate.updated_datetime         = crate_date
                                            usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                        else:
                                            usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
                                            usercrate.normal_credit            = 0
                                            usercrate.normal_debit             = dnormal_crate  
                                            if userlst:
                                                usercrate.normal_balance       = userlst.normal_balance+int(dnormal_crate) 
                                            else:
                                                usercrate.normal_balance       = int(dnormal_crate) 
                                            usercrate.jumbo_credit             = 0
                                            usercrate.jumbo_debit              = djumbo_crate
                                            if userlst:
                                                usercrate.jumbo_balance        = userlst.jumbo_balance+int(djumbo_crate)
                                            else:
                                                usercrate.jumbo_balance        = int(djumbo_crate)
                                            usercrate.updated_datetime         = crate_date
                                            usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                            usercrate.save()
                                            response['error'] = False
                                            response['message'] = "Record has been saved successfully."
                                    elif receive_check!="0" and dispatch_check=="0":
                                        if rnormal_crate!=0 and rjumbo_crate=="":
                                            dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
                                            
                                            if dispatch.normal_balance !=0 :
                                                usercrate                   =SpUserCrateLedger()
                                                usercrate.normal_credit     = int(rnormal_crate) 
                                                usercrate.user_id           = user_id 
                                                usercrate.normal_debit      = 0
                                                usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
                                                usercrate.jumbo_credit      = 0
                                                usercrate.jumbo_debit       = 0
                                                usercrate.jumbo_balance     = dispatch.jumbo_balance
                                                usercrate.updated_datetime  = crate_date
                                                usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                                usercrate.save()
                                                response['error'] = False
                                                response['message'] = "Record has been saved successfully."
                                            else:
                                                response['error'] = True
                                                response['message'] = "No opening balance of this create type.."
                                        elif rjumbo_crate!=0 and rnormal_crate=="":
                                            dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
                                            if dispatch.jumbo_balance!=0:
                                                usercrate                   =SpUserCrateLedger()
                                                usercrate.user_id           = user_id
                                                usercrate.normal_credit     = 0
                                                usercrate.normal_debit      = 0
                                                usercrate.normal_balance    = dispatch.normal_balance
                                                usercrate.jumbo_credit      = int(rjumbo_crate)
                                                usercrate.jumbo_debit       = 0
                                                usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
                                                usercrate.updated_datetime  = crate_date
                                                usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                                usercrate.save()
                                                response['error'] = False
                                                response['message'] = "Record has been saved successfully."
                                            else:
                                                response['error'] = True
                                                response['message'] = "No opening balance of this create type.."
                                        else:
                                            dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
                                            if dispatch.jumbo_balance!=0 and dispatch.normal_balance!=0: 
                                                usercrate                   = SpUserCrateLedger()
                                                usercrate.user_id           = user_id
                                                usercrate.normal_credit     = int(rnormal_crate)
                                                usercrate.normal_debit      = 0
                                                usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
                                                usercrate.jumbo_credit      = int(rjumbo_crate)
                                                usercrate.jumbo_debit       = 0
                                                usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
                                                usercrate.updated_datetime  = crate_date
                                                usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                                usercrate.save()
                                                response['error'] = False
                                                response['message'] = "Record has been saved successfully."
                                            else:
                                                response['error'] = True
                                                response['message'] = "No opening balance of this create type.."
                                    else:
                                        usercrate                               = SpUserCrateLedger.objects.get(id = dispatch[0])
                                        usercrate.normal_credit                 = 0
                                        usercrate.normal_debit                  = dnormal_crate
                                        if dnormal_crate !='':
                                            usercrate.normal_debit              = dnormal_crate
                                            if userlst:  
                                                usercrate.normal_balance        = userlst.normal_balance+int(dnormal_crate) 
                                            else:
                                                usercrate.normal_balance        = int(dnormal_crate)
                                        else:
                                            usercrate.normal_debit              = 0
                                            if userlst:  
                                                usercrate.normal_balance        = userlst.normal_balance
                                            else:
                                                usercrate.normal_balance        = 0
                                        usercrate.jumbo_credit                  = 0
                                        if djumbo_crate !='':
                                            usercrate.jumbo_debit               = djumbo_crate
                                            if userlst:
                                                usercrate.jumbo_balance         = userlst.jumbo_balance+int(djumbo_crate)
                                            else:
                                                usercrate.jumbo_balance         = djumbo_crate
                                        else:
                                            usercrate.jumbo_debit               = 0
                                            if userlst:
                                                usercrate.jumbo_balance         = userlst.jumbo_balance
                                            else:
                                                usercrate.jumbo_balance         = 0
                                        usercrate.updated_datetime              = crate_date
                                        usercrate.updated_at                    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        
    
                                        dispatch                     = SpUserCrateLedger.objects.get(id = dispatch[0])
                                        usercrate1                   = SpUserCrateLedger()
                                        usercrate1.user_id           = user_id
                                        if rnormal_crate!='':
                                            usercrate1.normal_credit     = int(rnormal_crate)
                                            usercrate1.normal_debit      = 0
                                            usercrate1.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
                                        else:
                                            usercrate1.normal_credit     = 0
                                            usercrate1.normal_debit      = 0
                                            usercrate1.normal_balance    = dispatch.normal_balance
                                        
                                        if rjumbo_crate!="":
                                            usercrate1.jumbo_credit      = int(rjumbo_crate)
                                            usercrate1.jumbo_debit       = 0
                                            usercrate1.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
                                        else:
                                            usercrate1.jumbo_credit      = 0
                                            usercrate1.jumbo_debit       = 0
                                            usercrate1.jumbo_balance     = dispatch.jumbo_balance
                                        usercrate1.updated_datetime      = crate_date
                                        usercrate1.updated_at            = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate1.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                        else:
                            usercrate                               =   SpUserCrateLedger()
                            usercrate.user_id                       =   user_id
                            if dispatch_check!="0" and receive_check=="0":
                                if dnormal_crate!=0 and djumbo_crate=="":
                                    if userlst:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = dnormal_crate 
                                        usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = 0
                                        usercrate.jumbo_balance     = 0
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    else:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = dnormal_crate 
                                        usercrate.normal_balance    = dnormal_crate
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = 0
                                        usercrate.jumbo_balance     = 0
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                elif dnormal_crate=="" and djumbo_crate!="0":
                                    if userlst:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = 0  
                                        usercrate.normal_balance    = userlst.normal_balance
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = djumbo_crate
                                        usercrate.jumbo_balance     = userlst.jumbo_balance+int(djumbo_crate)
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    else:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = 0  
                                        usercrate.normal_balance    = 0
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = djumbo_crate
                                        usercrate.jumbo_balance     = djumbo_crate
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                else:
                                    if userlst:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = dnormal_crate  
                                        usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = djumbo_crate
                                        usercrate.jumbo_balance     = userlst.jumbo_balance+int(djumbo_crate)
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    else:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      =dnormal_crate    
                                        usercrate.normal_balance    =dnormal_crate  
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = djumbo_crate
                                        usercrate.jumbo_balance     = djumbo_crate
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
    
                            elif receive_check!="0" and dispatch_check=="0" and userlst:
                                if rnormal_crate!=0 and rjumbo_crate=="":
                                    if userlst:
                                        usercrate.normal_credit     = int(rnormal_crate) 
                                        usercrate.normal_debit      = 0
                                        usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
                                        usercrate.jumbo_credit      = 0
                                        usercrate.jumbo_debit       = 0
                                        usercrate.jumbo_balance     = userlst.jumbo_balance
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    else:
                                        response['error'] = True
                                        response['message'] = "Party have no opening balance.. You can Dispatch only.."
                                elif rjumbo_crate!=0 and rnormal_crate=="":
                                    if userlst:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_debit      = 0
                                        usercrate.normal_balance    = userlst.normal_balance
                                        usercrate.jumbo_credit      = int(rjumbo_crate)
                                        usercrate.jumbo_debit       = 0
                                        usercrate.jumbo_balance     = userlst.jumbo_balance-int(rjumbo_crate)
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    else:
                                        response['error'] = True
                                        response['message'] = "Party have no opening balance.. You can Dispatch only.."
                                else:
                                    if userlst:
                                        usercrate.normal_credit     = int(rnormal_crate)
                                        usercrate.normal_debit      = 0
                                        usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
                                        usercrate.jumbo_credit      = int(rjumbo_crate)
                                        usercrate.jumbo_debit       = 0
                                        usercrate.jumbo_balance     = userlst.jumbo_balance-int(rjumbo_crate)
                                        usercrate.updated_datetime  = crate_date
                                        usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        usercrate.save()
                                        response['error'] = False
                                        response['message'] = "Record has been saved successfully."
                                    else:
                                        response['error'] = True
                                        response['message'] = "Party have no opening balance.. You can Dispatch only.."           
                            else:
                                if userlst:
                                    dispatch_usercrate     =SpUserCrateLedger()
                                    if dnormal_crate!='':
                                        dispatch_usercrate.normal_credit             = 0
                                        dispatch_usercrate.normal_debit              = dnormal_crate  
                                        dispatch_usercrate.normal_balance            = userlst.normal_balance+int(dnormal_crate) 
                                    else:
                                        dispatch_usercrate.normal_credit             = 0
                                        dispatch_usercrate.normal_debit              = 0
                                        dispatch_usercrate.normal_balance            = userlst.normal_balance
    
                                    if djumbo_crate!='':
                                        dispatch_usercrate.jumbo_credit              = 0
                                        dispatch_usercrate.jumbo_debit               = dnormal_crate
                                        dispatch_usercrate.jumbo_balance             = userlst.jumbo_balance+int(dnormal_crate)
                                    else:
                                        dispatch_usercrate.jumbo_credit              = 0
                                        dispatch_usercrate.jumbo_debit               = 0
                                        dispatch_usercrate.jumbo_balance             = userlst.jumbo_balance
                                    dispatch_usercrate.updated_datetime              = crate_date
                                    dispatch_usercrate.user_id                       =   user_id
                                    dispatch_usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    dispatch_usercrate.save()
                                    
                                    dispatch_obj = SpUserCrateLedger.objects.get(id=dispatch_usercrate.id)
                                    usercrate.normal_debit      = 0
                                    if rnormal_crate!='':
                                        usercrate.normal_credit     = int(rnormal_crate)
                                        usercrate.normal_balance    = dispatch_obj.normal_balance-int(rnormal_crate)
                                    else:
                                        usercrate.normal_credit     = 0
                                        usercrate.normal_balance    = dispatch_obj.normal_balance
                                    if rjumbo_crate!='':
                                        usercrate.jumbo_credit             = int(rjumbo_crate)
                                        usercrate.jumbo_debit              = 0
                                        usercrate.jumbo_balance        = dispatch_obj.jumbo_balance-int(rjumbo_crate)
                                    else:
                                        usercrate.jumbo_credit             = 0
                                        usercrate.jumbo_debit              = 0
                                        usercrate.jumbo_balance        = dispatch_obj.jumbo_balance
                                    usercrate.updated_datetime         = crate_date
                                    usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    usercrate.save()
                                    response['error'] = False
                                    response['message'] = "Record has been saved successfully." 
                                else:
                                    response['error'] = True
                                    response['message'] = "Party have no opening balance.. You can Dispatch only.."
            
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            print(e)
            response['error'] = True
            response['message'] = str(e)
        return JsonResponse(response)
    return redirect('/add-crate')

    
# @login_required
# def saveCrates(request):
#     today   = date.today() 
#     response = {}
#     error_response = {}
#     if request.method == "POST":
#         try:
#             user_id = request.POST.get('user_id')
#             receive_check = request.POST.get('receive_check')
#             dispatch_check = request.POST.get('dispatch_check')
#             crate_date = today.strftime("%Y-%m-%d")
#             rnormal_crate = request.POST.get('rnormal_crate')
#             rjumbo_crate = request.POST.get('rjumbo_crate')  
#             dnormal_crate = request.POST.get('dnormal_crate')
#             djumbo_crate = request.POST.get('djumbo_crate') 
#             error_count = 0 
#             if rnormal_crate=="0" or  dnormal_crate=="0":
#                 error_count = 1
#                 error_response = "Please select a valid crate number."
#             userlst = SpUserCrateLedger.objects.filter(user_id=user_id,normal_balance__gte=0).order_by('-id').exclude(updated_datetime__icontains=crate_date).first()
#             if user_id =='':
#                 error_count = 1
#                 error_response = "Please select user"
#             else:
#                 if receive_check!='0':
#                     if rnormal_crate == "" and rjumbo_crate == "":
#                         error_count = 1
#                         error_response = "Please fill atleast normal or jumbo in receive crate."
#                 if dispatch_check!='0':
#                     if dnormal_crate == "" and djumbo_crate == "":
#                         error_count = 1
#                         error_response = "Please fill atleast normal or jumbo in dispatch crate."
                   
#                 if(error_count > 0):
#                     response['error'] = True
#                     response['message'] = error_response
#                 else:
#                     current_crate=SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date)
#                     if current_crate:
#                         if len(current_crate)==2:
#                             receive = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).values_list('id',flat=True)
#                             dispatch = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).values_list('id',flat=True)
#                             if dispatch_check!="0" and receive_check=="0":
#                                 SpUserCrateLedger.objects.filter(id=dispatch[0]).delete()
#                                 if dnormal_crate!=0 and djumbo_crate=="":
#                                     receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                     temp                        =receive.normal_credit
#                                     receive.delete()
#                                     # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).delete()
#                                     usercrate                   = SpUserCrateLedger()
#                                     usercrate.user_id           = user_id
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate 
#                                     usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
                                    
#                                     usercraterecive                  = SpUserCrateLedger()
#                                     usercraterecive.user_id          =user_id
#                                     usercraterecive.normal_credit    =temp
#                                     usercraterecive.normal_debit     =0
#                                     if userlst:
#                                         usercraterecive.normal_balance   =userlst.normal_balance+int(dnormal_crate)-temp
#                                     else:
#                                         usercraterecive.normal_balance   =int(dnormal_crate)-temp
#                                     usercraterecive.jumbo_credit     =0
#                                     usercraterecive.jumbo_debit      =0
#                                     if userlst:
#                                         usercraterecive.jumbo_balance    =userlst.jumbo_balance
#                                     else:
#                                         usercraterecive.jumbo_balance    =0
#                                     usercraterecive.updated_datetime  = crate_date
#                                     usercraterecive.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercraterecive.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 elif dnormal_crate=="" and djumbo_crate!="0":
#                                     receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate                   = SpUserCrateLedger()
#                                     usercrate.user_id           =  user_id
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0  
#                                     usercrate.normal_balance    = receive.normal_balance
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate                   = SpUserCrateLedger()
#                                     usercrate.user_id           =  user_id
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate  
#                                     usercrate.normal_balance    = receive.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                             elif receive_check!="0" and dispatch_check=="0":
#                                 SpUserCrateLedger.objects.filter(id=receive[0]).delete()
#                                 if rnormal_crate!=0 and rjumbo_crate=="":
#                                     dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     temp                            = dispatch.normal_debit
#                                     dispatch.delete()
#                                     # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).delete()
#                                     usercrate                       = SpUserCrateLedger()
#                                     usercrate.user_id               = user_id
#                                     # if dispatch.normal_balance != 0:
#                                     usercrate.normal_credit     = int(rnormal_crate) 
#                                     usercrate.normal_debit      = 0
#                                     usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     if userlst:
#                                         usercrate.jumbo_balance     = userlst.jumbo_balance
#                                     else:
#                                         usercrate.jumbo_balance     = 0
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     usercratedispatch           =  SpUserCrateLedger()
#                                     usercratedispatch.user_id           = user_id
#                                     usercratedispatch.normal_credit     = 0
#                                     usercratedispatch.normal_debit      = temp
#                                     if userlst:
#                                         usercratedispatch.normal_balance    = userlst.normal_balance-int(rnormal_crate)+temp
#                                     else:
#                                         usercratedispatch.normal_balance    = int(rnormal_crate)+temp
#                                     usercratedispatch.jumbo_credit      = 0
#                                     usercratedispatch.jumbo_debit       = 0
#                                     if userlst:
#                                         usercratedispatch.jumbo_balance     = userlst.jumbo_balance
#                                     else:
#                                         usercratedispatch.jumbo_balance     = 0
#                                     usercratedispatch.updated_datetime  = crate_date
#                                     usercratedispatch.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercratedispatch.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 # else:
#                                     #     response['error'] = True
#                                     #     response['message'] = "No opening balace of this create type."
                                    
#                                 elif rjumbo_crate!=0 and rnormal_crate=="":
#                                     dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     usercrate                       = SpUserCrateLedger()
#                                     usercrate.user_id               =  user_id
#                                     if dispatch.jumbo_balance != 0:
#                                         usercrate.normal_credit     = 0
#                                         usercrate.normal_debit      = 0
#                                         usercrate.normal_balance    = dispatch.normal_balance
#                                         usercrate.jumbo_credit      = int(rjumbo_crate)
#                                         usercrate.jumbo_debit       = 0
#                                         usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         response['error'] = True
#                                         response['message'] = "No opening balace of this create type."
#                                 else:
#                                     dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     if dispatch.normal_balance != 0 and dispatch.jumbo_balance:
#                                         usercrate                   = SpUserCrateLedger()
#                                         usercrate.user_id           =  user_id
#                                         usercrate.normal_credit     = int(rnormal_crate)
#                                         usercrate.normal_debit      = 0
#                                         usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                         usercrate.jumbo_credit      = int(rjumbo_crate)
#                                         usercrate.jumbo_debit       = 0
#                                         usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         response['error'] = True
#                                         response['message'] = "No opening balace of this create type."
#                             else:
#                                 SpUserCrateLedger.objects.filter(id=dispatch[0]).delete()
#                                 # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).delete()
#                                 usercrate                          = SpUserCrateLedger()
#                                 if dnormal_crate!='':
#                                     usercrate.user_id              = user_id
#                                     usercrate.normal_credit        = 0
#                                     usercrate.normal_debit         = int(dnormal_crate)
#                                     if userlst:  
#                                         usercrate.normal_balance   = userlst.normal_balance+int(dnormal_crate) 
#                                     else:
#                                         usercrate.normal_balance   = int(dnormal_crate)
#                                 else:
#                                     usercrate.user_id              = user_id
#                                     usercrate.normal_debit         = 0
#                                     usercrate.normal_credit        = 0
#                                     if userlst:  
#                                         usercrate.normal_balance   = userlst.normal_balance
#                                     else:
#                                         usercrate.normal_balance   = 0
                            
#                                 if djumbo_crate!="":
#                                     usercrate.jumbo_credit             = int(dnormal_crate)
#                                     usercrate.jumbo_debit              = 0 
#                                     if userlst:
#                                         usercrate.jumbo_balance    = userlst.jumbo_balance+int(djumbo_crate)
#                                     else:
#                                         usercrate.jumbo_balance    = int(djumbo_crate)
#                                 else:
#                                     usercrate.jumbo_debit               = 0
#                                     usercrate.jumbo_credit             = 0
#                                     if userlst:
#                                         usercrate.jumbo_balance    = userlst.jumbo_balance
#                                     else:
#                                         usercrate.jumbo_balance    = 0
#                                 usercrate.updated_datetime         = crate_date
#                                 usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 usercrate.save()
#                                 SpUserCrateLedger.objects.filter(id=receive[0]).delete()
#                                 # SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).delete()
#                                 receive                       = SpUserCrateLedger()
#                                 if rnormal_crate!="":
#                                     receive.user_id           = user_id
#                                     receive.normal_credit     = int(rnormal_crate)
#                                     receive.normal_debit      = 0
#                                     if userlst:
#                                         receive.normal_balance    = int(userlst.normal_balance)+int(dnormal_crate)-int(rnormal_crate)
#                                     else:
#                                         receive.normal_balance    = int(dnormal_crate)-int(rnormal_crate)
#                                 else:
#                                     receive.normal_credit     = 0
#                                     receive.normal_debit      = 0
#                                     receive.normal_balance    = userlst.normal_balance
                                
#                                 if rjumbo_crate!="":
#                                     receive.jumbo_credit      = int(rjumbo_crate)
#                                     receive.jumbo_debit       = 0
#                                     if userlst:
#                                         receive.jumbo_balance     = int(userlst.jumbo_balance)+int(djumbo_crate)-int(rjumbo_crate)
#                                 else:
#                                     receive.jumbo_credit      = 0
#                                     receive.jumbo_debit       = 0
#                                     if userlst:
#                                         receive.jumbo_balance    = userlst.jumbo_balance
#                                     else:
#                                         receive.jumbo_balance    = 0
                                
#                                 receive.updated_datetime      = crate_date
#                                 usercrate.updated_at          = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 receive.save()
#                                 response['error'] = False
#                                 response['message'] = "Record has been saved successfully."
                                
#                         if len(current_crate)==1:
#                             receive = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).values_list('id',flat=True)
#                             dispatch = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).values_list('id',flat=True)
#                             count = 0
#                             if receive:
#                                 if dispatch_check!="0" and receive_check=="0":
#                                     if dnormal_crate!=0 and djumbo_crate=="":
#                                         receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                         usercrate                   = SpUserCrateLedger()
#                                         usercrate.user_id           =   user_id
#                                         usercrate.normal_credit     = 0
#                                         usercrate.normal_debit      = dnormal_crate 
#                                         usercrate.normal_balance    = receive.normal_balance+int(dnormal_crate) 
#                                         usercrate.jumbo_credit      = 0
#                                         usercrate.jumbo_debit       = 0
#                                         usercrate.jumbo_balance     = 0
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     elif dnormal_crate=="" and djumbo_crate!="0":
#                                         receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                         usercrate                   = SpUserCrateLedger()
#                                         usercrate.user_id           =   user_id
#                                         usercrate.normal_credit     = 0
#                                         usercrate.normal_debit      = 0  
#                                         usercrate.normal_balance    = receive.normal_balance
#                                         usercrate.jumbo_credit      = 0
#                                         usercrate.jumbo_debit       = djumbo_crate
#                                         usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         receive                            = SpUserCrateLedger.objects.get(id = receive[0])
#                                         usercrate                          = SpUserCrateLedger()
#                                         usercrate.normal_credit            = 0
#                                         usercrate.user_id                  =   user_id
#                                         usercrate.normal_debit             = dnormal_crate  
#                                         usercrate.normal_balance           = receive.normal_balance+int(dnormal_crate) 
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = djumbo_crate
#                                         usercrate.jumbo_balance            = receive.jumbo_balance+int(djumbo_crate)
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                 elif receive_check!="0" and dispatch_check=="0":
#                                     if rnormal_crate!=0 and rjumbo_crate=="":
#                                         usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
#                                         if userlst.normal_balance !=0:
#                                             usercrate.normal_credit            = int(rnormal_crate) 
#                                             usercrate.normal_debit             = 0
#                                             if userlst:
#                                                 usercrate.normal_balance       = userlst.normal_balance-int(rnormal_crate)
#                                             else:
#                                                 usercrate.normal_balance       = int(rnormal_crate)
#                                             usercrate.jumbo_credit             = 0
#                                             usercrate.jumbo_debit              = 0
#                                             if userlst:
#                                                 usercrate.jumbo_balance        = userlst.jumbo_balance
#                                             else:
#                                                 usercrate.jumbo_balance        = 0 
#                                             usercrate.updated_datetime         = crate_date
#                                             usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     elif rjumbo_crate!=0 and rnormal_crate=="":
#                                         usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
#                                         if userlst.jumbo_balance!=0:
#                                             usercrate.normal_credit            = 0
#                                             usercrate.normal_debit             = 0
#                                             if userlst:
#                                                 usercrate.normal_balance       = userlst.normal_balance
#                                             else:
#                                                 usercrate.normal_balance       = 0 
#                                             usercrate.jumbo_credit             = int(rjumbo_crate)
#                                             usercrate.jumbo_debit              = 0
#                                             if userlst:
#                                                 usercrate.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
#                                             else:
#                                                 usercrate.jumbo_balance        = int(rjumbo_crate)
#                                             usercrate.updated_datetime         = crate_date
#                                             usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     else:
#                                         usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
#                                         if userlst.jumbo_balance!=0 and userlst.normal_balance!=0: 
#                                             usercrate.normal_credit            = int(rnormal_crate)
#                                             usercrate.normal_debit             = 0
#                                             if userlst:
#                                                 usercrate.normal_balance       = userlst.normal_balance-int(rnormal_crate)
#                                             else:
#                                                 usercrate.normal_balance       = int(rnormal_crate)
#                                             usercrate.jumbo_credit             = int(rjumbo_crate)
#                                             usercrate.jumbo_debit              = 0
#                                             if userlst:
#                                                 usercrate.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
#                                             else:
#                                                 usercrate.jumbo_balance        = int(rjumbo_crate)
#                                             usercrate.updated_datetime         = crate_date
#                                             usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
                                        
#                                 else:
#                                     usercrate1                              = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate1.normal_debit                 = 0
#                                     if rnormal_crate!='':
#                                         usercrate1.normal_credit            = int(rnormal_crate)
#                                         if userlst:
#                                             usercrate1.normal_balance       = userlst.normal_balance-int(rnormal_crate)
#                                         else:
#                                             usercrate1.normal_balance       =  int(rnormal_crate)
#                                     else:
#                                         usercrate1.normal_credit            = 0
#                                         if userlst:
#                                             usercrate1.normal_balance       = userlst.normal_balance
#                                         else:
#                                             usercrate1.normal_balance       =  0
                                    
#                                     if rjumbo_crate!='':
#                                         usercrate1.jumbo_credit             = int(rjumbo_crate)
#                                         usercrate1.jumbo_debit              = 0
#                                         if userlst:
#                                             usercrate1.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
#                                         else:
#                                             usercrate1.jumbo_balance        = int(rjumbo_crate)
#                                     else:
#                                         usercrate1.jumbo_credit             = 0
#                                         usercrate1.jumbo_debit              = 0
#                                         if userlst:
#                                             usercrate1.jumbo_balance        = userlst.jumbo_balance
#                                         else:
#                                             usercrate1.jumbo_balance        = 0
#                                     usercrate1.updated_datetime             = crate_date
#                                     usercrate1.updated_at                   = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate1.save()

#                                     receive                                 = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate                               = SpUserCrateLedger()
#                                     usercrate.user_id                       =   user_id
#                                     if dnormal_crate!='':
#                                         usercrate.normal_credit             = 0
#                                         usercrate.normal_debit              = dnormal_crate  
#                                         usercrate.normal_balance            = userlst.normal_balance+int(dnormal_crate) 
#                                     else:
#                                         usercrate.normal_credit             = 0
#                                         usercrate.normal_debit              = 0
#                                         usercrate.normal_balance            = userlst.normal_balance

#                                     if rjumbo_crate!='':
#                                         usercrate.jumbo_credit              = 0
#                                         usercrate.jumbo_debit               = dnormal_crate
#                                         usercrate.jumbo_balance             = userlst.jumbo_balance+int(dnormal_crate)
#                                     else:
#                                         usercrate.jumbo_credit              = 0
#                                         usercrate.jumbo_debit               = 0
#                                         usercrate.jumbo_balance             = userlst.jumbo_balance
#                                     usercrate.updated_datetime              = crate_date
#                                     usercrate.updated_at                    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     count+=1
#                                     response['message'] = "Record has been saved successfully."
                                                            
#                             if dispatch and count == 0:
#                                 if dispatch_check!="0" and receive_check=="0":
#                                     if dnormal_crate!=0 and djumbo_crate=="":
#                                         usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         usercrate.normal_credit            = 0
#                                         usercrate.normal_debit             = dnormal_crate
#                                         if userlst:
#                                             usercrate.normal_balance       = userlst.normal_balance+int(dnormal_crate)
#                                         else:
#                                             usercrate.normal_balance       = int(dnormal_crate)
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = 0
#                                         if userlst:
#                                             usercrate.jumbo_balance        = userlst.jumbo_balance
#                                         else:
#                                             usercrate.jumbo_balance        = 0
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     elif dnormal_crate=="" and djumbo_crate!="0":
#                                         usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         usercrate.normal_credit            = 0
#                                         usercrate.normal_debit             = 0
#                                         if userlst:
#                                             usercrate.normal_balance       = userlst.normal_balance
#                                         else:
#                                             usercrate.normal_balance       = 0
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = djumbo_crate
#                                         if userlst:
#                                             usercrate.jumbo_balance        = userlst.jumbo_balance+int(djumbo_crate)
#                                         else:
#                                             usercrate.jumbo_balance        = int(djumbo_crate)
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         usercrate.normal_credit            = 0
#                                         usercrate.normal_debit             = dnormal_crate  
#                                         if userlst:
#                                             usercrate.normal_balance       = userlst.normal_balance+int(dnormal_crate) 
#                                         else:
#                                             usercrate.normal_balance       = int(dnormal_crate) 
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = djumbo_crate
#                                         if userlst:
#                                             usercrate.jumbo_balance        = userlst.jumbo_balance+int(djumbo_crate)
#                                         else:
#                                             usercrate.jumbo_balance        = int(djumbo_crate)
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                 elif receive_check!="0" and dispatch_check=="0":
#                                     if rnormal_crate!=0 and rjumbo_crate=="":
#                                         dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
                                        
#                                         if dispatch.normal_balance !=0 :
#                                             usercrate                   =SpUserCrateLedger()
#                                             usercrate.normal_credit     = int(rnormal_crate) 
#                                             usercrate.user_id           = user_id 
#                                             usercrate.normal_debit      = 0
#                                             usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                             usercrate.jumbo_credit      = 0
#                                             usercrate.jumbo_debit       = 0
#                                             usercrate.jumbo_balance     = dispatch.jumbo_balance
#                                             usercrate.updated_datetime  = crate_date
#                                             usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     elif rjumbo_crate!=0 and rnormal_crate=="":
#                                         dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         if dispatch.jumbo_balance!=0:
#                                             usercrate                   =SpUserCrateLedger()
#                                             usercrate.user_id           = user_id
#                                             usercrate.normal_credit     = 0
#                                             usercrate.normal_debit      = 0
#                                             usercrate.normal_balance    = dispatch.normal_balance
#                                             usercrate.jumbo_credit      = int(rjumbo_crate)
#                                             usercrate.jumbo_debit       = 0
#                                             usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                             usercrate.updated_datetime  = crate_date
#                                             usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     else:
#                                         dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         if dispatch.jumbo_balance!=0 and dispatch.normal_balance!=0: 
#                                             usercrate                   = SpUserCrateLedger()
#                                             usercrate.user_id           = user_id
#                                             usercrate.normal_credit     = int(rnormal_crate)
#                                             usercrate.normal_debit      = 0
#                                             usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                             usercrate.jumbo_credit      = int(rjumbo_crate)
#                                             usercrate.jumbo_debit       = 0
#                                             usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                             usercrate.updated_datetime  = crate_date
#                                             usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                 else:
#                                     usercrate                               = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     usercrate.normal_credit                 = 0
#                                     usercrate.normal_debit                  = dnormal_crate
#                                     if dnormal_crate !='':
#                                         usercrate.normal_debit              = dnormal_crate
#                                         if userlst:  
#                                             usercrate.normal_balance        = userlst.normal_balance+int(dnormal_crate) 
#                                         else:
#                                             usercrate.normal_balance        = int(dnormal_crate)
#                                     else:
#                                         usercrate.normal_debit              = 0
#                                         if userlst:  
#                                             usercrate.normal_balance        = userlst.normal_balance
#                                         else:
#                                             usercrate.normal_balance        = 0
#                                     usercrate.jumbo_credit                  = 0
#                                     if djumbo_crate !='':
#                                         usercrate.jumbo_debit               = djumbo_crate
#                                         if userlst:
#                                             usercrate.jumbo_balance         = userlst.jumbo_balance+int(djumbo_crate)
#                                         else:
#                                             usercrate.jumbo_balance         = djumbo_crate
#                                     else:
#                                         usercrate.jumbo_debit               = 0
#                                         if userlst:
#                                             usercrate.jumbo_balance         = userlst.jumbo_balance
#                                         else:
#                                             usercrate.jumbo_balance         = 0
#                                     usercrate.updated_datetime              = crate_date
#                                     usercrate.updated_at                    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
                                    

#                                     dispatch                     = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     usercrate1                   = SpUserCrateLedger()
#                                     usercrate1.user_id           = user_id
#                                     if rnormal_crate!='':
#                                         usercrate1.normal_credit     = int(rnormal_crate)
#                                         usercrate1.normal_debit      = 0
#                                         usercrate1.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                     else:
#                                         usercrate1.normal_credit     = 0
#                                         usercrate1.normal_debit      = 0
#                                         usercrate1.normal_balance    = dispatch.normal_balance
                                    
#                                     if rjumbo_crate!="":
#                                         usercrate1.jumbo_credit      = int(rjumbo_crate)
#                                         usercrate1.jumbo_debit       = 0
#                                         usercrate1.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                     else:
#                                         usercrate1.jumbo_credit      = 0
#                                         usercrate1.jumbo_debit       = 0
#                                         usercrate1.jumbo_balance     = dispatch.jumbo_balance
#                                     usercrate1.updated_datetime      = crate_date
#                                     usercrate1.updated_at            = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate1.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                     else:
#                         usercrate                               =   SpUserCrateLedger()
#                         usercrate.user_id                       =   user_id
#                         if dispatch_check!="0" and receive_check=="0":
#                             if dnormal_crate!=0 and djumbo_crate=="":
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate 
#                                     usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = 0
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate 
#                                     usercrate.normal_balance    = dnormal_crate
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = 0
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                             elif dnormal_crate=="" and djumbo_crate!="0":
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0  
#                                     usercrate.normal_balance    = userlst.normal_balance
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0  
#                                     usercrate.normal_balance    = 0
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = djumbo_crate
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                             else:
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate  
#                                     usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      =dnormal_crate    
#                                     usercrate.normal_balance    =dnormal_crate  
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = djumbo_crate
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."

#                         elif receive_check!="0" and dispatch_check=="0" and userlst:
#                             if rnormal_crate!=0 and rjumbo_crate=="":
#                                 if userlst:
#                                     usercrate.normal_credit     = int(rnormal_crate) 
#                                     usercrate.normal_debit      = 0
#                                     usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     response['error'] = True
#                                     response['message'] = "Party have no opening balance.. You can Dispatch only.."
#                             elif rjumbo_crate!=0 and rnormal_crate=="":
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0
#                                     usercrate.normal_balance    = userlst.normal_balance
#                                     usercrate.jumbo_credit      = int(rjumbo_crate)
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance-int(rjumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     response['error'] = True
#                                     response['message'] = "Party have no opening balance.. You can Dispatch only.."
#                             else:
#                                 if userlst:
#                                     usercrate.normal_credit     = int(rnormal_crate)
#                                     usercrate.normal_debit      = 0
#                                     usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
#                                     usercrate.jumbo_credit      = int(rjumbo_crate)
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance-int(rjumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     response['error'] = True
#                                     response['message'] = "Party have no opening balance.. You can Dispatch only.."           
#                         else:
#                             if userlst:
#                                 dispatch_usercrate     =SpUserCrateLedger()
#                                 if dnormal_crate!='':
#                                     dispatch_usercrate.normal_credit             = 0
#                                     dispatch_usercrate.normal_debit              = dnormal_crate  
#                                     dispatch_usercrate.normal_balance            = userlst.normal_balance+int(dnormal_crate) 
#                                 else:
#                                     dispatch_usercrate.normal_credit             = 0
#                                     dispatch_usercrate.normal_debit              = 0
#                                     dispatch_usercrate.normal_balance            = userlst.normal_balance

#                                 if djumbo_crate!='':
#                                     dispatch_usercrate.jumbo_credit              = 0
#                                     dispatch_usercrate.jumbo_debit               = dnormal_crate
#                                     dispatch_usercrate.jumbo_balance             = userlst.jumbo_balance+int(dnormal_crate)
#                                 else:
#                                     dispatch_usercrate.jumbo_credit              = 0
#                                     dispatch_usercrate.jumbo_debit               = 0
#                                     dispatch_usercrate.jumbo_balance             = userlst.jumbo_balance
#                                 dispatch_usercrate.updated_datetime              = crate_date
#                                 dispatch_usercrate.user_id                       =   user_id
#                                 dispatch_usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 dispatch_usercrate.save()
                                
#                                 dispatch_obj = SpUserCrateLedger.objects.get(id=dispatch_usercrate.id)
#                                 usercrate.normal_debit      = 0
#                                 if rnormal_crate!='':
#                                     usercrate.normal_credit     = int(rnormal_crate)
#                                     usercrate.normal_balance    = dispatch_obj.normal_balance-int(rnormal_crate)
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_balance    = dispatch_obj.normal_balance
#                                 if rjumbo_crate!='':
#                                     usercrate.jumbo_credit             = int(rjumbo_crate)
#                                     usercrate.jumbo_debit              = 0
#                                     usercrate.jumbo_balance        = dispatch_obj.jumbo_balance-int(rjumbo_crate)
#                                 else:
#                                     usercrate.jumbo_credit             = 0
#                                     usercrate.jumbo_debit              = 0
#                                     usercrate.jumbo_balance        = dispatch_obj.jumbo_balance
#                                 usercrate.updated_datetime         = crate_date
#                                 usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 usercrate.save()
#                                 response['error'] = False
#                                 response['message'] = "Record has been saved successfully." 
#                             else:
#                                 response['error'] = True
#                                 response['message'] = "Party have no opening balance.. You can Dispatch only.."
        
#         except ObjectDoesNotExist:
#             response['error'] = True
#             response['message'] = "Method not allowed"
#         except Exception as e:
#             print(e)
#             response['error'] = True
#             response['message'] = str(e)
#         return JsonResponse(response)
#     return redirect('/add-crate')

    
# @login_required
# def saveCrates(request):
#     today   = date.today() 
#     response = {}
#     error_response = {}
#     if request.method == "POST":
#         try:
#             user_id = request.POST.get('user_id')
#             receive_check = request.POST.get('receive_check')
#             dispatch_check = request.POST.get('dispatch_check')
#             crate_date = today.strftime("%Y-%m-%d")
#             rnormal_crate = request.POST.get('rnormal_crate')
#             rjumbo_crate = request.POST.get('rjumbo_crate')  
#             dnormal_crate = request.POST.get('dnormal_crate')
#             djumbo_crate = request.POST.get('djumbo_crate')  
#             userlst = SpUserCrateLedger.objects.filter(user_id=user_id).order_by('-id').exclude(updated_datetime__icontains=crate_date).first()
            
#             error_count = 0
#             if user_id =='':
#                 error_count = 1
#                 error_response['message'] = "Please select user"
#             else:
#                 if receive_check!='0':
#                     if rnormal_crate == "" and rjumbo_crate == "":
#                         error_count = 1
#                         error_response['rjumbo_error_back'] = "Please fill atleast normal or jumbo in receive crate."
#                 if dispatch_check!='0':
#                     if dnormal_crate == "" and djumbo_crate == "":
#                         error_count = 1
#                         error_response['rjumbo_error_back'] = "Please fill atleast normal or jumbo in dispatch crate."
                   
#                 if(error_count > 0):
#                     response['error'] = True
#                     response['message'] = error_response
#                 else:
#                     current_crate=SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date) 
#                     if current_crate:
#                         if len(current_crate)==2:
#                             receive = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).values_list('id',flat=True)
#                             dispatch = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).values_list('id',flat=True)
#                             if dispatch_check!="0" and receive_check=="0":
#                                 SpUserCrateLedger.objects.filter(id=dispatch[0]).delete()
#                                 if dnormal_crate!=0 and djumbo_crate=="":
#                                     receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate                   = SpUserCrateLedger()
#                                     usercrate.user_id           =   user_id
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate 
#                                     usercrate.normal_balance    = receive.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = receive.jumbo_balance
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 elif dnormal_crate=="" and djumbo_crate!="0":
#                                     receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate                   = SpUserCrateLedger()
#                                     usercrate.user_id           =  user_id
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0  
#                                     usercrate.normal_balance    = receive.normal_balance
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate                   = SpUserCrateLedger()
#                                     usercrate.user_id           =  user_id
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate  
#                                     usercrate.normal_balance    = receive.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                             elif receive_check!="0" and dispatch_check=="0":
#                                 SpUserCrateLedger.objects.filter(id=receive[0]).delete()
#                                 if rnormal_crate!=0 and rjumbo_crate=="":
#                                     dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     usercrate                       = SpUserCrateLedger()
#                                     usercrate.user_id               =  user_id
#                                     if dispatch.normal_balance != 0:
#                                         usercrate.normal_credit     = int(rnormal_crate) 
#                                         usercrate.normal_debit      = 0
#                                         usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                         usercrate.jumbo_credit      = 0
#                                         usercrate.jumbo_debit       = 0
#                                         usercrate.jumbo_balance     = dispatch.jumbo_balance
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         response['error'] = True
#                                         response['message'] = "No opening balace of this create type."
                                    
#                                 elif rjumbo_crate!=0 and rnormal_crate=="":
#                                     dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     usercrate                       = SpUserCrateLedger()
#                                     usercrate.user_id               =  user_id
#                                     if dispatch.jumbo_balance != 0:
#                                         usercrate.normal_credit     = 0
#                                         usercrate.normal_debit      = 0
#                                         usercrate.normal_balance    = dispatch.normal_balance
#                                         usercrate.jumbo_credit      = int(rjumbo_crate)
#                                         usercrate.jumbo_debit       = 0
#                                         usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         response['error'] = True
#                                         response['message'] = "No opening balace of this create type."
#                                 else:
#                                     dispatch                        = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     if dispatch.normal_balance != 0 and dispatch.jumbo_balance:
#                                         usercrate                   = SpUserCrateLedger()
#                                         usercrate.user_id           =  user_id
#                                         usercrate.normal_credit     = int(rnormal_crate)
#                                         usercrate.normal_debit      = 0
#                                         usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                         usercrate.jumbo_credit      = int(rjumbo_crate)
#                                         usercrate.jumbo_debit       = 0
#                                         usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         response['error'] = True
#                                         response['message'] = "No opening balace of this create type."
#                             else:
#                                 usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                 usercrate.normal_credit            = 0
#                                 if dnormal_crate!='':
#                                     usercrate.normal_debit         = dnormal_crate
#                                     if userlst:  
#                                         usercrate.normal_balance   = userlst.normal_balance+int(dnormal_crate) 
#                                     else:
#                                         usercrate.normal_balance   = int(dnormal_crate)
#                                 else:
#                                     usercrate.normal_debit         = 0
#                                     if userlst:  
#                                         usercrate.normal_balance   = userlst.normal_balance
#                                     else:
#                                         usercrate.normal_balance   = 0
#                                 usercrate.jumbo_credit             = 0
#                                 if djumbo_crate!="":
#                                     usercrate.jumbo_debit          = djumbo_crate
#                                     if userlst:
#                                         usercrate.jumbo_balance    = userlst.jumbo_balance+int(djumbo_crate)
#                                     else:
#                                         usercrate.jumbo_balance    = int(djumbo_crate)
#                                 else:
#                                     usercrate.jumbo_debit          = 0
#                                     if userlst:
#                                         usercrate.jumbo_balance    = userlst.jumbo_balance
#                                     else:
#                                         usercrate.jumbo_balance    = 0
#                                 usercrate.updated_datetime         = crate_date
#                                 usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 usercrate.save()
                                
                                
#                                 usercrate                     = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                 receive                       = SpUserCrateLedger.objects.get(id = receive[0])
#                                 if rnormal_crate!="":
#                                     receive.normal_credit     = int(rnormal_crate)
#                                     receive.normal_debit      = 0
#                                     receive.normal_balance    = usercrate.normal_balance-int(rnormal_crate)
#                                 else:
#                                     receive.normal_credit     = 0
#                                     receive.normal_debit      = 0
#                                     receive.normal_balance    = usercrate.normal_balance
                                
#                                 if rjumbo_crate!="":
#                                     receive.jumbo_credit      = int(rjumbo_crate)
#                                     receive.jumbo_debit       = 0
#                                     receive.jumbo_balance     = usercrate.jumbo_balance-int(rjumbo_crate)
#                                 else:
#                                     receive.jumbo_credit      = 0
#                                     receive.jumbo_debit       = 0
#                                     receive.jumbo_balance     = usercrate.jumbo_balance
                                
#                                 receive.updated_datetime      = crate_date
#                                 usercrate.updated_at          = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 receive.save()
#                                 response['error'] = False
#                                 response['message'] = "Record has been saved successfully."
                                
#                         if len(current_crate)==1:
#                             receive = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_debit = 0 , jumbo_debit = 0).values_list('id',flat=True)
#                             dispatch = SpUserCrateLedger.objects.filter(user_id=user_id, updated_datetime__icontains=crate_date,normal_credit = 0,jumbo_credit = 0).values_list('id',flat=True)
#                             count = 0
#                             if receive:
#                                 if dispatch_check!="0" and receive_check=="0":
#                                     if dnormal_crate!=0 and djumbo_crate=="":
#                                         receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                         usercrate                   = SpUserCrateLedger()
#                                         usercrate.user_id           =   user_id
#                                         usercrate.normal_credit     = 0
#                                         usercrate.normal_debit      = dnormal_crate 
#                                         usercrate.normal_balance    = receive.normal_balance+int(dnormal_crate) 
#                                         usercrate.jumbo_credit      = 0
#                                         usercrate.jumbo_debit       = 0
#                                         usercrate.jumbo_balance     = 0
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     elif dnormal_crate=="" and djumbo_crate!="0":
#                                         receive                     = SpUserCrateLedger.objects.get(id = receive[0])
#                                         usercrate                   = SpUserCrateLedger()
#                                         usercrate.user_id           =   user_id
#                                         usercrate.normal_credit     = 0
#                                         usercrate.normal_debit      = 0  
#                                         usercrate.normal_balance    = receive.normal_balance
#                                         usercrate.jumbo_credit      = 0
#                                         usercrate.jumbo_debit       = djumbo_crate
#                                         usercrate.jumbo_balance     = receive.jumbo_balance+int(djumbo_crate)
#                                         usercrate.updated_datetime  = crate_date
#                                         usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         receive                            = SpUserCrateLedger.objects.get(id = receive[0])
#                                         usercrate                          = SpUserCrateLedger()
#                                         usercrate.normal_credit            = 0
#                                         usercrate.user_id                  =   user_id
#                                         usercrate.normal_debit             = dnormal_crate  
#                                         usercrate.normal_balance           = receive.normal_balance+int(dnormal_crate) 
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = djumbo_crate
#                                         usercrate.jumbo_balance            = receive.jumbo_balance+int(djumbo_crate)
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                 elif receive_check!="0" and dispatch_check=="0":
#                                     if rnormal_crate!=0 and rjumbo_crate=="":
#                                         usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
#                                         if userlst.normal_balance !=0:
#                                             usercrate.normal_credit            = int(rnormal_crate) 
#                                             usercrate.normal_debit             = 0
#                                             if userlst:
#                                                 usercrate.normal_balance       = userlst.normal_balance-int(rnormal_crate)
#                                             else:
#                                                 usercrate.normal_balance       = int(rnormal_crate)
#                                             usercrate.jumbo_credit             = 0
#                                             usercrate.jumbo_debit              = 0
#                                             if userlst:
#                                                 usercrate.jumbo_balance        = userlst.jumbo_balance
#                                             else:
#                                                 usercrate.jumbo_balance        = 0 
#                                             usercrate.updated_datetime         = crate_date
#                                             usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     elif rjumbo_crate!=0 and rnormal_crate=="":
#                                         usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
#                                         if userlst.jumbo_balance!=0:
#                                             usercrate.normal_credit            = 0
#                                             usercrate.normal_debit             = 0
#                                             if userlst:
#                                                 usercrate.normal_balance       = userlst.normal_balance
#                                             else:
#                                                 usercrate.normal_balance       = 0 
#                                             usercrate.jumbo_credit             = int(rjumbo_crate)
#                                             usercrate.jumbo_debit              = 0
#                                             if userlst:
#                                                 usercrate.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
#                                             else:
#                                                 usercrate.jumbo_balance        = int(rjumbo_crate)
#                                             usercrate.updated_datetime         = crate_date
#                                             usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     else:
#                                         usercrate                              = SpUserCrateLedger.objects.get(id = receive[0])
#                                         if userlst.jumbo_balance!=0 and userlst.normal_balance!=0: 
#                                             usercrate.normal_credit            = int(rnormal_crate)
#                                             usercrate.normal_debit             = 0
#                                             if userlst:
#                                                 usercrate.normal_balance       = userlst.normal_balance-int(rnormal_crate)
#                                             else:
#                                                 usercrate.normal_balance       = int(rnormal_crate)
#                                             usercrate.jumbo_credit             = int(rjumbo_crate)
#                                             usercrate.jumbo_debit              = 0
#                                             if userlst:
#                                                 usercrate.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
#                                             else:
#                                                 usercrate.jumbo_balance        = int(rjumbo_crate)
#                                             usercrate.updated_datetime         = crate_date
#                                             usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
                                        
#                                 else:
#                                     usercrate1                              = SpUserCrateLedger.objects.get(id = receive[0])
                                    
#                                     usercrate1.normal_debit                 = 0
#                                     if rnormal_crate!='':
#                                         usercrate1.normal_credit            = int(rnormal_crate)
#                                         if userlst:
#                                             usercrate1.normal_balance       = userlst.normal_balance-int(rnormal_crate)
#                                         else:
#                                             usercrate1.normal_balance       =  int(rnormal_crate)
#                                     else:
#                                         usercrate1.normal_credit            = 0
#                                         if userlst:
#                                             usercrate1.normal_balance       = userlst.normal_balance
#                                         else:
#                                             usercrate1.normal_balance       =  0
                                    
#                                     if rjumbo_crate!='':
#                                         usercrate1.jumbo_credit             = int(rjumbo_crate)
#                                         usercrate1.jumbo_debit              = 0
#                                         if userlst:
#                                             usercrate1.jumbo_balance        = userlst.jumbo_balance-int(rjumbo_crate)
#                                         else:
#                                             usercrate1.jumbo_balance        = int(rjumbo_crate)
#                                     else:
#                                         usercrate1.jumbo_credit             = 0
#                                         usercrate1.jumbo_debit              = 0
#                                         if userlst:
#                                             usercrate1.jumbo_balance        = userlst.jumbo_balance
#                                         else:
#                                             usercrate1.jumbo_balance        = 0
#                                     usercrate1.updated_datetime             = crate_date
#                                     usercrate1.updated_at                   = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate1.save()


#                                     receive                                 = SpUserCrateLedger.objects.get(id = receive[0])
#                                     usercrate                               = SpUserCrateLedger()
                                    
#                                     usercrate.user_id                       =   user_id
#                                     if dnormal_crate!='':
#                                         usercrate.normal_credit             = 0
#                                         usercrate.normal_debit              = dnormal_crate  
#                                         usercrate.normal_balance            = userlst.normal_balance+int(dnormal_crate) 
#                                     else:
#                                         usercrate.normal_credit             = 0
#                                         usercrate.normal_debit              = 0
#                                         usercrate.normal_balance            = userlst.normal_balance

#                                     if rjumbo_crate!='':
#                                         usercrate.jumbo_credit              = 0
#                                         usercrate.jumbo_debit               = dnormal_crate
#                                         usercrate.jumbo_balance             = userlst.jumbo_balance+int(dnormal_crate)
#                                     else:
#                                         usercrate.jumbo_credit              = 0
#                                         usercrate.jumbo_debit               = 0
#                                         usercrate.jumbo_balance             = userlst.jumbo_balance
#                                     usercrate.updated_datetime              = crate_date
#                                     usercrate.updated_at                    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     count+=1
#                                     response['message'] = "Record has been saved successfully."
                                                            
#                             if dispatch and count == 0:
#                                 if dispatch_check!="0" and receive_check=="0":
#                                     if dnormal_crate!=0 and djumbo_crate=="":
#                                         usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         usercrate.normal_credit            = 0
#                                         usercrate.normal_debit             = dnormal_crate
#                                         if userlst:
#                                             usercrate.normal_balance       = userlst.normal_balance+int(dnormal_crate)
#                                         else:
#                                             usercrate.normal_balance       = int(dnormal_crate)
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = 0
#                                         if userlst:
#                                             usercrate.jumbo_balance        = userlst.jumbo_balance
#                                         else:
#                                             usercrate.jumbo_balance        = 0
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     elif dnormal_crate=="" and djumbo_crate!="0":
#                                         usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         usercrate.normal_credit            = 0
#                                         usercrate.normal_debit             = 0
#                                         if userlst:
#                                             usercrate.normal_balance       = userlst.normal_balance
#                                         else:
#                                             usercrate.normal_balance       = 0
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = djumbo_crate
#                                         if userlst:
#                                             usercrate.jumbo_balance        = userlst.jumbo_balance+int(djumbo_crate)
#                                         else:
#                                             usercrate.jumbo_balance        = int(djumbo_crate)
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                     else:
#                                         usercrate                          = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         usercrate.normal_credit            = 0
#                                         usercrate.normal_debit             = dnormal_crate  
#                                         if userlst:
#                                             usercrate.normal_balance       = userlst.normal_balance+int(dnormal_crate) 
#                                         else:
#                                             usercrate.normal_balance       = int(dnormal_crate) 
#                                         usercrate.jumbo_credit             = 0
#                                         usercrate.jumbo_debit              = djumbo_crate
#                                         if userlst:
#                                             usercrate.jumbo_balance        = userlst.jumbo_balance+int(djumbo_crate)
#                                         else:
#                                             usercrate.jumbo_balance        = int(djumbo_crate)
#                                         usercrate.updated_datetime         = crate_date
#                                         usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                         usercrate.save()
#                                         response['error'] = False
#                                         response['message'] = "Record has been saved successfully."
#                                 elif receive_check!="0" and dispatch_check=="0":
#                                     if rnormal_crate!=0 and rjumbo_crate=="":
#                                         dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
                                        
#                                         if dispatch.normal_balance !=0 :
#                                             usercrate                   =SpUserCrateLedger()
#                                             usercrate.normal_credit     = int(rnormal_crate) 
#                                             usercrate.user_id           = user_id 
#                                             usercrate.normal_debit      = 0
#                                             usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                             usercrate.jumbo_credit      = 0
#                                             usercrate.jumbo_debit       = 0
#                                             usercrate.jumbo_balance     = dispatch.jumbo_balance
#                                             usercrate.updated_datetime  = crate_date
#                                             usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     elif rjumbo_crate!=0 and rnormal_crate=="":
#                                         dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         if dispatch.jumbo_balance!=0:
#                                             usercrate                   =SpUserCrateLedger()
#                                             usercrate.user_id           = user_id
#                                             usercrate.normal_credit     = 0
#                                             usercrate.normal_debit      = 0
#                                             usercrate.normal_balance    = dispatch.normal_balance
#                                             usercrate.jumbo_credit      = int(rjumbo_crate)
#                                             usercrate.jumbo_debit       = 0
#                                             usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                             usercrate.updated_datetime  = crate_date
#                                             usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                     else:
#                                         dispatch                    = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                         if dispatch.jumbo_balance!=0 and dispatch.normal_balance!=0: 
#                                             usercrate                   = SpUserCrateLedger()
#                                             usercrate.user_id           = user_id
#                                             usercrate.normal_credit     = int(rnormal_crate)
#                                             usercrate.normal_debit      = 0
#                                             usercrate.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                             usercrate.jumbo_credit      = int(rjumbo_crate)
#                                             usercrate.jumbo_debit       = 0
#                                             usercrate.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                             usercrate.updated_datetime  = crate_date
#                                             usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                             usercrate.save()
#                                             response['error'] = False
#                                             response['message'] = "Record has been saved successfully."
#                                         else:
#                                             response['error'] = True
#                                             response['message'] = "No opening balance of this create type.."
#                                 else:
#                                     usercrate                               = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     usercrate.normal_credit                 = 0
#                                     usercrate.normal_debit                  = dnormal_crate
#                                     if dnormal_crate !='':
#                                         usercrate.normal_debit              = dnormal_crate
#                                         if userlst:  
#                                             usercrate.normal_balance        = userlst.normal_balance+int(dnormal_crate) 
#                                         else:
#                                             usercrate.normal_balance        = int(dnormal_crate)
#                                     else:
#                                         usercrate.normal_debit              = 0
#                                         if userlst:  
#                                             usercrate.normal_balance        = userlst.normal_balance
#                                         else:
#                                             usercrate.normal_balance        = 0
#                                     usercrate.jumbo_credit                  = 0
#                                     if djumbo_crate !='':
#                                         usercrate.jumbo_debit               = djumbo_crate
#                                         if userlst:
#                                             usercrate.jumbo_balance         = userlst.jumbo_balance+int(djumbo_crate)
#                                         else:
#                                             usercrate.jumbo_balance         = djumbo_crate
#                                     else:
#                                         usercrate.jumbo_debit               = 0
#                                         if userlst:
#                                             usercrate.jumbo_balance         = userlst.jumbo_balance
#                                         else:
#                                             usercrate.jumbo_balance         = 0
#                                     usercrate.updated_datetime              = crate_date
#                                     usercrate.updated_at                    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
                                    

#                                     dispatch                     = SpUserCrateLedger.objects.get(id = dispatch[0])
#                                     usercrate1                   = SpUserCrateLedger()
#                                     usercrate1.user_id           = user_id
#                                     if rnormal_crate!='':
#                                         usercrate1.normal_credit     = int(rnormal_crate)
#                                         usercrate1.normal_debit      = 0
#                                         usercrate1.normal_balance    = dispatch.normal_balance-int(rnormal_crate)
#                                     else:
#                                         usercrate1.normal_credit     = 0
#                                         usercrate1.normal_debit      = 0
#                                         usercrate1.normal_balance    = dispatch.normal_balance
                                    
#                                     if rjumbo_crate!="":
#                                         usercrate1.jumbo_credit      = int(rjumbo_crate)
#                                         usercrate1.jumbo_debit       = 0
#                                         usercrate1.jumbo_balance     = dispatch.jumbo_balance-int(rjumbo_crate)
#                                     else:
#                                         usercrate1.jumbo_credit      = 0
#                                         usercrate1.jumbo_debit       = 0
#                                         usercrate1.jumbo_balance     = dispatch.jumbo_balance
#                                     usercrate1.updated_datetime      = crate_date
#                                     usercrate1.updated_at            = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate1.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                     else:
#                         usercrate                               =   SpUserCrateLedger()
#                         usercrate.user_id                       =   user_id
#                         if dispatch_check!="0" and receive_check=="0":
#                             if dnormal_crate!=0 and djumbo_crate=="":
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate 
#                                     usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = 0
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate 
#                                     usercrate.normal_balance    = dnormal_crate
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = 0
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                             elif dnormal_crate=="" and djumbo_crate!="0":
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0  
#                                     usercrate.normal_balance    = userlst.normal_balance
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0  
#                                     usercrate.normal_balance    = 0
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = djumbo_crate
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                             else:
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = dnormal_crate  
#                                     usercrate.normal_balance    = userlst.normal_balance+int(dnormal_crate) 
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance+int(djumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      =dnormal_crate    
#                                     usercrate.normal_balance    =dnormal_crate  
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = djumbo_crate
#                                     usercrate.jumbo_balance     = djumbo_crate
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."

#                         elif receive_check!="0" and dispatch_check=="0" and userlst:
#                             if rnormal_crate!=0 and rjumbo_crate=="":
#                                 if userlst:
#                                     usercrate.normal_credit     = int(rnormal_crate) 
#                                     usercrate.normal_debit      = 0
#                                     usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
#                                     usercrate.jumbo_credit      = 0
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     response['error'] = True
#                                     response['message'] = "Party have no opening balance.. You can Dispatch only.."
#                             elif rjumbo_crate!=0 and rnormal_crate=="":
#                                 if userlst:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_debit      = 0
#                                     usercrate.normal_balance    = userlst.normal_balance
#                                     usercrate.jumbo_credit      = int(rjumbo_crate)
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance-int(rjumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     response['error'] = True
#                                     response['message'] = "Party have no opening balance.. You can Dispatch only.."
#                             else:
#                                 if userlst:
#                                     usercrate.normal_credit     = int(rnormal_crate)
#                                     usercrate.normal_debit      = 0
#                                     usercrate.normal_balance    = userlst.normal_balance-int(rnormal_crate)
#                                     usercrate.jumbo_credit      = int(rjumbo_crate)
#                                     usercrate.jumbo_debit       = 0
#                                     usercrate.jumbo_balance     = userlst.jumbo_balance-int(rjumbo_crate)
#                                     usercrate.updated_datetime  = crate_date
#                                     usercrate.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                     usercrate.save()
#                                     response['error'] = False
#                                     response['message'] = "Record has been saved successfully."
#                                 else:
#                                     response['error'] = True
#                                     response['message'] = "Party have no opening balance.. You can Dispatch only.."           
#                         else:
#                             if userlst:
#                                 dispatch_usercrate     =SpUserCrateLedger()
#                                 if dnormal_crate!='':
#                                     dispatch_usercrate.normal_credit             = 0
#                                     dispatch_usercrate.normal_debit              = dnormal_crate  
#                                     dispatch_usercrate.normal_balance            = userlst.normal_balance+int(dnormal_crate) 
#                                 else:
#                                     dispatch_usercrate.normal_credit             = 0
#                                     dispatch_usercrate.normal_debit              = 0
#                                     dispatch_usercrate.normal_balance            = userlst.normal_balance

#                                 if djumbo_crate!='':
#                                     dispatch_usercrate.jumbo_credit              = 0
#                                     dispatch_usercrate.jumbo_debit               = dnormal_crate
#                                     dispatch_usercrate.jumbo_balance             = userlst.jumbo_balance+int(dnormal_crate)
#                                 else:
#                                     dispatch_usercrate.jumbo_credit              = 0
#                                     dispatch_usercrate.jumbo_debit               = 0
#                                     dispatch_usercrate.jumbo_balance             = userlst.jumbo_balance
#                                 dispatch_usercrate.updated_datetime              = crate_date
#                                 dispatch_usercrate.user_id                       =   user_id
#                                 dispatch_usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 dispatch_usercrate.save()
                                
#                                 dispatch_obj = SpUserCrateLedger.objects.get(id=dispatch_usercrate.id)
#                                 usercrate.normal_debit      = 0
#                                 if rnormal_crate!='':
#                                     usercrate.normal_credit     = int(rnormal_crate)
#                                     usercrate.normal_balance    = dispatch_obj.normal_balance-int(rnormal_crate)
#                                 else:
#                                     usercrate.normal_credit     = 0
#                                     usercrate.normal_balance    = dispatch_obj.normal_balance
#                                 if rjumbo_crate!='':
#                                     usercrate.jumbo_credit             = int(rjumbo_crate)
#                                     usercrate.jumbo_debit              = 0
#                                     usercrate.jumbo_balance        = dispatch_obj.jumbo_balance-int(rjumbo_crate)
#                                 else:
#                                     usercrate.jumbo_credit             = 0
#                                     usercrate.jumbo_debit              = 0
#                                     usercrate.jumbo_balance        = dispatch_obj.jumbo_balance
#                                 usercrate.updated_datetime         = crate_date
#                                 usercrate.updated_at               = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#                                 usercrate.save()
#                                 response['error'] = False
#                                 response['message'] = "Record has been saved successfully." 
#                             else:
#                                 response['error'] = True
#                                 response['message'] = "Party have no opening balance.. You can Dispatch only.."
        
        
        
        
#         except ObjectDoesNotExist:
#             response['error'] = True
#             response['message'] = "Method not allowed"
#         except Exception as e:
#             print(e)
#             response['error'] = True
#             response['message'] = str(e)
#         return JsonResponse(response)
#     return redirect('/add-crate')
    
# #user crate summary
# @login_required
# def userCrateReport(request):
#     today   = date.today()  
#     year  = today.year
#     month = today.month

#     user_list = SpUsers.objects.filter(user_type=2, status=1)
#     month_list = days_in_months(year,month)

#     try:
#         user = SpUsers.objects.filter(user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
#     except SpUsers.DoesNotExist:
#         user = None

#     crate_lists                     = []
#     if user:
#         for id, months in enumerate(month_list):
#             month_lists                         = {} 
#             month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
#             month_lists['month_date']               = str(months)
#             month_lists['opening_balance']          = getOpeningCratesSum(user['id'], month_date)
#             month_lists['dispatch']                 = getTotalDispatchedCratesSums(user['id'], month_date)
#             month_lists['receiving']                = getTotalReceivedCratesSums(user['id'], month_date)
#             total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
#             month_lists['total_outstanding']        = total_outstanding
#             month_lists['store_name'] = getModelColumnById(SpUsers, user['id'], 'store_name') 
#             month_lists['emp_sap_id'] = getModelColumnById(SpUsers, user['id'], 'emp_sap_id') 
#             crate_lists.append(month_lists)

    
#     if user['id']:
#         party_name = getUserName(user['id'])
#         city_name  = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user['id'], 'town_name')
#     else:
#         party_name = ''
#         city_name  = ''
           
#     context = {}
#     context['user_list']                        = user_list
#     context['party_name']                       = party_name
#     context['city_name']                        = city_name
#     context['crate_lists']                      = crate_lists
#     context['today_date']                       = today.strftime("%m/%Y")
#     context['month_list']                       = days_in_months(year,month)
#     context['page_title']                       = "User Crate Report"

#     template = 'crate/user-crate-report.html'
#     return render(request, template, context)

@login_required
def userCrateReport(request):
    today   = date.today()
    today=today.strftime('%d/%m/%Y')
    crate_date=str(datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d'))
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    users = SpUsers.objects.filter(user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name')
    crate_lists                     = []
    if users:
        for user in users:
            month_lists                             = {} 
            month_lists['opening_balance']          = getOpeningCratesSum(user['id'],crate_date)
            month_lists['dispatch']                 = getTotalDispatchedCratesSums(user['id'], crate_date)
            month_lists['receiving']                = getTotalReceivedCratesSums(user['id'], crate_date)
            total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
            month_lists['total_outstanding']        = total_outstanding
            month_lists['today']                    = today
            month_lists['store_name']               = getModelColumnByColumnId(SpUsers,'id', user['id'], 'store_name') 
            month_lists['emp_sap_id']               = getModelColumnById(SpUsers, user['id'], 'emp_sap_id')
            month_lists['party_name']               = getUserName(user['id'])
            if SpUserAreaAllocations.objects.filter(user_id =  user['id']).exists():
                month_lists['city_name']                = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user['id'], 'town_name')
            else:
                month_lists['city_name'] = '-'
            crate_lists.append(month_lists)
    context = {}
    context['crate_lists']                      = crate_lists
    context['user_list']                        = user_list
    context['page_title']                       = "User Crate Report"
    context['today']                            = today
    template = 'crate/user-crate-report.html'
    return render(request, template, context)

# ajax List View
# @login_required
# def ajaxUserCrateReport(request):
#     today   = date.today()  
#     if request.GET['crate_date']:
#         crate_date = request.GET['crate_date']
#         crate_date = crate_date.split('/')
#         year  = int(crate_date[1])
#         month = int(crate_date[0])
#     else:
#         year  = today.year
#         if int(month) > 9:
#             month = today.month 
#         else:
#             month = '0'+str(today.month)
#             month = int(month)
           
#     user_list = SpUsers.objects.filter(user_type=2, status=1)
#     month_list = days_in_months(year,month)

#     try:
#         user = SpUsers.objects.filter(id=request.GET['user_id'],user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
#     except SpUsers.DoesNotExist:
#         user = None

#     crate_lists                     = []
#     if user:
#         for id, months in enumerate(month_list):
#             month_lists                             = {} 
#             month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
#             month_lists['month_date']               = str(months)
#             month_lists['opening_balance']          = getOpeningCratesSum(request.GET['user_id'], month_date)
#             month_lists['dispatch']                 = getTotalDispatchedCratesSums(request.GET['user_id'], month_date)
#             month_lists['receiving']                = getTotalReceivedCratesSums(request.GET['user_id'], month_date)
#             total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
#             month_lists['total_outstanding']        = total_outstanding
#             month_lists['store_name'] = getModelColumnById(SpUsers, user['id'], 'store_name') 
#             month_lists['emp_sap_id'] = getModelColumnById(SpUsers, user['id'], 'emp_sap_id') 
            
#             crate_lists.append(month_lists)

    
#     if user['id']:
#         party_name = getUserName(user['id'])
#         city_name  = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user['id'], 'town_name')
#     else:
#         party_name = ''
#         city_name  = ''

#     context = {}
#     context['user_list']                        = user_list
#     context['party_name']                       = party_name
#     context['city_name']                        = city_name
#     context['crate_lists']                      = crate_lists
#     context['today_date']                       = today.strftime("%m/%Y")
#     context['month_list']                       = days_in_months(year,month)

#     template = 'crate/ajax-user-crate-report.html'
#     return render(request, template, context) 
@login_required
def ajaxUserCrateReport(request):
    crate_date = request.GET['crate_date']
    today = request.GET['crate_date']
    crate_date=str(datetime.strptime(str(crate_date), '%d/%m/%Y').strftime('%Y-%m-%d'))
    user_id = request.GET['user_id']
    if user_id:
        users = SpUsers.objects.filter(id=user_id,user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name')
    else:
        users = SpUsers.objects.filter(user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name')
    crate_lists                     = []
    if users:
        for user in users:
            month_lists                             = {} 
            month_lists['opening_balance']          = getOpeningCratesSum(user['id'],crate_date)
            month_lists['dispatch']                 = getTotalDispatchedCratesSums(user['id'], crate_date)
            month_lists['receiving']                = getTotalReceivedCratesSums(user['id'], crate_date)
            total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
            month_lists['total_outstanding']        = total_outstanding
            month_lists['today']                    = today
            month_lists['store_name']               = getModelColumnByColumnId(SpUsers,'id',user['id'], 'store_name') 
            month_lists['emp_sap_id']               = getModelColumnById(SpUsers, user['id'], 'emp_sap_id')
            month_lists['party_name']               = getUserName(user['id'])
            if SpUserAreaAllocations.objects.filter(user_id =  user['id']).exists():
                month_lists['city_name']                = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user['id'], 'town_name')
            else:
                month_lists['city_name'] = '-' 
            crate_lists.append(month_lists) 
    context = {}
    context['crate_lists']                      = crate_lists
    context['page_title']                       = "User Crate Report"
    template = 'crate/ajax-user-crate-report.html'
    return render(request, template, context) 

# export view
# @login_required
# def exportUserCrateReport(request, crate_date, user_id):
#     today   = date.today()  
#     # crates_details = SpUserCrates.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).order_by('id')
#     if crate_date:
#         crate_date = crate_date.split('-')
#         year  = int(crate_date[1])
#         month = int(crate_date[0])
#     else:
#         year  = today.year
#         if int(month)> 9:
#             month = today.month 
#         else:
#             month = '0'+str(today.month)
#             month = int(month)

#     user_id = int(user_id)       
#     user_list = SpUsers.objects.filter(user_type=2, status=1)
#     month_list = days_in_months(year,month)

#     try:
#         user = SpUsers.objects.filter(id=user_id,user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
#     except SpUsers.DoesNotExist:
#         user = None

#     crate_lists                     = []
#     if user:
#         for id, months in enumerate(month_list):
#             month_lists                             = {} 
#             month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
#             month_lists['month_date']               = str(months)
#             month_lists['opening_balance']          = getOpeningCratesSum(user_id, month_date)
#             month_lists['dispatch']                 = getTotalDispatchedCratesSums(user_id, month_date)
#             month_lists['receiving']                = getTotalReceivedCratesSums(user_id, month_date)
#             total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
#             month_lists['total_outstanding']        = total_outstanding
#             month_lists['store_name'] = getModelColumnById(SpUsers, user['id'], 'store_name') 
#             month_lists['emp_sap_id'] = getModelColumnById(SpUsers, user['id'], 'emp_sap_id')
            
#             crate_lists.append(month_lists)

    
#     if user_id:
#         party_name = getUserName(user_id)
#         city_name  = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user_id, 'town_name')
#     else:
#         party_name = ''
#         city_name  = ''

#     #export code
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=user-crate-report.xlsx'.format(
#         date=today,
#     )
#     workbook = Workbook()

#     # Define some styles and formatting that will be later used for cells
#     header_font = Font(name='Calibri', bold=True)
#     centered_alignment = Alignment(horizontal='center')
#     thin = Side(border_style="thin", color="303030") 
#     black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
#     wrapped_alignment = Alignment(
#         horizontal='center',
#         vertical='center',
#         wrap_text=True
#     )
    
#     # Get active worksheet/tab
#     worksheet = workbook.active
#     worksheet.title = 'CRATE STATEMENT'
#     worksheet.merge_cells('A1:A1') 
    
#     worksheet.page_setup.orientation = 'landscape'
#     worksheet.page_setup.paperSize = 9
#     worksheet.page_setup.fitToPage = True
    
#     worksheet = workbook.worksheets[0]
#     img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
#     img.height = 50
#     img.alignment = 'center'
#     img.anchor = 'A1'
#     worksheet.add_image(img)

#     # Define the titles for columns
#     # columns = []
#     row_num = 1
#     worksheet.row_dimensions[1].height = 40

#     # cell = worksheet.cell(row=1, column=2)  
#     # cell.value = str(year)+'/'+str(month)
#     # cell.font = header_font
#     # cell.alignment = wrapped_alignment
#     # cell.border = black_border
#     # cell.font = Font(size=12, color='FFFFFFFF', bold=True)
#     # cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

#     column_length = 7
    
#     worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
#     worksheet.cell(row=1, column=2).value = 'CRATE STATEMENT'
#     worksheet.cell(row=1, column=2).font = header_font
#     worksheet.cell(row=1, column=2).alignment = wrapped_alignment
#     worksheet.cell(row=1, column=column_length).border = black_border
#     worksheet.cell(row=1, column=2).font = Font(size=15, color='FFFFFFFF', bold=True)
#     worksheet.cell(row=1, column=2).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")


#     # color_codes = []
#     columns = []
#     columns += [ 'Date' ]
#     columns += [ 'Party Name' ]
#     columns += [ 'City' ]
#     columns += [ 'Opening Balance' ]
#     columns += [ 'Dispatch' ]
#     columns += [ 'Receiving' ]
#     columns += [ 'Total Outstanding' ]
#     row_num = 2
    
    

#     # Assign the titles for each cell of the header
#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title
#         cell.alignment = wrapped_alignment
#         cell.border = black_border
#         cell.fill = PatternFill()
#         cell.font = Font(size=11, color='000000', bold=True) 
#         column_letter = get_column_letter(col_num)
#         column_dimensions = worksheet.column_dimensions[column_letter]
#         column_dimensions.width = 20


#     for id, crate in enumerate(crate_lists):
#         row_num += 1
#         # Define the data for each cell in the row    
#         row = []
#         row += [ crate['month_date'] ]
#         row += [ crate['store_name']+" "+"("+party_name+"/"+crate['emp_sap_id']+")" ]
#         row += [ city_name ]
#         row += [ crate['opening_balance'] ]
#         row += [ crate['dispatch'] ]
#         row += [ crate['receiving'] ]
#         row += [ crate['total_outstanding'] ]

#         # Assign the data for each cell of the row 
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#             cell.alignment = wrapped_alignment
#             cell.border = black_border
#             cell.fill = PatternFill()
#             cell.font = Font(size=12, color='000000')


#     row_num += 1
#     last_row = row_num
    

#     worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

#     worksheet.row_dimensions[last_row].height = 20
#     worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
#     worksheet.cell(row=last_row, column=1).font = header_font
#     worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
#     worksheet.cell(row=last_row, column=1).font = Font(size=9, color='808080', bold=True, underline="single")
#     worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
#     workbook.save(response)

#     return response    
    
@login_required
def exportUserCrateReport(request, crate_date, user_id):
    if user_id!='0':
        print(type(user_id))
        users = SpUsers.objects.filter(id=user_id,user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name')
    else:
        users = SpUsers.objects.filter(user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name')
    crate_lists                     = []
    if users:
        for user in users:
            month_lists                             = {} 
            month_lists['opening_balance']          = getOpeningCratesSum(user['id'],crate_date)
            month_lists['dispatch']                 = getTotalDispatchedCratesSums(user['id'], crate_date)
            month_lists['receiving']                = getTotalReceivedCratesSums(user['id'], crate_date)
            total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
            month_lists['total_outstanding']        = total_outstanding
            month_lists['today']                    =  str(datetime.strptime(str(crate_date),'%Y-%m-%d').strftime('%d/%m/%Y'))
            month_lists['store_name']               = getModelColumnByColumnId(SpUsers,'id',user['id'], 'store_name') 
            month_lists['emp_sap_id']               = getModelColumnById(SpUsers, user['id'], 'emp_sap_id')
            month_lists['party_name']               = getUserName(user['id'])
            if SpUserAreaAllocations.objects.filter(user_id =  user['id']).exists():
                month_lists['city_name']                = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user['id'], 'town_name')
            else:
                month_lists['city_name'] = '-'
            crate_lists.append(month_lists) 

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=user-crate-report.xlsx'.format(
        date=crate_date,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'CRATE STATEMENT'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    column_length = 7
    
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'CRATE STATEMENT'
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(size=15, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")


    # color_codes = []
    columns = []
    columns += [ 'Date' ]
    columns += [ 'Party Name' ]
    columns += [ 'City' ]
    columns += [ 'Opening Balance' ]
    columns += [ 'Dispatch' ]
    columns += [ 'Receiving' ]
    columns += [ 'Total Outstanding' ]
    row_num = 2
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20


    for id, crate in enumerate(crate_lists):
        row_num += 1  
        row = []
        row += [ crate['today'] ]
        row += [ crate['store_name']+" "+"("+crate['party_name']+"/"+crate['emp_sap_id']+")" ]
        row += [ crate['city_name'] ]
        row += [ crate['opening_balance'] ]
        row += [ crate['dispatch'] ]
        row += [ crate['receiving'] ]
        row += [ crate['total_outstanding'] ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            cell.fill = PatternFill()
            cell.font = Font(size=12, color='000000')
    row_num += 1
    last_row = row_num
    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=9, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response    


#user crate summary
@login_required
def userMonthlyCrateReport(request):
    today   = date.today()  
    year  = today.year
    month = today.month

    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    if user:
        for id, months in enumerate(month_list):
            month_lists                         = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['opening_balance']          = getOpeningCratesSum(user['id'], month_date)
            month_lists['dispatch']                 = getTotalDispatchedCratesSums(user['id'], month_date)
            month_lists['receiving']                = getTotalReceivedCratesSums(user['id'], month_date)
            total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
            month_lists['total_outstanding']        = total_outstanding
            month_lists['store_name'] = getModelColumnById(SpUsers, user['id'], 'store_name') 
            month_lists['emp_sap_id'] = getModelColumnById(SpUsers, user['id'], 'emp_sap_id') 
            crate_lists.append(month_lists)

    
    if user['id']:
        party_name = getUserName(user['id'])
        city_name  = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user['id'], 'town_name')
    else:
        party_name = ''
        city_name  = ''
           
    context = {}
    context['user_list']                        = user_list
    context['party_name']                       = party_name
    context['city_name']                        = city_name
    context['crate_lists']                      = crate_lists
    context['today_date']                       = today.strftime("%m/%Y")
    context['month_list']                       = days_in_months(year,month)
    context['page_title']                       = "User Crate Report"

    template = 'crate/user-monthly-crate.html'
    return render(request, template, context)

@login_required
def ajaxUserMonthlyCrateReport(request):
    today   = date.today()  
    if request.GET['crate_date']:
        crate_date = request.GET['crate_date']
        crate_date = crate_date.split('/')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month) > 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)
           
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(id=request.GET['user_id'],user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    if user:
        for id, months in enumerate(month_list):
            month_lists                             = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['opening_balance']          = getOpeningCratesSum(request.GET['user_id'], month_date)
            month_lists['dispatch']                 = getTotalDispatchedCratesSums(request.GET['user_id'], month_date)
            month_lists['receiving']                = getTotalReceivedCratesSums(request.GET['user_id'], month_date)
            total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
            month_lists['total_outstanding']        = total_outstanding
            month_lists['store_name'] = getModelColumnById(SpUsers, user['id'], 'store_name') 
            month_lists['emp_sap_id'] = getModelColumnById(SpUsers, user['id'], 'emp_sap_id')             
            crate_lists.append(month_lists)

    
    if user['id']:
        party_name = getUserName(user['id'])
        city_name  = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user['id'], 'town_name')
    else:
        party_name = ''
        city_name  = ''

    context = {}
    context['user_list']                        = user_list
    context['party_name']                       = party_name
    context['city_name']                        = city_name
    context['crate_lists']                      = crate_lists
    context['today_date']                       = today.strftime("%m/%Y")
    context['month_list']                       = days_in_months(year,month)

    template = 'crate/ajax-user-monthly-crate.html'
    return render(request, template, context) 

#export view
@login_required
def exportUserMonthlyCrateReport(request, crate_date, user_id):
    today   = date.today()  
    # crates_details = SpUserCrates.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).order_by('id')
    if crate_date:
        crate_date = crate_date.split('-')
        year  = int(crate_date[1])
        month = int(crate_date[0])
    else:
        year  = today.year
        if int(month)> 9:
            month = today.month 
        else:
            month = '0'+str(today.month)
            month = int(month)

    user_id = int(user_id)       
    user_list = SpUsers.objects.filter(user_type=2, status=1)
    month_list = days_in_months(year,month)

    try:
        user = SpUsers.objects.filter(id=user_id,user_type=2, status=1).values('id', 'first_name', 'middle_name', 'last_name').first()
    except SpUsers.DoesNotExist:
        user = None

    crate_lists                     = []
    if user:
        for id, months in enumerate(month_list):
            month_lists                             = {} 
            month_date                              = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            month_lists['month_date']               = str(months)
            month_lists['opening_balance']          = getOpeningCratesSum(user_id, month_date)
            month_lists['dispatch']                 = getTotalDispatchedCratesSums(user_id, month_date)
            month_lists['receiving']                = getTotalReceivedCratesSums(user_id, month_date)
            total_outstanding                       = (int(month_lists['opening_balance'])+int(month_lists['dispatch']))-int(month_lists['receiving'])
            month_lists['total_outstanding']        = total_outstanding
            month_lists['store_name'] = getModelColumnById(SpUsers, user['id'], 'store_name') 
            month_lists['emp_sap_id'] = getModelColumnById(SpUsers, user['id'], 'emp_sap_id')
            
            crate_lists.append(month_lists)

    
    if user_id:
        party_name = getUserName(user_id)
        city_name  = getModelColumnByColumnId(SpUserAreaAllocations, 'user_id', user_id, 'town_name')
    else:
        party_name = ''
        city_name  = ''

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=user-crate-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'CRATE STATEMENT'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
    img.height = 50
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    column_length = 7
    
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'CRATE STATEMENT'
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(size=15, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")


    # color_codes = []
    columns = []
    columns += [ 'Date' ]
    columns += [ 'Party Name' ]
    columns += [ 'City' ]
    columns += [ 'Opening Balance' ]
    columns += [ 'Dispatch' ]
    columns += [ 'Receiving' ]
    columns += [ 'Total Outstanding' ]
    row_num = 2
    
    

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.fill = PatternFill()
        cell.font = Font(size=11, color='000000', bold=True) 
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20


    for id, crate in enumerate(crate_lists):
        row_num += 1
        # Define the data for each cell in the row    
        row = []
        row += [ crate['month_date'] ]
        row += [ crate['store_name']+" "+"("+party_name+"/"+crate['emp_sap_id']+")" ]
        row += [ city_name ]
        row += [ crate['opening_balance'] ]
        row += [ crate['dispatch'] ]
        row += [ crate['receiving'] ]
        row += [ crate['total_outstanding'] ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            cell.fill = PatternFill()
            cell.font = Font(size=12, color='000000')


    row_num += 1
    last_row = row_num
    

    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=column_length)

    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Shreedhi'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=9, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")
    
    workbook.save(response)

    return response    
    
 
 