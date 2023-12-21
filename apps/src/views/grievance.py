import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password,check_password
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
import time



from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage

# Create your views here.

@login_required
def index(request):
    context = {}
    page = request.GET.get('page')
    grievances =  list(SpGrievance.objects.raw('''SELECT sp_grievance.*,sp_orders.id as order_id,sp_orders.order_code,
    sp_users.first_name,sp_users.middle_name,sp_users.last_name,
    sp_users.emp_sap_id,sp_users.store_name,sp_users.store_image,sp_users.profile_image
    FROM sp_grievance left join sp_orders on sp_orders.id = sp_grievance.order_id  
    left join sp_users on sp_users.id = sp_grievance.user_id order by sp_grievance.id desc '''))
    paginator = Paginator(grievances, getConfigurationResult('page_limit'))

    try:
        drivers = paginator.page(page)
    except PageNotAnInteger:
        drivers = paginator.page(1)
    except EmptyPage:
        drivers = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))   

    if(paginator.count == 0):
        paginator.count = 1
        
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    
    context['grievances']          = grievances
    context['page_limit']             = getConfigurationResult('page_limit')
    context['page_title'] = "Grievance Management"

    first_grievance = SpGrievance.objects.raw('''SELECT sp_grievance.*,sp_orders.id as order_id,sp_orders.order_code,
    sp_users.first_name,sp_users.middle_name,sp_users.last_name,
    sp_users.emp_sap_id,sp_users.store_name,sp_users.store_image,sp_users.profile_image
    FROM sp_grievance left join sp_orders on sp_orders.id = sp_grievance.order_id  
    left join sp_users on sp_users.id = sp_grievance.user_id order by sp_grievance.id desc LIMIT 1 ''')

    if first_grievance :
        context['first_grievance'] = first_grievance[0]
    else : 
        context['first_grievance'] = []

    
    template = 'grievance/grievances.html'
    return render(request, template, context)



@login_required
def grievanceDetails(request,grievance_id):
    context = {}
    first_grievance = SpGrievance.objects.raw('''SELECT sp_grievance.*,sp_orders.id as order_id,sp_orders.order_code,
    sp_users.first_name,sp_users.middle_name,sp_users.last_name,
    sp_users.emp_sap_id,sp_users.store_name,sp_users.store_image,sp_users.profile_image
    FROM sp_grievance left join sp_orders on sp_orders.id = sp_grievance.order_id  
    left join sp_users on sp_users.id = sp_grievance.user_id WHERE sp_grievance.id=%s  ''',[grievance_id])

    if first_grievance :
        context['first_grievance'] = first_grievance[0]
    else : 
        context['first_grievance'] = []

    template = 'grievance/grievance-short-details.html'
    return render(request, template, context)


@login_required
def ajaxGrievanceList(request):
    page = request.GET.get('page')
    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))   
    
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpOrganizations.objects.order_by('-id').first()
    
    context = {}
    context['organizations']          = organizations
    context['total_pages']            = total_pages
    context['organization_details']   = organization_details

    template = 'grievance/ajax-grievances.html'
    return render(request, template, context)

@login_required
def exportGrievanceToXls(request, columns):
    column_list = columns.split (",")
    grievances =  list(SpGrievance.objects.raw('''SELECT sp_grievance.*,sp_orders.id as order_id,sp_orders.order_code,
    sp_users.first_name,sp_users.middle_name,sp_users.last_name,
    sp_users.emp_sap_id,sp_users.store_name,sp_users.store_image,sp_users.profile_image
    FROM sp_grievance left join sp_orders on sp_orders.id = sp_grievance.order_id  
    left join sp_users on sp_users.id = sp_grievance.user_id order by sp_grievance.id desc '''))
    

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=grievances.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Grievances'
    
    # Define the titles for columns
    columns = []

    if 'user_name' in column_list:
        columns += [ 'User Name' ]

    if 'order_id' in column_list:
        columns += [ 'Order#' ]
    
    if 'reason' in column_list:
        columns += [ 'Reason' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for grievance in grievances :
        row_num += 1
        # Define the data for each cell in the row 
        if grievance.order_code is not None:
            order_code = grievance.order_code
        else:
            order_code = ''
        if grievance.reason_name is not None:
            reason_name = grievance.reason_name
        else:
            reason_name = ''        
        row = []
        if 'user_name' in column_list:
            row += [ grievance.first_name + ' ' + grievance.middle_name + ' ' + grievance.last_name ]

        if 'order_id' in column_list:
            row += [ order_code ]
        
        if 'reason' in column_list:
            row += [ reason_name ] 
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response    

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None

@login_required
def exportGrievanceToPDF(request, columns):
    column_list = columns.split (",")
    grievances =  list(SpGrievance.objects.raw('''SELECT sp_grievance.*,sp_orders.id as order_id,sp_orders.order_code,
    sp_users.first_name,sp_users.middle_name,sp_users.last_name,
    sp_users.emp_sap_id,sp_users.store_name,sp_users.store_image,sp_users.profile_image
    FROM sp_grievance left join sp_orders on sp_orders.id = sp_grievance.order_id  
    left join sp_users on sp_users.id = sp_grievance.user_id order by sp_grievance.id desc '''))

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('grievance/grievances_template.html', {'grievances': grievances, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'grievances.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 