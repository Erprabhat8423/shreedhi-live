import sys
import os
import openpyxl
import math
import time,json
import calendar
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q,F
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from django.core import serializers

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.hashers import make_password

# Create your views here.

# User List View
@login_required
def index(request):
    page = request.GET.get('page')
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    usercount = SpUsers.objects.filter(user_type=2).exclude(id=1).count()
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = math.ceil(paginator.count/getConfigurationResult('page_limit')) 
    
    if(paginator.count == 0):
        paginator.count = 1

          
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages
        
    oganizations    = SpOrganizations.objects.filter(status=1)
    context = {}
    context['users'] = users
    context['oganizations'] = oganizations
    context['usercount'] = usercount
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')

    #non operational
    page = request.GET.get('non_page')
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = math.ceil(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages
        
    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages

    #employee
    page = request.GET.get('employee_page')
    users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = math.ceil(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages
        
    context['employee_users'] = users
    context['employee_total_pages'] = total_pages

    first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
    ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    left join sp_addresses on sp_addresses.user_id = sp_users.id  
    where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id!=%s order by id desc LIMIT 1 ''',[1,'correspondence', 1])

    if first_employee :
        context['first_employee'] = first_employee[0]
        first_employee_permanent_address = SpAddresses.objects.get(user_id=first_employee[0].id,type='permanent')
    else : 
        context['first_employee'] = []
        first_employee_permanent_address = None

    context['first_employee_permanent_address'] = first_employee_permanent_address
    context['total_distributor'] = SpUsers.objects.filter(is_distributor=1).count()
    context['total_super_stockist'] = SpUsers.objects.filter(is_super_stockist=1).count()
    context['total_retailer'] = SpUsers.objects.filter(is_retailer=1).count()

    context['total_tagged_distributor'] = SpUsers.objects.filter(is_distributor=1,is_tagged=1).count()
    context['total_tagged_super_stockist'] = SpUsers.objects.filter(is_super_stockist=1,is_tagged=1).count()
    context['total_tagged_retailer'] = SpUsers.objects.filter(is_retailer=1,is_tagged=1).count()

    town_data = []
    towns = SpTowns.objects.all()
    for town in towns:

        distributors = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as distributor_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_distributor = %s and sp_user_area_allocations.town_id = %s ''',[1,town.id])[0]
        town.distributor_count = distributors.distributor_count

        super_stockist = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as super_stockist_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_super_stockist = %s and sp_user_area_allocations.town_id = %s ''',[1,town.id])[0]
        town.super_stockist_count = super_stockist.super_stockist_count

        retailers = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as retailers_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_retailer = %s and sp_user_area_allocations.town_id = %s ''',[1,town.id])[0]

        town.retailer_count = retailers.retailers_count
    
        town_data.append(town)
        
    ho_roles =  SpRoles.objects.filter(status=1).exclude(id=8).exclude(id=9).exclude(id=10)
    context['ho_roles'] = ho_roles
    
    context['towns'] = town_data
    context['page_title'] = "Manage Users"
    template = 'user-management/index.html'
    return render(request, template, context)

#ajax operational user list
@login_required
def ajaxOperationalUsersList(request):
    page = request.GET.get('page')
    employee_search = request.GET.get('employee_search')
    Organisation = request.GET.get('Organisation')

    users = SpUsers.objects.all().filter(user_type=2)
    if employee_search:
        users = users.filter(store_name__icontains=employee_search)
    if Organisation:
        users = users.filter(organization_id=Organisation)

    users = users.exclude(id=1).order_by('-id')    
    usercount = users.count()
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = math.ceil(paginator.count/getConfigurationResult('page_limit')) 
    
    if(paginator.count == 0):
        paginator.count = 1

          
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context = {}
    context['users'] = users
    context['total_pages'] = total_pages
    context['usercount'] = usercount
    context['page_limit'] = getConfigurationResult('page_limit')
    context['page_loading_type'] = request.GET.get('page_loading_type')
    template = 'user-management/ajax-operational-users-list.html'
    return render(request, template, context)

#ajax non operational user list
@login_required
def ajaxNonOperationalUsersList(request):
    page = request.GET.get('non_page')
    employee_search = request.GET.get('employee_search')
    
    users = SpUsers.objects.all().filter(user_type=3)
    if employee_search:
        users = users.filter(store_name__icontains=employee_search)
            
    users = users.exclude(id=1).order_by('-id')

    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = math.ceil(paginator.count/getConfigurationResult('page_limit')) 
    
    if(paginator.count == 0):
        paginator.count = 1

          
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context = {}
    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages
    context['page_loading_type'] = request.GET.get('page_loading_type')
    template = 'user-management/ajax-non-operational-users-list.html'
    return render(request, template, context)


#ajax employee user list
@login_required
def ajaxEmployeeUsersList(request):
    page = request.GET.get('employee_page')
    role_id = request.GET.get('role_id')
    employee_search = request.GET.get('employee_search')

    users = SpUsers.objects.all().filter(user_type=1)
    if role_id:
        users = users.filter(role_id=role_id)
    if employee_search:
        name = employee_search.split()
        if len(name) == 1:
            users = users.filter(Q(first_name__icontains=employee_search) | Q(middle_name__contains=employee_search) | Q(last_name__icontains=employee_search))
        if len(name) == 3:
            users = users.filter(first_name__icontains=name[0], middle_name__icontains=name[1], last_name__icontains=name[2])
        if len(name) == 2:
            users = users.filter(Q(middle_name__contains=name[1]) | Q(last_name__icontains=name[1]))
            
    users = users.exclude(id=1).order_by('-id')   
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()
    
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = math.ceil(paginator.count/getConfigurationResult('page_limit')) 
    
    if(paginator.count == 0):
        paginator.count = 1

          
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context = {}
    context['employee_users'] = users
    context['employee_total_pages'] = total_pages
    context['page_loading_type'] = request.GET.get('page_loading_type')
    template = 'user-management/ajax-employee-users-list.html'
    
    return render(request, template, context)

@login_required
def userGeoTagged(request):
    context = {}
  
    try:
        user_coordinates = SpUsers.objects.get(id=request.GET['id'])
    except SpUserAttendanceLocations.DoesNotExist:
        user_coordinates = None
           
    context['user_coordinates'] = user_coordinates
    context['google_app_key']   = getConfigurationResult('google_app_key')
    template = 'user-management/user-geo-tagged.html'

    return render(request, template,context)

# User basic details View
@login_required
def addUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    context = {}
    context['contact_types']    = contact_types
    context['countries']        = countries
    context['country_codes']    = country_codes

    template = 'user-management/add-user-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = '123456'
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password    

            error_count = 0
            if request.POST['last_user_id'] != '' and request.POST['official_email'] !='':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            elif request.POST['official_email'] !='':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
            else:
                user_exists = 0

            if user_exists:
                error_count = 1
                error_response['official_email_error'] = "Email already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                if bool(request.FILES.get('store_image', False)) == True:
                    if request.POST['previous_store_image'] != '':
                        deleteMediaFile(request.POST['previous_store_image'])
                    uploaded_store_image = request.FILES['store_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    store_image_name = uploaded_store_image.name
                    temp = store_image_name.split('.')
                    store_image_name = 'store_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    store_image = storage.save(store_image_name, uploaded_store_image)
                    store_image = storage.url(store_image)

                else:
                    if request.POST['previous_store_image'] != '':
                        store_image = request.POST['previous_store_image'] 
                    else:
                        store_image = None    
                
                if bool(request.FILES.get('profile_image', False)) == True:
                    if request.POST['previous_profile_image'] != '':
                        deleteMediaFile(request.POST['previous_profile_image'])
                    uploaded_profile_image = request.FILES['profile_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    profile_image_name = uploaded_profile_image.name
                    temp = profile_image_name.split('.')
                    profile_image_name = 'profile_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    profile_image = storage.save(profile_image_name, uploaded_profile_image)
                    profile_image = storage.url(profile_image)
                else:
                    if request.POST['previous_profile_image'] != '':
                        profile_image = request.POST['previous_profile_image'] 
                    else:
                        profile_image = None        
                
                if request.POST['last_user_id'] != '':
                    SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactPersons.objects.filter(user_id=request.POST['last_user_id']).delete()

                    user = SpUsers.objects.get(id=request.POST['last_user_id'])
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    if request.POST['store_name']:
                        user.user_type = 2
                    user.save()
                    last_user_id = request.POST['last_user_id']
                    user_inserted = 1
                else:
                    user = SpUsers()
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    user.password       = make_password(str(password))
                    if request.POST['store_name']:
                        user.user_type = 2
                    user.save()
                    last_user_id = user.id
                    user_inserted = 0
                    #sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
                    
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    user_contact_no         = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
                    if int(is_primary[id]) > 0:
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

            
                contact_person_names = request.POST.getlist('contact_person_name[]')
                designations = request.POST.getlist('designation[]')
                contact_numbers = request.POST.getlist('contact_number[]')

                for id, val in enumerate(contact_person_names):
                    user_contact_person         = SpContactPersons()
                    user_contact_person.user_id = last_user_id
                    if contact_person_names[id]!='':
                        user_contact_person.contact_person_name = contact_person_names[id]
                    if designations[id]!='':
                        user_contact_person.designation         = designations[id]
                    if contact_numbers[id]!='':
                        user_contact_person.contact_number      = contact_numbers[id]      
                    user_contact_person.save()

                if request.POST['last_user_id'] != '':
                    basic = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save()
                else:
                    basic = SpBasicDetails()
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save() 

                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id           = last_user_id
                permanent.type              = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                oganizations    = SpOrganizations.objects.filter(status=1)
                working_shifts  = SpWorkingShifts.objects.all()
                zones           = SpZones.objects.all()
                routes          = SpRoutes.objects.all()
                user_details    = SpUsers.objects.get(id=last_user_id)

                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=last_user_id)
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None
                
                try:
                    user_basic_details = SpBasicDetails.objects.get(user_id=last_user_id)
                except SpBasicDetails.DoesNotExist:
                    user_basic_details = None
                
                try:
                    departments = SpDepartments.objects.filter(organization_id=user_details.organization_id)
                except SpDepartments.DoesNotExist:
                    departments = None
                
                if user_details.department_id == 33:
                    try:
                        roles = SpRoles.objects.filter(department_id=3).filter(id__in=[8,9])
                    except SpRoles.DoesNotExist:
                        roles = None
                else:
                    try:
                        roles = SpRoles.objects.filter(department_id=user_details.department_id).filter(id__in=[8,9])
                    except SpRoles.DoesNotExist:
                        roles = None
                
                if user_area_allocations is None:
                    towns = None
                else:
                    towns = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)

                if user_inserted == 0:
                    user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' added'
                    activity    = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' added by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    
                    saveActivity('Users Management', 'User', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
                else:
                    user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' updated'
                    activity    = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    
                    saveActivity('Users Management', 'User', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')

                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,last_user_id,'firebase_token')
                employee_name = getUserName(last_user_id)

                message_title = "Profile updated"
                message_body = "Your profile has been updated by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    # send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(last_user_id,'SpUsers','User Management','Profile updated',message_title,message_body,notification_image,request.user.id,user_name,last_user_id,employee_name,'profile.png',2,'app.png',1)
                #-----------------------------save notification block----------------------------#
                if user_area_allocations:
                    vehicle         = SpVehicles.objects.filter(route_id = user_area_allocations.route_id)
                else:
                    vehicle         = SpVehicles.objects.all()
                context = {}
                context['oganizations']             = oganizations
                context['working_shifts']           = working_shifts
                context['vehicle'] = vehicle
                context['zones']                    = zones
                context['routes']                   = routes
                context['user_details']             = user_details
                context['user_area_allocations']     = user_area_allocations
                context['user_basic_details']       = user_basic_details
                context['departments']              = departments
                context['roles']                    = roles
                context['towns']                    = towns
                context['last_user_id']             = last_user_id
                context['production_unit']          = SpProductionUnit.objects.all()
                
                template = 'user-management/add-user-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            print(e)
            response['message'] = e
            return HttpResponse(e)
    return render(request, template, context)

# Edit User basic details View
@login_required
def editUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    try:
        user_basic_details = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    except SpBasicDetails.DoesNotExist:
        user_basic_details = None

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_contact_details    = SpContactNumbers.objects.filter(user_id=request.GET['last_user_id'])
    user_basic_details      = user_basic_details
    user_store_address      = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='correspondence').first()
    user_permanent_address  = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='permanent').first()
    user_contact_person     = SpContactPersons.objects.filter(user_id=request.GET['last_user_id'])

    if user_store_address:
        store_states = SpStates.objects.filter(country_id=user_store_address.country_id)
        store_cities = SpCities.objects.filter(state_id=user_store_address.state_id)
        if user_permanent_address:
            permanent_states = SpStates.objects.filter(country_id=user_permanent_address.country_id)
            permanent_cities = SpCities.objects.filter(state_id=user_permanent_address.state_id)
        else:
            permanent_states = None
            permanent_cities = None    
    else:
        store_states = None
        store_cities = None
        permanent_states = None
        permanent_cities = None

    context = {}
    context['contact_types']            = contact_types
    context['countries']                = countries
    context['country_codes']            = country_codes
    context['user_details']             = user_details
    context['user_contact_details']     = user_contact_details
    context['user_basic_details']       = user_basic_details
    context['user_store_address']       = user_store_address
    context['user_permanent_address']   = user_permanent_address
    context['user_contact_person']      = user_contact_person
    context['store_states']             = store_states
    context['store_cities']             = store_cities
    context['permanent_states']         = permanent_states
    context['permanent_cities']         = permanent_cities
    context['last_user_id']             = request.GET['last_user_id']
    template = 'user-management/add-user-basic-detail.html'
    
    return render(request, template, context)    

# User offical details View
@login_required
def addUserOfficalDetail(request):
    template = 'user-management/add-user-offical-detail.html'
   
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
            else:
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "SAP ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role & map product
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                    mapProductToUser(user_data.id)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']
                if request.POST['role_id'] == '8':
                    user_data.is_distributor = 1
                    user_data.is_super_stockist = 0
                else:
                    user_data.is_distributor = 0
                    user_data.is_super_stockist = 1    
                user_data.save()

                

                try:
                    user_basic_detail = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                except SpBasicDetails.DoesNotExist:
                    user_basic_detail = None

                if user_basic_detail is None:        
                    user_basic_details                  = SpBasicDetails()
                else:
                    user_basic_details                  = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])

                user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                user_basic_details.pan_number           = request.POST['pan_number']
                user_basic_details.cin                  = request.POST['cin']
                user_basic_details.gstin                = request.POST['gstin']
                user_basic_details.fssai                = request.POST['fssai']
                user_basic_details.working_shift_id     = request.POST['working_shift_id']
                user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                user_basic_details.order_timing         = request.POST['order_timing']
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_basic_details.outstanding_amount   = request.POST['outstanding_amount']
                user_basic_details.security_amount      = request.POST['security_amount']
                user_basic_details.production_unit_id   = request.POST['production_unit_id']
                user_basic_details.tcs_applicable       = request.POST['tcs_applicable']
                user_basic_details.vehicle_id = request.POST['vehicle_id']
                user_basic_details.vehilcle_number = getModelColumnById(SpVehicles, request.POST['vehicle_id'], 'registration_number')
                user_basic_details.tcs_value            = 0 if (request.POST['tcs_applicable'] == "0" or request.POST['tcs_value'] == "") else request.POST['tcs_value']
                user_basic_details.per_crate_incentive  = 0 if (request.POST['per_crate_incentive'] == "") else request.POST['per_crate_incentive']
                user_basic_details.save()
                if SpUserLedger.objects.filter(user_id = request.POST['last_user_id']).exists():
                    pass
                else:
                    ledger = SpUserLedger()
                    ledger.user_id = request.POST['last_user_id']
                    ledger.order_id = '0'
                    ledger.invoice_no = '0'
                    ledger.organization_id = request.POST['organization_id']
                    ledger.credit = 0
                    ledger.debit = 0
                    ledger.particulars = 'Opening Balanace'
                    ledger.order_date = date.today().strftime("%Y-%m-%d")
                    ledger.balance = request.POST['outstanding_amount']
                    ledger.created_by = request.user.id
                    ledger.save()
                
                
                
                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None

                if user_area_allocations is None:        
                    area_allocation = SpUserAreaAllocations()
                else:
                    area_allocation = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])

                area_allocation.user_id     = request.POST['last_user_id']
                area_allocation.state_id     = getModelColumnById(SpZones,request.POST['zone_id'],'state_id')
                area_allocation.state_name   = getModelColumnById(SpZones,request.POST['zone_id'],'state_name')
                area_allocation.zone_id     = request.POST['zone_id']
                area_allocation.zone_name   = getModelColumnById(SpZones,request.POST['zone_id'],'zone')
                area_allocation.town_id     = request.POST['town_id']
                area_allocation.town_name   = getModelColumnById(SpTowns,request.POST['town_id'],'town')
                area_allocation.route_id    = request.POST['route_id']
                area_allocation.route_name  = getModelColumnById(SpRoutes,request.POST['route_id'],'route')
                area_allocation.save()
                
                try:
                    user_variants = SpUserProductVariants.objects.filter(user_id=request.POST['last_user_id'])
                except SpUserProductVariants.DoesNotExist:
                    user_variants = None

                user_details = SpUsers.objects.filter(id=request.POST['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]

                # message = "Your Password is 123456 for sales port app login. Kindly reset your password"
                # sendSMS('ENQARY',getModelColumnById(SpUsers,request.POST['last_user_id'],'primary_contact_number'),message)

                context                  = {}
                context['last_user_id']  = request.POST['last_user_id']
                context['user_variants'] = user_variants
                context['user_details']  = user_details
                template                 = 'user-management/add-user-product-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)

# Edit User basic details View
@login_required
def editUserOfficalDetail(request):
    oganizations    = SpOrganizations.objects.filter(status=1)
    working_shifts  = SpWorkingShifts.objects.all()
    zones           = SpZones.objects.all()
    routes          = SpRoutes.objects.all()

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_area_allocations   = SpUserAreaAllocations.objects.get(user_id=request.GET['last_user_id'])
    user_basic_details      = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    
    departments     = SpDepartments.objects.filter(organization_id=user_details.organization_id)
    if user_details.department_id == 33:
        roles           = SpRoles.objects.filter(department_id=3).filter(id__in=[8,9])
    else:
        roles           = SpRoles.objects.filter(department_id=user_details.department_id).filter(id__in=[8,9])
    towns           = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)
    if user_area_allocations:
        vehicle         = SpVehicles.objects.filter(route_id = user_area_allocations.route_id)
    else:
        vehicle         = SpVehicles.objects.all()
    context = {}
    context['oganizations']             = oganizations
    context['working_shifts']           = working_shifts
    context['zones']                    = zones
    context['vehicle']                  = vehicle
    context['routes']                   = routes
    context['user_details']             = user_details
    context['user_area_allocations']     = user_area_allocations
    context['user_basic_details']       = user_basic_details
    context['departments']              = departments
    context['roles']                    = roles
    context['towns']                    = towns
    context['last_user_id']             = request.GET['last_user_id']
    context['production_unit']          = SpProductionUnit.objects.all()
    template = 'user-management/add-user-offical-detail.html'
    
    return render(request, template, context)      

@login_required
def updateProductStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpUserProductVariants.objects.get(id=id)
            data.status = is_active
            data.save()                
            
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/add-user-product-detail')

# User product details View
@login_required
def addUserProductDetail(request):
    context = {}
    template = 'user-management/add-user-product-detail.html'
    if request.method == "POST":
        try:
            user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
        except SpUserDocuments.DoesNotExist:
            user_documents = None
        context['user_documents']   = user_documents
        context['last_user_id']     = request.POST['last_user_id']
        template = 'user-management/add-user-document-detail.html'
    return render(request, template, context)

# User product details View
@login_required
def editUserProductDetail(request):
    try:
        user_variants = SpUserProductVariants.objects.filter(user_id=request.GET['last_user_id'])
    except SpUserProductVariants.DoesNotExist:
        user_variants = None

    user_details = SpUsers.objects.filter(id=request.GET['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]

    context = {}
    context['last_user_id']  = request.GET['last_user_id']
    context['user_variants'] = user_variants
    context['user_details']  = user_details
    template = 'user-management/add-user-product-detail.html'
    return render(request, template, context) 

# User document details View
@login_required
def addUserDocumentDetail(request):
    template = 'user-management/add-user-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                if request.POST['previous_aadhaar_card'] != '':
                        deleteMediaFile(request.POST['previous_aadhaar_card'])
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                aadhaar_card_name = uploaded_aadhaar_card.name
                temp = aadhaar_card_name.split('.')
                aadhaar_card_name = 'aadhaar_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                aadhaar_card = storage.save(aadhaar_card_name, uploaded_aadhaar_card)
                aadhaar_card = storage.url(aadhaar_card)
            else:
                if request.POST['previous_aadhaar_card'] != '':
                        aadhaar_card = request.POST['previous_aadhaar_card'] 
                else:
                    aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:
                if request.POST['previous_pan_card'] != '':
                        deleteMediaFile(request.POST['previous_pan_card'])        
                uploaded_pan_card = request.FILES['pan_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                pan_card_name = uploaded_pan_card.name
                temp = pan_card_name.split('.')
                pan_card_name = 'pan_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                pan_card = storage.save(pan_card_name, uploaded_pan_card)
                pan_card = storage.url(pan_card)
            else:
                if request.POST['previous_pan_card'] != '':
                        pan_card = request.POST['previous_pan_card'] 
                else:
                    pan_card = None

            if bool(request.FILES.get('cin', False)) == True:
                if request.POST['previous_cin'] != '':
                        deleteMediaFile(request.POST['previous_cin'])
                uploaded_cin = request.FILES['cin']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                cin_name = uploaded_cin.name
                temp = cin_name.split('.')
                cin_name = 'cin_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                cin = storage.save(cin_name, uploaded_cin)
                cin = storage.url(cin)
            else:
                if request.POST['previous_cin'] != '':
                        cin = request.POST['previous_cin'] 
                else:
                    cin = None

            if bool(request.FILES.get('gstin', False)) == True:
                if request.POST['previous_gstin'] != '':
                        deleteMediaFile(request.POST['previous_gstin'])
                uploaded_gstin = request.FILES['gstin']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                gstin_name = uploaded_gstin.name
                temp = gstin_name.split('.')
                gstin_name = 'gstin_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                gstin = storage.save(gstin_name, uploaded_gstin)
                gstin = storage.url(gstin)
            else:
                if request.POST['previous_gstin'] != '':
                        gstin = request.POST['previous_gstin'] 
                else:
                    gstin = None

            if bool(request.FILES.get('fssai', False)) == True:
                if request.POST['previous_fssai'] != '':
                        deleteMediaFile(request.POST['previous_fssai'])
                uploaded_fssai = request.FILES['fssai']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                fssai_name = uploaded_fssai.name
                temp = fssai_name.split('.')
                fssai_name = 'fssai_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                fssai = storage.save(fssai_name, uploaded_fssai)
                fssai = storage.url(fssai)
            else:
                if request.POST['previous_fssai'] != '':
                        fssai = request.POST['previous_fssai'] 
                else:
                    fssai = None        
            
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None

            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.fssai         = fssai
                documents.save()

                response['error'] = False
                response['message'] = "Record has been saved successfully."
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.fssai         = fssai
                documents.save()

                response['error'] = False
                response['message'] = "Record has been updated successfully."

            return JsonResponse(response)
        except Exception as e:
            response['error']            = True
            response['message']          = e
            return HttpResponse(e)
    return render(request, template)
    

# Employee basic details View
@login_required
def addEmployeeBasicDetail(request):
    contact_types = SpContactTypes.objects.filter(status=1)
    countries = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    context = {}
    context['contact_types'] = contact_types
    context['countries']     = countries
    context['country_codes'] = country_codes
    template = 'user-management/add-employee-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = '123456'
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password
            
            error_count = 0 
            user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
                return JsonResponse(response)
            else:    
                if bool(request.FILES.get('profile_image', False)) == True:
                    uploaded_profile_image = request.FILES['profile_image']
                    pfs = FileSystemStorage()
                    profile_image = pfs.save(uploaded_profile_image.name, uploaded_profile_image)
                else:
                    profile_image = False

                user = SpUsers()
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                if profile_image:
                    user.profile_image = pfs.url(profile_image)
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.user_type = 1
                user.password       = make_password(str(password))
                user.save()
                last_user_id = user.id
                #sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
            
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if int(is_primary[id]) > 0:
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()

                basic = SpBasicDetails()
                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

            
                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                oganizations                = SpOrganizations.objects.filter(status=1)
                working_shifts              = SpWorkingShifts.objects.all()
                zones                       = SpZones.objects.all()
                routes                      = SpRoutes.objects.all()

                context                     = {}
                context['oganizations']     = oganizations
                context['working_shifts']   = working_shifts
                context['zones']            = zones
                context['routes']           = routes
                context['last_user_id']     = last_user_id
                context['production_unit']  = SpProductionUnit.objects.all()

                template = 'user-management/add-employee-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))
    return render(request, template, context)

# Employee offical details View
@login_required
def addEmployeeOfficalDetail(request):
    template = 'user-management/add-employee-offical-detail.html'
    
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else: 
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']

                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

                user_data.save()

                user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                user_basic_details.pan_number           = request.POST['pan_number']
                user_basic_details.working_shift_id     = request.POST['working_shift_id']
                user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['leave_count']:
                    user_basic_details.leave_count          = request.POST['leave_count']
                else:
                    user_basic_details.leave_count          = 0
                user_basic_details.week_of_day          = request.POST['week_of_day']   
                user_basic_details.production_unit_id   = ','.join([str(elem) for elem in request.POST.getlist('production_unit_id[]')]) 
                user_basic_details.save()

                towns    = request.POST.getlist('town_id[]')
                for id, val in enumerate(towns):
                    area_allocation = SpUserAreaAllocations()
                    area_allocation.user_id = request.POST['last_user_id']
                    if towns[id] != '':
                        zone_id = getModelColumnById(SpTowns,towns[id],'zone_id')
                        area_allocation.zone_id                 =   zone_id
                        area_allocation.zone_name               = getModelColumnById(SpZones,zone_id,'zone')
                        area_allocation.state_id     = getModelColumnById(SpZones,zone_id,'state_id')
                        area_allocation.state_name   = getModelColumnById(SpZones,zone_id,'state_name')
                        area_allocation.town_id = towns[id]
                        area_allocation.town_name               = getModelColumnById(SpTowns,towns[id],'town')
                    area_allocation.save()

                context = {}
                context['last_user_id'] = request.POST['last_user_id']
                
                distributors = SpUsers.objects.raw(''' select sp_users.id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                from sp_user_area_allocations 
                left join sp_users on sp_users.id = sp_user_area_allocations.user_id
                where sp_user_area_allocations.town_id in 
                (select town_id from sp_user_area_allocations as sura where sura.user_id = %s ) 
                and  (sp_users.is_distributor = %s or sp_users.is_super_stockist = %s)
                ''',[request.POST['last_user_id'], 1,1])
                
                # distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
                # from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
                if distributors:
                    context['distributors'] = distributors
                else:
                    context['distributors'] = None
                template = 'user-management/add-employee-attendance.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template,response)

# Employee attendance details View
@login_required
def addEmployeeAttendanceDetail(request):
    if request.method == "POST":

        user                = SpUsers.objects.get(id=request.POST['last_user_id'])
        user.periphery      = request.POST['periphery']
        user.timing         = request.POST['timing']
        user.save()

        context = {}
        context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-document-detail.html'
        return render(request, template, context)
    else:
        context = {}
        user_attendance_locations = SpUsers.objects.get(id=request.POST['last_user_id'])
        context['user_attendance_locations'] = user_attendance_locations
       
        context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-attendance.html'
        return render(request, template, context)

    

# Employee document details View
@login_required
def addEmployeeDocumentDetail(request):
    template = 'user-management/add-employee-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                aadhaar = FileSystemStorage()
                aadhaar_card = aadhaar.save(uploaded_aadhaar_card.name, uploaded_aadhaar_card)
                aadhaar_card = aadhaar.url(aadhaar_card)
            else:
                aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:        
                uploaded_pan_card = request.FILES['pan_card']
                pan = FileSystemStorage()
                pan_card = pan.save(uploaded_pan_card.name, uploaded_pan_card)
                pan_card = pan.url(pan_card)
            else:
                pan_card = None

            documents               = SpUserDocuments()
            documents.user_id       = request.POST['last_user_id']
            documents.aadhaar_card  = aadhaar_card
            documents.pan_card      = pan_card
            documents.save()
            response['error'] = False
            response['message'] = "Record has been updated successfully."

            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)



# Employee basic details View
@login_required
def editEmployeeBasicDetail(request, employee_id):
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            else:
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
  
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
            
                if bool(request.FILES.get('profile_image', False)) == True:
                    uploaded_profile_image = request.FILES['profile_image']
                    pfs = FileSystemStorage()
                    profile_image = pfs.save(uploaded_profile_image.name, uploaded_profile_image)
                    profile_image = pfs.url(profile_image)
                else:
                    if request.POST['previous_profile_image'] != '':
                        profile_image = request.POST['previous_profile_image'] 
                    else:
                        profile_image = None

                SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()

                user = SpUsers.objects.get(id=request.POST['last_user_id'])
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                user.profile_image = profile_image
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.save()
                last_user_id = request.POST['last_user_id']

            
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if int(is_primary[id]) > 0:
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
            
                basic                       = SpBasicDetails.objects.get(user_id=last_user_id)
                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

            
                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                
                user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,request.POST['last_user_id'],'firebase_token')
                employee_name = getUserName(request.POST['last_user_id'])

                message_title = "Profile updated"
                message_body = "Your profile has been updated by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(request.POST['last_user_id'],'SpUsers','User Management','Profile updated',message_title,message_body,notification_image,request.user.id,user_name,request.POST['last_user_id'],employee_name,'profile.png',2,'app.png',1,1)
                #-----------------------------save notification block----------------------------#

                response['error'] = False
                response['last_user_id'] = last_user_id
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))

    else : 
        contact_types = SpContactTypes.objects.filter(status=1)
        countries = SpCountries.objects.all()
        country_codes   = SpCountryCodes.objects.filter(status=1)

        context = {}
        context['contact_types'] = contact_types
        context['countries']     = countries
        context['country_codes'] = country_codes

        employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
        sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        where sp_users.id = %s''',[employee_id])
        
        try:
            employee_correspondence_address = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
        except SpAddresses.DoesNotExist:
            employee_correspondence_address = None

        try:
            employee_permanent_address = SpAddresses.objects.get(user_id=employee_id,type='permanent')
        except SpAddresses.DoesNotExist:
            employee_permanent_address = None

        try:
            employee_contact_numbers = SpContactNumbers.objects.filter(user_id=employee_id)
        except SpContactNumbers.DoesNotExist:
            employee_contact_numbers = None    

        try:
            employee_user_allocation = SpUserAreaAllocations.objects.filter(user_id=employee_id)
        except SpUserAreaAllocations.DoesNotExist:
            employee_user_allocation = None

        try:
            if employee_correspondence_address is not None:
                employee_store_states = SpStates.objects.filter(country_id=employee_correspondence_address.country_id)
            else:
                employee_store_states = None    
        except SpStates.DoesNotExist:
            employee_store_states = None    

        try:
            if employee_correspondence_address is not None:
                employee_store_cities = SpCities.objects.filter(state_id=employee_correspondence_address.state_id)
            else:
                employee_store_cities = None
        except SpCities.DoesNotExist:
            employee_store_cities = None 
        
        try:
            if employee_permanent_address is not None:
                employee_permanent_states = SpStates.objects.filter(country_id=employee_permanent_address.country_id)
            else:
                employee_permanent_states = None    
        except SpStates.DoesNotExist:
            employee_permanent_states = None

        try:
            if employee_permanent_address is not None:
                employee_permanent_cities = SpCities.objects.filter(state_id=employee_permanent_address.state_id)
            else:
                employee_permanent_cities = None    
        except SpCities.DoesNotExist:
            employee_permanent_cities = None

        if employee:
            context['employee'] = employee[0]
            context['employee_correspondence_address']  = employee_correspondence_address
            context['employee_permanent_address']       = employee_permanent_address
            context['user_contacts']                    = employee_contact_numbers
            context['user_areas']                       = employee_user_allocation
            context['store_states']                     = employee_store_states
            context['store_cities']                     = employee_store_cities
            context['permanent_states']                 = employee_permanent_states
            context['permanent_cities']                 = employee_permanent_cities
            context['last_user_id']                     = employee_id
            try:
                user_documents                              = SpUserDocuments.objects.get(user_id=employee_id)
                context['user_documents'] = user_documents
            except SpUserDocuments.DoesNotExist:
                context['user_documents'] = None

            context['user_attendance_locations'] = SpUserAttendanceLocations.objects.filter(user_id=employee_id,status=1)
            template = 'user-management/edit-employee/employee-basic-detail.html'
            return render(request, template, context)
        else:
            return HttpResponse('Employee not found')

# Employee offical details View
@login_required
def editEmployeeOfficalDetail(request,employee_id):
    error_response = {}
    response = {}
    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['SAPID_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:  
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)
                    
                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name')
                user_data.emp_sap_id            = request.POST['emp_sap_id']  
                

                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

                user_data.save()

                if request.POST['last_user_id'] != '':
                    user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    user_basic_details.pan_number           = request.POST['pan_number']
                    user_basic_details.working_shift_id     = request.POST['working_shift_id']
                    user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    if request.POST['leave_count']:
                        user_basic_details.leave_count          = request.POST['leave_count']
                    else:
                        user_basic_details.leave_count          = 0
                    user_basic_details.week_of_day          = request.POST['week_of_day'] 
                    user_basic_details.production_unit_id   = ','.join([str(elem) for elem in request.POST.getlist('production_unit_id[]')]) 
                    user_basic_details.save()
                else:
                    user_basic_details                      = SpBasicDetails()
                    user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    user_basic_details.pan_number           = request.POST['pan_number']
                    user_basic_details.working_shift_id     = request.POST['working_shift_id']
                    user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    if request.POST['leave_count']:
                        user_basic_details.leave_count          = request.POST['leave_count']
                    else:
                        user_basic_details.leave_count          = 0
                    user_basic_details.week_of_day          = request.POST['week_of_day']    
                    user_basic_details.production_unit_id   = ','.join([str(elem) for elem in request.POST.getlist('production_unit_id[]')]) 
                    user_basic_details.save()
                
                SpUserAreaAllocations.objects.filter(user_id=request.POST['last_user_id']).delete()
                
                towns    = request.POST.getlist('town_id[]')
                if(len(towns)):
                    for id, val in enumerate(towns):
                        if towns[id] != '':
                            area_allocation                         = SpUserAreaAllocations()
                            area_allocation.user_id                 = request.POST['last_user_id']
                            zone_id                                 = getModelColumnById(SpTowns,towns[id],'zone_id')
                            area_allocation.zone_id                 =   zone_id
                            area_allocation.zone_name               = getModelColumnById(SpZones,zone_id,'zone')
                            area_allocation.state_id                = getModelColumnById(SpZones,zone_id,'state_id')
                            area_allocation.state_name              = getModelColumnById(SpZones,zone_id,'state_name')
                            area_allocation.town_id                 = towns[id]
                            area_allocation.town_name               = getModelColumnById(SpTowns,towns[id],'town')
                            area_allocation.save()

                
                response = {}
                response['error'] = False
                response['last_user_id'] = request.POST['last_user_id']
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    else:
        
        template = 'user-management/edit-employee/employee-offical-detail.html'
        context = {}

        employee_details = SpUsers.objects.get(id=employee_id)
        employee_area_allocations = SpUserAreaAllocations.objects.filter(user_id=employee_id)
        emp_zones = []
        emp_towns = []
        for area_allocation in employee_area_allocations:
            emp_zones.append(area_allocation.zone_id)
            emp_towns.append(area_allocation.town_id)
        
        context['emp_zones'] = emp_zones = list(set(emp_zones))
        context['emp_towns'] = emp_towns

        employee_basic_details = SpBasicDetails.objects.get(user_id=employee_id)
        working_shifts = SpWorkingShifts.objects.all()
        departments = SpDepartments.objects.filter(organization_id=employee_details.organization_id)
        roles = SpRoles.objects.filter(department_id=employee_details.department_id)
        reporting_users = SpUsers.objects.filter(role_id=employee_details.role_id)
        zones = SpZones.objects.filter()
        for zone in zones:
            zone.towns = SpTowns.objects.filter(zone_id=zone.id)

        oganizations = SpOrganizations.objects.filter(status=1)
        working_shifts = SpWorkingShifts.objects.all()

        

        context['oganizations']                 = oganizations
        context['working_shifts']               = working_shifts
        context['employee_details']             = employee_details
        context['employee_area_allocations']    = employee_area_allocations
        context['employee_basic_details']       = employee_basic_details
        context['departments']                  = departments
        context['roles']                        = roles
        context['reporting_users']              = reporting_users 
        context['zones']                        = zones
        context['last_user_id']                 = employee_id
        context['production_unit']              = SpProductionUnit.objects.all() 
        
        return render(request, template,context)

# Employee attendance details View
@login_required
def editEmployeeAttendanceDetail(request,employee_id):
    if request.method == "POST":
        user                = SpUsers.objects.get(id=request.POST['last_user_id'])
        user.periphery      = request.POST['periphery']
        user.timing         = request.POST['timing']
        user.save()
                
        response = {}
        response['error'] = False
        response['last_user_id'] = request.POST['last_user_id']
        return JsonResponse(response)
    else:
        context = {}
        user_attendance_locations = SpUsers.objects.get(id=employee_id)
        context['user_attendance_locations'] = user_attendance_locations
       
        context['last_user_id'] = employee_id
        template = 'user-management/edit-employee/employee-attendance.html'
        return render(request, template, context)

# Employee document details View
@login_required
def editEmployeeDocumentDetail(request,employee_id):
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                if request.POST['previous_aadhar_card']:
                    deleteMediaFile(request.POST['previous_aadhar_card'])
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                aadhaar_card_name = uploaded_aadhaar_card.name
                temp = aadhaar_card_name.split('.')
                aadhaar_card_name = 'aadhaar_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                aadhaar_card = storage.save(aadhaar_card_name, uploaded_aadhaar_card)
                aadhaar_card = storage.url(aadhaar_card)
            else:
                if request.POST['previous_aadhar_card'] != '':
                    aadhaar_card = request.POST['previous_aadhar_card'] 
                else:
                    aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:
                if request.POST['previous_pan_card']:
                    deleteMediaFile(request.POST['previous_pan_card'])  
                uploaded_pan_card = request.FILES['pan_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                pan_card_name = uploaded_pan_card.name
                temp = pan_card_name.split('.')
                pan_card_name = 'pan_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                pan_card = storage.save(pan_card_name, uploaded_pan_card)
                pan_card = storage.url(pan_card)
            else:
                if request.POST['previous_pan_card'] != '':
                    pan_card = request.POST['previous_pan_card'] 
                else:
                    pan_card = None
                    
           
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None
            
            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.save()

                response['error'] = False
                response['message'] = "Record has been saved successfully."
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.save()

                response['error'] = False
                response['message'] = "Record has been updated successfully."
            
            return JsonResponse(response)

        except Exception as e:
            return HttpResponse(e)
    else:
        try:
            employee_documents = SpUserDocuments.objects.get(user_id=employee_id)
        except SpUserDocuments.DoesNotExist:
            employee_documents = None

        template = 'user-management/edit-employee/employee-document-detail.html'
        response = {}
        response['last_user_id'] = employee_id
        response['employee_documents'] = employee_documents
        return render(request, template,response)




@login_required
def userShortDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender,
    sp_basic_details.outstanding_amount,sp_basic_details.opening_crates,
    sp_addresses.address_line_1,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,
    sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_addresses on sp_addresses.user_id = sp_users.id   
    where sp_addresses.type=%s  and sp_users.id = %s ''',['correspondence',user_id])
    if user :
        context['user'] = user[0]
        context['contact_persons'] = SpContactPersons.objects.filter(user_id=user_id)
    else : 
        context['user'] = []
      
    template = 'user-management/user-short-details.html'
    return render(request, template,context)

def getFreeSchemes(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'variant_name')+''
        if free_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'product_id')
            free += str(free_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if free_scheme.container_quantity>0:
            free += ' and '     
        if free_scheme.pouch_quantity>0:
            if free_scheme.pouch_quantity == 1:
                free += ' '+str(free_scheme.pouch_quantity)+' free Pouch'
            else:    
                free += ' '+str(free_scheme.pouch_quantity)+' free Pouches'
              
        free_scheme = free
    else:
        free_scheme = None  
    return free_scheme


def getQuantitativeScheme(order_id, user_id):
    try:
        quantitative_scheme = SpOrderSchemes.objects.get(order_id=order_id, scheme_type='quantitative', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        quantitative_scheme = None
    if quantitative_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, quantitative_scheme.free_variant_id, 'variant_name')+'-'
        if quantitative_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, quantitative_scheme.free_variant_id, 'product_id')
            free += str(quantitative_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if quantitative_scheme.pouch_quantity>0:
            free += ' and '+str(quantitative_scheme.pouch_quantity)+' free Pouches'
        free += ' under the '+getModelColumnById(SpSchemes, quantitative_scheme.scheme_id, 'name')+' Scheme'      
        quantitative_scheme = free
    else:
        quantitative_scheme = None  
    return quantitative_scheme

def getFlatScheme(order_id, user_id):
    try:
        flat_scheme = SpOrderSchemes.objects.get(order_id=order_id, scheme_type='flat', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        free = ''
        free += str(flat_scheme.incentive_amount)+' Incentive amount has been applied under the '+getModelColumnById(SpFlatSchemes, flat_scheme.scheme_id, 'name')+' Scheme'      
        flat_scheme = free
    else:
        flat_scheme = None  
    return flat_scheme    

def getFlatSchemeByVariant(product_variant_id, order_id, user_id):
    try:
        flat_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='flat', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        flat = 'Discount of Rs. '+str(flat_scheme.incentive_amount)+' applied.'
        flat_scheme = flat
    else:
        flat_scheme = None  
    return flat_scheme

def getFlatSchemeByVariants(product_variant_id, order_id, user_id):
    try:
        flat_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='flat', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        flat_scheme = None
    if flat_scheme:
        flat = flat_scheme.incentive_amount
        flat_scheme = flat
    else:
        flat_scheme = None  
    return flat_scheme 
    
def getBulkpackScheme(order_id, user_id):
    try:
        bulk_pack_scheme = SpOrderSchemes.objects.get(order_id=order_id, scheme_type='bulkpack', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        bulk_pack_scheme = None
    if bulk_pack_scheme:
        free = ''
        free += str(bulk_pack_scheme.incentive_amount)+' Incentive amount has been applied under the '+getModelColumnById(SpBulkpackSchemes, bulk_pack_scheme.scheme_id, 'name')+' Scheme'      
        bulk_pack_scheme = free
    else:
        bulk_pack_scheme = None  
    return bulk_pack_scheme

def getOrderFreeSchemes(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free = 'Free '+getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'variant_name')+'-'
        if free_scheme.container_quantity>0:
            product_id             = getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'product_id')
            free += str(free_scheme.container_quantity)+' free '+getModelColumnById(SpProducts, product_id, 'container_name')+''
        if free_scheme.container_quantity>0:
            free += ' and '     
        if free_scheme.pouch_quantity>0:
            if free_scheme.pouch_quantity == 1:
                free += ' '+str(free_scheme.pouch_quantity)+' free Pouch'
            else:    
                free += ' '+str(free_scheme.pouch_quantity)+' free Pouches'
              
        free_scheme = free
    else:
        free_scheme = None  
    return free_scheme

def getOrderFreeScheme(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.get(order_id=order_id, variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        free_scheme = (int(getModelColumnById(SpProductVariants, free_scheme.free_variant_id, 'no_of_pouch'))*int(free_scheme.container_quantity))+int(free_scheme.pouch_quantity)  
    else:
        free_scheme = 0
    return free_scheme

def getFreeScheme(product_variant_id, order_id, user_id):
    try:
        free_scheme = SpOrderSchemes.objects.filter(order_id=order_id, free_variant_id=product_variant_id, scheme_type='free', user_id=user_id)
    except SpOrderSchemes.DoesNotExist:
        free_scheme = None
    if free_scheme:
        sum_free_scheme = []
        for scheme in free_scheme:
            scheme = (int(getModelColumnById(SpProductVariants, scheme.free_variant_id, 'no_of_pouch'))*int(scheme.container_quantity))+int(scheme.pouch_quantity)  
        sum_free_scheme.append(scheme) 
        free_scheme = sum(sum_free_scheme)
    else:
        free_scheme = 0
    return free_scheme

#get order details view
@login_required
def getOrderDetails(request):
    id                          = request.GET.get('id')

    if request.user.role_id == 0:
        order_details  = SpOrders.objects.get(id=id)
    else:
        order_details = SpOrders.objects.raw('''SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id 
    where sp_orders.id = %s order by id desc LIMIT 1 ''',[request.GET.get('id')])[0] 

    order_details.profile_image = getModelColumnById(SpUsers, order_details.user_id, 'profile_image')
    order_details.user_details  = SpBasicDetails.objects.get(user_id=order_details.user_id)
    order_details.user_address  = SpAddresses.objects.get(user_id=order_details.user_id, type='permanent')

    order_item_list = SpOrderDetails.objects.filter(order_id=id)
    quantitative_scheme           = getQuantitativeScheme(id, getModelColumnById(SpOrders, id, 'user_id'))      
    flat_scheme                   = 0 
    bulk_scheme                   = getBulkpackScheme(id, getModelColumnById(SpOrders, id, 'user_id'))         
    for order_item in order_item_list:
        order_item.free_schemes        = getOrderFreeSchemes(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
        order_item.free_scheme         = getOrderFreeScheme(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
        order_item.flat_scheme         = getFlatSchemeByVariant(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
        order_item.flat_schemes         = getFlatSchemeByVariants(order_item.product_variant_id, order_item.order_id, getModelColumnById(SpOrders, order_item.order_id, 'user_id'))
        
    crate_sum  = SpOrderDetails.objects.filter(order_id=id,product_container_type='Crate').aggregate(Sum('quantity'))['quantity__sum']
    matki_sum  = SpOrderDetails.objects.filter(order_id=id,product_container_type='Matki').aggregate(Sum('quantity'))['quantity__sum']
    
    total_incentive = getFlatBulkIncentive(id, getModelColumnById(SpOrders, id, 'user_id'))
    find_order_amount = round(float(order_details.order_total_amount)-float(total_incentive), 2)
    
    context = {}
    context['order_details']             = order_details
    context['order_item_list']           = order_item_list
    context['quantitative_scheme']       = quantitative_scheme
    context['flat_scheme']               = flat_scheme
    context['bulk_scheme']               = bulk_scheme
    context['crate_sum']                 = crate_sum
    context['matki_sum']                 = matki_sum
    context['total_incentive']           = total_incentive
    context['find_order_amount']         = find_order_amount
    context['role_id']                   = request.user.role_id

    template = 'user-management/get-user-order-details.html'
    return render(request, template, context)

@login_required
def userDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
    sp_basic_details.gender,sp_basic_details.personal_email,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.date_of_joining,sp_basic_details.working_shift_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number,sp_basic_details.vehilcle_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    where sp_users.id = %s''',[user_id])[0]
    
    try:
        documents = SpUserDocuments.objects.get(user_id=user_id)
    except SpUserDocuments.DoesNotExist:
        documents = None

    try:
        allocation = SpUserAreaAllocations.objects.get(user_id=user_id)
    except SpUserAreaAllocations.DoesNotExist:
        allocation = None
            
    try:
        user_variants = SpUserProductVariants.objects.filter(user_id=user_id)
    except SpUserProductVariants.DoesNotExist:
        user_variants = None

    try:
        permanent_address = SpAddresses.objects.get(user_id=user_id,type='permanent')
    except SpAddresses.DoesNotExist:
        permanent_address = None

    try:
        user_order_details = SpOrders.objects.filter(user_id=user_id).order_by('-id')
    except SpOrders.DoesNotExist:
        user_order_details = None         
    if user_order_details:
        order_details = SpOrders.objects.filter(user_id=user_id).order_by('-id')[0]
        quantitative_scheme           = getQuantitativeScheme(order_details.id, user_id)      
        flat_scheme                   = 0 
        bulk_scheme                   = getBulkpackScheme(order_details.id, user_id)         
        user_order_item_details = SpOrderDetails.objects.filter(order_id=order_details.id)  
        for order_item in user_order_item_details:
            order_item.free_schemes        = getOrderFreeSchemes(order_item.product_variant_id, order_item.order_id, user_id)
            order_item.free_scheme         = getOrderFreeScheme(order_item.product_variant_id, order_item.order_id, user_id)
            order_item.flat_scheme         = getFlatSchemeByVariant(order_item.product_variant_id, order_item.order_id, user_id)
            order_item.flat_schemes         = getFlatSchemeByVariants(order_item.product_variant_id, order_item.order_id, user_id)
            
        crate_sum  = SpOrderDetails.objects.filter(order_id=order_details.id,product_container_type='Crate').aggregate(Sum('quantity'))['quantity__sum']
        matki_sum  = SpOrderDetails.objects.filter(order_id=order_details.id,product_container_type='Matki').aggregate(Sum('quantity'))['quantity__sum']
        total_incentive = getFlatBulkIncentive(order_details.id, user_id)
        find_order_amount = round(float(order_details.order_total_amount)-float(total_incentive), 2)
        
    else:
        order_details           = None
        user_order_item_details = None
        crate_sum               = None
        matki_sum               = None
        quantitative_scheme     = None
        flat_scheme             = None
        bulk_scheme             = None
        total_incentive         = None
        find_order_amount       = None

    try:
        user_scheme_details = SpOrderSchemes.objects.filter(user_id=user_id).order_by('-id')
    except SpOrders.DoesNotExist:
        user_scheme_details = None
    if user_scheme_details:
        for scheme_details in user_scheme_details:
            scheme_details.order_code       = getModelColumnById(SpOrders, scheme_details.order_id, 'order_code')
            if scheme_details.variant_id:
                scheme_details.product_variant  = getModelColumnById(SpProductVariants, scheme_details.variant_id, 'variant_name')
            if scheme_details.free_variant_id:
                scheme_details.free_variant_name  = getModelColumnById(SpProductVariants, scheme_details.free_variant_id, 'variant_name')

    
    if SpIncentive.objects.filter(user_id=user_id).exists():
        incentive = SpIncentive.objects.filter(user_id=user_id).first()
        productClassList = SpProductClass.objects.filter(status=1)
        for productClass in productClassList:
            
            productClass.levelFlag = 1
            incentiveDetails = SpIncentiveDetails.objects.filter(incentive_id=incentive.id,incentive_type=1,class_product_variant_id=productClass.id).first()
            if incentiveDetails is None:
                is_slab = 0
            else:
                is_slab = incentiveDetails.is_slab
            productClass.is_slab = is_slab
            if is_slab == 1:
                incentiveSlabDetails = SpIncentiveSlabDetails.objects.filter(incentive_detail_id=incentiveDetails.id).all()
                productClass.incentiveSlabDetails = incentiveSlabDetails
            productClass.ss_incentive = incentiveDetails.ss_incentive if incentiveDetails else 0
            productClass.distributor_incentive = incentiveDetails.distributor_incentive if incentiveDetails else 0
            
            produClassSlabDetails = SpSlabMasterList.objects.filter(product_class_id=productClass.id).all()
            productClass.produClassSlabDetails = produClassSlabDetails
             
            productList = SpProducts.objects.filter(status=1,product_class_id=productClass.id)
            productClass.productlevelFlag = 0
            for product in productList:
                incentiveDetails = SpIncentiveDetails.objects.filter(incentive_id=incentive.id,incentive_type=2,class_product_variant_id=product.id).first()
                if incentiveDetails:
                        if incentiveDetails.ss_incentive > 0 or incentiveDetails.distributor_incentive > 0:
                            productClass.productlevelFlag = 1
                            
                if incentiveDetails is None:
                    is_slab = 0
                else:
                    productClass.levelFlag = 2
                    is_slab = incentiveDetails.is_slab
                product.is_slab = is_slab
                if is_slab == 1:
                    incentiveSlabDetails = SpIncentiveSlabDetails.objects.filter(incentive_detail_id=incentiveDetails.id).all()
                    product.incentiveSlabDetails = incentiveSlabDetails
                else:
                    product.ss_incentive = incentiveDetails.ss_incentive if incentiveDetails else 0
                    product.distributor_incentive = incentiveDetails.distributor_incentive if incentiveDetails else 0
                
                product.variant_levelFlag = 0
                productVariantList = SpProductVariants.objects.filter(status=1,product_id=product.id)
                for productVariant in productVariantList:
                    incentiveDetails = SpIncentiveDetails.objects.filter(incentive_id=incentive.id,incentive_type=3,class_product_variant_id=productVariant.id).first()
                    if incentiveDetails:
                        if incentiveDetails.ss_incentive > 0 or incentiveDetails.distributor_incentive > 0:
                            product.variant_levelFlag = 1
                            
                        
                    if incentiveDetails is None:
                        is_slab = 0
                    else:
                        productClass.levelFlag = 3
                        is_slab = incentiveDetails.is_slab
                    productVariant.is_slab = is_slab
                    if is_slab == 1:
                        incentiveSlabDetails = SpIncentiveSlabDetails.objects.filter(incentive_detail_id=incentiveDetails.id).all()
                        productVariant.incentiveSlabDetails = incentiveSlabDetails
                    else:
                        productVariant.ss_incentive = incentiveDetails.ss_incentive if incentiveDetails else 0
                        productVariant.distributor_incentive = incentiveDetails.distributor_incentive if incentiveDetails else 0
                    
                product.variant=productVariantList
            productClass.product=productList
    else: 
        productClassList = SpProductClass.objects.filter(status=1)
        for productClass in productClassList:
            productList = SpProducts.objects.filter(status=1,product_class_id=productClass.id)
            for product in productList:
                productVariantList = SpProductVariants.objects.filter(status=1,product_id=product.id)
                product.variant=productVariantList
            productClass.product=productList
        
    incentive=[]
    if SpIncentive.objects.filter(user_id=user_id).exists():
        incentive = SpIncentive.objects.filter(user_id=user_id).first()
        
    context['user'] = user
    context['user_correspondence_address'] = SpAddresses.objects.get(user_id=user_id,type='correspondence')
    context['user_permanent_address']      = permanent_address
    context['contact_persons']             = SpContactPersons.objects.filter(user_id=user_id)
    context['user_contacts']               = SpContactNumbers.objects.filter(user_id=user_id)
    context['area_allocated']              = allocation 
    context['user_documents']              = documents
    context['user_variants']               = user_variants
    context['user_order_details']          = user_order_details
    context['user_order_item_details']     = user_order_item_details
    context['crate_sum']                   = crate_sum
    context['matki_sum']                   = matki_sum
    context['order_details']               = order_details
    context['quantitative_scheme']         = quantitative_scheme
    context['flat_scheme']                 = flat_scheme
    context['bulk_scheme']                 = bulk_scheme
    context['total_incentive']             = total_incentive 
    context['find_order_amount']           = find_order_amount 
    context['user_attendance_locations']   = SpUserAttendanceLocations.objects.filter(user_id=user_id,status=1)
    context['user_scheme_details']         = user_scheme_details
    context['productClassList']            = productClassList
    context['incentive']                   = incentive
    template = 'user-management/user-details.html'

    return render(request, template,context)


@login_required
def employeeShortDetail(request,employee_id):
    context = {}
    employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender,sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
    ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    left join sp_addresses on sp_addresses.user_id = sp_users.id  
    where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id = %s ''',[1,'correspondence',employee_id])
    if employee :
        context['employee'] = employee[0]
    else : 
        context['employee'] = []

    context['employee_permanent_address'] = SpAddresses.objects.get(user_id=employee_id,type='permanent')    
    template = 'user-management/employee-short-details.html'
    return render(request, template,context)

@login_required
def employeeDetail(request,employee_id):
    context = {}
    context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
    sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    where sp_users.id = %s''',[employee_id])[0]

    try:
        documents = SpUserDocuments.objects.get(user_id=employee_id)
    except SpUserDocuments.DoesNotExist:
        documents = None

    try:
        user_coordinates = SpUserAttendanceLocations.objects.filter(user_id=employee_id,status=1)
        for user_coordinate in user_coordinates:
            user_coordinate.coordinate = SpUsers.objects.filter(id=user_coordinate.distributor_ss_id).values('latitude', 'longitude').first()
    except SpUserAttendanceLocations.DoesNotExist:
        user_coordinates = None
        
    context['employee_correspondence_address']  = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
    context['employee_permanent_address']       = SpAddresses.objects.get(user_id=employee_id,type='permanent')
    context['user_contacts']                    = SpContactNumbers.objects.filter(user_id=employee_id)
    context['user_areas']                       = SpUserAreaAllocations.objects.filter(user_id=employee_id)
    context['user_documents']                   = documents
    context['user_coordinates']                 = user_coordinates
    context['user_attendance_locations']        = SpUserAttendanceLocations.objects.filter(user_id=employee_id,status=1)
    context['google_app_key']                   = getConfigurationResult('google_app_key')
    template = 'user-management/employee-details.html'

    return render(request, template,context)

def getGroupedTownOptions(request):
    options = ''
    zone_ids = request.POST['zone_ids'].split(',')
    zones = SpZones.objects.raw(''' select * from sp_zones where id in %s ''',[zone_ids])
    for zone in zones:
        towns = SpTowns.objects.filter(zone_id=zone.id)
        if towns:
            options += '<optgroup label="' + zone.zone + '">'
            for town in towns : 
                options += "<option value="+str(town.id)+">"+town.town+"</option>"
            options += '</optgroup>'
    
    return HttpResponse(options)

def getReportingUserOptions(request, role_id):
    role = SpRoles.objects.get(id=role_id)
    options = '<option value="">Select role</option>'
    reporting_users = SpUsers.objects.filter(user_type=1,role_id=role.reporting_role_id)
    for reporting_user in reporting_users:
        options += "<option value="+str(reporting_user.id)+">"+reporting_user.first_name+" "+reporting_user.last_name+"</option>"

    return HttpResponse(options)

#update user status
@login_required
def updateUserStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpUsers.objects.get(id=id)
            data.status = is_active
            data.save()

            if is_active == '1':
                status = 'Unblock'
            else:
                AuthtokenToken.objects.filter(user_id=id).delete()
                status = 'Block'
                
            user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status
            activity    = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status+' by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            saveActivity('Users Management', 'Users', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')


#update user status
@login_required
def updateUsersStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpUsers.objects.get(id=id)
            data.purchase_milk_from_org = is_active
            data.save()                
            
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

#update user status
@login_required
def updateUserVariantPrice(request):

    response = {}
    if request.method == "POST":
        try:
            id              = request.POST.get('id')
            price           = request.POST.get('price')
            user_type       = request.POST.get('user_type')
            data = SpUserProductVariants.objects.get(id=id)
            data.sp_user = price
            previous_price = data.sp_user 
            data.save()
            container_price = float(float(price) * int(data.no_of_pouch))
            
            SpUserProductVariants.objects.filter(id=id).update(container_sp_user=F('sp_user')*F('no_of_pouch'))

            response['error'] = False
            response['message'] = "Record has been updated successfully."
            response['id'] = id
            response['price'] = price
            response['container_price'] = container_price
            response['user_type'] = user_type
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = str(e)
        return JsonResponse(response)
    return redirect('/users')

#export to excel operational user list
@login_required
def exportOperationalUserToXlsx(request, columns,Organisation):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    if Organisation!='0':
        users = users.filter(organization_id=Organisation)
    for user in users :
        try:
            user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)
        except SpBasicDetails.DoesNotExist:
            user.outstanding_amount = None
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Operational User'
    
    # Define the titles for columns
    columns = []

    if 'SAP_ID' in column_list:
        columns += [ 'SAP ID' ]
        
    if 'store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'role' in column_list:
        columns += [ 'Role' ]
    
    if 'contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'contact_no' in column_list:
        columns += [ 'Contact No.' ]

    if 'outstanding_amount' in column_list:
        columns += [ 'Outstanding Amount' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 
        if user.outstanding_amount:
            outstanding_amount = user.outstanding_amount.outstanding_amount 
        else:
            outstanding_amount = ''   
        row = []
        if 'SAP_ID' in column_list:
            # row += [ user.store_name ]
            row +=[ user.emp_sap_id ]
        if 'store_name' in column_list:
            # row += [ user.store_name ]
            row +=[ user.store_name]

        if 'role' in column_list:
            row += [ user.role_name ]
        
        if 'contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'contact_no' in column_list:
            row += [ user.primary_contact_number ]

        if 'outstanding_amount' in column_list:
            row += [ outstanding_amount ]           
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#export to pdf operational user list
@login_required
def exportOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

#export to excel non-operational user list
@login_required
def exportNonOperationalUserToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=non-operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Non-Operational-Users'
    
    # Define the titles for columns
    columns = []

    if 'non_store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'non_role' in column_list:
        columns += [ 'Role' ]
    
    if 'non_contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'non_contact_no' in column_list:
        columns += [ 'Contact No.' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'non_store_name' in column_list:
            row += [ user.store_name ]

        if 'non_role' in column_list:
            row += [ user.role_name ]
        
        if 'non_contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'non_contact_no' in column_list:
            row += [ user.primary_contact_number ]         
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

#export to pdf non operational user list
@login_required
def exportNonOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/non_operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Non-Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response

#export to excel employee list
@login_required
def exportEmployeeToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Employees'
    
    # Define the titles for columns
    columns = []

    if 'employee_name' in column_list:
        columns += [ 'Employee Name' ]

    if 'employee_role' in column_list:
        columns += [ 'Role' ]
    
    if 'employee_dep_org' in column_list:
        columns += [ 'Dept./Org.' ] 

    if 'employee_platform' in column_list:
        columns += [ 'Platform(web/mobile)' ]    

    if 'employee_last_sign_in' in column_list:
        columns += [ 'Last Login' ]

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        if user.last_login is not None:
            if user.web_auth_token is None:
                employee_platform = 'Web'
            else:
                employee_platform = 'APP'
        else:
            employee_platform = ''               
         
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'employee_name' in column_list:
            row += [ user.first_name + user.middle_name + user.last_name ]

        if 'employee_role' in column_list:
            row += [ user.role_name ]
        
        if 'employee_dep_org' in column_list:
            row += [ user.department_name + '/' + user.organization_name ] 

        if 'employee_platform' in column_list:
            row += [ employee_platform ]

        if 'employee_last_sign_in' in column_list:
            row += [ user.last_login ]             
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response

    #export to pdf non operational user list
@login_required
def exportEmployeeToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/employee_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Employees.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response

@login_required
def getUserMap(request):
    user_coordinates = SpUsers.objects.filter(id=request.GET.get('distributor_id')).values('latitude', 'longitude').first()
    
    context = {}
    context['user_coordinates'] = user_coordinates
    context['periphery']        = request.GET.get('periphery')
    context['google_app_key']   = getConfigurationResult('google_app_key')
    template = 'user-management/user-map.html'
    return render(request, template, context) 

@login_required
def importProductVariant(request):
    # workbook object is created 
    # wb_obj = load_workbook('media/operatonal-users.xlsx')

    # sheet_obj = wb_obj.active 
    # m_row = sheet_obj.max_row 
    
    # # Loop will print all values 
    # # of first column  
    
    # for i in range(1, m_row + 1): 
    #     row = [cell.value for cell in sheet_obj[i]] 
    #     print(row)
    #     template = ProductVariantTemplate()
    #     template.store_name = row[0]
    #     template.role = row[1]
    #     template.save()
            
    return HttpResponse('row')     


def updateUserRole(user_id,params):
    role_permissions = SpRolePermissions.objects.filter(role_id=params.POST['role_id'])
    SpUserRolePermissions.objects.filter(user_id=user_id).delete()
    SpUserRoleWorkflowPermissions.objects.filter(user_id=user_id).delete()
    if len(role_permissions):
        for role_permission in role_permissions:
            user_role_permission = SpUserRolePermissions()
            user_role_permission.user_id = user_id
            user_role_permission.role_id = params.POST['role_id']
            user_role_permission.module_id = role_permission.module_id
            user_role_permission.sub_module_id = role_permission.sub_module_id
            user_role_permission.permission_id = role_permission.permission_id
            user_role_permission.permission_slug = getModelColumnById(SpPermissions,role_permission.permission_id,'slug')
            
            user_role_permission.save()

        
        role_permission_workflows = SpRoleWorkflowPermissions.objects.filter(role_id=params.POST['role_id'])
        if len(role_permission_workflows):
            for role_permission_workflow in role_permission_workflows : 
                user_role_permission_wf = SpUserRoleWorkflowPermissions()
                user_role_permission_wf.user_id = user_id
                user_role_permission_wf.role_id = role_permission_workflow.role_id
                user_role_permission_wf.sub_module_id = role_permission_workflow.sub_module_id
                user_role_permission_wf.permission_id = role_permission_workflow.permission_id
                user_role_permission_wf.permission_slug = getModelColumnById(SpPermissions,role_permission_workflow.permission_id,'slug')
                user_role_permission_wf.level_id = role_permission_workflow.level_id
                user_role_permission_wf.level = role_permission_workflow.level
                user_role_permission_wf.description = role_permission_workflow.description
                user_role_permission_wf.workflow_level_dept_id = role_permission_workflow.workflow_level_dept_id
                user_role_permission_wf.workflow_level_role_id = role_permission_workflow.workflow_level_role_id
                user_role_permission_wf.status = role_permission_workflow.status
                user_role_permission_wf.save()


def mapProductToUser(user_id):
    product_variants = SpProductVariants.objects.all()
    if len(product_variants):
        
        SpUserProductVariants.objects.filter(user_id=user_id).delete()

        for product_variant in product_variants:
            user_product_variant                            = SpUserProductVariants()
            user_product_variant.user_id                    = user_id
            user_product_variant.product_id                 = product_variant.product_id
            user_product_variant.product_name               = product_variant.product_name
            user_product_variant.product_class_id           = getModelColumnById(SpProducts,product_variant.product_id,'product_class_id')
            user_product_variant.product_variant_id         = product_variant.id
            user_product_variant.item_sku_code              = product_variant.item_sku_code
            user_product_variant.variant_quantity           = product_variant.variant_quantity
            user_product_variant.variant_unit_id            = product_variant.variant_unit_id
            user_product_variant.variant_name               = product_variant.variant_name
            user_product_variant.variant_unit_name          = product_variant.variant_unit_name
            user_product_variant.largest_unit_name          = product_variant.largest_unit_name
            user_product_variant.variant_size               = product_variant.variant_size
            user_product_variant.no_of_pouch                = product_variant.no_of_pouch
            user_product_variant.container_size             = product_variant.container_size
            user_product_variant.is_bulk_pack               = product_variant.is_bulk_pack
            user_product_variant.included_in_scheme         = product_variant.included_in_scheme
            user_product_variant.mrp                        = product_variant.mrp
            user_product_variant.container_mrp              = float(product_variant.mrp) * float(product_variant.no_of_pouch)
            user_product_variant.sp_user                    = product_variant.sp_employee
            user_product_variant.container_sp_user          = float(product_variant.sp_employee) * float(product_variant.no_of_pouch)
            user_product_variant.valid_from                 = product_variant.valid_from
            user_product_variant.valid_to                   = product_variant.valid_to
            user_product_variant.status                     = product_variant.status
            user_product_variant.save()


@login_required
def viewUserRolePermission(request,user_id):
    if request.method == "POST":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        context = {}
        role_id = getModelColumnById(SpUsers,user_id,'role_id')
        role = SpRoles.objects.get(id=role_id)
        permissions = SpPermissions.objects.filter(status=1)
        organizations = SpOrganizations.objects.filter(status=1)
        departments = SpDepartments.objects.filter(status=1)

        other_departments = SpDepartments.objects.filter(status=1,organization_id=role.organization_id)
        for department in other_departments : 
            department.other_roles = SpRoles.objects.filter(status=1,department_id=department.id).exclude(id=role.id)

        modules = SpModules.objects.filter(status=1)
        for module in modules : 
            module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)
        
        context['permissions'] = permissions
        context['organizations'] = organizations
        context['modules'] = modules
        context['user_id'] = user_id
        context['user_name'] = getModelColumnById(SpUsers, user_id, 'first_name')+' '+getModelColumnById(SpUsers, user_id, 'middle_name')+' '+getModelColumnById(SpUsers, user_id, 'last_name')
        context['role'] = role
        context['other_departments'] = other_departments
        context['first_workflow_level'] =  SpWorkflowLevels.objects.get(priority='first')
        context['last_workflow_level'] =  SpWorkflowLevels.objects.get(priority='last')
        context['middle_workflow_level'] =  SpWorkflowLevels.objects.get(priority='middle')

        
        template = 'user-management/user-role-permission.html'
        return render(request,template,context)

@login_required
def updateUserRolePermission(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
       
        permissions = SpPermissions.objects.filter(status=1)
        sub_modules = SpSubModules.objects.filter(status=1)

        for sub_module in sub_modules :
            for permission in permissions :
                SpUserModulePermissions.objects.filter(user_id=request.POST['user_id'],sub_module_id=sub_module.id,permission_id=permission.id).delete()
                SpUserRolePermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'],sub_module_id=sub_module.id,permission_id=permission.id).delete()
                SpUserRoleWorkflowPermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'],sub_module_id=sub_module.id,permission_id=permission.id).delete()

                var_name = 'permission_'+ str(sub_module.id) + '_' + str(permission.id)
                if var_name in request.POST:
                    if request.POST['user_id']!='0':
                        module_permission = SpUserModulePermissions()
                        module_permission.user_id = request.POST['user_id']
                        module_permission.module_id = getModelColumnById(SpSubModules,sub_module.id,'module_id')
                        module_permission.sub_module_id = sub_module.id
                        module_permission.permission_id = permission.id
                        module_permission.permission_slug = getModelColumnById(SpPermissions,permission.id,'slug')
                        module_permission.save()
                        total_work_flows_var = 'workflow_'+str(sub_module.id)+'_'+str(permission.id)
                        if request.POST[total_work_flows_var] :

                            module_permission.workflow = request.POST[total_work_flows_var]
                            module_permission.save()

                    role_permission = SpUserRolePermissions()
                    role_permission.user_id = request.POST['user_id']
                    role_permission.role_id = request.POST['role_id']
                    role_permission.module_id = getModelColumnById(SpSubModules,sub_module.id,'module_id')
                    role_permission.sub_module_id = sub_module.id
                    role_permission.permission_id = permission.id
                    role_permission.permission_slug = getModelColumnById(SpPermissions,permission.id,'slug')
                    role_permission.save()
                    total_work_flows_var = 'workflow_'+str(sub_module.id)+'_'+str(permission.id)
                    from_date   = 'from_date_'+str(sub_module.id)+'_'+str(permission.id)
                    to_date     = 'to_date_'+str(sub_module.id)+'_'+str(permission.id)
                    if request.POST[total_work_flows_var] :

                        role_permission.workflow    = request.POST[total_work_flows_var]
                        if request.POST[from_date]:
                            role_permission.from_date   = datetime.strptime(str(request.POST[from_date]), '%d/%m/%Y').strftime('%Y-%m-%d')
                        if request.POST[to_date]:    
                            role_permission.to_date     = datetime.strptime(str(request.POST[to_date]), '%d/%m/%Y').strftime('%Y-%m-%d')
                        role_permission.save()

                        total_work_flows = json.loads(request.POST[total_work_flows_var])
                        
                        for total_work_flow in total_work_flows :

                            level_roles = total_work_flow['role_id'].split(',')
                            for level_role in level_roles:
                                if int(level_role) == int(request.POST['role_id']) :
                                    role_permission_level = SpUserRoleWorkflowPermissions()
                                    role_permission_level.user_id = request.POST['user_id']
                                    role_permission_level.role_id = request.POST['role_id']
                                    role_permission_level.sub_module_id = sub_module.id
                                    role_permission_level.permission_id = permission.id
                                    role_permission_level.permission_slug = getModelColumnById(SpPermissions,permission.id,'slug')
                                    role_permission_level.level_id = total_work_flow['level_id']
                                    role_permission_level.level = getModelColumnById(SpWorkflowLevels,total_work_flow['level_id'],'level')
                                    role_permission_level.description = total_work_flow['description']
                                    role_permission_level.workflow_level_dept_id = getModelColumnById(SpRoles,level_role,'department_id')
                                    role_permission_level.workflow_level_role_id = level_role
                                    if 'status' in total_work_flow :
                                        role_permission_level.status = total_work_flow['status']
                                    else:
                                        role_permission_level.status = 1

                                    role_permission_level.save()
                                    if request.POST[from_date]:
                                        role_permission_level.from_date   = datetime.strptime(str(request.POST[from_date]), '%d/%m/%Y').strftime('%Y-%m-%d')
                                    if request.POST[to_date]:    
                                        role_permission_level.to_date     = datetime.strptime(str(request.POST[to_date]), '%d/%m/%Y').strftime('%Y-%m-%d')
                                    role_permission_level.save()
        response['flag']    = True
        response['message'] = "Record has been updated successfully."
        return JsonResponse(response)
        
@login_required
def checkRolePermision(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
        if SpPermissionWorkflows.objects.filter(permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).exists():
            response['flag'] = True
            response['message'] = "Workflow is already applied to this permission."
        else :
            response['flag'] = False
            response['message'] = "Workflow is not applied to this permission. Please contact administator."
            
        return JsonResponse(response)
        
@login_required
def saveRolePermisionValidity(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
        role_permission = SpRolePermissions.objects.filter(permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).first()
        role_permission_workflows = SpRoleWorkflowPermissions.objects.filter(role_id=role_permission.role_id, permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id'])

        SpUserRolePermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'],permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).delete()
        
        user_role_permission = SpUserRolePermissions()
        user_role_permission.user_id = request.POST['user_id']
        user_role_permission.role_id = request.POST['role_id']
        user_role_permission.module_id = role_permission.module_id
        user_role_permission.sub_module_id = role_permission.sub_module_id
        user_role_permission.permission_id = role_permission.permission_id
        user_role_permission.permission_slug = getModelColumnById(SpPermissions,role_permission.permission_id,'slug')
        
        user_role_permission.from_date = datetime.strptime(request.POST['from_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        user_role_permission.to_date = datetime.strptime(request.POST['to_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        user_role_permission.save()

        if len(role_permission_workflows) :

            SpUserRoleWorkflowPermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'], permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).delete()

            for role_permission_workflow in role_permission_workflows :
                user_role_permission_level = SpUserRoleWorkflowPermissions()
                user_role_permission_level.user_id = request.POST['user_id']
                user_role_permission_level.role_id = request.POST['role_id']
                user_role_permission_level.sub_module_id = role_permission_workflow.sub_module_id
                user_role_permission_level.permission_id = role_permission_workflow.permission_id
                user_role_permission_level.permission_slug = getModelColumnById(SpPermissions,role_permission_workflow.permission_id,'slug')
                user_role_permission_level.level_id = role_permission_workflow.level_id
                user_role_permission_level.level = role_permission_workflow.level
                user_role_permission_level.description = role_permission_workflow.description
                if int(role_permission_workflow.level_id) > 0 :
                    user_role_permission_level.workflow_level_dept_id = role_permission_workflow.workflow_level_dept_id
                else:
                    user_role_permission_level.workflow_level_dept_id = None

                user_role_permission_level.workflow_level_role_id = role_permission_workflow.workflow_level_role_id
                user_role_permission_level.status = role_permission_workflow.status
                user_role_permission_level.from_date = datetime.strptime(request.POST['from_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_role_permission_level.to_date = datetime.strptime(request.POST['to_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_role_permission_level.save()

        response['flag'] = True
        response['message'] = "Record has been updated successfully."
        response['workflow'] = user_role_permission.workflow
            
        return JsonResponse(response)


@login_required
def resetCredential(request,user_id):
    if request.method == "POST":
        response = {}
        try:
            user = SpUsers.objects.get(id=request.POST['user_id'])
            password = make_password(request.POST['new_password'])
            user.password = password
            user.save()

            if user.id :

                #Save Activity
                user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'User credentials updated'
                activity    = 'User credentials updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('User Management', 'User Management', heading, activity, request.user.id, user_name, 'updateVehiclePass.png', '1', 'web.png')
                
                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,user.id,'firebase_token')
                employee_name = getUserName(user.id)

                message_title = "Password reset"
                message_body = "You password has been changed by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(user.id,'SpUsers','User Management','Password reset',message_title,message_body,notification_image,request.user.id,user_name,user.id,employee_name,'password.png',2,'app.png',1,1)
                #-----------------------------save notification block----------------------------#

                response['flag'] = True
                response['user_id'] = request.POST['user_id']
                response['message'] = "Record has been updated successfully."
            else:
                response['flag'] = False
                response['message'] = "Failed to save"
        except Exception as e:
            response['error'] = False
            response['message'] = str(e)
        return JsonResponse(response)
    else:
        context = {}
        context['user']     = SpUsers.objects.get(id=user_id)
        template = 'user-management/reset-user-credential.html'
        return render(request, template, context)

@login_required
def userTracking(request,user_id):

    if 'track_date' in request.GET and request.GET['track_date'] != "" :
            today                   = request.GET['track_date']
            today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()
    
    context = {}
    tracks = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today)
    if len(tracks):
        context['tracks'] = tracks
        context['first_track'] = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today).first
        context['last_track'] = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today).last

    context['user'] = SpUsers.objects.get(id=user_id)
    template = 'user-management/user-tracking.html'
    return render(request, template, context)

@login_required
def ajaxUserTracking(request,user_id):

    if 'track_date' in request.GET and request.GET['track_date'] != "" :
            today                   = request.GET['track_date']
            today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()
    
    context = {}
    tracks = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today)
    if len(tracks):
        context['tracks'] = tracks
        context['first_track'] = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today).first
        context['last_track'] = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today).last

    context['user'] = SpUsers.objects.get(id=user_id)
    template = 'user-management/ajax-user-tracking.html'
    return render(request, template, context)    



def locateUsersOnMap(request):
    context = {}
    users      = SpUsers.objects.filter(is_tagged=1,latitude__isnull=False,longitude__isnull=False).exclude(user_type=1)
    for user in users:
        last_order = SpOrders.objects.filter(user_id=user.id).values('created_at').first()
        shipping_address = SpAddresses.objects.filter(user_id=user.id,type="correspondence").values('address_line_1').first()
        user.address = shipping_address['address_line_1']

        if last_order:
            user.last_order_place_date = last_order['created_at']
        else:
            user.last_order_place_date = '-'
        
    context['users'] = users
    template = 'user-management/locate-user-on-map.html'

    return render(request, template, context)
    

@login_required
def userTrackingReport(request):
    context = {}
    context['users'] = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
    context['page_title'] = "Users Tracking Report"
    context['org_latitude'] = getConfigurationResult('org_latitude')
    context['org_longitude'] = getConfigurationResult('org_longitude')
    template = 'user-management/user-tracking-report.html'
    return render(request, template, context)

@login_required
def userTravelSummary(request):
    today       = date.today()
    users   = SpUserTracking.objects.filter().values('user_id').distinct().values('user_id', 'distance_travelled')
    for user in users:
        user['name'] = getUserName(user['user_id'])

    user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).values('user_id').distinct().values('user_id', 'travel_charges')
    for user_tracking in user_tracking_details:
        distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today.strftime("%Y-%m-%d")).aggregate(Sum('distance_travelled'))
        user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
        if user_tracking['distance_travelled'] > 0:
            user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
            user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
        else:
            user_tracking['charges']            = 0
            user_tracking['total_charges']      = 0
        user_tracking['user_name']          = getUserName(user_tracking['user_id'])


    context = {}
    context['today_date']               = today.strftime("%d/%m/%Y")
    context['users']                    = users
    context['user_tracking_details']    = user_tracking_details
    context['month_date']               = date.today().strftime("%m/%Y")
    context['page_title']               = "User Travel Summary"
    template = 'user-management/user-travel-summary.html'
    return render(request, template, context)

@login_required
def ajaxuserTravelSummary(request):
    today                   = request.GET['travel_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    context = {}
    
    if request.GET['time_period'] == '1':
        user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today)
        if request.GET['user_id']:
            user_tracking_details = user_tracking_details.filter(user_id=request.GET['user_id'])
        user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
        for user_tracking in user_tracking_details:
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today).aggregate(Sum('distance_travelled'))
            user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            if user_tracking['distance_travelled'] > 0:
                user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
                user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
            else:
                user_tracking['charges']            = 0
                user_tracking['total_charges']      = 0
            user_tracking['user_name']          = getUserName(user_tracking['user_id'])
        context['user_tracking_details']    = user_tracking_details       
    else:
        user_tracking_detail = []
        if request.GET['user_id']:
            travel_month = request.GET['travel_month_picker']
            travel_month = travel_month.split('/')
            year  = int(travel_month[1])
            month = int(travel_month[0])
            
            last_day_of_month = calendar.monthrange(year,month)[1]
            if int(month) < 10:
                month = '0'+str(month)
                 
            for x in range(last_day_of_month):
                x = x+1
                if x < 10:
                    x = '0'+str(x)
                travel_date  = str(year)+'-'+str(month)+'-'+str(x)
                travel_dates = str(x)+'/'+str(month)+'/'+str(year)
    
                user_trackings = {}
                
                try:
                    distance_travelled  = SpUserTracking.objects.filter(user_id=request.GET['user_id'], created_at__icontains=travel_date).aggregate(Sum('distance_travelled'))
                except SpUserTracking.DoesNotExist:
                    distance_travelled = None
    
                    
                user_trackings['travel_date']          = travel_dates
                if distance_travelled['distance_travelled__sum']:
                    user_trackings['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
                else:
                    user_trackings['distance_travelled'] = 0.0
                if user_trackings['distance_travelled'] > 0:
                    travel_charges       = SpUserTracking.objects.filter(user_id=request.GET['user_id'], created_at__icontains=travel_date).values('travel_charges').first()        
                    user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
                    user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
                else:
                    user_trackings['charges']            = 0
                    user_trackings['total_charges']      = 0
                user_trackings['user_name']          = getUserName(request.GET['user_id'])
                user_tracking_detail.append(user_trackings)  
        context['user_tracking_details']    = user_tracking_detail
        
    if request.GET['travel_month_picker']:
        context['month_date']               = request.GET['travel_month_picker']
    else:
        context['month_date']               = date.today().strftime("%m/%Y")
    context['time_period']              = request.GET['time_period']    
    template = 'user-management/ajax-user-travel-summary-report.html'
    return render(request, template, context)

#get export user summary
@login_required
def exportUserTravelSummary(request, travel_date, user_id, travel_month_picker, time_period):
    today                   = travel_date
          
    if time_period == '1':
        user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today)
        if user_id!='0':
            user_tracking_details = user_tracking_details.filter(user_id=user_id)
        user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
        for user_tracking in user_tracking_details:
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today).aggregate(Sum('distance_travelled'))
            user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            if user_tracking['distance_travelled'] > 0:
                user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
                user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
            else:
                user_tracking['charges']            = 0
                user_tracking['total_charges']      = 0   
            user_tracking['user_name']          = getUserName(user_tracking['user_id'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=user_daily_travel_summary.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='left')
        thin = Side(border_style="thin", color="303030") 
        black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wrapped_alignment = Alignment(
            vertical='top',
            horizontal='left',
            wrap_text=True
        )

        header_alignment = Alignment(
            vertical='top',
            horizontal='center',
            wrap_text=True
        )
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'User Daily Travel Summary'
        worksheet.merge_cells('A1:A1') 
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = 9
        worksheet.page_setup.fitToPage = True
        
        worksheet = workbook.worksheets[0]
        img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
        img.height = 50
        img.alignment = 'center'
        img.anchor = 'A1'
        worksheet.add_image(img)
        
        column_length = 4
        
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
        worksheet.cell(row=1, column=2).value = 'User Daily Travel Summary as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = header_alignment
        worksheet.cell(row=1, column=column_length).border = black_border
        worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
        worksheet.cell(row=1, column=2).fill = PatternFill()

        # Define the titles for columns
        # columns = []
        row_num = 1
        worksheet.row_dimensions[1].height = 40
        
        # Define the titles for columns
        columns = []

        columns += [ 'Employee Name' ]
        columns += [ 'Distance Travelled in Kilometer' ]
        columns += [ 'Charges' ]
        columns += [ 'Total Charges' ]

        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32

        for user_tracking in user_tracking_details:
            row_num += 1
            # Define the data for each cell in the row 
            row = []
            row += [ user_tracking['user_name'] ]
            row += [ user_tracking['distance_travelled'] ]
            row += [ user_tracking['charges'] ]
            row += [ user_tracking['total_charges'] ]           
        
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border  

        wrapped_alignment = Alignment(
            horizontal='center',
            wrap_text=True
        )

        row_num += 1
        last_row = row_num
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
        worksheet.row_dimensions[last_row].height = 20
        worksheet.cell(row=last_row, column=1).value = 'Generated By SAAHAJ MILK PRODUCER COMPANY LIMITED'
        worksheet.cell(row=last_row, column=1).font = header_font
        worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
        worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
        worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

        workbook.save(response)
        return response

    else:
        travel_month = travel_month_picker
        travel_month = travel_month.split('-')
        year  = int(travel_month[1])
        month = int(travel_month[0])
        
        last_day_of_month = calendar.monthrange(year,month)[1]
        if int(month) < 10:
            month = '0'+str(month)
        user_tracking_detail = []     
        for x in range(last_day_of_month):
            x = x+1
            if x < 10:
                x = '0'+str(x)
            travel_date  = str(year)+'-'+str(month)+'-'+str(x)
            travel_dates = str(x)+'/'+str(month)+'/'+str(year)

            user_trackings = {}
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_id, created_at__icontains=travel_date).aggregate(Sum('distance_travelled'))
            user_trackings['travel_date']          = travel_dates
            if distance_travelled['distance_travelled__sum']:
                user_trackings['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            else:
                user_trackings['distance_travelled'] = 0    
            if user_trackings['distance_travelled'] > 0:
                travel_charges       = SpUserTracking.objects.filter(user_id=user_id, created_at__icontains=travel_date).values('travel_charges').first()        
                user_trackings['charges']            = float(travel_charges['travel_charges'])
                user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
            else:
                user_trackings['charges']            = 0
                user_trackings['total_charges']      = 0
            user_trackings['user_name']          = getUserName(user_id)
            user_tracking_detail.append(user_trackings)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=user_monthly_travel_summary.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='left')
        thin = Side(border_style="thin", color="303030") 
        black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wrapped_alignment = Alignment(
            vertical='top',
            horizontal='left',
            wrap_text=True
        )

        header_alignment = Alignment(
            vertical='top',
            horizontal='center',
            wrap_text=True
        )
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'User Monthly Travel Summary'
        worksheet.merge_cells('A1:A1') 
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = 9
        worksheet.page_setup.fitToPage = True
        
        worksheet = workbook.worksheets[0]
        img = openpyxl.drawing.image.Image('static/img/sahaaj.png')
        img.height = 50
        img.alignment = 'center'
        img.anchor = 'A1'
        worksheet.add_image(img)
        
        column_length = 4
        month_name = datetime(int(travel_month[1]),int(travel_month[0]),1).strftime( '%B' )
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
        worksheet.cell(row=1, column=2).value = ''+getUserName(user_id)+' Monthly Travel Summary as on '+str(month_name)+'-'+str(int(travel_month[1]))+''
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = header_alignment
        worksheet.cell(row=1, column=column_length).border = black_border
        worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
        worksheet.cell(row=1, column=2).fill = PatternFill()

        # Define the titles for columns
        # columns = []
        row_num = 1
        worksheet.row_dimensions[1].height = 40
        
        # Define the titles for columns
        columns = []

        columns += [ 'Date' ]
        columns += [ 'Distance Travelled in Kilometer' ]
        columns += [ 'Charges' ]
        columns += [ 'Total Charges' ]

        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32

        for user_tracking in user_tracking_detail:
            row_num += 1
            # Define the data for each cell in the row 
            row = []
            row += [ user_tracking['travel_date'] ]
            row += [ user_tracking['distance_travelled'] ]
            row += [ user_tracking['charges'] ]
            row += [ user_tracking['total_charges'] ]           
        
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border  

        wrapped_alignment = Alignment(
            horizontal='center',
            wrap_text=True
        )

        row_num += 1
        last_row = row_num
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
        worksheet.row_dimensions[last_row].height = 20
        worksheet.cell(row=last_row, column=1).value = 'Generated By SAAHAJ MILK PRODUCER COMPANY LIMITED'
        worksheet.cell(row=last_row, column=1).font = header_font
        worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
        worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
        worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

        workbook.save(response)
        return response 

def makeIncentiveHistory(user_id):
    incentiveID = SpIncentive.objects.filter(user_id=user_id).first()
    incentiveHistory = SpIncentiveHistory()
    incentiveHistory.user_id = incentiveID.user_id
    incentiveHistory.payment_cycle = incentiveID.payment_cycle
    incentiveHistory.status = incentiveID.status
    incentiveHistory.created_by = incentiveID.created_by
    incentiveHistory.created_at = incentiveID.created_at
    incentiveHistory.updated_at = incentiveID.updated_at
    incentiveHistory.save()
    
    incentiveDetails = SpIncentiveDetails.objects.filter(incentive_id=incentiveID.id).all()
    for details in incentiveDetails:
        incentiveDetailsHistory = SpIncentiveDetailsHistory()
        incentiveDetailsHistory.incentive_id = details.incentive_id
        incentiveDetailsHistory.incentive_type = details.incentive_type
        incentiveDetailsHistory.class_product_variant_id = details.class_product_variant_id
        incentiveDetailsHistory.is_slab = details.is_slab
        incentiveDetailsHistory.ss_incentive = details.ss_incentive
        incentiveDetailsHistory.distributor_incentive = details.distributor_incentive
        incentiveDetailsHistory.created_at = details.created_at
        incentiveDetailsHistory.updated_at = details.updated_at
        incentiveDetailsHistory.save()
    
    incentiveSlabDetails = SpIncentiveSlabDetails.objects.filter(incentive_id=incentiveID.id).all()
    for slabDetails in incentiveSlabDetails:
        incentiveSlabDetailsHistory = SpIncentiveSlabDetailsHistory()
        incentiveSlabDetailsHistory.incentive_id = slabDetails.incentive_id
        incentiveSlabDetailsHistory.incentive_detail_id = slabDetails.incentive_detail_id
        incentiveSlabDetailsHistory.slab_id = slabDetails.slab_id
        incentiveSlabDetailsHistory.ss_incentive = slabDetails.ss_incentive
        incentiveSlabDetailsHistory.distributor_incentive = slabDetails.distributor_incentive
        incentiveSlabDetailsHistory.created_at = slabDetails.created_at
        incentiveSlabDetailsHistory.updated_at = slabDetails.updated_at
        incentiveSlabDetailsHistory.save()


@login_required          
def createIncentive1(request):
    productClassList = SpProductClass.objects.filter(status=1)
    for productClass in productClassList:
        productList = SpProducts.objects.filter(status=1,product_class_id=productClass.id)
        for product in productList:
            productVariantList = SpProductVariants.objects.filter(status=1,product_id=product.id)
            product.variant=productVariantList
        productClass.product=productList
    
    if SpIncentive.objects.filter(user_id=request.POST['user_id']).exists():
        makeIncentiveHistory(request.POST['user_id'])
        incentiveID = SpIncentive.objects.filter(user_id=request.POST['user_id']).values_list('id', flat=True)
        incentiveDetailsID = SpIncentiveDetails.objects.filter(incentive_id__in=incentiveID).values_list('id', flat=True)
        for idd in incentiveDetailsID:
            SpIncentiveDetails.objects.filter(id=idd).delete()
        SpIncentiveSlabDetails.objects.filter(incentive_detail_id__in=incentiveDetailsID).delete()
        incentive = SpIncentive.objects.get(user_id=request.POST['user_id'])
    else:
        incentive = SpIncentive()
    incentive.user_id = request.POST['user_id']
    incentive.payment_cycle = float(request.POST['payment_cycle'])
    incentive.created_by = request.user.id
    incentive.save()
    incentive_id = incentive.id
    for productClass in productClassList:
        product_class_id = productClass.id
        incentive_type = request.POST['level_flag_'+str(product_class_id)]
        
        slab_flag = request.POST['slab_flag_'+str(product_class_id)]
        if (incentive_type == "1" or slab_flag == "1"):
            incentive_details = SpIncentiveDetails()
            incentive_details.incentive_id = incentive_id
            incentive_details.incentive_type = 1
            incentive_details.class_product_variant_id = product_class_id
            incentive_details.is_slab = slab_flag
            incentive_details.ss_incentive = 0 if (request.POST['ss_incentive_'+str(product_class_id)] == "" ) else request.POST['ss_incentive_'+str(product_class_id)]
            incentive_details.distributor_incentive = 0 if (request.POST['distributor_incentive_'+str(product_class_id)] == "" ) else request.POST['distributor_incentive_'+str(product_class_id)]
            incentive_details.save()
            if slab_flag == "1":
                slab_id = request.POST.getlist('slab_id_productClass_'+str(product_class_id)+'[]')
                ss_incentive = request.POST.getlist('ss_incentive_productClass_'+str(product_class_id)+'[]')
                for i, more_than in enumerate(slab_id):
                    if (ss_incentive[i] != ""):
                        if ((slab_id[i]) != "" ) and ((ss_incentive[i]) != "") and (int(slab_id[i]) > 0 ) and (float(ss_incentive[i]) > 0):
                            incentive_slab_details = SpIncentiveSlabDetails()
                            incentive_slab_details.incentive_id = incentive_id
                            incentive_slab_details.incentive_detail_id = incentive_details.id
                            incentive_slab_details.distributor_incentive = 0 
                            incentive_slab_details.slab_id = 0 if (slab_id[i] == "" ) else slab_id[i]
                            incentive_slab_details.ss_incentive = 0 if (ss_incentive[i] == "" ) else ss_incentive[i]
                            incentive_slab_details.save()
            
                
        if (incentive_type != "1"):
            for product in productClass.product:
                product_id = product.id
                if (incentive_type == '2'):
                    incentive_details = SpIncentiveDetails()
                    incentive_details.incentive_id = incentive_id
                    incentive_details.incentive_type = incentive_type
                    slab_flag = request.POST['slab_flag_'+str(product_class_id)+'_'+str(product_id)]
                    incentive_details.class_product_variant_id = product_id
                    incentive_details.ss_incentive = 0 if (request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)] == "" ) else request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)]
                    incentive_details.distributor_incentive = 0 if (request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)] == "" ) else request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)]
                    incentive_details.save()
                
                elif (incentive_type == '3'):
                    for variant in product.variant:
                        variant_id = variant.id
                        incentive_details = SpIncentiveDetails()
                        incentive_details.incentive_id = incentive_id
                        incentive_details.incentive_type = incentive_type
                        incentive_details.class_product_variant_id = variant_id
                        slab_flag = request.POST['slab_flag_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)]
                        incentive_details.ss_incentive = 0 if (request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)] == "" ) else request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)]
                        incentive_details.distributor_incentive = 0 if (request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)] == "" ) else request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)]
                        incentive_details.save()
        
    message ="Incentive details has been updated successfully."
    response = {}
    response['error'] = False
    response['message'] = message
    # messages.success(request, message, extra_tags='success')
    return JsonResponse(response)


@login_required          
def createIncentive(request):
    productClassList = SpProductClass.objects.filter(status=1)
    for productClass in productClassList:
        productList = SpProducts.objects.filter(status=1,product_class_id=productClass.id)
        for product in productList:
            productVariantList = SpProductVariants.objects.filter(status=1,product_id=product.id)
            product.variant=productVariantList
        productClass.product=productList
    
    if SpIncentive.objects.filter(user_id=request.POST['user_id']).exists():
        makeIncentiveHistory(request.POST['user_id'])
        incentiveID = SpIncentive.objects.filter(user_id=request.POST['user_id']).values_list('id', flat=True)
        incentiveDetailsID = SpIncentiveDetails.objects.filter(incentive_id__in=incentiveID).values_list('id', flat=True)
        for idd in incentiveDetailsID:
            SpIncentiveDetails.objects.filter(id=idd).delete()
        SpIncentiveSlabDetails.objects.filter(incentive_detail_id__in=incentiveDetailsID).delete()
        incentive = SpIncentive.objects.get(user_id=request.POST['user_id'])
    else:
        incentive = SpIncentive()
    incentive.user_id = request.POST['user_id']
    incentive.payment_cycle = float(request.POST['payment_cycle'])
    incentive.created_by = request.user.id
    incentive.save()
    incentive_id = incentive.id
    for productClass in productClassList:
        product_class_id = productClass.id
        incentive_type = request.POST['level_flag_'+str(product_class_id)]
        
        slab_flag = request.POST['slab_flag_'+str(product_class_id)]
        if (incentive_type == "1" or slab_flag == "1"):
            incentive_details = SpIncentiveDetails()
            incentive_details.incentive_id = incentive_id
            incentive_details.incentive_type = 1
            incentive_details.class_product_variant_id = product_class_id
            incentive_details.is_slab = slab_flag
            incentive_details.ss_incentive = 0 if (request.POST['ss_incentive_'+str(product_class_id)] == "" ) else request.POST['ss_incentive_'+str(product_class_id)]
            incentive_details.distributor_incentive = 0 if (request.POST['distributor_incentive_'+str(product_class_id)] == "" ) else request.POST['distributor_incentive_'+str(product_class_id)]
            incentive_details.save()
            if slab_flag == "1":
                slab_id = request.POST.getlist('slab_id_productClass_'+str(product_class_id)+'[]')
                ss_incentive = request.POST.getlist('ss_incentive_productClass_'+str(product_class_id)+'[]')
                for i, more_than in enumerate(slab_id):
                    if (ss_incentive[i] != ""):
                        if ((slab_id[i]) != "" ) and ((ss_incentive[i]) != "") and (int(slab_id[i]) > 0 ) and (float(ss_incentive[i]) > 0):
                            incentive_slab_details = SpIncentiveSlabDetails()
                            incentive_slab_details.incentive_id = incentive_id
                            incentive_slab_details.incentive_detail_id = incentive_details.id
                            incentive_slab_details.distributor_incentive = 0 
                            incentive_slab_details.slab_id = 0 if (slab_id[i] == "" ) else slab_id[i]
                            incentive_slab_details.ss_incentive = 0 if (ss_incentive[i] == "" ) else ss_incentive[i]
                            incentive_slab_details.save()
            
                
        for product in productClass.product:
            product_id = product.id
            # if (incentive_type == '2'):
            if (request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)] != "") or (request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)] != ""):
                incentive_details = SpIncentiveDetails()
                incentive_details.incentive_id = incentive_id
                incentive_details.incentive_type = 2
                slab_flag = request.POST['slab_flag_'+str(product_class_id)+'_'+str(product_id)]
                incentive_details.class_product_variant_id = product_id
                incentive_details.ss_incentive = 0 if (request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)] == "" ) else request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)]
                incentive_details.distributor_incentive = 0 if (request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)] == "" ) else request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)]
                incentive_details.save()
                
                # elif (incentive_type == '3'):
            for variant in product.variant:
                variant_id = variant.id
                if request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)] != "" or request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)] != "":
                    incentive_details = SpIncentiveDetails()
                    incentive_details.incentive_id = incentive_id
                    incentive_details.incentive_type = 3
                    incentive_details.class_product_variant_id = variant_id
                    slab_flag = request.POST['slab_flag_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)]
                    incentive_details.ss_incentive = 0 if (request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)] == "" ) else request.POST['ss_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)]
                    incentive_details.distributor_incentive = 0 if (request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)] == "" ) else request.POST['distributor_incentive_'+str(product_class_id)+'_'+str(product_id)+'_'+str(variant_id)]
                    incentive_details.save()
        
    message ="Incentive details has been updated successfully."
    response = {}
    response['error'] = False
    response['message'] = message
    # messages.success(request, message, extra_tags='success')
    return JsonResponse(response)



@login_required
def generateIncentive(request):
    values = []
    users_list = SpUsers.objects.raw(''' Select sp_users.*, sp_incentive.id as incentive_id, sp_incentive.payment_cycle from sp_users left join sp_incentive on sp_users.id=sp_incentive.user_id where is_super_stockist=1''')
    for user_list in users_list:
        if user_list.payment_cycle is not None:
            
            start_date = ""
            end_date = ""
            today = len(days_in_months(int(datetime.now().strftime('%Y')), int(datetime.now().strftime('%m'))))
            if user_list.payment_cycle == 1:
                if int(datetime.now().strftime('%m')) == 2:
                    mid_date=14
                else:
                    mid_date=15
                # mid_date=16
                if int(datetime.now().strftime('%d')) == mid_date:
                    start_date = str(datetime.now().strftime('%Y-%m-'))+"01"
                    end_date = str(datetime.now().strftime('%Y-%m-'))+str(mid_date)
                    
                elif int(datetime.now().strftime('%d')) == today:
                    start_date = str(datetime.now().strftime('%Y-%m-'))+str(mid_date+1)
                    end_date = str(datetime.now().strftime('%Y-%m-'))+str(today)
                    
                
            elif user_list.payment_cycle == 2:
                today = 15
                if int(datetime.now().strftime('%d')) == today:
                    start_date = str(datetime.now().strftime('%Y-%m-'))+"01"
                    end_date = str(datetime.now().strftime('%Y-%m-'))+str(today)
                    # start_date = "2021-07-01"
                    # end_date = "2021-07-31"
                
            if start_date != "" and end_date != "":
                    
                ss_liter_kg_incentive = 0
                dist_liter_kg_incentive = 0
                primary_transpoter_crate_incentive = 0
                secondary_transpoter_crate_incentive = 0
                net_slab_amount = 0
                total_net_amount = 0
                user_slab_list = []
                ss_crate_incentive = getModelColumnByColumnId(SpBasicDetails,'user_id',user_list.id,'per_crate_incentive')
                if ss_crate_incentive is None:
                    ss_crate_incentive = 0
                    
                incentive_ids = SpIncentive.objects.filter(user_id=user_list.id).values_list('id',flat=True)
                                    
                # orders_id = SpOrders.objects.filter(order_date__gte=start_date,order_date__lte=end_date,reporting_to_user_id=user_list.id).values_list('id',flat=True)
                orders = SpOrders.objects.filter(order_date__gte=start_date,order_date__lte=end_date,reporting_to_user_id=user_list.id).all()
                if orders:
                    for order in orders:
                        num_of_crate = 0 
                        # order_details = SpOrderDetails.objects.filter(order_id__in=orders_id).all()
                        order_details = SpOrderDetails.objects.filter(order_id=order.id).all()
                        if order_details:
                            for order_detail in order_details:
                                if order_detail.product_container_type == 'Crate':
                                    if order_detail.packaging_type == "0":
                                        num_of_crate +=  order_detail.quantity
                                    elif order_detail.packaging_type == "1":
                                        num_of_crate +=(order_detail.quantity/order_detail.product_no_of_pouch)
                                product_class_id = getModelColumnById(SpProducts,order_detail.product_id,'product_class_id')
                                incentive_details = SpIncentiveDetails.objects.filter((Q(incentive_type=1,class_product_variant_id=product_class_id))|(Q(incentive_type=2,class_product_variant_id=order_detail.product_id))|(Q(incentive_type=3,class_product_variant_id=order_detail.product_variant_id)),incentive_id__in=incentive_ids).all()
                                # incentive_detail_class = SpIncentiveDetails.objects.filter(incentive_type=1,class_product_variant_id=product_class_id).all()
                                # if SpIncentiveDetails.objects.filter(incentive_type=1,class_product_variant_id=product_class_id).exists():
                                #     incentive_details = SpIncentiveDetails.objects.filter(incentive_type=1,class_product_variant_id=product_class_id).first()
                                # else:
                                #     # incentive_detail_product = SpIncentiveDetails.objects.filter(incentive_type=2,class_product_variant_id=order_detail.product_id).all()
                                #     if SpIncentiveDetails.objects.filter(incentive_type=2,class_product_variant_id=order_detail.product_id).exists():
                                #         incentive_details = SpIncentiveDetails.objects.filter(incentive_type=2,class_product_variant_id=order_detail.product_id).first()
                                #     else:
                                #         # incentive_detail_variant = SpIncentiveDetails.objects.filter(incentive_type=3,class_product_variant_id=order_detail.product_variant_id).all()
                                #         if SpIncentiveDetails.objects.filter(incentive_type=3,class_product_variant_id=order_detail.product_variant_id).exists():
                                #             incentive_details = SpIncentiveDetails.objects.filter(incentive_type=3,class_product_variant_id=order_detail.product_variant_id).first()
                                            
                                # incentive_details = SpIncentiveDetails.objects.filter(incentive_type=1,class_product_variant_id=product_class_id).all()        
                                # if incentive_details:
                                #     for incentive_detail in incentive_details:
                                #         ss_liter = (order_detail.quantity_in_ltr * incentive_detail.ss_incentive)
                                #         ss_liter_kg_incentive += ss_liter
                                #         dist_liter = (order_detail.quantity_in_ltr * incentive_detail.distributor_incentive)
                                #         dist_liter_kg_incentive += dist_liter
                                        
                                #         if incentive_detail.is_slab == 1:
                                #             incentiveSlabDetails = SpIncentiveSlabDetails.objects.filter(incentive_detail_id=incentive_detail.id).all()
                                #             if incentiveSlabDetails:
                                #                 for incentiveSlabDetail in incentiveSlabDetails:
                                #                     more_than_quantity = getModelColumnById(SpSlabMasterList, incentiveSlabDetail.slab_id,'more_than_quantity')
                                #                     upto_quantity = getModelColumnById(SpSlabMasterList, incentiveSlabDetail.slab_id,'upto_quantity')
                                #                     if order_detail.quantity_in_ltr > more_than_quantity and order_detail.quantity_in_ltr <= upto_quantity:  
                                #                         slab_amount = (order_detail.quantity_in_ltr * incentiveSlabDetail.ss_incentive)
                                #                         net_slab_amount += slab_amount
                                #                         user_slab_data = {}
                                #                         user_slab_data[incentiveSlabDetail.slab_id] = slab_amount
                                #                         user_slab_list.append(user_slab_data)
                                                        
                                # incentive_details = SpIncentiveDetails.objects.filter(incentive_id__in=incentive_ids,incentive_type=2,class_product_variant_id=order_detail.product_id).all()        
                                
                                if incentive_details:
                                    for incentive_detail in incentive_details:
                                        ss_liter = (order_detail.quantity_in_ltr * incentive_detail.ss_incentive)
                                        ss_liter_kg_incentive += ss_liter
                                        dist_liter = (order_detail.quantity_in_ltr * incentive_detail.distributor_incentive)
                                        dist_liter_kg_incentive += dist_liter
                                        if incentive_detail.is_slab == 1:
                                            
                                            user_slab_data = {}
                                            user_slab_data[product_class_id] = order_detail.quantity_in_ltr
                                            user_slab_list.append(user_slab_data)
                                                            
                                                    
                                dist_crate_incentive = getModelColumnByColumnId(SpBasicDetails,'user_id',order.user_id,'per_crate_incentive')
                                if dist_crate_incentive is None:
                                    dist_crate_incentive = 0
                                secondary_transpoter_crate_incentive += (num_of_crate * dist_crate_incentive)
                
                slab_total_values = {}
                slab_total_value_total = 0
                if len(user_slab_list) > 0:
                    for d in user_slab_list:
                        for k in d.keys():
                            slab_total_values[k] = (slab_total_values.get(k, 0) + d[k])
                    slab_total_value_total = slab_total_values.values()
                    slab_total_value_total = sum(slab_total_value_total)
                
                slab = []
                for key, value in slab_total_values.items():
                    slab_amount = 0
                    slab_id = 0
                    incentive_detail_ids = SpIncentiveDetails.objects.filter(incentive_type=1,class_product_variant_id=key,incentive_id__in=incentive_ids,is_slab=1).values_list('id', flat=True)
                    incentiveSlabDetails = SpIncentiveSlabDetails.objects.filter(incentive_detail_id__in=incentive_detail_ids).all()
                    if incentiveSlabDetails:
                        for incentiveSlabDetail in incentiveSlabDetails:
                            more_than_quantity = getModelColumnById(SpSlabMasterList, incentiveSlabDetail.slab_id,'more_than_quantity')
                            upto_quantity = getModelColumnById(SpSlabMasterList, incentiveSlabDetail.slab_id,'upto_quantity')
                            # slab.append(more_than_quantity)
                            # slab.append(upto_quantity)
                            if value > more_than_quantity or value >= upto_quantity:  
                                slab_amount = (value * incentiveSlabDetail.ss_incentive)
                                slab_id = incentiveSlabDetail.slab_id
                        net_slab_amount += slab_amount
                        user_slab_data = {}
                        user_slab_data['key'] = slab_id
                        user_slab_data['value'] = slab_amount
                        slab.append(user_slab_data)
                                            
                primary_transpoter_crate_incentive += (num_of_crate * ss_crate_incentive)
                total_net_amount += (net_slab_amount + ss_liter_kg_incentive + dist_liter_kg_incentive + primary_transpoter_crate_incentive + secondary_transpoter_crate_incentive)
                                          
                UserIncentive = SpUserIncentive()
                UserIncentive.user_id = user_list.id
                UserIncentive.ss_incentive = ss_liter_kg_incentive
                UserIncentive.distributor_incentive = dist_liter_kg_incentive
                UserIncentive.primary_transporter_amount = primary_transpoter_crate_incentive
                UserIncentive.secondary_transporter_amount = secondary_transpoter_crate_incentive
                UserIncentive.net_amount = total_net_amount
                UserIncentive.payment_cycle = user_list.payment_cycle
                UserIncentive.save()
                 
                    
                for data in slab:
                    values.append(str(data['key'])+" "+str(data['value']))
                    UserIncentiveDetails = SpUserIncentiveDetails()
                    UserIncentiveDetails.user_incentive_id = UserIncentive.id
                    UserIncentiveDetails.master_slab_id = data['key']
                    UserIncentiveDetails.slab_amount = data['value']
                    UserIncentiveDetails.save()
                     
                print(slab_total_values)                                                   
                values22 = str(num_of_crate)+" "+str(ss_liter_kg_incentive)+" "+str(dist_liter_kg_incentive)+" "+str(primary_transpoter_crate_incentive)+" "+str(secondary_transpoter_crate_incentive)+" "+str(total_net_amount)
                values.append(values22)
    message ="Cron incentive updated successfully."
    response = {}
    response['error'] = False
    response['message'] = message
    response['values'] = values
    messages.success(request, message, extra_tags='success')
    return JsonResponse(response)
        

            



